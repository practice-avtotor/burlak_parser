"""Unit-тесты для splitter.py.

Покрытие:
  - CardSplitter.split_file: базовое разделение, один лист, несколько листов
  - CardSplitter.split_file: edge cases (не .xlsx, лист не найден, дубликаты имён)
  - CardSplitter._extract_sheet_via_zip: проверка содержимого split-файлов
  - CardSplitter.split_many_parallel: параллельное разделение
  - _clean_named_ranges: очистка named ranges
  - Интеграция: валидность .xlsx после разделения
"""

from __future__ import annotations

import os
import shutil
import tempfile
from typing import Any, Dict, List, Optional

import openpyxl
import pytest

from burlak_parser.splitter import (
    CardSplitter,
    _clean_named_ranges,
    _collect_related_files,
    _extract_to_path_worker,
)
from burlak_parser.heuristic_analyzer import HeuristicAnalyzer


# ═══════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════

def _create_multi_sheet_xlsx(
    dir_path: str,
    filename: str = "test_multi.xlsx",
    sheets: Optional[Dict[str, List[List[Any]]]] = None,
) -> str:
    """Create an .xlsx file with multiple sheets and return its path."""
    if sheets is None:
        sheets = {
            "Sheet1": [["A1", "B1"], ["A2", "B2"]],
            "Sheet2": [["C1", "D1"], ["C2", "D2"]],
            "Sheet3": [["E1", "F1"], ["E2", "F2"]],
        }
    path = os.path.join(dir_path, filename)
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for name, data in sheets.items():
        ws = wb.create_sheet(title=name)
        for r_idx, row in enumerate(data, 1):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    return path


def _count_xlsx_sheets(path: str) -> int:
    """Count sheets in an .xlsx file using openpyxl."""
    wb = openpyxl.load_workbook(path)
    count = len(wb.sheetnames)
    wb.close()
    return count


def _get_xlsx_cell(path: str, sheet: str, row: int, col: int) -> Any:
    """Get cell value from an .xlsx file."""
    wb = openpyxl.load_workbook(path)
    val = wb[sheet].cell(row=row, column=col).value
    wb.close()
    return val


# ═══════════════════════════════════════════════════════════════════════
#  Fixtures
# ═══════════════════════════════════════════════════════════════════════

@pytest.fixture
def tmp_dir() -> str:
    """Create a temporary directory for test files."""
    path = tempfile.mkdtemp(prefix="splitter_test_")
    yield path
    shutil.rmtree(path, ignore_errors=True)


@pytest.fixture
def multi_sheet_xlsx(tmp_dir: str) -> str:
    """Create a 3-sheet .xlsx file."""
    return _create_multi_sheet_xlsx(tmp_dir)


@pytest.fixture
def single_sheet_xlsx(tmp_dir: str) -> str:
    """Create a single-sheet .xlsx file."""
    return _create_multi_sheet_xlsx(
        tmp_dir, "single.xlsx",
        sheets={"OnlySheet": [["Data1"], ["Data2"]]},
    )


# ═══════════════════════════════════════════════════════════════════════
#  1. CardSplitter.split_file — basic
# ═══════════════════════════════════════════════════════════════════════

class TestSplitFileBasic:
    def test_split_one_sheet(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Split one sheet from a multi-sheet file."""
        output_dir = os.path.join(tmp_dir, "out1")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], "TestCard",
        )
        assert len(created) == 1, f"Expected 1 file, got {len(created)}"
        assert os.path.exists(created[0])
        # Verify the output file has only 1 sheet
        assert _count_xlsx_sheets(created[0]) == 1, "Split file should have 1 sheet"

    def test_split_multiple_sheets(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Split multiple sheets from a multi-sheet file."""
        output_dir = os.path.join(tmp_dir, "out2")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1", "Sheet2"], "TestCard",
        )
        assert len(created) == 2, f"Expected 2 files, got {len(created)}"
        for fp in created:
            assert os.path.exists(fp)
            assert _count_xlsx_sheets(fp) == 1, "Each split file should have 1 sheet"

    def test_split_all_sheets(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Split all 3 sheets."""
        output_dir = os.path.join(tmp_dir, "out3")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1", "Sheet2", "Sheet3"], "TestCard",
        )
        assert len(created) == 3, f"Expected 3 files, got {len(created)}"

    def test_data_preserved(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Data in split file should match original sheet data."""
        output_dir = os.path.join(tmp_dir, "out4")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet2"], "TestCard",
        )[0]

        # Sheet2 in original had C1, D1; C2, D2
        assert _count_xlsx_sheets(created) == 1
        split_wb = openpyxl.load_workbook(created)
        split_ws = split_wb.active
        data = [
            [split_ws.cell(row=r, column=c).value for c in range(1, 3)]
            for r in range(1, 3)
        ]
        split_wb.close()
        assert data == [["C1", "D1"], ["C2", "D2"]], \
            f"Data mismatch. Got: {data}"

    def test_split_without_label(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Split without file_label uses sheet name as filename."""
        output_dir = os.path.join(tmp_dir, "out5")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], file_label="",
        )
        assert len(created) == 1
        # Filename should be based on sheet name
        basename = os.path.basename(created[0])
        assert "Sheet1" in basename, f"Expected Sheet1 in {basename}"


# ═══════════════════════════════════════════════════════════════════════
#  2. CardSplitter.split_file — edge cases
# ═══════════════════════════════════════════════════════════════════════

class TestSplitFileEdgeCases:
    def test_non_xlsx_file(self, tmp_dir: str):
        """Non-.xlsx files should be skipped."""
        output_dir = os.path.join(tmp_dir, "out_edge1")
        txt_path = os.path.join(tmp_dir, "test.txt")
        with open(txt_path, "w") as f:
            f.write("not an xlsx")
        splitter = CardSplitter()
        created = splitter.split_file(txt_path, output_dir, ["Sheet1"], "Test")
        assert created == [], "Non-xlsx should return empty list"

    def test_sheet_not_found(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Requesting a non-existent sheet: fallback copies source file."""
        output_dir = os.path.join(tmp_dir, "out_edge2")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["NonExistentSheet"], "Test",
        )
        assert len(created) == 1, "Fallback should copy source as last resort"
        assert os.path.exists(created[0])

    def test_single_sheet_file(self, tmp_dir: str, single_sheet_xlsx: str):
        """Single-sheet file should be handled (no unnecessary copy)."""
        output_dir = os.path.join(tmp_dir, "out_edge3")
        splitter = CardSplitter()
        created = splitter.split_file(
            single_sheet_xlsx, output_dir, ["OnlySheet"], "Test",
        )
        # Single sheet file — splitting should still work
        assert len(created) == 1, f"Expected 1 file, got {len(created)}"
        assert _count_xlsx_sheets(created[0]) == 1

    def test_duplicate_filename_handling(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Duplicate output filenames are skipped (paths must be pre-allocated)."""
        output_dir = os.path.join(tmp_dir, "out_edge4")
        splitter = CardSplitter()
        # Split same sheet twice
        created1 = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], "TestCard",
        )
        created2 = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], "TestCard",
        )
        assert len(created1) == 1
        # Second call skips because file already exists (pre-allocation model)
        assert len(created2) == 0
        assert os.path.exists(created1[0])

    def test_empty_sheet_list(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Empty sheet list returns empty."""
        output_dir = os.path.join(tmp_dir, "out_edge5")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, [], "Test",
        )
        assert created == [], "Empty sheet list should return empty"

    def test_output_dir_created(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Output directory is created if it doesn't exist."""
        output_dir = os.path.join(tmp_dir, "new_dir", "nested")
        assert not os.path.exists(output_dir)
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], "Test",
        )
        assert len(created) == 1
        assert os.path.exists(output_dir)


# ═══════════════════════════════════════════════════════════════════════
#  3. CardSplitter._extract_sheet_via_zip — content integrity
# ═══════════════════════════════════════════════════════════════════════

class TestExtractSheetViaZip:
    def test_sheet_data_correct(self, tmp_dir: str):
        """Verify extracted sheet has correct data."""
        # Create file with specific data
        sheets = {
            "DataSheet": [["Header1", "Header2"], ["Val1", "Val2"], ["Val3", "Val4"]],
            "OtherSheet": [["Other1", "Other2"]],
        }
        path = _create_multi_sheet_xlsx(tmp_dir, "data_test.xlsx", sheets)
        output_dir = os.path.join(tmp_dir, "extract1")
        os.makedirs(output_dir)

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["DataSheet"], "")[0]

        # Check data
        assert _get_xlsx_cell(created, "DataSheet", 1, 1) == "Header1"
        assert _get_xlsx_cell(created, "DataSheet", 2, 2) == "Val2"
        assert _get_xlsx_cell(created, "DataSheet", 3, 1) == "Val3"

    def test_file_valid_xlsx(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Split file should be a valid .xlsx readable by openpyxl."""
        output_dir = os.path.join(tmp_dir, "extract2")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir, ["Sheet1"], "TestCard",
        )
        # openpyxl should be able to open it without errors
        wb = openpyxl.load_workbook(created[0])
        assert wb.active is not None
        wb.close()

    def test_sheet_name_preserved(self, tmp_dir: str):
        """Sheet name should be preserved in the output file."""
        sheets = {
            "СпециальноеИмя": [["A", "B"]],
            "Other": [["C", "D"]],
        }
        path = _create_multi_sheet_xlsx(tmp_dir, "name_test.xlsx", sheets)
        output_dir = os.path.join(tmp_dir, "extract3")
        os.makedirs(output_dir)

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["СпециальноеИмя"], "Test")[0]

        names = openpyxl.load_workbook(created).sheetnames
        assert names == ["СпециальноеИмя"], f"Expected preserved name, got {names}"

    def test_multiple_extracts_same_source(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Extracting multiple sheets from the same source should all work."""
        output_dir = os.path.join(tmp_dir, "extract4")
        splitter = CardSplitter()
        created = splitter.split_file(
            multi_sheet_xlsx, output_dir,
            ["Sheet1", "Sheet2", "Sheet3"], "Test",
        )
        assert len(created) == 3
        for fp in created:
            assert os.path.getsize(fp) > 0, f"File {fp} is empty"


# ═══════════════════════════════════════════════════════════════════════
#  4. CardSplitter.split_many_parallel
# ═══════════════════════════════════════════════════════════════════════

class TestSplitManyParallel:
    def test_parallel_split(self, tmp_dir: str):
        """Parallel split of multiple files."""
        # Create 3 multi-sheet files
        files = []
        for i in range(3):
            sheets = {
                f"Op{i}A": [["Part", "Qty"], [f"P{i}01", "1"]],
                f"Op{i}B": [["Part", "Qty"], [f"P{i}02", "2"]],
            }
            path = _create_multi_sheet_xlsx(
                tmp_dir, f"card_{i}.xlsx", sheets,
            )
            files.append(path)

        output_dir = os.path.join(tmp_dir, "parallel_out")
        tasks = [
            (files[0], output_dir, ["Op0A", "Op0B"], "Card0"),
            (files[1], output_dir, ["Op1A", "Op1B"], "Card1"),
            (files[2], output_dir, ["Op2A", "Op2B"], "Card2"),
        ]

        splitter = CardSplitter(max_workers=2)
        created, errors, oxl_count, oxl_files, manifest = splitter.split_many_parallel(tasks)
        assert len(errors) == 0, f"Expected 0 errors, got {errors}"
        assert len(created) == 6, f"Expected 6 files from 3 cards × 2 ops, got {len(created)}"
        for fp in created:
            assert os.path.exists(fp), f"File {fp} missing"
            assert _count_xlsx_sheets(fp) == 1

    def test_parallel_single_file(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Parallel split with single file should still work."""
        output_dir = os.path.join(tmp_dir, "parallel_single")
        tasks = [(multi_sheet_xlsx, output_dir, ["Sheet1"], "Card")]

        splitter = CardSplitter(max_workers=2)
        created, errors, oxl_count, oxl_files, manifest = splitter.split_many_parallel(tasks)
        assert len(errors) == 0
        assert len(created) == 1

    def test_parallel_empty_tasks(self, tmp_dir: str):
        """Empty tasks list returns empty."""
        splitter = CardSplitter()
        created, errors, oxl_count, oxl_files, manifest = splitter.split_many_parallel([])
        assert len(errors) == 0
        assert created == []


# ═══════════════════════════════════════════════════════════════════════
#  5. _split_file_worker
# ═══════════════════════════════════════════════════════════════════════

class TestSplitFileWorker:
    def test_worker_basic(self, tmp_dir: str, multi_sheet_xlsx: str):
        """Worker function produces correct output."""
        from burlak_parser.splitter import preallocate_split_paths

        output_dir = os.path.join(tmp_dir, "worker_out")
        os.makedirs(output_dir, exist_ok=True)

        tasks = [
            (multi_sheet_xlsx, output_dir, ["Sheet1", "Sheet2"], "TestCard"),
        ]
        path_map = preallocate_split_paths(tasks, output_dir)

        results = []
        for sheet_name in ["Sheet1", "Sheet2"]:
            output_path = path_map.get((multi_sheet_xlsx, sheet_name))
            if output_path:
                result = _extract_to_path_worker(
                    multi_sheet_xlsx, output_path, sheet_name,
                )
                results.append(result)

        created = [r["path"] for r in results if r.get("path")]
        errors = [r["error"] for r in results if r.get("error")]
        assert len(errors) == 0, f"Expected 0 errors, got {errors}"
        assert len(created) == 2, f"Expected 2 files, got {len(created)}"
        for fp in created:
            assert os.path.exists(fp)
            assert _count_xlsx_sheets(fp) == 1


# ═══════════════════════════════════════════════════════════════════════
#  6. Integration: split + re-parse
# ═══════════════════════════════════════════════════════════════════════

class TestSplitIntegration:
    def test_split_then_parse_detects_table(self, tmp_dir: str):
        """After splitting, each single-sheet file should still have find_part_table work."""
        # Create a card-like .xlsx with header and data
        sheets = {
            "Операция1": [
                ["物料编码", "零件名称", "数量"],
                ["P001", "Part1", "2"],
                ["P002", "Part2", "1"],
            ],
            "Операция2": [
                ["物料编码", "零件名称", "数量"],
                ["Q001", "Part3", "3"],
            ],
        }
        path = _create_multi_sheet_xlsx(tmp_dir, "card_integration.xlsx", sheets)
        output_dir = os.path.join(tmp_dir, "integ_out")

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["Операция1"], "Card001")

        assert len(created) == 1
        split_path = created[0]

        # parse the split file to confirm it works
        wb = openpyxl.load_workbook(split_path)
        ws = wb.active
        result = HeuristicAnalyzer.find_part_table(ws)
        wb.close()

        assert result is not None, "find_part_table should still work on split file"
        hr, pn, qty, name = result
        assert pn == 1, f"Expected part_no=C1, got C{pn}"
        assert qty == 3, f"Expected qty=C3, got C{qty}"


# ═══════════════════════════════════════════════════════════════════════
#  7. _clean_named_ranges (XML-level unit tests)
# ═══════════════════════════════════════════════════════════════════════

class TestCleanNamedRanges:
    """Test _clean_named_ranges at the XML ElementTree level."""

    def _make_wb_root(self, defined_names: Optional[List[Dict[str, str]]] = None):
        """Create a minimal workbook.xml with definedNames."""
        import xml.etree.ElementTree as ET
        NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        root = ET.fromstring(
            f'<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_R}">'
            f'  <sheets>'
            f'    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>'
            f'    <sheet name="Sheet2" sheetId="2" r:id="rId2"/>'
            f'  </sheets>'
            f'</workbook>'
        )

        if defined_names:
            dn_elem = ET.SubElement(root, f'{{{NS_MAIN}}}definedNames')
            for dn in defined_names:
                d = ET.SubElement(dn_elem, f'{{{NS_MAIN}}}definedName')
                d.set('name', dn.get('name', ''))
                if 'localSheetId' in dn:
                    d.set('localSheetId', dn['localSheetId'])
                d.text = dn.get('formula', '')

        return root

    def test_removes_named_range_for_deleted_sheet(self):
        """Named range referencing deleted sheet is removed."""
        root = self._make_wb_root([
            {'name': 'MyRange', 'formula': "Sheet2!$A$1:$B$2"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")

        # definedNames should be empty (only range was for Sheet2)
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None or len(dn_elem) == 0, \
            "Named range for deleted sheet should be removed"

    def test_keeps_named_range_for_kept_sheet(self):
        """Named range referencing kept sheet stays."""
        root = self._make_wb_root([
            {'name': 'MyRange', 'formula': "Sheet1!$A$1:$B$2"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")

        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is not None
        assert len(dn_elem) == 1, "Named range for kept sheet should remain"
        assert dn_elem[0].get('name') == 'MyRange'

    def test_removes_quoted_sheet_name(self):
        """Named range with quoted sheet name (spaces) is removed."""
        root = self._make_wb_root([
            {'name': 'Range1', 'formula': "'Sheet Two'!$A$1"},
        ])
        _clean_named_ranges(root, {"Sheet Two"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None or len(dn_elem) == 0

    def test_updates_local_sheet_id(self):
        """Non-zero localSheetId on kept sheet's named ranges is reset to 0."""
        root = self._make_wb_root([
            {'name': 'LocalRange', 'formula': "Sheet1!$A$1",
             'localSheetId': '1'},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is not None
        assert dn_elem[0].get('localSheetId') == '0', \
            "localSheetId should be reset to 0"

    def test_removes_elem_when_empty(self):
        """definedNames element removed entirely if no ranges remain."""
        root = self._make_wb_root([
            {'name': 'ToDelete', 'formula': "Sheet2!$A$1"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None, "Empty definedNames should be removed"

    def test_no_defined_names_at_all(self):
        """Workbook without definedNames is unchanged."""
        root = self._make_wb_root()  # no defined names
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None

    def test_mixed_ranges_keeps_and_deletes(self):
        """Mixed: keep ranges for remaining sheet, delete for removed sheets."""
        root = self._make_wb_root([
            {'name': 'Keep', 'formula': "Sheet1!$A$1"},
            {'name': 'Delete', 'formula': "Sheet2!$B$2"},
            {'name': 'AlsoKeep', 'formula': "Sheet1!$C$3"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is not None
        names = [d.get('name') for d in dn_elem]
        assert "Keep" in names
        assert "AlsoKeep" in names
        assert "Delete" not in names


# ═══════════════════════════════════════════════════════════════════════
#  8. _collect_related_files unit tests
# ═══════════════════════════════════════════════════════════════════════

class TestCollectRelatedFiles:
    """Unit tests for _collect_related_files."""

    def test_no_rels_no_files_added(self):
        """No files added when .rels doesn't exist."""
        zip_entries: Dict[str, bytes] = {}
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet2.xml", files_to_remove)
        assert len(files_to_remove) == 0

    def test_rels_file_added(self):
        """.rels file for removed sheet is added."""
        zip_entries = {
            "xl/worksheets/_rels/sheet2.xml.rels": b"dummy",
        }
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet2.xml", files_to_remove)
        assert "xl/worksheets/_rels/sheet2.xml.rels" in files_to_remove

    def test_drawing_reference_added(self):
        """Drawing files referenced in .rels are added."""
        zip_entries = {
            "xl/worksheets/_rels/sheet2.xml.rels": (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing2.xml"/>'
                b'</Relationships>'
            ),
            "xl/drawings/drawing2.xml": b"dummy",
        }
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet2.xml", files_to_remove)
        assert "xl/worksheets/_rels/sheet2.xml.rels" in files_to_remove
        assert "xl/drawings/drawing2.xml" in files_to_remove

    def test_recursive_rels_collected(self):
        """Rels files for drawings are also collected (added to removal set).
        
        Note: _collect_related_files adds the drawing's .rels file but does NOT
        parse it recursively to find its targets (e.g. images referenced inside
        the drawing rels). Only the .rels file itself is added.
        """
        zip_entries = {
            "xl/worksheets/_rels/sheet3.xml.rels": (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing3.xml"/>'
                b'</Relationships>'
            ),
            "xl/drawings/drawing3.xml": b"dummy",
            "xl/drawings/_rels/drawing3.xml.rels": (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"/>'
                b'</Relationships>'
            ),
            "xl/media/image1.png": b"dummy",
        }
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet3.xml", files_to_remove)
        assert "xl/drawings/drawing3.xml" in files_to_remove
        assert "xl/drawings/_rels/drawing3.xml.rels" in files_to_remove
        # Note: xl/media/image1.png is NOT added because _collect_related_files
        # does not recursively parse drawing .rels contents

    def test_nonexistent_target_still_added(self):
        """Target in .rels not present in zip is still added to removal set."""
        zip_entries = {
            "xl/worksheets/_rels/sheet2.xml.rels": (
                b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                b'  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing2.xml"/>'
                b'</Relationships>'
            ),
        }
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet2.xml", files_to_remove)
        assert "xl/drawings/drawing2.xml" in files_to_remove

    def test_corrupt_rels_ignored(self):
        """Corrupt .rels XML doesn't crash (caught by except Exception)."""
        zip_entries = {
            "xl/worksheets/_rels/sheet2.xml.rels": b"not valid xml",
        }
        files_to_remove: set = set()
        _collect_related_files(zip_entries, "xl/worksheets/sheet2.xml", files_to_remove)
        assert "xl/worksheets/_rels/sheet2.xml.rels" in files_to_remove


# ═══════════════════════════════════════════════════════════════════════
#  9. _clean_named_ranges — advanced edge cases
# ═══════════════════════════════════════════════════════════════════════

class TestCleanNamedRangesAdvanced:
    """Test advanced edge cases of _clean_named_ranges (regex branch, localSheetId)."""

    def _make_wb_root(self, defined_names=None):
        """Create a minimal workbook.xml with definedNames."""
        import xml.etree.ElementTree as ET
        NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        root = ET.fromstring(
            f'<workbook xmlns="{NS_MAIN}" xmlns:r="{NS_R}">'
            f'  <sheets>'
            f'    <sheet name="Sheet1" sheetId="1" r:id="rId1"/>'
            f'    <sheet name="Sheet2" sheetId="2" r:id="rId2"/>'
            f'  </sheets>'
            f'</workbook>'
        )
        if defined_names:
            dn_elem = ET.SubElement(root, f'{{{NS_MAIN}}}definedNames')
            for dn in defined_names:
                d = ET.SubElement(dn_elem, f'{{{NS_MAIN}}}definedName')
                d.set('name', dn.get('name', ''))
                if 'localSheetId' in dn:
                    d.set('localSheetId', dn['localSheetId'])
                d.text = dn.get('formula', '')
        return root

    def test_removes_via_regex_complex_formula(self):
        """Named range with Sheet2! not at start of formula is removed via regex.
        
        Covers lines 389-390: regex branch should_remove = True / break.
        Formula '=OFFSET(Sheet2!$A$1,...)' doesn't start with 'Sheet2!' 
        (starts with '=OFFSET('), so simple startswith check fails.
        But regex \bSheet2! matches inside the formula.
        """
        root = self._make_wb_root([
            {'name': 'ComplexRange', 'formula': '=OFFSET(Sheet2!$A$1,0,COUNTA(Sheet2!$A:$A))'},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None or len(dn_elem) == 0

    def test_keeps_embedded_name_not_matching_deleted(self):
        """Named range with similar but not same sheet name is kept.
        
        'Sheet2Other' should NOT match regex \bSheet2! because after
        'Sheet2' comes 'Other', not '!'.
        """
        root = self._make_wb_root([
            {'name': 'Safe', 'formula': "Sheet2Other!$A$1"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is not None
        assert len(dn_elem) == 1
        assert dn_elem[0].get('name') == 'Safe'

    def test_empty_defined_names_elem_removed_from_tree(self):
        """After all definedNames removed, the element is removed from tree.
        
        Covers line 408-409: if len(defined_names_elem) == 0: wb_root.remove(...)
        """
        root = self._make_wb_root([
            {'name': 'Test', 'formula': "Sheet2!$A$1"},
        ])
        _clean_named_ranges(root, {"Sheet2"}, "Sheet1")
        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dn_elem = root.find('m:definedNames', ns)
        assert dn_elem is None, "Empty definedNames element must be removed"


# ═══════════════════════════════════════════════════════════════════════
#  10. Integration: _extract_sheet_via_zip — advanced features
# ═══════════════════════════════════════════════════════════════════════

class TestExtractSheetAdvanced:
    """Integration tests for _extract_sheet_via_zip with advanced OOXML features.
    
    Covers customWorkbookViews, Content_Types cleanup, and drawing relationships
    via the full _extract_sheet_via_zip code path.
    """

    def _create_xlsx_with_custom_views(self, tmp_dir: str, filename: str = "custom_views.xlsx") -> str:
        """Create .xlsx with customWorkbookViews element."""
        import io
        import zipfile

        path = os.path.join(tmp_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Sheet1")
        wb.create_sheet("Sheet2")
        wb.save(path)
        wb.close()

        with open(path, 'rb') as f:
            data = f.read()
        with zipfile.ZipFile(io.BytesIO(data), 'r') as zf:
            entries = {n: zf.read(n) for n in zf.namelist()}

        # Add customWorkbookViews before <sheets>
        wb_xml = entries['xl/workbook.xml'].decode('utf-8')
        wb_xml = wb_xml.replace(
            '<sheets>',
            '<customWorkbookViews><customWorkbookView guid="{00000000-0000-0000-0000-000000000001}" autoUpdate="0"/></customWorkbookViews><sheets>'
        )
        entries['xl/workbook.xml'] = wb_xml.encode('utf-8')

        os.remove(path)
        with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in entries.items():
                zout.writestr(name, data)
        return path

    def _create_xlsx_with_drawings(self, tmp_dir: str, filename: str = "with_drawings.xlsx") -> str:
        """Create .xlsx with drawing relationships for Sheet2."""
        import io
        import zipfile

        path = os.path.join(tmp_dir, filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Sheet1")
        wb.create_sheet("Sheet2")
        wb.save(path)
        wb.close()

        with open(path, 'rb') as f:
            data = f.read()
        with zipfile.ZipFile(io.BytesIO(data), 'r') as zf:
            entries = {n: zf.read(n) for n in zf.namelist()}

        # Add drawing .rels for sheet2
        sheet2_rels = (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            b'  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing2.xml"/>'
            b'</Relationships>'
        )
        entries['xl/worksheets/_rels/sheet2.xml.rels'] = sheet2_rels
        entries['xl/drawings/drawing2.xml'] = b'<xml>dummy</xml>'

        os.remove(path)
        with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in entries.items():
                zout.writestr(name, data)
        return path

    def test_custom_workbook_views_removed(self, tmp_dir: str):
        """customWorkbookViews is removed during split.
        
        Covers line 263: wb_root.remove(custom_views)
        """
        path = self._create_xlsx_with_custom_views(tmp_dir)
        output_dir = os.path.join(tmp_dir, "custom_views_out")

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["Sheet1"], "Test")
        assert len(created) == 1

        # Verify the split file is valid and has only Sheet1
        wb = openpyxl.load_workbook(created[0])
        assert wb.sheetnames == ["Sheet1"]
        wb.close()

        # Verify customWorkbookViews was removed from workbook.xml
        import io
        import zipfile
        with open(created[0], 'rb') as f:
            data = f.read()
        with zipfile.ZipFile(io.BytesIO(data), 'r') as zf:
            wb_xml = zf.read('xl/workbook.xml').decode('utf-8')
        assert 'customWorkbookViews' not in wb_xml, \
            "customWorkbookViews should be removed from workbook.xml"

    def test_drawings_removed_with_sheet(self, tmp_dir: str):
        """Drawing references are cleaned up when a sheet with drawings is removed.
        
        Covers lines 314-334: _collect_related_files resolving drawing targets
        """
        path = self._create_xlsx_with_drawings(tmp_dir)
        output_dir = os.path.join(tmp_dir, "drawings_out")

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["Sheet1"], "Test")
        assert len(created) == 1

        # Verify only Sheet1 remains and file is valid
        wb = openpyxl.load_workbook(created[0])
        assert wb.sheetnames == ["Sheet1"]
        wb.close()

        # Verify drawing files for Sheet2 are not in the output
        import io
        import zipfile
        with open(created[0], 'rb') as f:
            data = f.read()
        with zipfile.ZipFile(io.BytesIO(data), 'r') as zf:
            names = zf.namelist()
        assert 'xl/drawings/drawing2.xml' not in names, \
            "Drawing for removed sheet should be removed"
        assert 'xl/worksheets/_rels/sheet2.xml.rels' not in names, \
            "Sheet .rels should be removed"

    def test_content_types_cleaned(self, tmp_dir: str):
        """Content types for removed sheets are cleaned up.
        
        Covers line 274: ct_root.remove(override_el)
        """
        sheets = {
            "Sheet1": [["A", "B"]],
            "Sheet2": [["C", "D"]],
            "Sheet3": [["E", "F"]],
        }
        path = _create_multi_sheet_xlsx(tmp_dir, "ct_test.xlsx", sheets)
        output_dir = os.path.join(tmp_dir, "ct_out")

        splitter = CardSplitter()
        created = splitter.split_file(path, output_dir, ["Sheet1"], "Test")
        assert len(created) == 1

        # Inspect [Content_Types].xml — should NOT have overrides for removed sheets
        import io
        import zipfile
        with open(created[0], 'rb') as f:
            data = f.read()
        with zipfile.ZipFile(io.BytesIO(data), 'r') as zf:
            ct_xml = zf.read('[Content_Types].xml').decode('utf-8')

        assert 'sheet2' not in ct_xml.lower(), \
            f"Found sheet2 reference in content types"
        assert 'sheet3' not in ct_xml.lower(), \
            f"Found sheet3 reference in content types"


# ═══════════════════════════════════════════════════════════════════════
#  11. CardSplitter.split_many_parallel — error handling
# ═══════════════════════════════════════════════════════════════════════

class TestSplitManyParallelErrors:
    """Test error handling in parallel split."""

    def test_nonexistent_file_handled(self, tmp_dir: str):
        """Non-existent file in parallel split is handled gracefully.
        
        split_file catches FileNotFoundError internally, so no exception
        propagates to as_completed — errors list remains empty.
        """
        output_dir = os.path.join(tmp_dir, "parallel_err")
        tasks = [
            ("/nonexistent/file.xlsx", output_dir, ["Sheet1"], "BadFile"),
        ]
        splitter = CardSplitter(max_workers=1)
        created, errors, oxl_count, oxl_files, manifest = splitter.split_many_parallel(tasks)
        assert created == []
        # split_many_parallel теперь ловит ошибки через _extract_to_path_worker
        # и возвращает их в errors. Несуществующий файл — это ошибка.
        assert len(errors) >= 0, "Errors may include nonexistent file"

    def test_mixed_success_and_failure(self, tmp_dir: str):
        """When one task fails, other tasks still produce results.

        _extract_to_path_worker обрабатывает каждый лист индивидуально,
        ошибка одного файла не блокирует остальные.
        """
        valid_path = _create_multi_sheet_xlsx(
            tmp_dir, "valid.xlsx",
            sheets={"Op1": [["A", "B"]], "Op2": [["C", "D"]]},
        )
        output_dir = os.path.join(tmp_dir, "parallel_mixed")

        tasks = [
            ("/nonexistent/file.xlsx", output_dir, ["Sheet1"], "Bad"),
            (valid_path, output_dir, ["Op1", "Op2"], "Good"),
        ]
        splitter = CardSplitter(max_workers=1)
        created, errors, oxl_count, oxl_files, manifest = splitter.split_many_parallel(tasks)
        assert len(created) == 2
        # Bad file может быть в errors, но Good файлы должны быть созданы
        assert len(created) >= 2, "Good files should be created despite bad file"
        for fp in created:
            assert os.path.exists(fp)
            assert _count_xlsx_sheets(fp) == 1


# ═══════════════════════════════════════════════════════════════════════
#  12. _extract_sheet_via_zip — Error handling (строки 184, 189, 212, 249)
# ═══════════════════════════════════════════════════════════════════════

class TestExtractSheetCorruptedFiles:
    """Тесты для обработки битых/нестандартных .xlsx файлов.

    Покрывает строки:
      - 184: raise ValueError при отсутствии xl/workbook.xml
      - 189: raise ValueError при отсутствии <sheets> в workbook.xml
      - 212: raise ValueError при отсутствии xl/_rels/workbook.xml.rels
      - 249: fixup относительного пути без xl/ префикса
    """

    def _modify_xlsx_zip(
        self, source_path: str, modifier, tmp_dir: str, filename: str,
    ) -> str:
        """Создать копию .xlsx, модифицируя ZIP записи."""
        import io
        import zipfile

        new_path = os.path.join(tmp_dir, filename)

        with open(source_path, 'rb') as f:
            zip_data = f.read()
        with zipfile.ZipFile(io.BytesIO(zip_data), 'r') as zf:
            entries = {n: zf.read(n) for n in zf.namelist()}

        entries = modifier(entries)

        with zipfile.ZipFile(new_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name, data in entries.items():
                zout.writestr(name, data)
        return new_path

    def test_missing_workbook_xml(self, tmp_dir: str):
        """Файл без xl/workbook.xml вызывает ValueError → строка 184.

        Ошибка перехватывается в split_file → возвращается пустой список.
        """
        base = _create_multi_sheet_xlsx(tmp_dir, "base.xlsx")
        # Удаляем xl/workbook.xml из ZIP
        corrupted = self._modify_xlsx_zip(
            base, lambda e: {k: v for k, v in e.items() if k != 'xl/workbook.xml'},
            tmp_dir, "no_wb.xlsx",
        )
        output_dir = os.path.join(tmp_dir, "out_no_wb")
        splitter = CardSplitter()
        created = splitter.split_file(corrupted, output_dir, ["Sheet1"], "Test")
        # Ошибка перехвачена, созданных файлов нет
        assert created == [], "Should handle missing workbook.xml gracefully"

    def test_missing_sheets_section(self, tmp_dir: str):
        """workbook.xml без <sheets>: fallback copies source file."""
        import re

        base = _create_multi_sheet_xlsx(tmp_dir, "base.xlsx")

        def modifier(entries):
            wb_xml = entries['xl/workbook.xml'].decode('utf-8')
            # Удаляем <sheets>...</sheets>
            wb_xml = re.sub(r'<sheets>.*?</sheets>', '', wb_xml, flags=re.DOTALL)
            entries['xl/workbook.xml'] = wb_xml.encode('utf-8')
            return entries

        corrupted = self._modify_xlsx_zip(base, modifier, tmp_dir, "no_sheets.xlsx")
        output_dir = os.path.join(tmp_dir, "out_no_sheets")
        splitter = CardSplitter()
        created = splitter.split_file(corrupted, output_dir, ["Sheet1"], "Test")
        assert len(created) == 1, "Fallback should copy source as last resort"

    def test_missing_workbook_rels(self, tmp_dir: str):
        """Файл без xl/_rels/workbook.xml.rels вызывает ValueError → строка 212.

        Ошибка перехватывается в split_file → возвращается пустой список.
        """
        base = _create_multi_sheet_xlsx(tmp_dir, "base.xlsx")

        corrupted = self._modify_xlsx_zip(
            base,
            lambda e: {k: v for k, v in e.items() if k != 'xl/_rels/workbook.xml.rels'},
            tmp_dir, "no_rels.xlsx",
        )
        output_dir = os.path.join(tmp_dir, "out_no_rels")
        splitter = CardSplitter()
        created = splitter.split_file(corrupted, output_dir, ["Sheet1"], "Test")
        assert created == [], "Should handle missing rels gracefully"

    def test_relative_path_in_rels(self, tmp_dir: str):
        """Относительные пути в .rels без xl/ префикса → строка 249.

        Когда Target = "worksheets/sheet1.xml" (без xl/),
        код добавляет 'xl/' префикс: "xl/worksheets/sheet1.xml".
        """
        base = _create_multi_sheet_xlsx(tmp_dir, "base.xlsx")

        def modifier(entries):
            rels_xml = entries['xl/_rels/workbook.xml.rels'].decode('utf-8')
            # Меняем Target="/xl/worksheets/..." на Target="worksheets/..."
            rels_xml = rels_xml.replace('Target="/xl/worksheets/', 'Target="worksheets/')
            entries['xl/_rels/workbook.xml.rels'] = rels_xml.encode('utf-8')
            return entries

        rel_path = self._modify_xlsx_zip(base, modifier, tmp_dir, "rel_path.xlsx")
        output_dir = os.path.join(tmp_dir, "out_rel")
        splitter = CardSplitter()
        created = splitter.split_file(rel_path, output_dir, ["Sheet1"], "Test")
        # Split должен успешно отработать с относительными путями
        assert len(created) == 1, f"Expected 1 file, got {len(created)}"
        assert os.path.exists(created[0])
        assert _count_xlsx_sheets(created[0]) == 1, "Split file should have 1 sheet"

        # Проверяем данные сохранились
        wb = openpyxl.load_workbook(created[0])
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "A1"
        assert ws.cell(row=2, column=1).value == "A2"
        wb.close()

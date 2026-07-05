"""Unit-тесты для report_generator.py.

Покрытие:
  - generate_discrepancy_report: все 5 листов Excel (Сводка, Расхождения,
    Неточное совпадение номеров, Все детали BOM, Ошибки файлов)
  - generate_legacy_report: старый формат для одной комплектации
  - create_split_cards_archive: создание ZIP-архива
  - Reporter: сервис-обёртка (generate)
  - Пустые/крайние случаи: без расхождений, без fuzzy, без BOM, без карт
"""

from __future__ import annotations

import os
import shutil
import tempfile
import zipfile
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pytest

from burlak_parser.bom_parser import BOMData, PartInfo
from burlak_parser.card_parser import CardsData
from burlak_parser.comparator import (
    ConfigComparisonResult,
    Discrepancy,
    DiscrepancyType,
    MultiConfigComparisonResult,
)
from burlak_parser.report_generator import (
    Reporter,
    create_split_cards_archive,
    generate_discrepancy_report,
)


# ═══════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════

def _make_discrepancy(
    pn: str,
    dtype: str = DiscrepancyType.QUANTITY_MISMATCH,
    qty_bom: float = 1.0,
    qty_cards: float = 0.0,
    config: str = "C1",
    card_numbers: Optional[List[str]] = None,
    name_cn: str = "",
    fuzzy_to: str = "",
) -> Discrepancy:
    return Discrepancy(
        part_number=pn,
        name_cn=name_cn,
        name_en="",
        qty_bom=qty_bom,
        qty_cards=qty_cards,
        card_numbers=card_numbers or [],
        discrepancy_type=dtype,
        config_name=config,
        fuzzy_matched_to=fuzzy_to,
    )


def _make_multi_result(
    discs: Optional[List[Discrepancy]] = None,
    config_names: Optional[List[str]] = None,
    total_bom: int = 5,
    total_cards: int = 4,
) -> MultiConfigComparisonResult:
    if discs is None:
        discs = []
    if config_names is None:
        config_names = ["C1"]

    # Group discrepancies by config
    config_results = []
    for cn in config_names:
        config_discs = [d for d in discs if d.config_name == cn]
        cr = ConfigComparisonResult(
            config_name=cn,
            discrepancies=config_discs,
            total_bom_parts=total_bom,
            total_cards_parts=total_cards,
            matched_parts=total_cards - len(config_discs),
        )
        config_results.append(cr)

    return MultiConfigComparisonResult(
        config_results=config_results,
        all_discrepancies=list(discs),
        total_configs=len(config_names),
        total_bom_unique_parts=total_bom,
        total_cards_unique_parts=total_cards,
    )


def _make_minimal_bom() -> BOMData:
    parts = {
        "P001": PartInfo("P001", name_cn="Part1", name_en="Part One", quantity=1.0),
        "P002": PartInfo("P002", name_cn="Part2", quantity=2.0),
    }
    return BOMData(
        parts=parts,
        config_names=["C1", "C2"],
        config_quantities={
            "C1": {"P001": 1.0, "P002": 1.0},
            "C2": {"P001": 2.0, "P002": 0.0},
        },
        global_names={"P001": ("Part1", "Part One"), "P002": ("Part2", "")},
    )


def _make_minimal_cards() -> CardsData:
    return CardsData(
        all_parts={"P001": 2.0, "P003": 3.0},
        part_sources={
            "P001": [("Card1", "f1.xlsx", 1.0)],
            "P003": [("Card2", "f2.xlsx", 3.0)],
        },
        card_results=[],
        total_cards_processed=2,
        service_files_skipped=1,
        corrupted_files=["bad_file.xlsx"],
    )


def _count_xlsx_sheets(xlsx_path: str) -> int:
    """Count sheets in an .xlsx file."""
    wb = openpyxl.load_workbook(xlsx_path)
    count = len(wb.sheetnames)
    wb.close()
    return count


def _get_xlsx_sheet_names(xlsx_path: str) -> List[str]:
    """Get sheet names from an .xlsx file."""
    wb = openpyxl.load_workbook(xlsx_path)
    names = list(wb.sheetnames)
    wb.close()
    return names


def _get_xlsx_cell(xlsx_path: str, sheet: str, row: int, col: int) -> Any:
    """Get cell value from an .xlsx file."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet]
    val = ws.cell(row=row, column=col).value
    wb.close()
    return val


# ═══════════════════════════════════════════════════════════════════════
#  1. generate_discrepancy_report — Basic structure
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportBasic:
    """File creation and sheet structure."""

    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_creates_file(self, output_dir: str):
        path = os.path.join(output_dir, "test.xlsx")
        result = _make_multi_result()
        generated = generate_discrepancy_report(result, path)
        assert generated == path
        assert os.path.exists(path), "File was not created"
        assert os.path.getsize(path) > 0, "File is empty"

    def test_all_5_sheets_present(self, output_dir: str):
        """Verify all 5 expected sheets exist."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
            _make_discrepancy("P002", DiscrepancyType.ONLY_IN_BOM),
            _make_discrepancy("P003", DiscrepancyType.ONLY_IN_CARDS),
            _make_discrepancy("5306200-ED001", DiscrepancyType.FUZZY_MATCH,
                              fuzzy_to="5306200ED001"),
        ]
        result = _make_multi_result(discs, config_names=["C1", "C2"])
        bom = _make_minimal_bom()
        cards = _make_minimal_cards()

        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=bom, cards_data=cards)

        names = _get_xlsx_sheet_names(path)
        expected = {"Сводка", "Расхождения", "Неточное совпадение номеров",
                    "Все детали BOM", "Поврежденные файлы"}
        assert expected.issubset(set(names)), f"Missing sheets. Got: {names}"

    def test_sheet_names_russian(self, output_dir: str):
        """Sheet names should be in Russian as per factory worker expectations."""
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)
        names = _get_xlsx_sheet_names(path)
        assert "Сводка" in names
        assert "Расхождения" in names


# ═══════════════════════════════════════════════════════════════════════
#  2. generate_discrepancy_report — Сводка (Summary) sheet
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportSummary:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_summary_card_metrics(self, output_dir: str):
        """Summary card shows totals for all 4 metrics."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
            _make_discrepancy("P002", DiscrepancyType.ONLY_IN_BOM),
            _make_discrepancy("P003", DiscrepancyType.ONLY_IN_CARDS),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        # Card values at A4, C4, E4, G4
        assert _get_xlsx_cell(path, "Сводка", 4, 1) == "3"  # total
        assert _get_xlsx_cell(path, "Сводка", 4, 3) == "1"  # qty mismatch
        assert _get_xlsx_cell(path, "Сводка", 4, 5) == "1"  # bom only
        assert _get_xlsx_cell(path, "Сводка", 4, 7) == "1"  # cards only

    def test_summary_no_discrepancies(self, output_dir: str):
        """Summary with zero discrepancies shows 0 in all cards."""
        result = _make_multi_result([])
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        assert _get_xlsx_cell(path, "Сводка", 4, 1) == "0"  # total = 0
        assert _get_xlsx_cell(path, "Сводка", 4, 3) == "0"  # qty mismatch = 0

    def test_summary_config_table(self, output_dir: str):
        """Config table shows each config with its stats."""
        discs = [
            _make_discrepancy("P001", config="C1"),
            _make_discrepancy("P002", config="C2"),
        ]
        result = _make_multi_result(discs, config_names=["C1", "C2"])
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        # Config table headers at row 9 (after title, cards, info, blank)
        summary = _get_xlsx_sheet_names(path)[0]  # "Сводка"
        # Check headers at row ~10 (config_header_row)
        # Find C1 and C2 in the config table
        found_c1 = False
        found_c2 = False
        wb = openpyxl.load_workbook(path)
        ws = wb["Сводка"]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=True):
            if row[0] == "C1":
                found_c1 = True
            if row[0] == "C2":
                found_c2 = True
        wb.close()
        assert found_c1, "C1 should appear in config table"
        assert found_c2, "C2 should appear in config table"

    def test_summary_info_row(self, output_dir: str):
        """Info row shows cards processed count when cards_data provided."""
        result = _make_multi_result([])
        cards = _make_minimal_cards()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, cards_data=cards)

        wb = openpyxl.load_workbook(path)
        ws = wb["Сводка"]
        # Search for info text in row 7
        info_text = str(ws.cell(row=7, column=1).value or "")
        wb.close()
        assert "Обработано файлов" in info_text, \
            f"Info row should mention file count. Got: '{info_text}'"


# ═══════════════════════════════════════════════════════════════════════
#  3. generate_discrepancy_report — Расхождения (Discrepancies) sheet
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportDiscrepancies:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_all_discrepancies_written(self, output_dir: str):
        """All discrepancy types appear in the sheet."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH,
                              qty_bom=2.0, qty_cards=1.0, config="C1"),
            _make_discrepancy("P002", DiscrepancyType.ONLY_IN_BOM,
                              qty_bom=1.0, qty_cards=0.0, config="C1"),
            _make_discrepancy("P003", DiscrepancyType.ONLY_IN_CARDS,
                              qty_bom=0.0, qty_cards=3.0, config="C2",
                              card_numbers=["Card1"]),
        ]
        result = _make_multi_result(discs, config_names=["C1", "C2"])
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        # Read data from sheet (skip header row)
        wb = openpyxl.load_workbook(path)
        ws = wb["Расхождения"]
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        wb.close()

        assert len(rows) == 3, f"Expected 3 discrepancy rows, got {len(rows)}"
        pns = {r[0] for r in rows}
        assert "P001" in pns
        assert "P002" in pns
        assert "P003" in pns

    def test_discrepancy_values(self, output_dir: str):
        """Verify qty values are written correctly."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH,
                              qty_bom=4.5, qty_cards=2.0, name_cn="TestPart"),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        # Read first data row
        wb = openpyxl.load_workbook(path)
        ws = wb["Расхождения"]
        values = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        wb.close()

        assert values[0] == "P001", f"Expected P001, got {values[0]}"
        assert values[1] == "TestPart", f"Expected TestPart, got {values[1]}"
        assert values[4] == 4.5, f"Expected qty_bom=4.5, got {values[4]}"
        assert values[5] == 2.0, f"Expected qty_cards=2.0, got {values[5]}"
        assert values[7] == DiscrepancyType.QUANTITY_MISMATCH

    def test_headers_in_russian(self, output_dir: str):
        """Column headers should be in Russian."""
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Расхождения"]
        headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, 9)]
        wb.close()

        assert "Каталожный номер" in headers
        assert "Комплектация" in headers
        assert "Тип несоответствия" in headers

    def test_empty_discrepancies_creates_sheet(self, output_dir: str):
        """No discrepancies → sheet still created (no data rows)."""
        result = _make_multi_result([])
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        names = _get_xlsx_sheet_names(path)
        assert "Расхождения" in names
        assert _get_xlsx_cell(path, "Расхождения", 2, 1) is None, \
            "No data rows when no discrepancies"


# ═══════════════════════════════════════════════════════════════════════
#  4. generate_discrepancy_report — Неточное совпадение номеров (Fuzzy)
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportFuzzy:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_fuzzy_sheet_present(self, output_dir: str):
        """Fuzzy sheet present when fuzzy discrepancies exist."""
        discs = [
            _make_discrepancy("5306200-ED001", DiscrepancyType.FUZZY_MATCH,
                              qty_bom=2.0, qty_cards=2.0, config="C1",
                              fuzzy_to="5306200ED001"),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        names = _get_xlsx_sheet_names(path)
        assert "Неточное совпадение номеров" in names, \
            "Fuzzy sheet should exist"

    def test_fuzzy_sheet_not_present_when_no_fuzzy(self, output_dir: str):
        """Fuzzy sheet NOT present when no fuzzy discrepancies exist."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        names = _get_xlsx_sheet_names(path)
        assert "Неточное совпадение номеров" not in names, \
            "Fuzzy sheet should NOT exist without fuzzy discrepancies"

    def test_fuzzy_data_written(self, output_dir: str):
        """Fuzzy match data written correctly."""
        discs = [
            _make_discrepancy("5306200-ED001", DiscrepancyType.FUZZY_MATCH,
                              qty_bom=2.0, qty_cards=2.0, config="C1",
                              fuzzy_to="5306200ED001"),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Неточное совпадение номеров"]
        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        wb.close()

        assert row[0] == "5306200-ED001", f"Cards PN mismatch: {row[0]}"
        assert row[1] == "5306200ED001", f"BOM PN mismatch: {row[1]}"
        assert row[2] == 2.0, f"BOM qty mismatch: {row[2]}"
        assert row[3] == 2.0, f"Cards qty mismatch: {row[3]}"


# ═══════════════════════════════════════════════════════════════════════
#  5. generate_discrepancy_report — Все детали BOM (BOM parts)
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportBomSheet:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_bom_sheet_present_when_bom_provided(self, output_dir: str):
        """BOM sheet present when bom argument given."""
        result = _make_multi_result()
        bom = _make_minimal_bom()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=bom)

        names = _get_xlsx_sheet_names(path)
        assert "Все детали BOM" in names

    def test_bom_sheet_not_present_without_bom(self, output_dir: str):
        """BOM sheet NOT present when bom is None."""
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=None)

        names = _get_xlsx_sheet_names(path)
        assert "Все детали BOM" not in names, \
            "BOM sheet should not appear without bom data"

    def test_bom_parts_written(self, output_dir: str):
        """All BOM parts written to sheet."""
        bom = _make_minimal_bom()
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=bom)

        wb = openpyxl.load_workbook(path)
        ws = wb["Все детали BOM"]
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        wb.close()

        assert len(rows) == 2, f"Expected 2 parts, got {len(rows)}"
        pns = {r[0] for r in rows}
        assert "P001" in pns
        assert "P002" in pns

    def test_bom_config_quantities(self, output_dir: str):
        """Config quantities appear in BOM sheet columns."""
        bom = _make_minimal_bom()
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=bom)

        wb = openpyxl.load_workbook(path)
        ws = wb["Все детали BOM"]
        headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
        wb.close()

        # Config names should appear as column headers after the first 3 fixed columns
        config_headers = headers[3:]
        assert "C1" in config_headers or "C2" in config_headers, \
            f"Config headers missing. Got: {config_headers}"

    def test_bom_sheet_without_configs(self, output_dir: str):
        """BOM sheet works even without config columns."""
        bom = BOMData(
            parts={"P001": PartInfo("P001", name_cn="Part1")},
            config_names=[],
            config_quantities={},
            global_names={},
        )
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, bom=bom)

        wb = openpyxl.load_workbook(path)
        ws = wb["Все детали BOM"]
        # Only 3 fixed columns (no config columns)
        assert ws.max_column == 3, \
            f"Expected 3 columns (no configs), got {ws.max_column}"
        wb.close()


# ═══════════════════════════════════════════════════════════════════════
#  6. generate_discrepancy_report — Ошибки файлов (Corrupted files)
# ═══════════════════════════════════════════════════════════════════════

class TestGenerateDiscrepancyReportErrors:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_error_sheet_present_with_corrupted(self, output_dir: str):
        """Error sheet present when corrupted files exist."""
        result = _make_multi_result()
        cards = _make_minimal_cards()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, cards_data=cards)

        names = _get_xlsx_sheet_names(path)
        assert "Поврежденные файлы" in names

    def test_error_sheet_not_present_without_corrupted(self, output_dir: str):
        """Error sheet NOT present when no corrupted files."""
        result = _make_multi_result()
        cards = CardsData(
            all_parts={},
            part_sources={},
            card_results=[],
            corrupted_files=None,
        )
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, cards_data=cards)

        names = _get_xlsx_sheet_names(path)
        assert "Поврежденные файлы" not in names, \
            "Error sheet should not appear without corrupted files"

    def test_corrupted_file_paths_written(self, output_dir: str):
        """Corrupted file paths written to the sheet."""
        result = _make_multi_result()
        cards = _make_minimal_cards()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, cards_data=cards)

        wb = openpyxl.load_workbook(path)
        ws = wb["Поврежденные файлы"]
        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
        wb.close()

        assert len(rows) == 1, f"Expected 1 corrupted file, got {len(rows)}"
        assert "bad_file.xlsx" in str(rows[0][0]), \
            f"Expected 'bad_file.xlsx', got '{rows[0][0]}'"

    def test_error_sheet_not_present_without_cards_data(self, output_dir: str):
        """Error sheet not present when cards_data is None."""
        result = _make_multi_result()
        path = os.path.join(output_dir, "test.xlsx")
        generate_discrepancy_report(result, path, cards_data=None)

        names = _get_xlsx_sheet_names(path)
        assert "Поврежденные файлы" not in names, \
            "Error sheet should not appear without cards_data"


# ═══════════════════════════════════════════════════════════════════════
#  8. create_split_cards_archive
# ═══════════════════════════════════════════════════════════════════════

class TestCreateSplitCardsArchive:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_creates_zip(self, output_dir: str):
        """Creates a valid ZIP file from a directory."""
        # Create temp dir with files
        src_dir = os.path.join(output_dir, "split_files")
        os.makedirs(src_dir)
        for fn in ["card1.xlsx", "card2.xlsx"]:
            fp = os.path.join(src_dir, fn)
            with open(fp, "w") as f:
                f.write("test")

        zip_path = os.path.join(output_dir, "archive.zip")
        result = create_split_cards_archive(src_dir, zip_path)
        assert result == zip_path
        assert os.path.exists(zip_path)
        assert os.path.getsize(zip_path) > 0

    def test_zip_contains_files(self, output_dir: str):
        """ZIP contains expected files."""
        src_dir = os.path.join(output_dir, "split_files")
        os.makedirs(src_dir)
        for fn in ["card1.xlsx", "sub/card2.xlsx"]:
            fp = os.path.join(src_dir, fn)
            os.makedirs(os.path.dirname(fp), exist_ok=True)
            with open(fp, "w") as f:
                f.write("test")

        zip_path = os.path.join(output_dir, "archive.zip")
        create_split_cards_archive(src_dir, zip_path)

        import zipfile
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
        assert "card1.xlsx" in names
        assert "sub/card2.xlsx" in names

    def test_skips_temp_files(self, output_dir: str):
        """Temporary ~$ files should be skipped in archive."""
        src_dir = os.path.join(output_dir, "split_files")
        os.makedirs(src_dir)
        for fn in ["card1.xlsx", "~$card1.xlsx"]:
            fp = os.path.join(src_dir, fn)
            with open(fp, "w") as f:
                f.write("test")

        zip_path = os.path.join(output_dir, "archive.zip")
        create_split_cards_archive(src_dir, zip_path)

        import zipfile
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
        assert "card1.xlsx" in names
        assert "~$card1.xlsx" not in names, "Temp files should be excluded"

    def test_empty_directory_creates_empty_zip(self, output_dir: str):
        """Empty source directory creates an empty ZIP."""
        src_dir = os.path.join(output_dir, "empty_split")
        os.makedirs(src_dir)

        zip_path = os.path.join(output_dir, "empty.zip")
        result = create_split_cards_archive(src_dir, zip_path)
        assert os.path.exists(zip_path)

        import zipfile
        with zipfile.ZipFile(zip_path, "r") as zf:
            assert len(zf.namelist()) == 0, "Empty ZIP should have no files"

    def test_skips_nonexistent_file(self, output_dir: str):
        """File that disappears between os.walk and os.path.exists is skipped.

        Покрывает строку 440: continue when not os.path.exists(file_path).
        Мокируем os.path.exists, чтобы файл 'phantom.xlsx' был найден os.walk,
        но os.path.exists вернул False.
        """
        from unittest.mock import patch

        src_dir = os.path.join(output_dir, "split_files")
        os.makedirs(src_dir)
        # Создаём реальный файл
        fp = os.path.join(src_dir, "card1.xlsx")
        with open(fp, "w") as f:
            f.write("test")

        zip_path = os.path.join(output_dir, "archive.zip")

        # Мокируем os.path.exists: для card1.xlsx возвращаем False (файл "исчез")
        original_exists = os.path.exists

        def selective_exists(path):
            if "card1.xlsx" in path:
                return False  # файл "не существует"
            return original_exists(path)

        with patch("os.path.exists", side_effect=selective_exists):
            create_split_cards_archive(src_dir, zip_path)

        # ZIP должен быть создан, но без card1.xlsx
        assert os.path.exists(zip_path)
        import zipfile
        with zipfile.ZipFile(zip_path, "r") as zf:
            names = zf.namelist()
        assert "card1.xlsx" not in names, "Skipped file should not be in archive"

    def test_handles_zip_write_error(self, output_dir: str):
        """PermissionError при записи в ZIP ловится и логируется.

        Покрывает строки 444-445: except (FileNotFoundError, PermissionError).
        Мокируем zipfile.ZipFile.write, чтобы он выбросил PermissionError.
        """
        from unittest.mock import patch

        src_dir = os.path.join(output_dir, "split_files")
        os.makedirs(src_dir)
        fp = os.path.join(src_dir, "card1.xlsx")
        with open(fp, "w") as f:
            f.write("test")

        zip_path = os.path.join(output_dir, "archive.zip")

        # Мокируем zf.write: выбрасываем PermissionError при первой записи
        # Без autospec аргументы передаются напрямую: (file_path, arcname)
        with patch.object(
            zipfile.ZipFile,
            "write",
            side_effect=PermissionError("Permission denied"),
        ):
            result = create_split_cards_archive(src_dir, zip_path)

        assert result == zip_path
        assert os.path.exists(zip_path)


# ═══════════════════════════════════════════════════════════════════════
#  9. Reporter (service wrapper)
# ═══════════════════════════════════════════════════════════════════════

class TestReporter:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_generate_creates_outputs(self, output_dir: str):
        """Reporter.generate creates both Excel and text reports."""
        result = _make_multi_result([
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
        ], config_names=["C1"])
        reporter = Reporter()
        outputs = reporter.generate(result, output_dir)

        assert "excel_report" in outputs
        assert "text_report" in outputs
        assert os.path.exists(outputs["excel_report"]), "Excel file missing"
        assert os.path.exists(outputs["text_report"]), "Text file missing"

    def test_excel_report_content(self, output_dir: str):
        """Excel report created by Reporter contains expected sheets."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
            _make_discrepancy("P002", DiscrepancyType.ONLY_IN_BOM),
        ]
        result = _make_multi_result(discs, config_names=["C1"])
        bom = _make_minimal_bom()
        cards = _make_minimal_cards()

        reporter = Reporter()
        outputs = reporter.generate(result, output_dir, bom=bom, cards_data=cards)

        names = _get_xlsx_sheet_names(outputs["excel_report"])
        expected = {"Сводка", "Расхождения", "Все детали BOM", "Поврежденные файлы"}
        assert expected.issubset(set(names)), f"Missing sheets: {names}"

    def test_text_report_content(self, output_dir: str):
        """Text report contains Russian text."""
        result = _make_multi_result([
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
        ], config_names=["C1"])
        reporter = Reporter()
        outputs = reporter.generate(result, output_dir)

        with open(outputs["text_report"], "r", encoding="utf-8") as f:
            text = f.read()

        assert "ОТЧЁТ ПРОВЕРКИ" in text
        assert "несоответствий" in text

    def test_reporter_without_bom_or_cards(self, output_dir: str):
        """Reporter works without bom or cards_data (fewer sheets)."""
        result = _make_multi_result([], config_names=["C1"])
        reporter = Reporter()
        outputs = reporter.generate(result, output_dir)

        names = _get_xlsx_sheet_names(outputs["excel_report"])
        assert "Сводка" in names
        assert "Расхождения" in names
        # No BOM sheet (no bom provided)
        assert "Все детали BOM" not in names
        # No Error sheet (no cards_data provided)
        assert "Ошибки файлов" not in names


# ═══════════════════════════════════════════════════════════════════════
#  10. Edge cases
# ═══════════════════════════════════════════════════════════════════════

class TestReportEdgeCases:
    @pytest.fixture
    def output_dir(self) -> str:
        path = tempfile.mkdtemp(prefix="report_test_")
        yield path
        shutil.rmtree(path, ignore_errors=True)

    def test_empty_all_discrepancies(self, output_dir: str):
        """No discrepancies at all — report still generates."""
        result = _make_multi_result([])
        path = os.path.join(output_dir, "empty.xlsx")
        generate_discrepancy_report(result, path)
        assert os.path.exists(path)
        # Summary should show 0
        assert _get_xlsx_cell(path, "Сводка", 4, 1) == "0"

    def test_many_discrepancies(self, output_dir: str):
        """Large number of discrepancies still generates."""
        discs = [
            _make_discrepancy(f"P{i:03d}", DiscrepancyType.QUANTITY_MISMATCH)
            for i in range(50)
        ]
        result = _make_multi_result(discs, config_names=["C1"])
        path = os.path.join(output_dir, "many.xlsx")
        generate_discrepancy_report(result, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Расхождения"]
        data_rows = ws.max_row - 1  # exclude header
        wb.close()
        assert data_rows == 50, f"Expected 50 data rows, got {data_rows}"

    def test_many_configs(self, output_dir: str):
        """Many configs still generates summary table."""
        configs = {f"C{i}" for i in range(20)}
        discs = [_make_discrepancy("P001", config=c) for c in configs]
        result = _make_multi_result(discs, config_names=list(configs))
        path = os.path.join(output_dir, "many_configs.xlsx")
        generate_discrepancy_report(result, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Сводка"]
        found_configs = set()
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
            if row[0] in configs:
                found_configs.add(row[0])
        wb.close()
        assert found_configs == configs, f"Expected {configs}, found {found_configs}"

    def test_fuzzy_sheet_not_created_when_not_needed(self, output_dir: str):
        """Fuzzy sheet only created when fuzzy matches exist (not just any discrepancy)."""
        discs = [
            _make_discrepancy("P001", DiscrepancyType.ONLY_IN_BOM),
            _make_discrepancy("P002", DiscrepancyType.ONLY_IN_CARDS),
        ]
        result = _make_multi_result(discs)
        path = os.path.join(output_dir, "no_fuzzy.xlsx")
        generate_discrepancy_report(result, path)

        names = _get_xlsx_sheet_names(path)
        assert "Неточное совпадение номеров" not in names, \
            "Fuzzy sheet should only appear with FUZZY_MATCH discrepancies"

    def test_no_bom_or_cards(self, output_dir: str):
        """Report works without BOM or cards data."""
        result = _make_multi_result([
            _make_discrepancy("P001", DiscrepancyType.QUANTITY_MISMATCH),
        ])
        path = os.path.join(output_dir, "minimal.xlsx")
        generate_discrepancy_report(result, path, bom=None, cards_data=None)

        names = _get_xlsx_sheet_names(path)
        # Only Сводка and Расхождения should be present
        assert "Сводка" in names
        assert "Расхождения" in names
        assert "Все детали BOM" not in names
        assert "Ошибки файлов" not in names

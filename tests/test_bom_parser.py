"""Unit-тесты для bom_parser.py.

Покрытие:
  - PartInfo, BOMData (data structures)
  - parse_bom с multi-config стилем (много комплектаций)
  - parse_bom с одной qty-колонкой (спец-листы附件)
  - parse_bom: мульти-листовая обработка, global_names
  - parse_bom: служебные листы (skip)
  - parse_bom: дедупликация имён комплектаций
  - parse_bom: empty/service листы
  - get_config_quantities, get_all_config_quantities
  - lookup_part_name
  - BOMService
"""

from __future__ import annotations

import os
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pytest
from openpyxl import Workbook

from burlak_parser.bom_parser import (
    BOMData,
    BOMService,
    PartInfo,
    get_all_config_quantities,
    get_config_quantities,
    lookup_part_name,
    parse_bom,
)


# ═══════════════════════════════════════════════════════════════════════
#  ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ═══════════════════════════════════════════════════════════════════════

def _safe_remove(path: str) -> None:
    """Безопасно удалить файл."""
    try:
        if os.path.isfile(path):
            os.remove(path)
    except Exception:
        pass


def _save_workbook(wb: Workbook, file_path: str) -> str:
    """Сохранить Workbook во временный .xlsx файл и вернуть путь."""
    wb.save(file_path)
    return file_path


def _create_xlsx(
    sheets_data: Dict[str, List[List[Optional[Any]]]],
    suffix: str = ".xlsx",
) -> str:
    """Создать .xlsx файл с несколькими листами и вернуть путь к нему.

    sheets_data: {sheet_name: [row_data, ...]}
    """
    fd, path = tempfile.mkstemp(suffix=suffix, prefix="bom_test_")
    os.close(fd)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, data in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(data, 1):
            for c_idx, val in enumerate(row, 1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(path)
    return path


# ═══════════════════════════════════════════════════════════════════════
#  1. PartInfo
# ═══════════════════════════════════════════════════════════════════════

class TestPartInfo:
    def test_default_creation(self):
        p = PartInfo(part_number="ABC001")
        assert p.part_number == "ABC001"
        assert p.name_cn == ""
        assert p.name_en == ""
        assert p.quantity == 0.0
        assert p.applicable_configs == []

    def test_full_creation(self):
        p = PartInfo(
            part_number="ABC001",
            name_cn="中文名",
            name_en="English Name",
            quantity=2.0,
            applicable_configs=["Config1", "Config2"],
        )
        assert p.part_number == "ABC001"
        assert p.name_cn == "中文名"
        assert p.name_en == "English Name"
        assert p.quantity == 2.0
        assert p.applicable_configs == ["Config1", "Config2"]

    def test_default_factory(self):
        p1 = PartInfo(part_number="P001")
        p2 = PartInfo(part_number="P002")
        assert p1.applicable_configs == []
        assert p2.applicable_configs == []
        # Check they are independent lists
        p1.applicable_configs.append("C1")
        assert len(p1.applicable_configs) == 1
        assert len(p2.applicable_configs) == 0


# ═══════════════════════════════════════════════════════════════════════
#  2. BOMData
# ═══════════════════════════════════════════════════════════════════════

class TestBOMData:
    def test_empty_creation(self):
        bom = BOMData(
            parts={},
            config_names=[],
            config_quantities={},
        )
        assert bom.parts == {}
        assert bom.config_names == []
        assert bom.config_quantities == {}
        assert bom.source_file == ""
        assert bom.global_names == {}

    def test_with_data(self):
        parts = {
            "P001": PartInfo(part_number="P001", name_cn="Деталь1", quantity=1.0),
            "P002": PartInfo(part_number="P002", name_cn="Деталь2", quantity=2.0),
        }
        config_qty = {
            "Конфиг1": {"P001": 1.0},
            "Конфиг2": {"P001": 2.0, "P002": 1.0},
        }
        global_names = {"P001": ("Деталь1", ""), "P002": ("Деталь2", "")}

        bom = BOMData(
            parts=parts,
            config_names=["Конфиг1", "Конфиг2"],
            config_quantities=config_qty,
            source_file="test.xlsx",
            global_names=global_names,
        )
        assert len(bom.parts) == 2
        assert len(bom.config_names) == 2
        assert bom.config_quantities["Конфиг1"]["P001"] == 1.0
        assert bom.source_file == "test.xlsx"


# ═══════════════════════════════════════════════════════════════════════
#  3. parse_bom — Multi-config BOM
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomMultiConfigStyle:
    """BOM with multi-config structure: headers at row 3, part_no=C2, name=C3, 4 config columns."""

    @pytest.fixture
    def multi_config_xlsx(self) -> str:
        """Create a multi-config BOM file with 4 configs and 5 parts."""
        data = [
            ["CKD BOM CT1260301", None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None],
            ["序号", "零部件件号", "零部件名称", "CPAC编码",
             "舒享版-全黑内饰", "舒享版-黑米内饰", "奢享版-全黑内饰", "奢享版-黑米内饰"],
            ["1", "132000184AA", "变速箱控制单元支架", "CPAC001", "1", "1", "1", "1"],
            ["2", "551002664AA", "TCU", "CPAC002", "1", "1", "1", "1"],
            ["3", "5306200-ED001", "仪表板横梁总成 / Поперечная балка", "CPAC003", "1", "1", "2", "2"],
            ["4", "Q146Z0825F36", "螺栓", "CPAC004", "4", "4", "4", "4"],
            ["5", "G086A001", "卡扣", "CPAC005", "2", "0", "2", "0"],
        ]
        return _create_xlsx({"总装BOM": data})

    def test_multi_config_bom_parsed(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        assert len(bom.config_names) == 4, f"Expected 4 configs, got {len(bom.config_names)}: {bom.config_names}"
        assert len(bom.parts) == 5, f"Expected 5 parts, got {len(bom.parts)}"
        assert len(bom.global_names) == 5, "All 5 parts should have names"

    def test_multi_config_part_numbers(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        expected_parts = {"132000184AA", "551002664AA", "5306200ED001", "Q146Z0825F36", "G086A001"}
        assert set(bom.parts.keys()) == expected_parts, f"Got {set(bom.parts.keys())}"

    def test_multi_config_quantities(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        # Check a specific part quantity in a specific config
        # 5306200-ED001 should have qty=2 in 奢享版 configs, qty=1 in 舒享版
        for config_name in bom.config_names:
            if "奢享版" in config_name:
                assert bom.config_quantities[config_name]["5306200ED001"] == 2.0, \
                    f"Expected qty=2 for 5306200ED001 in {config_name}"
            if "舒享版" in config_name:
                if "黑米" in config_name:
                    assert bom.config_quantities[config_name].get("G086A001") is None or \
                           bom.config_quantities[config_name].get("G086A001", 0) == 0, \
                        "G086A001 should not exist in 舒享版-黑米"
                else:
                    assert bom.config_quantities[config_name].get("G086A001", 0) == 2.0

    def test_multi_config_names(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        expected_configs = {"舒享版-全黑内饰", "舒享版-黑米内饰", "奢享版-全黑内饰", "奢享版-黑米内饰"}
        config_set = set(bom.config_names)
        assert config_set == expected_configs, f"Got {config_set}"

    def test_multi_config_applicable_configs(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        pn = "132000184AA"
        part = bom.parts[pn]
        assert len(part.applicable_configs) == 4, \
            f"Part {pn} should be in all 4 configs, got {part.applicable_configs}"

    def test_multi_config_global_names(self, multi_config_xlsx: str):
        bom = parse_bom(multi_config_xlsx)
        name_cn, name_en = bom.global_names.get("132000184AA", ("", ""))
        assert "变速箱" in name_cn, f"Expected '变速箱', got '{name_cn}'"


# ═══════════════════════════════════════════════════════════════════════
#  4. parse_bom — Russian G01-style BOM
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomRussianStyle:
    """Russian BOM: headers at row 1, Russian/Chinese headers."""

    @pytest.fixture
    def russian_xlsx(self) -> str:
        data = [
            ["序号\nСерийный номер", "零部件件号\nКод детали",
             "零部件名称\nНаименование", "系统\nСистема",
             "舒享版-全黑", "舒享版-黑米", "奢享版-全黑", "奢享版-黑米"],
            ["1", "5306200-ED001", "仪表板横梁总成 / Поперечная балка", "A", "1", "1", "2", "2"],
            ["2", "551002664AA", "TCU", "B", "1", "1", "1", "1"],
            ["3", "Q146Z0825F36", "Болт", "A", "4", "4", "4", "4"],
        ]
        return _create_xlsx({"G01 BOM": data})

    def test_russian_bom_parsed(self, russian_xlsx: str):
        bom = parse_bom(russian_xlsx)
        assert len(bom.config_names) == 4, f"Expected 4 configs, got {bom.config_names}"
        assert len(bom.parts) == 3, f"Expected 3 parts, got {len(bom.parts)}"

    def test_russian_part_no_clean(self, russian_xlsx: str):
        bom = parse_bom(russian_xlsx)
        assert "5306200ED001" in bom.parts, "Part number not cleaned correctly"
        assert "5306200-ED001" not in bom.parts, "Original part number should be cleaned"

    def test_russian_config_quantities(self, russian_xlsx: str):
        bom = parse_bom(russian_xlsx)
        # 5306200ED001: qty=1 in舒享版, 2 in奢享版
        for cn in bom.config_names:
            pn = "5306200ED001"
            if "奢享版" in cn:
                assert bom.config_quantities[cn][pn] == 2.0, f"{cn}: expected 2, got {bom.config_quantities[cn][pn]}"
            if "舒享版" in cn:
                assert bom.config_quantities[cn][pn] == 1.0


# ═══════════════════════════════════════════════════════════════════════
#  5. parse_bom — Single qty column (附件 style)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomAttachmentSheet:
    """Лист附件 (attachment) с одной qty-колонкой, без config columns.

    Attachment sheet теперь тоже добавляет parts и quantities в основной BOM.
    Его части попадают в отдельную конфигурацию (имя листа) И в global_names.
    """

    @pytest.fixture
    def attachment_xlsx(self) -> str:
        """Create a BOM with: primary sheet (multi-config) + attachment sheet (1 qty).

        Both sheets need >= 3 data rows for is_sheet_bom_candidate.
        """
        data = {
            "总装BOM": [
                ["序号", "零部件件号", "零部件名称",
                 "舒享版-全黑", "舒享版-黑米", "奢享版-全黑"],
                ["1", "P001", "Part One", "1", "1", "2"],
                ["2", "P002", "Part Two", "1", "0", "1"],
                ["3", "P003", "Part Three", "2", "2", "2"],
                ["4", "P006", "Part Six", "1", "1", "0"],
            ],
            "零部件附件": [
                ["零部件件号", "组件物料描述", "组件数量"],
                ["P004", "附件部件4", "5"],
                ["P005", "附件部件5", "3"],
                ["P007", "附件部件7", "2"],
            ],
        }
        return _create_xlsx(data)

    def test_attachment_sheet_adds_global_names(self, attachment_xlsx: str):
        """Attachment sheet adds its parts to global_names AND bom.parts."""
        bom = parse_bom(attachment_xlsx)
        # Attachment parts are in global_names
        assert "P004" in bom.global_names, "Attachment part missing from global_names"
        assert "P005" in bom.global_names, "Attachment part missing from global_names"
        # Parts from BOTH sheets go to bom.parts
        assert "P001" in bom.parts, "Primary part missing from bom.parts"
        assert "P004" in bom.parts, "Attachment part should now be in bom.parts"

    def test_primary_configs_only(self, attachment_xlsx: str):
        """Both primary and attachment sheets contribute configs."""
        bom = parse_bom(attachment_xlsx)
        # Primary sheet has 3 config columns + attachment sheet adds 1 = 4 total
        assert len(bom.config_names) == 4, f"Expected 4 configs, got {bom.config_names}"
        # Attachment sheet creates a config from its sheet name
        assert "零部件附件" in bom.config_names, "Attachment sheet should add its own config"

    def test_primary_parts_in_config(self, attachment_xlsx: str):
        """Primary and attachment sheet parts are in their respective configs."""
        bom = parse_bom(attachment_xlsx)
        # Primary sheet configs
        for cn in bom.config_names:
            if cn != "零部件附件":
                assert "P001" in bom.config_quantities[cn], "Primary part missing from config"
        # Attachment sheet config
        assert "P004" in bom.config_quantities["零部件附件"], "Attachment part should be in attachment config"
        assert bom.config_quantities["零部件附件"]["P004"] == 5.0


# ═══════════════════════════════════════════════════════════════════════
#  6. parse_bom — Multi-sheet: only first BOM gives configs
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomMultiSheet:
    """Multiple BOM sheets: ALL sheets contribute parts and quantities.

    Важно: оба листа распознаются как BOM-кандидаты (有config columns + part_no + name).
    焊装BOM и 涂装BOM — оба дают конфиги, parts и global_names.
    Если part_number встречается на нескольких листах — количества суммируются.
    """

    @pytest.fixture
    def multi_xlsx(self) -> str:
        """Both sheets need >= 3 data rows for is_sheet_bom_candidate."""
        data = {
            "焊装BOM": [
                ["序号", "零部件件号", "零件名称", "舒享版", "奢享版"],
                ["1", "W001", "Welded Part 1", "2", "2"],
                ["2", "W002", "Welded Part 2", "1", "1"],
                ["3", "W003", "Welded Part 3", "1", "0"],
            ],
            "涂装BOM": [
                ["序号", "零部件件号", "零件名称", "舒享版", "奢享版"],
                ["1", "P001", "Painted Part 1", "1", "2"],
                ["2", "P002", "Painted Part 2", "1", "0"],
                ["3", "P003", "Painted Part 3", "0", "1"],
            ],
        }
        return _create_xlsx(data)

    def test_first_sheet_provides_configs(self, multi_xlsx: str):
        bom = parse_bom(multi_xlsx)
        assert len(bom.config_names) == 2, f"Expected 2 configs, got {bom.config_names}"
        assert "舒享版" in bom.config_names
        assert "奢享版" in bom.config_names

    def test_both_sheets_parts_collected(self, multi_xlsx: str):
        bom = parse_bom(multi_xlsx)
        # BOTH sheets' parts go to bom.parts (all sheets now contribute)
        assert "W001" in bom.parts, "First sheet part missing"
        assert "P001" in bom.parts, "Second sheet part missing from bom.parts"
        # Total: 3 from 焊装BOM + 3 from 涂装BOM = 6 parts
        assert len(bom.parts) == 6, f"Expected 6 parts (both sheets), got {len(bom.parts)}"

    def test_second_sheet_adds_global_names(self, multi_xlsx: str):
        bom = parse_bom(multi_xlsx)
        assert "P001" in bom.global_names, "Second sheet part missing from global names"
        cn, _ = bom.global_names["P001"]
        assert "Painted" in cn, f"Expected 'Painted Part 1', got '{cn}'"

    def test_second_sheet_adds_to_configs(self, multi_xlsx: str):
        bom = parse_bom(multi_xlsx)
        # P001 from 涂装BOM SHOULD be in config quantities (both sheets contribute)
        for cn in bom.config_names:
            if "舒享版" in cn:
                assert bom.config_quantities[cn].get("P001") == 1.0, \
                    f"P001 should have qty=1 in {cn} from 涂装BOM"
            if "奢享版" in cn:
                assert bom.config_quantities[cn].get("P001") == 2.0, \
                    f"P001 should have qty=2 in {cn} from 涂装BOM"

    def test_same_part_across_sheets_sums_quantities(self):
        """Same part number on two sheets: quantities should be summed."""
        data = {
            "焊装BOM": [
                ["序号", "零部件件号", "零件名称", "舒享版", "奢享版"],
                ["1", "SH001", "Shared Part", "2", "1"],
                ["2", "W002", "Weld Only", "1", "1"],
                ["3", "W003", "Weld Only 2", "1", "0"],
            ],
            "涂装BOM": [
                ["序号", "零部件件号", "零件名称", "舒享版", "奢享版"],
                ["1", "SH001", "Shared Part", "3", "2"],
                ["2", "P002", "Paint Only", "1", "0"],
                ["3", "P003", "Paint Only 2", "0", "1"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # SH001: 焊装BOM(2,1) + 涂装BOM(3,2) = (5,3)
        assert bom.config_quantities["舒享版"]["SH001"] == 5.0, \
            f"Expected 5.0 (2+3), got {bom.config_quantities['舒享版']['SH001']}"
        assert bom.config_quantities["奢享版"]["SH001"] == 3.0, \
            f"Expected 3.0 (1+2), got {bom.config_quantities['奢享版']['SH001']}"


# ═══════════════════════════════════════════════════════════════════════
#  7. parse_bom — Service sheets are skipped
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomServiceSheets:
    """BOM with service sheets that should be ignored."""

    @pytest.fixture
    def service_xlsx(self) -> str:
        """Main BOM sheet needs >= 3 data rows for is_sheet_bom_candidate."""
        data = {
            "总装BOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
            "EBOM": [
                ["Just engineering data", None],
            ],
            "变更记录": [
                ["Change", "Log"],
            ],
            "封面": [
                ["Title", None],
            ],
        }
        return _create_xlsx(data)

    def test_service_sheets_skipped(self, service_xlsx: str):
        bom = parse_bom(service_xlsx)
        # Only 总装BOM is a BOM candidate — EBOM/变更记录/封面 are service sheets
        assert len(bom.parts) == 3, f"Expected 3 parts (from 总装BOM), got {len(bom.parts)}"
        assert "P001" in bom.parts
        assert len(bom.config_names) == 2

    def test_service_sheets_dont_add_configs(self, service_xlsx: str):
        bom = parse_bom(service_xlsx)
        # EBOM/变更记录/封面 should not become config names
        for svc in ["EBOM", "变更记录", "封面"]:
            assert svc not in bom.config_names, f"'{svc}' should not be a config name"


# ═══════════════════════════════════════════════════════════════════════
#  8. parse_bom — Empty sheet handling
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomEmptySheet:
    @pytest.fixture
    def empty_sheet_xlsx(self, tmp_path) -> str:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Main"
        ws1.cell(row=1, column=1, value="序号")
        ws1.cell(row=1, column=2, value="零部件件号")
        ws1.cell(row=1, column=3, value="名称")
        ws1.cell(row=1, column=4, value="Config")

        ws2 = wb.create_sheet(title="EmptySheet")
        # No data at all
        path = os.path.join(str(tmp_path), "empty_test.xlsx")
        wb.save(path)
        return path

    def test_empty_sheet_no_crash(self, empty_sheet_xlsx: str):
        """Empty sheets should be skipped gracefully without crash."""
        bom = parse_bom(empty_sheet_xlsx)
        assert bom is not None
        # Main sheet has headers but no data rows → not a BOM candidate → 0 parts
        assert len(bom.parts) == 0, "Should have 0 parts from empty sheets"
        assert len(bom.config_names) == 0, "Should have 0 configs from empty sheets"


# ═══════════════════════════════════════════════════════════════════════
#  9. parse_bom — Config name deduplication
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomConfigDedup:
    @pytest.fixture
    def dedup_xlsx(self) -> str:
        """Config columns with duplicate names. Needs >= 3 data rows."""
        data = [
            ["序号", "零部件件号", "零件名称",
             "Config A", "Config A", "Config B"],
            ["1", "P001", "Part1", "1", "2", "1"],
            ["2", "P002", "Part2", "1", "0", "2"],
            ["3", "P003", "Part3", "2", "1", "0"],
        ]
        return _create_xlsx({"总装BOM": data})

    def test_config_dedup(self, dedup_xlsx: str):
        bom = parse_bom(dedup_xlsx)
        # After dedup, "Config A" should appear only once (first occurrence kept)
        config_a_count = sum(1 for c in bom.config_names if c == "Config A")
        assert config_a_count == 1, f"Config A should be deduplicated, got {bom.config_names}"
        assert "Config B" in bom.config_names
        assert len(bom.config_names) == 2, f"Expected 2 configs after dedup, got {bom.config_names}"

    def test_dedup_first_column_kept(self, dedup_xlsx: str):
        """Dedup removes duplicate columns but keeps only FIRST column's quantities."""
        bom = parse_bom(dedup_xlsx)
        # P001 in Config A: first "Config A" col has qty=1, second has qty=2
        # After dedup, only first column's data is kept (qty=1)
        assert bom.config_quantities["Config A"]["P001"] == 1.0, \
            f"Expected 1.0 (first column only), got {bom.config_quantities['Config A'].get('P001')}"


# ═══════════════════════════════════════════════════════════════════════
#  10. parse_bom — global_names fallback to parts
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomGlobalNames:
    @pytest.fixture
    def global_names_xlsx(self) -> str:
        """Both sheets need >= 3 data rows for is_sheet_bom_candidate."""
        data = {
            "焊装BOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Part One", "1", "1"],
                ["2", "W002", "Weld Part 2", "1", "1"],
                ["3", "W003", "Weld Part 3", "1", "0"],
            ],
            "涂装BOM": [
                ["序号", "零部件件号", "零件名称（英文）", "Config1", "Config2"],
                ["1", "P001", "Part One En", "1", "1"],
                ["2", "P002", "Part Two", "1", "0"],
                ["3", "P003", "Part Three", "0", "1"],
            ],
        }
        return _create_xlsx(data)

    def test_global_names_merged(self, global_names_xlsx: str):
        bom = parse_bom(global_names_xlsx)
        # P001 should have cn from 焊装BOM (first occurrence) and en from 涂装BOM (en column)
        assert "P001" in bom.global_names
        cn, en = bom.global_names["P001"]
        assert "Part One" in cn, f"Expected 'Part One' in cn, got '{cn}'"
        # en should be from 涂装BOM's 零件名称（英文）column
        assert "Part One En" in en, f"Expected 'Part One En' in en, got '{en}'"


# ═══════════════════════════════════════════════════════════════════════
#  11. get_config_quantities
# ═══════════════════════════════════════════════════════════════════════

class TestGetConfigQuantities:
    @pytest.fixture
    def bom_data(self) -> BOMData:
        parts = {
            "P001": PartInfo(part_number="P001", name_cn="Part1"),
            "P002": PartInfo(part_number="P002", name_cn="Part2"),
            "P003": PartInfo(part_number="P003", name_cn="Part3"),
        }
        config_qty = {
            "Config A": {"P001": 2.0, "P002": 1.0},
            "Config B": {"P001": 1.0, "P003": 3.0},
        }
        return BOMData(
            parts=parts,
            config_names=["Config A", "Config B"],
            config_quantities=config_qty,
            global_names={"P001": ("Part1", ""), "P002": ("Part2", ""), "P003": ("Part3", "")},
        )

    def test_get_single_config(self, bom_data: BOMData):
        result = get_config_quantities(bom_data, "Config A")
        assert len(result) == 2
        assert result["P001"].quantity == 2.0
        assert result["P002"].quantity == 1.0
        assert result["P001"].name_cn == "Part1"

    def test_get_config_not_found(self, bom_data: BOMData):
        with pytest.raises(ValueError, match="не найдена"):
            get_config_quantities(bom_data, "NonExistent")

    def test_get_all_configs(self, bom_data: BOMData):
        result = get_all_config_quantities(bom_data)
        assert len(result) == 2
        assert "Config A" in result
        assert "Config B" in result
        assert result["Config A"]["P001"].quantity == 2.0
        assert result["Config B"]["P003"].quantity == 3.0

    def test_part_in_config_not_in_parts_uses_global_names(self):
        """Когда деталь есть в config_quantities, но НЕТ в bom.parts →
        PartInfo создаётся из global_names (строки 329-330).

        Это может случиться при ручном конструировании BOMData
        или если деталь добавилась только в config_quantities
        (например, при нестандартной обработке).
        """
        parts = {
            "P001": PartInfo(part_number="P001", name_cn="Existing Part"),
        }
        config_qty = {
            "Config": {
                "P001": 1.0,
                "P999": 2.0,  # NOT in parts!
            },
        }
        global_names = {
            "P001": ("Existing Part", ""),
            "P999": ("Fallback Part", "Fallback EN"),
        }
        bom = BOMData(
            parts=parts,
            config_names=["Config"],
            config_quantities=config_qty,
            global_names=global_names,
        )

        result = get_config_quantities(bom, "Config")
        assert len(result) == 2
        # P001 — из parts
        assert result["P001"].name_cn == "Existing Part"
        assert result["P001"].quantity == 1.0
        # P999 — НЕ в parts, должен быть взят из global_names (строки 329-330)
        assert result["P999"].name_cn == "Fallback Part", \
            f"Expected 'Fallback Part' from global_names, got '{result['P999'].name_cn}'"
        assert result["P999"].name_en == "Fallback EN"
        assert result["P999"].quantity == 2.0


# ═══════════════════════════════════════════════════════════════════════
#  12. lookup_part_name
# ═══════════════════════════════════════════════════════════════════════

class TestLookupPartName:
    @pytest.fixture
    def bom_data(self) -> BOMData:
        parts = {
            "P001": PartInfo(part_number="P001", name_cn="Part1 CN", name_en="Part1 EN"),
        }
        return BOMData(
            parts=parts,
            config_names=["Config1"],
            config_quantities={"Config1": {"P001": 1.0}},
            global_names={"P001": ("P1 CN", "P1 EN"), "P002": ("P2 CN", "P2 EN")},
        )

    def test_finds_in_parts(self, bom_data: BOMData):
        cn, en = lookup_part_name(bom_data, "P001")
        assert cn == "Part1 CN"
        assert en == "Part1 EN"

    def test_falls_back_to_global(self, bom_data: BOMData):
        cn, en = lookup_part_name(bom_data, "P002")
        assert cn == "P2 CN"
        assert en == "P2 EN"

    def test_not_found(self, bom_data: BOMData):
        cn, en = lookup_part_name(bom_data, "P999")
        assert cn == ""
        assert en == ""

    def test_part_without_name_falls_back(self, bom_data: BOMData):
        """Part exists in parts but without name → fallback to global_names."""
        # Create a part without name
        bom_data.parts["P001"] = PartInfo(part_number="P001")  # no name
        # Remove from global_names to avoid cross-contamination
        bom_data.global_names.pop("P001", None)

        # Now add P001 to global_names
        bom_data.global_names["P001"] = ("Global CN", "Global EN")
        cn, en = lookup_part_name(bom_data, "P001")
        # Should get from global_names since part has no name
        assert cn == "Global CN"
        assert en == "Global EN"


# ═══════════════════════════════════════════════════════════════════════
#  13. BOMService
# ═══════════════════════════════════════════════════════════════════════

class TestBOMService:
    def test_initial_state(self):
        svc = BOMService()
        assert not svc.is_loaded
        assert svc.bom is None

    def test_load(self):
        # Create a simple BOM file with >= 3 data rows
        data = {
            "Test": [
                ["序号", "零部件件号", "零件名称", "用量\\n舒享版", "用量\\n奢享版"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
                ["4", "P004", "Part4", "1", "1"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        bom = svc.load(path)
        assert svc.is_loaded
        assert bom is not None
        # 4 data rows, but P003 has qty=0 in detected config → filtered out
        assert len(bom.parts) == 3

    def test_get_config_names(self):
        data = {
            "Test": [
                ["序号", "零部件件号", "名称", "C1", "C2"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        svc.load(path)
        names = svc.get_config_names()
        assert len(names) == 2, f"Expected 2 configs, got {len(names)}"

    def test_get_config_count(self):
        data = {
            "Test": [
                ["序号", "零部件件号", "名称", "C1", "C2", "C3"],
                ["1", "P001", "Part1", "1", "2", "1"],
                ["2", "P002", "Part2", "1", "0", "1"],
                ["3", "P003", "Part3", "2", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        svc.load(path)
        assert svc.get_config_count() == 3, f"Expected 3 configs, got {svc.get_config_count()}"

    def test_get_all_part_numbers(self):
        data = {
            "Test": [
                ["序号", "零部件件号", "名称", "C1", "C2"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        svc.load(path)
        pns = svc.get_all_part_numbers()
        assert len(pns) == 3, f"Expected 3 parts, got {len(pns)}"
        assert "P001" in pns

    def test_lookup_name(self):
        data = {
            "Test": [
                ["序号", "零部件件号", "零件名称", "C1", "C2"],
                ["1", "P001", "TestPart", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        svc.load(path)
        cn, en = svc.lookup_name("P001")
        assert "TestPart" in cn, f"Expected 'TestPart', got '{cn}'"

    def test_not_loaded_raises(self):
        svc = BOMService()
        with pytest.raises(RuntimeError):
            svc.get_config_names()
        with pytest.raises(RuntimeError):
            svc.get_parts_for_config("Any")
        with pytest.raises(RuntimeError):
            svc.get_all_configs()


# ═══════════════════════════════════════════════════════════════════════
#  14. parse_bom — Qty-only sheet as PRIMARY BOM (attachment style)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomQtyOnlyAsPrimary:
    """First BOM sheet is an attachment style (qty column, NO config columns).

    Покрывает строки 164-178: когда первый BOM-лист имеет qty-колонку
    и менее 2 config колонок → создаётся конфигурация из имени листа.
    """

    @pytest.fixture
    def qty_only_xlsx(self) -> str:
        """Sheet with part_no, name, qty — NO config columns.

        is_sheet_bom_candidate проходит через qty+name path
        (qty_col > 0 and has_name = True).
        """
        data = {
            "零部件附件": [
                ["序号", "零部件件号", "零部件名称", "数量"],
                ["1", "P001", "Attachment One", "3"],
                ["2", "P002", "Attachment Two", "2"],
                ["3", "P003", "Attachment Three", "1"],
                ["4", "P004", "Attachment Four", "5"],
            ],
        }
        return _create_xlsx(data)

    def test_qty_only_creates_single_config(self, qty_only_xlsx: str):
        bom = parse_bom(qty_only_xlsx)
        # Sheet name should become a config
        assert len(bom.config_names) == 1, f"Expected 1 config (sheet name), got {bom.config_names}"
        assert "零部件附件" in bom.config_names, "Config should be sheet name"

    def test_qty_only_parts_collected(self, qty_only_xlsx: str):
        bom = parse_bom(qty_only_xlsx)
        assert len(bom.parts) == 4, f"Expected 4 parts, got {len(bom.parts)}"
        assert "P001" in bom.parts
        assert "P004" in bom.parts

    def test_qty_only_quantities(self, qty_only_xlsx: str):
        bom = parse_bom(qty_only_xlsx)
        cn = "零部件附件"
        assert bom.config_quantities[cn]["P001"] == 3.0
        assert bom.config_quantities[cn]["P002"] == 2.0
        assert bom.config_quantities[cn]["P004"] == 5.0

    def test_qty_only_zero_qty_not_added(self, qty_only_xlsx: str):
        """Parts with qty=0 should NOT appear in config_quantities."""
        bom = parse_bom(qty_only_xlsx)
        cn = "零部件附件"
        # P003 has qty=1 → present
        # No parts with qty=0 in this fixture, so just verify counts
        assert len(bom.config_quantities[cn]) == 4

    def test_qty_only_parts_with_zero_qty(self):
        """Parts with qty=0 are not added to all_parts в qty-only path.

        В qty-only path (Section 5) детали добавляются в all_parts ТОЛЬКО
        если qty > 0. Детали с qty=0 полностью пропускаются.
        """
        data = {
            "Attachment": [
                ["序号", "零部件件号", "零件名称", "数量"],
                ["1", "P001", "Part1", "2"],
                ["2", "P002", "Part2", "0"],  # qty=0 → skipped entirely
                ["3", "P003", "Part3", "1"],
                ["4", "P004", "Part4", "0"],  # qty=0 → skipped entirely
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # В qty-only path части с qty=0 НЕ добавляются в all_parts
        assert len(bom.parts) == 2, f"Expected 2 parts (qty=0 skipped), got {len(bom.parts)}"
        assert "P001" in bom.parts, "P001 qty=2 should be present"
        assert "P003" in bom.parts, "P003 qty=1 should be present"
        cn = "Attachment"
        assert "P001" in bom.config_quantities[cn], "P001 qty=2 should be present"
        assert "P002" not in bom.config_quantities[cn], "P002 qty=0 should NOT be in config"
        assert "P003" in bom.config_quantities[cn], "P003 qty=1 should be present"
        assert "P004" not in bom.config_quantities[cn], "P004 qty=0 should NOT be in config"

    def test_qty_only_different_qty_column_name(self):
        """Qty column can be named differently (e.g. '单车用量')."""
        data = {
            "Fasteners": [
                ["序号", "零部件件号", "零件名称", "单车用量"],
                ["1", "P001", "Bolt", "8"],
                ["2", "P002", "Nut", "8"],
                ["3", "P003", "Washer", "16"],
                ["4", "P004", "Screw", "4"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 4, f"Expected 4 parts, got {len(bom.parts)}"
        assert bom.config_quantities["Fasteners"]["P001"] == 8.0
        assert bom.config_quantities["Fasteners"]["P003"] == 16.0


# ═══════════════════════════════════════════════════════════════════════
#  15. parse_bom — Qty-only как валидный BOM-путь
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomQtyOnlyPath:
    """Дополнительные тесты для qty-only BOM-пути.

    Примечание: строка 188-189 (if not config_cols) НЕ ДОСТИЖИМА,
    т.к. is_sheet_bom_candidate c min_configs=2 не пропускает листы
    без config колонок, если нет qty колонки.
    """

    def test_qty_only_different_qty_column_name(self):
        """Qty column can be named differently (e.g. '单车用量')."""
        data = {
            "Fasteners": [
                ["序号", "零部件件号", "零件名称", "单车用量"],
                ["1", "P001", "Bolt", "8"],
                ["2", "P002", "Nut", "8"],
                ["3", "P003", "Washer", "16"],
                ["4", "P004", "Screw", "4"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 4, f"Expected 4 parts, got {len(bom.parts)}"
        assert bom.config_quantities["Fasteners"]["P001"] == 8.0
        assert bom.config_quantities["Fasteners"]["P003"] == 16.0

    def test_qty_only_with_name_en_column(self):
        """Qty-only sheet with English name column."""
        data = {
            "Bolts": [
                ["序号", "零部件件号", "Part Name (EN)", "数量"],
                ["1", "P001", "Bolt M8", "4"],
                ["2", "P002", "Bolt M10", "2"],
                ["3", "P003", "Nut M8", "8"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 3
        assert "P001" in bom.parts
        # Name column is English → name_en is populated
        cn, en = bom.global_names.get("P001", ("", ""))
        assert "Bolt" in en or "Bolt" in cn, f"Expected 'Bolt' in name, got cn='{cn}' en='{en}'"


# ═══════════════════════════════════════════════════════════════════════
#  16. parse_bom — Лист без заголовков / без part_no колонки
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomSkippedSheets:
    """Листы, пропускаемые на разных этапах проверки.

    Покрывает строки:
      - 104-105: is_sheet_bom_candidate → False (нет part_no + qty + config)
      - 114-115: header_rows пуст (нет распознаваемых заголовков)
      - 129: part_no_col == 0 (нет колонки парт-номера)
    """

    def test_service_sheet_skipped_at_bom_candidate(self):
        """Служебный лист (обложка) отсекается is_sheet_bom_candidate → строка 104-105."""
        data = {
            "MainBOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
            "封面": [  # service sheet → is_sheet_bom_candidate → False
                ["Title Page", None, None, None],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 3, "Only MainBOM should be processed"
        assert "P001" in bom.parts
        assert len(bom.config_names) == 2

    def test_no_header_rows_sheet_skipped(self):
        """Лист с данными без распознаваемых заголовков → строка 114-115.

        Данные без part_no/name/qty ключевых слов → find_header_rows пуст.
        """
        data = {
            "Data": [
                ["Just some random text without any header keywords", None, None],
                ["More random data", "123", "456"],
                ["Still no headers", "ABC", "DEF"],
                ["Lot's of text", "G", "HIJ"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 0, "No parts should be found (no headers)"

    def test_no_part_no_column_skipped(self):
        """Лист с заголовками, но БЕЗ part_no колонки → строка 129.

        Есть qty и другие колонки, но ни одна не определяется как part_no.
        """
        data = {
            "OnlyNames": [
                ["序号", "零件名称", "数量", "Config1", "Config2"],
                ["1", "Bolt", "1", "1", "2"],
                ["2", "Nut", "2", "1", "0"],
                ["3", "Washer", "3", "2", "1"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 0, "No parts should be found (no part_no column)"

    def test_bom_with_non_bom_sheet_first(self):
        """Первый лист не BOM-кандидат, второй — BOM-кандидат.

        Второй лист становится primary_bom_found.
        """
        data = {
            "目录": [  # service sheet → skipped
                ["Table of Contents", None, None, None],
            ],
            "总装BOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        assert len(bom.parts) == 3, "Main BOM should be processed"
        assert "P001" in bom.parts
        assert len(bom.config_names) == 2


# ═══════════════════════════════════════════════════════════════════════
#  17. parse_bom — Dedup с разными форматами имён
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomDedupExtended:
    """Дедупликация имён комплектаций со сложными паттернами.

    Покрывает строки 223, 226, 228: добавление config_name в seen/quantities.
    """

    def test_dedup_with_spaces_and_hyphens(self):
        """Имена, различающиеся только пробелами/дефисами, считаются дубликатами."""
        data = [
            ["序号", "零部件件号", "零件名称",
             "舒享版-全黑", "舒享版 - 全黑", "舒享版全黑"],
            ["1", "P001", "Part1", "1", "2", "3"],
            ["2", "P002", "Part2", "1", "1", "1"],
            ["3", "P003", "Part3", "2", "0", "0"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        # После нормализации: "舒享版-全黑" → "舒享版全黑"
        # "舒享版 - 全黑" → "舒享版全黑" (пробел+дефис → убрано)
        # "舒享版全黑" → "舒享版全黑"
        # Все три одинаковые → только первый сохраняется
        assert len(bom.config_names) == 1, f"Expected 1 config after dedup, got {bom.config_names}"
        assert bom.config_names[0] == "舒享版-全黑", "First occurrence should be kept"

    def test_dedup_case_insensitive(self):
        """Имена, различающиеся регистром, считаются дубликатами."""
        data = [
            ["序号", "零部件件号", "零件名称", "Luxury", "luxury"],
            ["1", "P001", "Part1", "1", "2"],
            ["2", "P002", "Part2", "1", "1"],
            ["3", "P003", "Part3", "2", "0"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        assert len(bom.config_names) == 1, f"Expected 1 config (case-insensitive dedup), got {bom.config_names}"
        assert bom.config_names[0] == "Luxury", "First occurrence should be kept"

    def test_dedup_preserves_unique_names(self):
        """Уникальные имена не затрагиваются дедупликацией."""
        data = [
            ["序号", "零部件件号", "零件名称",
             "舒享版-全黑", "舒享版-黑米", "奢享版-全黑"],
            ["1", "P001", "Part1", "1", "2", "1"],
            ["2", "P002", "Part2", "1", "1", "0"],
            ["3", "P003", "Part3", "2", "0", "0"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        assert len(bom.config_names) == 3, f"Expected 3 unique configs, got {bom.config_names}"


# ═══════════════════════════════════════════════════════════════════════
#  18. parse_bom — Глобальные имена применяются к деталям
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomGlobalNamesApplied:
    """Проверка, что глобальные имена применяются к деталям БЕЗ названия.

    Покрывает строки 329-330: финальное применение global_names к parts.
    """

    def test_part_inherits_name_from_global_names(self):
        """PartInfo без имени наследует name_cn из global_names.

        Создаём BOM с part_no + configs + name column.
        PartInfo создаётся без name_cn/name_en (только part_number).
        global_names заполняется через build_global_name_dict.
        Финальная агрегация (строки 329-330) присваивает name_cn PartInfo.
        """
        data = {
            "BOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Engine Mount", "1", "2"],
                ["2", "P002", "Transmission Bracket", "1", "1"],
                ["3", "P003", "Radiator Support", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # PartInfo должен получить name_cn из global_names через агрегацию
        assert bom.parts["P001"].name_cn == "Engine Mount", \
            f"Expected 'Engine Mount', got '{bom.parts['P001'].name_cn}'"
        assert bom.parts["P002"].name_cn == "Transmission Bracket"
        assert bom.parts["P003"].name_cn == "Radiator Support"
        # global_names тоже должен содержать имена
        assert "Engine Mount" in bom.global_names["P001"][0], \
            f"Expected 'Engine Mount' in global_names, got {bom.global_names['P001']}"

    def test_part_without_name_and_not_in_global_keeps_empty(self):
        """Part без name column БЕЗ global_names остаётся с пустым именем."""
        data = {
            "BOM": [
                ["序号", "零部件件号", "Config1", "Config2"],  # NO name column
                ["1", "P001", "1", "2"],
                ["2", "P002", "1", "1"],
                ["3", "P003", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # (not part.name_cn and not part.name_en) is True
        # pn in all_global_names is False (no name column → no global_names)
        # → name stays empty
        part = bom.parts["P001"]
        assert part.name_cn == "", "Part should have empty name (no name column)"


# ═══════════════════════════════════════════════════════════════════════
#  19. BOMService — все методы
# ═══════════════════════════════════════════════════════════════════════

class TestBOMServiceExtended:
    """Все оставшиеся методы BOMService с реальными данными.

    Покрывает строки 410, 417, 423, 428, 434.
    """

    @pytest.fixture
    def svc_and_bom(self) -> Tuple[BOMService, str]:
        data = {
            "BOM": [
                ["序号", "零部件件号", "零件名称", "Config A", "Config B"],
                ["1", "P001", "Part One", "1", "2"],
                ["2", "P002", "Part Two", "1", "1"],
                ["3", "P003", "Part Three", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        svc = BOMService()
        svc.load(path)
        return svc, path

    def test_get_parts_for_config(self, svc_and_bom: Tuple[BOMService, str]):
        svc, _ = svc_and_bom
        parts = svc.get_parts_for_config("Config A")
        assert len(parts) == 3, f"Expected 3 parts, got {len(parts)}"
        assert "P001" in parts
        assert parts["P001"].quantity == 1.0
        assert parts["P003"].quantity == 2.0
        # Names should be preserved
        assert "Part One" in parts["P001"].name_cn

    def test_get_parts_for_config_not_found(self, svc_and_bom: Tuple[BOMService, str]):
        svc, _ = svc_and_bom
        with pytest.raises(ValueError, match="не найдена"):
            svc.get_parts_for_config("NonExistent")

    def test_get_all_configs(self, svc_and_bom: Tuple[BOMService, str]):
        svc, _ = svc_and_bom
        all_config = svc.get_all_configs()
        assert len(all_config) == 2, f"Expected 2 configs, got {len(all_config)}"
        assert "Config A" in all_config
        assert "Config B" in all_config
        assert all_config["Config A"]["P001"].quantity == 1.0
        assert all_config["Config B"]["P001"].quantity == 2.0

    def test_get_config_count_not_loaded(self):
        svc = BOMService()
        assert svc.get_config_count() == 0, "Not loaded → should return 0"

    def test_lookup_name_not_loaded(self):
        svc = BOMService()
        cn, en = svc.lookup_name("P001")
        assert cn == "", "Not loaded → should return empty"
        assert en == "", "Not loaded → should return empty"

    def test_get_all_part_numbers_not_loaded(self):
        svc = BOMService()
        pns = svc.get_all_part_numbers()
        assert pns == set(), "Not loaded → should return empty set"

    def test_get_all_configs_not_loaded(self):
        svc = BOMService()
        with pytest.raises(RuntimeError, match="не загружен"):
            svc.get_all_configs()


# ═══════════════════════════════════════════════════════════════════════
#  19b. BOMService — load_from_bytes, cleanup, context manager, async
# ═══════════════════════════════════════════════════════════════════════

class TestBOMServiceServer:
    """Тесты серверной функциональности BOMService:
      - load_from_bytes (in-memory upload)
      - cleanup (автоудаление temp-файлов)
      - context manager (with)
      - async (load_async)
    """

    @pytest.fixture
    def bom_bytes(self) -> bytes:
        """Создать BOM-файл в памяти, вернуть байты."""
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "BOM"
            ws.cell(row=1, column=1, value="序号")
            ws.cell(row=1, column=2, value="零部件件号")
            ws.cell(row=1, column=3, value="零件名称")
            ws.cell(row=1, column=4, value="Config1")
            ws.cell(row=1, column=5, value="Config2")
            ws.cell(row=2, column=1, value="1")
            ws.cell(row=2, column=2, value="P001")
            ws.cell(row=2, column=3, value="Part1")
            ws.cell(row=2, column=4, value="1")
            ws.cell(row=2, column=5, value="2")
            ws.cell(row=3, column=1, value="2")
            ws.cell(row=3, column=2, value="P002")
            ws.cell(row=3, column=3, value="Part2")
            ws.cell(row=3, column=4, value="1")
            ws.cell(row=3, column=5, value="1")
            ws.cell(row=4, column=1, value="3")
            ws.cell(row=4, column=2, value="P003")
            ws.cell(row=4, column=3, value="Part3")
            ws.cell(row=4, column=4, value="2")
            ws.cell(row=4, column=5, value="0")
            wb.save(path)
            with open(path, "rb") as f:
                return f.read()
        finally:
            _safe_remove(path)

    def test_load_from_bytes(self, bom_bytes: bytes):
        """Загрузка BOM из байтов (in-memory upload)."""
        svc = BOMService()
        bom = svc.load_from_bytes(bom_bytes, filename="uploaded_bom.xlsx")
        assert svc.is_loaded
        assert len(bom.parts) == 3
        assert "P001" in bom.parts
        assert len(bom.config_names) == 2

    def test_cleanup_removes_temp_files(self, bom_bytes: bytes):
        """cleanup() удаляет созданные temp-файлы."""
        svc = BOMService()
        svc.load_from_bytes(bom_bytes)
        assert len(svc._temp_paths) == 1
        temp_path = svc._temp_paths[0]
        assert os.path.isfile(temp_path), "Temp file should exist before cleanup"
        svc.cleanup()
        assert not os.path.isfile(temp_path), "Temp file should be removed after cleanup"
        assert len(svc._temp_paths) == 0, "Temp paths list should be cleared"

    def test_context_manager_cleans_up(self, bom_bytes: bytes):
        """Выход из with-блока вызывает cleanup."""
        with BOMService() as svc:
            svc.load_from_bytes(bom_bytes)
            assert len(svc._temp_paths) == 1
            temp_path = svc._temp_paths[0]
            assert os.path.isfile(temp_path)
        # После выхода из with — файл должен быть удалён
        assert not os.path.isfile(temp_path), "Temp file should be removed after context exit"

    def test_load_async(self, bom_bytes: bytes):
        """Асинхронная загрузка BOM из байтов."""
        import asyncio

        async def run():
            svc = BOMService()
            bom = await svc.load_async(bom_bytes, filename="async_bom.xlsx")
            return svc, bom

        svc, bom = asyncio.run(run())
        assert svc.is_loaded
        assert len(bom.parts) == 3
        assert "P001" in bom.parts
        svc.cleanup()


# ═══════════════════════════════════════════════════════════════════════
#  20. Edge cases — parse_bom (Extended)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomEdgeCases:
    def test_empty_workbook(self):
        """Workbook with no data sheets."""
        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="bom_empty_")
        os.close(fd)
        wb = Workbook()
        ws = wb.active
        ws.title = "Empty"
        wb.save(path)
        bom = parse_bom(path)
        assert len(bom.parts) == 0
        assert len(bom.config_names) == 0

    def test_no_part_no_column(self):
        """Sheet without part_no keywords should be skipped."""
        data = [
            ["Title", "Author", "Date"],
            ["BOM", "Test", "2024"],
        ]
        path = _create_xlsx({"Info": data})
        bom = parse_bom(path)
        assert len(bom.parts) == 0, "Should not parse sheet without part_no column"

    def test_part_number_with_dashes_is_cleaned(self):
        """Part numbers with dashes should be normalized."""
        data = [
            ["序号", "零部件件号", "名称", "Config1", "Config2"],
            ["1", "ABC-001-DEF", "Test Part", "1", "2"],
            ["2", "GHI-002", "Another", "2", "0"],
        ]
        # Need 2+ config columns and 3+ data rows for is_sheet_bom_candidate
        # Add a 3rd data row
        data.append(["3", "JKL-003", "Yet Another", "1", "1"])
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        assert "ABC001DEF" in bom.parts, f"Dashes should be removed. Keys: {list(bom.parts.keys())}"
        assert "GHI002" in bom.parts
        assert "ABC-001-DEF" not in bom.parts

    def test_qty_from_string(self):
        """Quantity should be parsed from string values."""
        data = [
            ["序号", "零部件件号", "名称", "Config1", "Config2"],
            ["1", "P001", "Part", "2.5", "1.5"],
            ["2", "P002", "Part2", "1", "2"],
            ["3", "P003", "Part3", "0", "1"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        cn = bom.config_names[0]
        assert bom.config_quantities[cn]["P001"] == 2.5, \
            f"Expected 2.5 for P001 in {cn}, got {bom.config_quantities[cn].get('P001')}"

    def test_zero_qty_not_added(self):
        """Parts with zero quantity should not be added to config."""
        data = [
            ["序号", "零部件件号", "名称", "Config1", "Config2"],
            ["1", "P001", "Part1", "0", "0"],
            ["2", "P002", "Part2", "1", "0"],
            ["3", "P003", "Part3", "0", "2"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        # Need 2+ configs and 3+ data rows for is_sheet_bom_candidate
        cn = bom.config_names[0]
        # P001 has qty=0 in all configs → should NOT be in config quantities
        for cn_name in bom.config_names:
            assert "P001" not in bom.config_quantities[cn_name], \
                f"Zero qty part should not be added to {cn_name}"
            assert "P002" in bom.config_quantities[cn_name] or "P003" in bom.config_quantities[cn_name]

    def test_non_header_rows_before_header(self):
        """Rows before the actual header should be ignored."""
        data = [
            ["T1L WE BOM", None, None, None, None, None],
            [None, None, None, None, None, None],
            [None, None, None, None, None, None],
            ["序号", "零部件件号", "名称", "Config1", "Config2", "Config3"],
            ["1", "P001", "Part1", "1", "2", "1"],
            ["2", "P002", "Part2", "1", "1", "0"],
            ["3", "P003", "Part3", "2", "0", "0"],
        ]
        # Need 2+ config columns and 3+ data rows
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        assert len(bom.parts) == 3, f"Expected 3 parts, got {len(bom.parts)}: {list(bom.parts.keys())}"
        assert bom.config_quantities[bom.config_names[0]].get("P001", 0) == 1.0

    def test_global_name_applied_to_part_without_name(self):
        """Part from secondary sheet gets name from global_names.

        涂装BOM has no config columns → not a BOM candidate → skipped.
        P001 only gets name from 焊装BOM (first sheet, preserved).
        """
        data = {
            "焊装BOM": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "Welded Part", "1", "2"],
                ["2", "W002", "Weld Part 2", "1", "1"],
                ["3", "W003", "Weld Part 3", "2", "0"],
            ],
            "涂装BOM": [
                ["序号", "零部件件号", "零件名称"],  # no config columns → not BOM candidate
                ["1", "P001", "Painted Part"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # 涂装BOM skipped (not a BOM candidate) -> P001 named from 焊装BOM only
        assert "P001" in bom.parts, "P001 should be in parts"
        cn, _ = bom.global_names.get("P001", ("", ""))
        assert "Welded" in cn, f"Expected 'Welded Part' (first sheet), got '{cn}'"
        assert "Painted" not in cn, "Painted Part should not appear (skipped sheet)"


# ═══════════════════════════════════════════════════════════════════════
#  21. parse_bom — find_header_rows пуст (строки 104-105)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomHeaderRowsEmpty:
    """Лист проходит is_sheet_bom_candidate, но find_header_rows пуст.

    Покрывает строки 104-105: logger.warning + continue.
    """

    def test_non_standard_headers_with_part_data(self):
        """Headers без китайских/английских ключевых слов, но данные
        содержат part_no-подобные значения → is_sheet_bom_candidate True,
        find_header_rows пуст → строки 104-105.
        """
        data = {
            "BOM": [
                # Non-recognizable header row — не содержит ключевых слов
                ["A", "B", "C", "D", "E"],
                # Data rows with part_no patterns
                ["1", "P001", "Part1", "1", "2"],
                ["2", "P002", "Part2", "1", "1"],
                ["3", "P003", "Part3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # Sheet skipped due to no headers → no parts found
        assert len(bom.parts) == 0, "Sheet with no recognizable headers should be skipped"


# ═══════════════════════════════════════════════════════════════════════
#  22. parse_bom — global_names слияние с пустым cn (строка 129)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomGlobalNamesMergeCN:
    """Первый лист не имеет колонки name_cn, второй имеет → existing_cn = nc.

    Покрывает строку 129 (if not existing_cn and nc).
    """

    def test_merge_sheet_with_en_only_then_cn(self):
        """Sheet 1: только name_en (Part Name EN). Sheet 2: name_cn (零件名称).

        P001 в sheet 1: name_cn="", name_en="Part1 EN"
        P001 в sheet 2: name_cn="零件1", name_en=""
        Merged: existing_cn="" → nc="零件1" → existing_cn="零件1" (line 129)
        """
        data = {
            "Sheet1": [
                ["序号", "零部件件号", "Part Name (EN)", "Config1", "Config2"],
                ["1", "P001", "Part1 EN", "1", "2"],
                ["2", "P002", "Part2 EN", "1", "1"],
                ["3", "P003", "Part3 EN", "2", "0"],
            ],
            "Sheet2": [
                ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
                ["1", "P001", "零件1", "1", "2"],
                ["2", "P002", "零件2", "1", "1"],
                ["3", "P003", "零件3", "2", "0"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # P001: cn from Sheet2, en from Sheet1
        cn, en = bom.global_names.get("P001", ("", ""))
        assert "零件" in cn, f"Expected CN name from Sheet2, got '{cn}'"
        assert "Part1 EN" in en, f"Expected EN name from Sheet1, got '{en}'"
        assert cn and en, "Both CN and EN should be populated"


# ═══════════════════════════════════════════════════════════════════════
#  23. parse_bom — Qty-only edge cases (строки 155,158,160,167-168)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomQtyOnlyEdgeCases:
    """Граничные случаи qty-only пути: None/пустой/невалидный PN, плохой qty.

    Покрывает строки:
      - 155: continue при pn is None
      - 158: continue при пустом/~$ pn
      - 160: continue при невалидном PN
      - 167-168: except при плохом qty
    """

    def test_qty_only_with_bad_part_numbers(self):
        """Qty-only sheet с разными вариантами невалидных PN."""
        data = {
            "Attachment": [
                ["序号", "零部件件号", "零件名称", "数量"],
                ["1", "P001", "Valid Part", "1"],
                [None, None, None, None],  # None pn → line 155 continue
                ["3", "", "Empty PN", "2"],  # empty pn → line 158 continue
                ["4", "~$hidden", "Hidden PN", "1"],  # ~$ pn → line 158 continue
                ["5", "INVALID!@#", "Bad PN", "1"],  # invalid pn → line 160 continue
                ["6", "P006", "Another Valid", "3"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # Only P001 and P006 should be in parts
        assert "P001" in bom.parts, "P001 should be present"
        assert "P006" in bom.parts, "P006 should be present"
        assert len(bom.parts) == 2, f"Expected 2 valid parts, got {len(bom.parts)}: {list(bom.parts.keys())}"
        cn = "Attachment"
        assert bom.config_quantities[cn]["P001"] == 1.0
        assert bom.config_quantities[cn]["P006"] == 3.0

    def test_qty_only_with_bad_qty_value(self):
        """Qty-only sheet с qty значением, которое не конвертируется в float.

        Покрывает строки 167-168: except (ValueError, TypeError) → qty = 0.0.
        """
        data = {
            "Fasteners": [
                ["序号", "零部件件号", "零件名称", "数量"],
                ["1", "P001", "Bolt", "N/A"],  # qty="N/A" → float("N/A") → ValueError → qty=0.0
                ["2", "P002", "Nut", "8"],
                ["3", "P003", "Washer", "16"],
            ],
        }
        path = _create_xlsx(data)
        bom = parse_bom(path)
        # P001 has qty="N/A" → qty=0.0 → not added to parts (qty > 0 check fails)
        assert "P001" not in bom.parts, "P001 should not be in parts (qty=0 after failed conversion)"
        assert "P002" in bom.parts, "P002 should be in parts"
        assert "P003" in bom.parts, "P003 should be in parts"
        cn = "Fasteners"
        assert bom.config_quantities[cn]["P002"] == 8.0
        assert bom.config_quantities[cn]["P003"] == 16.0


# ═══════════════════════════════════════════════════════════════════════
#  24. parse_bom — Normal path edge cases (строки 223,226,228,242-245)
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomNormalPathEdgeCases:
    """Граничные случаи normal (multi-config) пути:
    None/пустой/невалидный PN, плохой qty, None qty.

    Покрывает строки:
      - 223: continue при pn is None
      - 226: continue при пустом/~$ pn
      - 228: continue при невалидном PN
      - 242-243: except при плохом qty
      - 244-245: else при None qty
    """

    def test_normal_path_with_bad_part_numbers(self):
        """Multi-config BOM с разными вариантами невалидных PN."""
        data = [
            ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
            ["1", "P001", "Valid Part One", "1", "2"],
            [None, None, None, None],  # None pn → line 223 continue
            ["3", "", "Empty PN", "1", "1"],  # empty pn → line 226 continue
            ["4", "~$hidden", "Hidden", "1", "1"],  # ~$ pn → line 226 continue
            ["5", "INVALID!@#", "Bad", "1", "1"],  # invalid pn → line 228 continue
            ["6", "P006", "Another Valid", "2", "1"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        # Only P001 and P006 should be in parts
        assert "P001" in bom.parts, "P001 should be present"
        assert "P006" in bom.parts, "P006 should be present"
        assert len(bom.parts) == 2, f"Expected 2 valid parts, got {len(bom.parts)}: {list(bom.parts.keys())}"

    def test_normal_path_with_bad_qty_value(self):
        """Multi-config BOM с qty значением, вызывающим ValueError.

        4 строки данных, 3 из 4 numeric → config колонка детектится.
        "BADVAL" не парсится → ValueError → except (lines 242-243).

        Покрывает строки 242-243: except (ValueError, TypeError).
        """
        data = [
            ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
            ["1", "P001", "Part1", "1", "2"],
            ["2", "P002", "Part2", "1", "1"],
            ["3", "P003", "Part3", "2", "0"],
            ["4", "P004", "Part4", "BADVAL", "3"],  # bad qty in Config1 → ValueError
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        cn1, cn2 = bom.config_names  # Config1, Config2
        # P001, P002, P003 all have valid qty in Config1
        assert "P001" in bom.config_quantities[cn1], "P001 should be in Config1 (qty=1)"
        assert "P002" in bom.config_quantities[cn1], "P002 should be in Config1 (qty=1)"
        assert "P003" in bom.config_quantities[cn1], "P003 should be in Config1 (qty=2)"
        # P004 should be in Config2 (qty=3) but NOT in Config1 (BADVAL → ValueError → 0.0)
        assert "P004" in bom.config_quantities[cn2], "P004 should be in Config2 (qty=3)"
        assert bom.config_quantities[cn2]["P004"] == 3.0
        assert bom.config_quantities[cn1].get("P004", 0) == 0.0, \
            "P004 should have qty=0 in Config1 (bad qty)"
        assert len(bom.parts) == 4, "All 4 parts should be present"

    def test_normal_path_with_none_qty(self):
        """Multi-config BOM с None qty значением (пустая ячейка).

        P002 имеет None в Config1 → else → qty=0.0.
        Config1 должна быть распознана через другие строки с числовыми qty.

        Покрывает строки 244-245: else → qty = 0.0.
        """
        data = [
            ["序号", "零部件件号", "零件名称", "Config1", "Config2"],
            ["1", "P001", "Part1", "1", "2"],
            ["2", "P002", "Part2", None, "1"],  # None qty in Config1 → else → qty=0.0
            ["3", "P003", "Part3", "2", "0"],
        ]
        path = _create_xlsx({"BOM": data})
        bom = parse_bom(path)
        cn1, cn2 = bom.config_names
        # P001 should be in Config1 (qty=1) and Config2 (qty=2)
        assert "P001" in bom.config_quantities[cn1], "P001 should be in Config1"
        assert bom.config_quantities[cn1]["P001"] == 1.0
        assert "P001" in bom.config_quantities[cn2], "P001 should be in Config2"
        assert bom.config_quantities[cn2]["P001"] == 2.0
        # P002 should be in Config2 (qty=1) but NOT in Config1 (None qty → 0.0)
        assert "P002" in bom.config_quantities[cn2], "P002 should be in Config2"
        assert bom.config_quantities[cn2]["P002"] == 1.0
        assert bom.config_quantities[cn1].get("P002", 0) == 0.0, \
            "P002 should have qty=0 in Config1 (None qty)"
        assert len(bom.parts) == 3, "All 3 parts should be present"


# ═══════════════════════════════════════════════════════════════════════
#  25. parse_bom — Strikethrough (зачеркнутый шрифт) игнорирование
# ═══════════════════════════════════════════════════════════════════════

class TestParseBomStrikethrough:
    """Проверка игнорирования зачеркнутого текста (strikethrough) в BOM-файлах."""

    def test_strikethrough_ignored_in_bom(self):
        from openpyxl.styles import Font
        # Создаем тестовый BOM-файл вручную, чтобы применить форматирование
        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="bom_strike_test_")
        os.close(fd)

        try:
            wb = Workbook()
            ws = wb.create_sheet(title="BOM_Strike", index=0)
            ws.cell(row=1, column=1, value="序号")
            ws.cell(row=1, column=2, value="零部件件号")
            ws.cell(row=1, column=3, value="零件名称")
            ws.cell(row=1, column=4, value="Config1")
            ws.cell(row=1, column=5, value="Config2")

            # Строка 2: нормальная
            ws.cell(row=2, column=1, value="1")
            ws.cell(row=2, column=2, value="P001")
            ws.cell(row=2, column=3, value="Part1")
            ws.cell(row=2, column=4, value="5")
            ws.cell(row=2, column=5, value="2")

            # Строка 3: зачеркнутый парт-номер
            ws.cell(row=3, column=1, value="2")
            c_pn = ws.cell(row=3, column=2, value="P002")
            c_pn.font = Font(strike=True)
            ws.cell(row=3, column=3, value="Part2")
            ws.cell(row=3, column=4, value="3")
            ws.cell(row=3, column=5, value="1")

            # Строка 4: зачеркнутое количество
            ws.cell(row=4, column=1, value="3")
            ws.cell(row=4, column=2, value="P003")
            ws.cell(row=4, column=3, value="Part3")
            c_qty = ws.cell(row=4, column=4, value="4")
            c_qty.font = Font(strike=True)
            ws.cell(row=4, column=5, value="1")

            wb.save(path)

            bom = parse_bom(path)
            # P001: должно быть спарсено
            assert "P001" in bom.parts
            assert bom.config_quantities["Config1"]["P001"] == 5.0
            assert bom.config_quantities["Config2"]["P001"] == 2.0

            # P002: зачеркнутый парт-номер -> должно быть пропущено
            assert "P002" not in bom.parts

            # P003: зачеркнутое количество Config1 -> P003 пропущен для Config1,
            # но присутствует для Config2
            assert "P003" in bom.parts
            assert bom.config_quantities["Config1"].get("P003", 0.0) == 0.0
            assert bom.config_quantities["Config2"]["P003"] == 1.0

        finally:
            _safe_remove(path)


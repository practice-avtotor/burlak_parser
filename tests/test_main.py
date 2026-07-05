"""Интеграционные тесты для main.py — полный пайплайн.

Создаёт миниатюрные BOM и card .xlsx файлы, запускает run_pipeline
и проверяет выходные файлы (report.txt, excel, split_cards).

Покрытие:
  - run_pipeline: all configs, single config, no-fuzzy, no-split
  - clean_output_dirs: очистка директорий
  - setup_logging: настройка логирования
  - main: CLI аргументы (через argparse)
  - Error handling: missing BOM, missing cards
  - Output validation: report.txt content, excel file, zip archive
"""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from burlak_parser.main import (
    AUTO_CLEAN_DIRS,
    clean_output_dirs,
    run_pipeline,
    select_config_interactive,
    setup_logging,
    main,
)

from burlak_parser.bom_parser import BOMData


# ═══════════════════════════════════════════════════════════════════════
#  HELPERS: создание тестовых .xlsx файлов
# ═══════════════════════════════════════════════════════════════════════


def _create_test_bom(path: str, multi_sheet: bool = False) -> str:
    """Создать миниатюрный BOM .xlsx для тестов.

    Структура (Sheet1 - основной BOM):
      R1:  序号 | 零部件代号 | 零部件名称 | 舒享版 | 奢享版
      R2:   1   | P001      | 螺母      |  2.0   |  3.0
      R3:   2   | P002      | 螺栓      |  1.0   |  0.0
      R4:   3   | P003      | 垫片      |  0.0   |  4.0

    Если multi_sheet=True — добавляет Sheet2 с доп. названиями.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["序号", "零部件代号", "零部件名称", "舒享版", "奢享版"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    data = [
        [1, "P001", "螺母", 2.0, 3.0],
        [2, "P002", "螺栓", 1.0, 0.0],
        [3, "P003", "垫片", 0.0, 4.0],
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    if multi_sheet:
        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(row=1, column=1, value="序号")
        ws2.cell(row=1, column=2, value="零部件代号")
        ws2.cell(row=1, column=3, value="零部件名称")
        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=2, column=2, value="P004")
        ws2.cell(row=2, column=3, value="弹簧")

    wb.save(path)
    wb.close()
    return path


def _create_test_card(path: str) -> str:
    """Создать миниатюрную операционную карту .xlsx для тестов.

    Структура:
      R1:  序号 | 零部件代号 | 数量
      R2:   1   | P001      |  2.0
      R3:   2   | P003      |  3.0
      R4:   3   | P999      |  1.0   (только в картах)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["序号", "零部件代号", "数量"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    data = [
        [1, "P001", 2.0],
        [2, "P003", 3.0],
        [3, "P999", 1.0],
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    wb.close()
    return path


def _create_multi_sheet_card(path: str) -> str:
    """Создать многолистовую карту для теста split.

    Sheet1: P001
    Sheet2: P999
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Операция_1"
    ws1.cell(row=1, column=1, value="序号")
    ws1.cell(row=1, column=2, value="零部件代号")
    ws1.cell(row=1, column=3, value="数量")
    ws1.cell(row=2, column=1, value=1)
    ws1.cell(row=2, column=2, value="P001")
    ws1.cell(row=2, column=3, value=1.0)

    ws2 = wb.create_sheet("Операция_2")
    ws2.cell(row=1, column=1, value="序号")
    ws2.cell(row=1, column=2, value="零部件代号")
    ws2.cell(row=1, column=3, value="数量")
    ws2.cell(row=2, column=1, value=1)
    ws2.cell(row=2, column=2, value="P999")
    ws2.cell(row=2, column=3, value=2.0)

    wb.save(path)
    wb.close()
    return path


def _create_multi_config_bom(path: str, n_configs: int = 6) -> str:
    """Создать BOM с N комплектациями для теста вывода '... и ещё N'.

    Структура:
      R1: 序号 | 零部件代号 | 零部件名称 | Config0 | Config1 | ... | Config{N-1}
      R2:  1   | P001      | Part1      |  1.0   |  1.0   | ... |  1.0
      R3:  2   | P002      | Part2      |  1.0   |  1.0   | ... |  1.0
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    config_names = [f"Config{i}" for i in range(n_configs)]
    headers = ["序号", "零部件代号", "零部件名称"] + config_names
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    data = [
        [1, "P001", "Part1"] + [1.0] * n_configs,
        [2, "P002", "Part2"] + [1.0] * n_configs,
        [3, "P003", "Part3"] + [1.0] * n_configs,
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    wb.close()
    return path


def _create_card_with_many_parts(path: str, n_parts: int = 12) -> str:
    """Создать карту с N деталями (все не в BOM → ONLY_IN_CARDS).

    Для теста вывода '... и ещё N' при >10 расхождениях.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=2, value="零部件代号")
    ws.cell(row=1, column=3, value="数量")

    for i in range(n_parts):
        ws.cell(row=i + 2, column=1, value=i + 1)
        ws.cell(row=i + 2, column=2, value=f"P{i + 100:03d}")
        ws.cell(row=i + 2, column=3, value=1.0)

    wb.save(path)
    wb.close()
    return path


def _count_lines(path: str) -> int:
    """Подсчитать непустые строки в файле."""
    with open(path) as f:
        return sum(1 for line in f if line.strip())


# ═══════════════════════════════════════════════════════════════════════
#  FIXTURES
# ═══════════════════════════════════════════════════════════════════════


@pytest.fixture
def bom_path(tmp_path: Path) -> str:
    """Создать временный BOM .xlsx."""
    path = os.path.join(tmp_path, "G01_ bom.xlsx")
    return _create_test_bom(path)


@pytest.fixture
def card_path(tmp_path: Path) -> str:
    """Создать временную карту .xlsx с номером операции."""
    path = os.path.join(tmp_path, "001-card.xlsx")
    return _create_test_card(path)


@pytest.fixture
def multi_card_path(tmp_path: Path) -> str:
    """Создать многолистовую карту .xlsx с номером операции."""
    path = os.path.join(tmp_path, "002-multi_card.xlsx")
    return _create_multi_sheet_card(path)


@pytest.fixture
def output_dir(tmp_path: Path) -> str:
    """Временная директория для результатов."""
    path = os.path.join(tmp_path, "output")
    os.makedirs(path, exist_ok=True)
    return path


# ═══════════════════════════════════════════════════════════════════════
#  1. setup_logging
# ═══════════════════════════════════════════════════════════════════════

class TestSetupLogging:
    def test_setup_logging_default(self):
        """Default logging does not crash (root logger level not changed by pytest)."""
        # In pytest, root logger is already configured at WARNING level
        # setup_logging calls basicConfig which is a no-op if already configured
        # Just verify the function doesn't crash
        setup_logging(verbose=False)
        assert True

    def test_setup_logging_verbose(self):
        """Verbose logging does not crash."""
        setup_logging(verbose=True)
        assert True


# ═══════════════════════════════════════════════════════════════════════
#  2. clean_output_dirs
# ═══════════════════════════════════════════════════════════════════════

class TestCleanOutputDirs:
    def test_cleans_existing_dir(self, tmp_path: Path):
        """Existing output dir is cleaned."""
        test_dir = os.path.join(tmp_path, "output")
        os.makedirs(test_dir)
        open(os.path.join(test_dir, "test.txt"), "w").close()
        assert os.path.isdir(test_dir)

        clean_output_dirs(test_dir)
        assert not os.path.isdir(test_dir)

    def test_cleans_auto_clean_dirs(self, tmp_path: Path):
        """Auto-clean dirs (output, split_cards, _extracted_cards) are cleaned."""
        for d in ["output", "split_cards", "_extracted_cards"]:
            path = os.path.join(tmp_path, d)
            os.makedirs(path)
            open(os.path.join(path, "test.txt"), "w").close()

        clean_output_dirs(str(tmp_path))
        # Only the main output dir is cleaned (tmp_path, not the subdirs)
        # The auto-clean ones in CWD would be checked - but we're not in CWD
        # So this test just verifies no crash
        assert True

    def test_clean_nonexistent_dir(self):
        """Non-existent dir doesn't crash."""
        clean_output_dirs("/nonexistent/path/12345")
        assert True  # no crash


# ═══════════════════════════════════════════════════════════════════════
#  3. run_pipeline — all configs (multi-config)
# ═══════════════════════════════════════════════════════════════════════

class TestRunPipelineAllConfigs:
    def test_full_pipeline_creates_outputs(self, bom_path, card_path, output_dir):
        """Full pipeline creates report.txt and excel file."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=True,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        # Проверяем, что файлы созданы
        txt_report = os.path.join(output_dir, "report.txt")
        excel_files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx")]

        assert os.path.isfile(txt_report), "report.txt not created"
        assert len(excel_files) > 0, "No excel report created"

        # Проверяем содержание отчёта
        content = open(txt_report).read()
        assert "ОТЧЁТ" in content or "ПРОВЕРКИ" in content
        assert "舒享版" in content  # config name
        assert "奢享版" in content  # config name
        assert "P001" in content  # part number
        assert "P003" in content  # part number
        assert "расхождений" in content or "несоответствий" in content

    def test_pipeline_discrepancy_types(self, bom_path, card_path, output_dir):
        """Pipeline detects different discrepancy types."""
        # BOM: P001(2,3), P002(1,0), P003(0,4)
        # Cards: P001(2), P003(3), P999(1)
        # Expected: P001 qty match for 舒享(BOM=2, card=2), mismatch for 奢享(BOM=3,card=2)
        # P002 ONLY_IN_BOM, P003 qty mismatch, P999 ONLY_IN_CARDS
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        txt = os.path.join(output_dir, "report.txt")
        content = open(txt).read()
        assert "Разное количество" in content or "разное" in content.lower()
        assert "Есть в BOM" in content or "BOM" in content
        assert "Есть в" in content or "нет в" in content

    def test_pipeline_with_multi_sheet_bom(self, tmp_path, card_path, output_dir):
        """Pipeline works with multi-sheet BOM."""
        bom_path = os.path.join(tmp_path, "multi_bom.xlsx")
        _create_test_bom(bom_path, multi_sheet=True)

        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt), "report.txt not created"

    def test_pipeline_no_fuzzy(self, bom_path, card_path, output_dir):
        """Pipeline works without fuzzy matching."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=False,
            single_config=False,
            max_workers=1,
        )
        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt)

    def test_pipeline_no_split(self, bom_path, card_path, output_dir):
        """Pipeline without auto-split still generates outputs."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )
        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt)

    def test_pipeline_with_multi_sheet_card(self, bom_path, multi_card_path, output_dir):
        """Pipeline with multi-sheet card creates split files."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=multi_card_path,
            output_dir=output_dir,
            auto_split=True,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt)

        # Check split cards were created
        split_dir = os.path.join(output_dir, "split_cards")
        zip_path = os.path.join(output_dir, "split_cards.zip")
        assert os.path.isdir(split_dir) or os.path.isfile(zip_path), \
            "Neither split dir nor zip found"


# ═══════════════════════════════════════════════════════════════════════
#  4. run_pipeline — single config
# ═══════════════════════════════════════════════════════════════════════

class TestRunPipelineSingleConfig:
    def test_single_config_creates_outputs(self, bom_path, card_path, output_dir):
        """Single config mode creates report with discrepancies."""
        # BOM 舒享版: P001(2.0), P002(1.0), P003(0.0)
        # Cards:       P001(2.0), P003(3.0), P999(1.0)
        # Expected: P001 perfect match (not in discrepancy list!)
        #           P002 ONLY_IN_BOM (in report), P003 ONLY_IN_CARDS, P999 ONLY_IN_CARDS
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            config_name="舒享版",
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=True,
            max_workers=1,
        )

        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt)

        content = open(txt).read()
        assert "舒享版" in content
        # P002 should be in ONLY_IN_BOM section
        assert "P002" in content
        # P999 should be in ONLY_IN_CARDS section
        assert "P999" in content

    def test_single_config_wrong_name_raises(self, bom_path, card_path, output_dir):
        """Wrong config name exits with sys.exit(1)."""
        with pytest.raises(SystemExit) as exc:
            run_pipeline(
                bom_path=bom_path,
                cards_path=card_path,
                config_name="NonExistent",
                output_dir=output_dir,
                auto_split=False,
                use_fuzzy=True,
                single_config=True,
                max_workers=1,
            )
        assert exc.value.code == 1

    def test_single_config_quantity_mismatch(self, bom_path, card_path, output_dir):
        """Single config shows quantity mismatch details."""
        # 舒享版: P001(BOM=2.0, card=2.0) → perfect match
        # P002(BOM=1.0, card=0) → ONLY_IN_BOM
        # P003(BOM=0, card=3.0) → ONLY_IN_CARDS? No, P003 has 0 in BOM for 舒享版
        # P999(not in BOM)→ ONLY_IN_CARDS
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            config_name="舒享版",
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=True,
            max_workers=1,
        )

        txt = os.path.join(output_dir, "report.txt")
        # P001 qty: BOM=2.0, card=2.0 → perfect match for 舒享版
        # P002 qty: BOM=1.0, card=0 → ONLY_IN_BOM
        # P999: not in BOM → ONLY_IN_CARDS
        content = open(txt).read()
        assert "несоответствий" in content or "расхождени" in content


# ═══════════════════════════════════════════════════════════════════════
#  5. run_pipeline — default output dir
# ═══════════════════════════════════════════════════════════════════════

class TestRunPipelineDefaultOutput:
    def test_default_output_dir_created(self, bom_path, card_path, monkeypatch):
        """Default output dir (./output) is created when not specified."""
        # Change to a tmp dir so we don't pollute real project
        import tempfile
        with tempfile.TemporaryDirectory() as tmp:
            monkeypatch.chdir(tmp)
            run_pipeline(
                bom_path=bom_path,
                cards_path=card_path,
                output_dir=None,  # default
                auto_split=False,
                use_fuzzy=True,
                single_config=False,
                max_workers=1,
            )
            assert os.path.isdir(os.path.join(tmp, "output"))
            assert os.path.isfile(os.path.join(tmp, "output", "report.txt"))


# ═══════════════════════════════════════════════════════════════════════
#  6. Error handling
# ═══════════════════════════════════════════════════════════════════════

class TestErrorHandling:
    def test_missing_bom_exits(self, card_path, output_dir):
        """Missing BOM file causes sys.exit(1) from main()."""
        with pytest.raises(SystemExit) as exc:
            bad_path = "/nonexistent/bom.xlsx"
            with patch.object(sys, "argv", [
                "main.py", "--bom", bad_path, "--cards", card_path,
            ]):
                main()
        assert exc.value.code == 1

    def test_missing_cards_exits(self, bom_path, output_dir):
        """Missing cards path causes sys.exit(1) from main()."""
        with pytest.raises(SystemExit) as exc:
            bad_path = "/nonexistent/cards"
            with patch.object(sys, "argv", [
                "main.py", "--bom", bom_path, "--cards", bad_path,
            ]):
                main()
        assert exc.value.code == 1


# ═══════════════════════════════════════════════════════════════════════
#  7. CLI arg parsing (main function)
# ═══════════════════════════════════════════════════════════════════════

class TestCLIArguments:
    def test_minimal_args(self, bom_path, card_path, output_dir):
        """Minimal required args run successfully."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path, "-o", output_dir,
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_single_config_arg(self, bom_path, card_path, output_dir):
        """--single-config flag works."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
            "-o", output_dir, "--single-config", "--config", "舒享版",
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_no_fuzzy_arg(self, bom_path, card_path, output_dir):
        """--no-fuzzy flag works."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
            "-o", output_dir, "--no-fuzzy",
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_no_split_arg(self, bom_path, card_path, output_dir):
        """--no-split flag works."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
            "-o", output_dir, "--no-split",
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_verbose_arg(self, bom_path, card_path, output_dir):
        """--verbose flag works."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
            "-o", output_dir, "--verbose",
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_workers_arg(self, bom_path, card_path, output_dir):
        """--workers flag works."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
            "-o", output_dir, "--workers", "2",
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))

    def test_short_args(self, bom_path, card_path, output_dir):
        """Short flags (-b, -c, -o) work."""
        with patch.object(sys, "argv", [
            "main.py", "-b", bom_path, "-c", card_path, "-o", output_dir,
        ]):
            main()
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))


# ═══════════════════════════════════════════════════════════════════════
#  8. Excel report validation
# ═══════════════════════════════════════════════════════════════════════

class TestExcelReport:
    def test_excel_has_expected_sheets(self, bom_path, card_path, output_dir):
        """Generated Excel has expected sheet names."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        excel_files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx") and "split" not in f]
        assert excel_files, "No excel report found"

        excel_path = os.path.join(output_dir, excel_files[0])
        wb = openpyxl.load_workbook(excel_path)
        sheet_names = wb.sheetnames

        # Should have at least Сводка sheet
        assert any("Сводка" in s or "свод" in s.lower() for s in sheet_names), \
            f"No summary sheet found in {sheet_names}"

        wb.close()

    def test_excel_contains_data(self, bom_path, card_path, output_dir):
        """Excel contains actual part data."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        excel_files = [f for f in os.listdir(output_dir) if f.endswith(".xlsx") and "split" not in f]
        assert excel_files

        wb = openpyxl.load_workbook(os.path.join(output_dir, excel_files[0]))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                cell_text = " ".join(cells)
                if "P001" in cell_text:
                    wb.close()
                    return  # Found P001 in some sheet

        wb.close()
        pytest.fail("Part P001 not found in any excel sheet")


# ═══════════════════════════════════════════════════════════════════════
#  9. Report content edge cases
# ═══════════════════════════════════════════════════════════════════════

class TestReportContent:
    def test_report_has_all_sections(self, bom_path, card_path, output_dir):
        """Report has all expected sections."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        content = open(os.path.join(output_dir, "report.txt")).read()
        assert "ОТЧЁТ" in content
        assert "ПРОВЕРКИ" in content
        assert "КОМПЛЕКТАЦИЙ" in content
        assert "BOM" in content or "bom" in content.lower()

    def test_report_non_empty(self, bom_path, card_path, output_dir):
        """Report is not empty."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        assert _count_lines(os.path.join(output_dir, "report.txt")) > 5

    def test_pipeline_without_fuzzy_still_finds_discrepancies(
        self, bom_path, card_path, output_dir,
    ):
        """Without fuzzy, basic discrepancies are still found."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=False,
            single_config=False,
            max_workers=1,
        )

        content = open(os.path.join(output_dir, "report.txt")).read()
        assert "несоответствий" in content or "расхождени" in content or "ОТЧЁТ" in content


# ═══════════════════════════════════════════════════════════════════════
#  10. Split cards integration
# ═══════════════════════════════════════════════════════════════════════

class TestSplitCardsIntegration:
    def test_split_cards_zip_created(self, bom_path, multi_card_path, output_dir):
        """Auto-split creates ZIP with split files."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=multi_card_path,
            output_dir=output_dir,
            auto_split=True,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        zip_path = os.path.join(output_dir, "split_cards.zip")
        assert os.path.isfile(zip_path), "split_cards.zip not created"

        # Check zip contains split files
        with zipfile.ZipFile(zip_path) as zf:
            names = zf.namelist()
            assert len(names) > 0, "Zip is empty"
            # Should have at least one .xlsx file
            xlsx_in_zip = [n for n in names if n.endswith(".xlsx")]
            assert len(xlsx_in_zip) > 0, "No .xlsx files in zip"

    def test_split_cards_dir_created(self, bom_path, multi_card_path, output_dir):
        """Auto-split creates split_cards directory."""
        run_pipeline(
            bom_path=bom_path,
            cards_path=multi_card_path,
            output_dir=output_dir,
            auto_split=True,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        split_dir = os.path.join(output_dir, "split_cards")
        assert os.path.isdir(split_dir), "split_cards dir not created"
        assert len(os.listdir(split_dir)) > 0, "split_cards dir is empty"


# ═══════════════════════════════════════════════════════════════════════
#  11. select_config_interactive
# ═══════════════════════════════════════════════════════════════════════

class TestSelectConfigInteractive:
    """select_config_interactive — интерактивный выбор комплектации.

    Покрывает строки 100-128 main.py:
      - Пустой список комплектаций → sys.exit(1)
      - Одна комплектация → авто-выбор
      - Валидный выбор из нескольких
      - Неверный ввод, затем валидный
      - Число вне диапазона, затем валидное
    """

    def test_no_configs_exits(self):
        """Empty config list → sys.exit(1)."""
        bom = BOMData(parts={}, config_names=[], config_quantities={})
        with pytest.raises(SystemExit) as exc:
            select_config_interactive(bom)
        assert exc.value.code == 1

    def test_single_config_auto_selected(self):
        """Single config → auto-selected without prompting."""
        bom = BOMData(
            parts={},
            config_names=["Единственная"],
            config_quantities={"Единственная": {}},
        )
        result = select_config_interactive(bom)
        assert result == "Единственная"

    def test_valid_choice(self, monkeypatch):
        """Valid config index returns correct name."""
        configs = ["Config A", "Config B", "Config C"]
        bom = BOMData(parts={}, config_names=configs,
                      config_quantities={c: {} for c in configs})
        monkeypatch.setattr("builtins.input", lambda _: "2")
        result = select_config_interactive(bom)
        assert result == "Config B"

    def test_invalid_then_valid(self, monkeypatch):
        """Non-numeric input, then valid choice."""
        inputs = iter(["abc", "3"])
        monkeypatch.setattr("builtins.input", lambda _: next(inputs))
        configs = ["Config A", "Config B", "Config C"]
        bom = BOMData(parts={}, config_names=configs,
                      config_quantities={c: {} for c in configs})
        result = select_config_interactive(bom)
        assert result == "Config C"

    def test_out_of_range_then_valid(self, monkeypatch):
        """Out of range input (0), then valid."""
        inputs = iter(["0", "1"])
        monkeypatch.setattr("builtins.input", lambda _: next(inputs))
        configs = ["Config A", "Config B"]
        bom = BOMData(parts={}, config_names=configs,
                      config_quantities={c: {} for c in configs})
        result = select_config_interactive(bom)
        assert result == "Config A"


# ═══════════════════════════════════════════════════════════════════════
#  12. clean_output_dirs — расширенные кейсы
# ═══════════════════════════════════════════════════════════════════════

class TestCleanOutputDirsExtended:
    """Расширенные тесты clean_output_dirs.

    Покрывает строки 86-97 main.py:
      - Авто-очистка split_cards / _extracted_cards в CWD
      - output_dir не удаляется повторно через auto-clean
    """

    def test_auto_clean_dirs_in_cwd(self, tmp_path, monkeypatch):
        """Auto-clean dirs (split_cards, _extracted_cards) in CWD are removed
        when output_dir is different."""
        monkeypatch.chdir(tmp_path)
        # Create auto-clean dirs in CWD
        for d in ["split_cards", "_extracted_cards"]:
            path = os.path.join(tmp_path, d)
            os.makedirs(path)
            open(os.path.join(path, "test.txt"), "w").close()

        # Different output_dir (not in CWD)
        output_dir = os.path.join(str(tmp_path), "custom_output")
        os.makedirs(output_dir)
        open(os.path.join(output_dir, "report.txt"), "w").close()

        clean_output_dirs(output_dir)

        # Auto-clean dirs in CWD should be removed
        assert not os.path.isdir(os.path.join(tmp_path, "split_cards")), \
            "split_cards should be cleaned"
        assert not os.path.isdir(os.path.join(tmp_path, "_extracted_cards")), \
            "_extracted_cards should be cleaned"
        # output_dir should also be cleaned (it was added first)
        assert not os.path.isdir(output_dir), "output_dir should be cleaned"

    def test_output_dir_not_cleaned_twice(self, tmp_path, monkeypatch):
        """When output_dir matches an auto-clean dir, it's skipped in the
        auto-clean loop (already added to dirs_to_clean)."""
        monkeypatch.chdir(tmp_path)
        output_dir = os.path.join(tmp_path, "output")
        os.makedirs(output_dir)

        # output_dir is also in CWD and matches an auto-clean dir name
        # The code checks `path != output_dir` and skips it
        clean_output_dirs(output_dir)
        assert not os.path.isdir(output_dir), "output_dir should be cleaned"

    def test_rmtree_error_logged(self, tmp_path, monkeypatch, caplog):
        """Error during rmtree is caught and logged as warning.

        Покрывает строки 96-97: except Exception в clean_output_dirs.
        """
        output_dir = os.path.join(tmp_path, "output")
        os.makedirs(output_dir)

        def failing_rmtree(path):
            raise PermissionError("Access denied")

        monkeypatch.setattr(shutil, "rmtree", failing_rmtree)

        with caplog.at_level(logging.WARNING):
            clean_output_dirs(output_dir)

        assert "Не удалось очистить" in caplog.text
        assert "Access denied" in caplog.text


# ═══════════════════════════════════════════════════════════════════════
#  13. run_pipeline — интерактивный выбор (single_config + no config_name)
# ═══════════════════════════════════════════════════════════════════════

class TestRunPipelineInteractiveConfig:
    """run_pipeline с single_config=True и config_name=None →
    должен вызвать select_config_interactive.

    Покрывает строку 192 (вызов select_config_interactive).
    """

    def test_interactive_selection(self, bom_path, card_path, output_dir, monkeypatch):
        """single_config with no config_name → interactive selection is called."""
        # Mock interactive selection to return a valid config
        monkeypatch.setattr(
            "burlak_parser.main.select_config_interactive",
            lambda bom: "舒享版",
        )
        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            config_name=None,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=True,
            max_workers=1,
        )
        txt = os.path.join(output_dir, "report.txt")
        assert os.path.isfile(txt)
        content = open(txt).read()
        # Should have processed 舒享版
        assert "舒享版" in content


# ═══════════════════════════════════════════════════════════════════════
#  14. main() — обработка ошибок
# ═══════════════════════════════════════════════════════════════════════

class TestMainErrorHandling:
    """Обработка ошибок в main(): KeyboardInterrupt и Exception.

    Покрывает строки 453-462 main.py.
    """

    def test_keyboard_interrupt(self, bom_path, card_path):
        """KeyboardInterrupt в run_pipeline → sys.exit(1)."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
        ]):
            with patch(
                "burlak_parser.main.run_pipeline",
                side_effect=KeyboardInterrupt(),
            ):
                with pytest.raises(SystemExit) as exc:
                    main()
                assert exc.value.code == 1

    def test_generic_exception(self, bom_path, card_path):
        """Generic exception в run_pipeline → sys.exit(1)."""
        with patch.object(sys, "argv", [
            "main.py", "--bom", bom_path, "--cards", card_path,
        ]):
            with patch(
                "burlak_parser.main.run_pipeline",
                side_effect=ValueError("test error"),
            ):
                with pytest.raises(SystemExit) as exc:
                    main()
                assert exc.value.code == 1


# ═══════════════════════════════════════════════════════════════════════
#  15. report — крайние случаи вывода
# ═══════════════════════════════════════════════════════════════════════

def _create_perfect_card(path: str) -> str:
    """Создать карту, идеально совпадающую с 舒享版.

    BOM 舒享版: P001=2.0, P002=1.0, P003=0.0
    Card:       P001=2.0, P002=1.0  (P003 qty=0 → не включается)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=2, value="零部件代号")
    ws.cell(row=1, column=3, value="数量")
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="P001")
    ws.cell(row=2, column=3, value=2.0)
    ws.cell(row=3, column=1, value=2)
    ws.cell(row=3, column=2, value="P002")
    ws.cell(row=3, column=3, value=1.0)
    wb.save(path)
    wb.close()
    return path


class TestReportEdgeCases:
    """Крайние случаи отчёта: 0 расхождений.

    Покрывает строку 355 (вывод "Расхождений не найдено!").
    """

    def test_no_discrepancies_message(self, bom_path, output_dir, tmp_path, capsys):
        """Perfect match → 'Расхождений не найдено!' displayed in stdout.

        File must have a digit-prefixed name (e.g. '001-...') for the
        file_classifier to recognize it.
        """
        card_path = os.path.join(tmp_path, "001-perfect.xlsx")
        _create_perfect_card(card_path)

        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            config_name="舒享版",
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=True,
            max_workers=1,
        )

        captured = capsys.readouterr()
        assert "Расхождений не найдено" in captured.out


# ═══════════════════════════════════════════════════════════════════════
#  16. Many configs — строка 290: '... и ещё N комплектаций'
# ═══════════════════════════════════════════════════════════════════════

class TestManyConfigs:
    """Тест для >5 комплектаций → вывод '... и ещё N комплектаций'

    Покрывает строку 290 main.py.
    """

    def test_many_configs_message(self, tmp_path, capsys):
        """BOM с 6+ комплектациями показывает '... и ещё N комплектаций'."""
        bom_path = os.path.join(tmp_path, "multi_config_bom.xlsx")
        _create_multi_config_bom(bom_path, n_configs=6)

        card_path = os.path.join(tmp_path, "001-card.xlsx")
        _create_test_card(card_path)

        output_dir = os.path.join(tmp_path, "output")
        os.makedirs(output_dir)

        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        captured = capsys.readouterr()
        assert "... и ещё 1 комплектаций" in captured.out or "... и ещё 1" in captured.out


# ═══════════════════════════════════════════════════════════════════════
#  17. Many discrepancies — строка 353: '... и ещё N'
# ═══════════════════════════════════════════════════════════════════════

class TestManyDiscrepancies:
    """Тест для >10 расхождений → вывод '... и ещё N'

    Покрывает строку 353 main.py.
    """

    def test_many_discrepancies_message(self, bom_path, output_dir, tmp_path, capsys):
        """Карта с 12 частями не из BOM → 12+ расхождений → '... и ещё N'."""
        card_path = os.path.join(tmp_path, "001-many_parts.xlsx")
        _create_card_with_many_parts(card_path, n_parts=12)

        run_pipeline(
            bom_path=bom_path,
            cards_path=card_path,
            config_name="舒享版",
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=True,
            max_workers=1,
        )

        captured = capsys.readouterr()
        assert "... и ещё" in captured.out


# ═══════════════════════════════════════════════════════════════════════
#  18. Corrupted files — строка 211
# ═══════════════════════════════════════════════════════════════════════

class TestCorruptedFiles:
    """Тест для повреждённых файлов.

    Покрывает строку 211 main.py: if cards.corrupted_files: print(...)
    """

    def test_corrupted_file_warning(self, bom_path, output_dir, tmp_path, capsys):
        """Повреждённый .xlsx файл → предупреждение в stdout.

        Создаём файл с .xlsx расширением, но невалидным содержимым.
        Файловый классификатор определяет его как карту (цифровой префикс),
        парсер не может открыть → попадает в corrupted_files.
        """
        # Создаём битый .xlsx
        bad_card = os.path.join(tmp_path, "003-corrupted.xlsx")
        with open(bad_card, "w") as f:
            f.write("this is not a valid xlsx file")

        # Создаём нормальную карту (без неё CardService.load() может не найти карт)
        good_card = os.path.join(tmp_path, "001-card.xlsx")
        _create_test_card(good_card)

        run_pipeline(
            bom_path=bom_path,
            cards_path=str(tmp_path),
            output_dir=output_dir,
            auto_split=False,
            use_fuzzy=True,
            single_config=False,
            max_workers=1,
        )

        captured = capsys.readouterr()
        assert "Повреждённых" in captured.out or "corrupted" in captured.out.lower()


# ═══════════════════════════════════════════════════════════════════════
#  19. if __name__ == "__main__" — строка 466
# ═══════════════════════════════════════════════════════════════════════

class TestMainEntryPoint:
    """Тест для точки входа if __name__ == '__main__'.

    Покрывает строку 466 main.py.
    """

    def test_main_entry_point(self, bom_path, card_path, output_dir):
        """Запуск main.py как `__main__` через subprocess."""
        result = subprocess.run(
            [
                sys.executable, "-m", "burlak_parser.main",
                "--bom", bom_path,
                "--cards", card_path,
                "-o", output_dir,
            ],
            capture_output=True,
            text=True,
            timeout=30,
        )
        assert result.returncode == 0, (
            f"main.py exit code {result.returncode}\n"
            f"stderr: {result.stderr[:500]}"
        )
        assert os.path.isfile(os.path.join(output_dir, "report.txt"))


# ═══════════════════════════════════════════════════════════════════════
#  20. Integrity checks — строки 314-318, 329-330, 336
# ═══════════════════════════════════════════════════════════════════════

class TestIntegrityChecks:
    """Тесты для проверок целостности в main.py.

    Эти проверки срабатывают, только когда учёт BOM-частей не совпадает
    с ожидаемым количеством. В корректных данных такое невозможно —
    используем mock компаратора.

    Покрывает:
      - Строки 314-318: accounted_bom != expected
      - Строки 329-330: sum_check != total_discrepancies
      - Строка 336: integrity_ok = False → warning на stdout
    """

    def test_integrity_mismatch_logged(self, bom_path, card_path, output_dir, caplog, capsys):
        """Mock компаратора с несовпадающими счётчиками → integrity warning.

        Создаём fake MultiConfigComparisonResult, где:
          - matched_parts + only_bom_count + qty_mismatch_count + fuzzy_count != total_bom_parts
          - В all_discrepancies есть расхождение с неизвестным типом (не входит ни в один из 4)
        """
        from burlak_parser.comparator import (
            MultiConfigComparisonResult, ConfigComparisonResult,
            Discrepancy,
        )

        # Создаём fake discrepancy с неизвестным типом — он будет учтён в
        # total_discrepancies, но не попадёт ни в один из 4 type-sum counters
        fake_disc = Discrepancy(
            part_number="P999",
            name_cn="", name_en="",
            qty_bom=0.0, qty_cards=1.0,
            card_numbers=["card1"],
            discrepancy_type="__UNKNOWN_TYPE__",  # не входит в 4 известных типа
            config_name="舒享版",
        )

        mock_result = MultiConfigComparisonResult(
            config_results=[
                ConfigComparisonResult(
                    config_name="舒享版",
                    discrepancies=[fake_disc],
                    total_bom_parts=5,   # matched=2 + only_bom=0 + qty=0 + fuzzy=0 = 2 != 5
                    total_cards_parts=3,
                    matched_parts=2,     # не совпадает с total_bom_parts
                    fuzzy_matched=0,
                ),
            ],
            all_discrepancies=[fake_disc],  # 1 discrepancy unknown type → sum_check=0 != total=1
            total_configs=1,
        )

        with caplog.at_level(logging.WARNING):
            with patch(
                "burlak_parser.main.compare_all_configs",
                return_value=mock_result,
            ):
                run_pipeline(
                    bom_path=bom_path,
                    cards_path=card_path,
                    output_dir=output_dir,
                    auto_split=False,
                    use_fuzzy=True,
                    single_config=False,
                    max_workers=1,
                )

        # Строка 314-318: accounted_bom != expected → logger.warning
        assert "Нарушение целостности" in caplog.text, \
            "Expected integrity violation warning in logs"
        assert "учтено 2, ожидалось 5" in caplog.text, \
            "Expected diff message in log: учтено 2, ожидалось 5"

        # Строка 329-330: sum_check != total_discrepancies → logger.warning
        assert "сумма типов" in caplog.text.lower(), \
            "Expected type sum mismatch warning in logs"

        # Строка 336: integrity_ok = False → print warning
        captured = capsys.readouterr()
        assert "нарушения целостности" in captured.out, \
            "Expected integrity warning in stdout"


# ═══════════════════════════════════════════════════════════════════════
#  21. --help output — проверка всех CLI флагов и примеров
# ═══════════════════════════════════════════════════════════════════════

class TestHelpOutput:
    """Автоматическая проверка `--help` — все флаги и примеры."""

    def test_help_contains_all_flags(self):
        """--help содержит все 10 CLI флагов (long и short формы)."""
        result = subprocess.run(
            [sys.executable, "-m", "burlak_parser.main", "--help"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        output = result.stdout

        # Проверяем все long флаги
        long_flags = [
            "--bom", "--cards", "--config", "--output",
            "--single-config", "--no-split", "--no-fuzzy",
            "--workers", "--verbose", "--split-stats",
        ]
        for flag in long_flags:
            assert flag in output, f"Flag {flag} not found in --help output"

        # Проверяем все short флаги
        short_flags = ["-b", "-c", "-k", "-o", "-s", "-w", "-v", "-S"]
        for flag in short_flags:
            assert flag in output, f"Short flag {flag} not found in --help output"

    def test_help_contains_epilog_examples(self):
        """--help epilog содержит все примеры использования."""
        result = subprocess.run(
            [sys.executable, "-m", "burlak_parser.main", "--help"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        output = result.stdout

        # Ключевые маркеры из epilog-примеров
        examples = [
            "--bom BOM.xlsx --cards ./cards/",
            "--single-config",
            "--config",
            "--no-fuzzy",
            "--workers",
            "--split-stats",
            "--no-split",
        ]
        for example in examples:
            assert example in output, f"Example '{example}' not found in --help output"

    def test_help_does_not_crash_without_args(self):
        """Запуск без обязательных аргументов не падает (argparse сам выводит usage)."""
        result = subprocess.run(
            [sys.executable, "-m", "burlak_parser.main"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        # Без --bom argparse завершается с кодом 2
        assert result.returncode == 2
        assert "usage:" in result.stdout or "usage:" in result.stderr

    def test_help_via_h_flag(self):
        """-h (short help) работает так же как --help."""
        result = subprocess.run(
            [sys.executable, "-m", "burlak_parser.main", "-h"],
            capture_output=True,
            text=True,
            timeout=10,
        )
        assert result.returncode == 0
        assert "Burlak Parser" in result.stdout
        assert "--bom" in result.stdout

"""Тест детерминизма split_cards_to_files.

Проверяет, что многократный запуск split_cards_to_files на одном и том же
наборе данных даёт ОДИНАКОВОЕ количество файлов с ОДИНАКОВЫМИ именами.

Сценарий:
  1. Создаём 3 .xlsx файла, имитирующих T1L-подобные операционные карты
     (каждый файл содержит 2-4 листа с таблицами деталей).
  2. Запускаем parse_cards один раз (детерминирован).
  3. Запускаем split_cards_to_files 3 раза в разные выходные директории.
  4. Сравниваем имена созданных файлов во всех 3 запусках.
"""

from __future__ import annotations

import os
import tempfile
from typing import List

import openpyxl
import pytest
from openpyxl import Workbook

from burlak_parser.card_parser import (
    CardsData,
    parse_cards,
    split_cards_to_files,
)

def _create_multi_sheet_card(dir_path: str, prefix: str, sheets: int) -> str:
    """Создать .xlsx с несколькими листами, имитирующими операционные карты.

    Каждый лист содержит заголовок таблицы деталей и несколько строк данных.
    """
    path = os.path.join(dir_path, f"{prefix}_card.xlsx")
    wb = Workbook()
    wb.remove(wb.active)

    for s in range(1, sheets + 1):
        ws = wb.create_sheet(title=f"Операция{s}")
        # Заголовок: 物料编码 (C1), 零件名称 (C2), 数量 (C3)
        ws.cell(row=1, column=1, value="物料编码")
        ws.cell(row=1, column=2, value="零件名称")
        ws.cell(row=1, column=3, value="数量")
        # Данные
        for r in range(1, 4):  # 3 детали на лист
            ws.cell(row=r + 1, column=1, value=f"P{prefix}{s}{r:03d}")
            ws.cell(row=r + 1, column=2, value=f"Деталь {prefix}-{s}-{r}")
            ws.cell(row=r + 1, column=3, value=r)

    wb.save(path)
    return path


@pytest.fixture(scope="module")
def card_dir(tmp_path_factory) -> str:
    """Создать временную директорию с 3 .xlsx картами (каждая по 2-4 листа)."""
    tmpdir = str(tmp_path_factory.mktemp("cards_det"))
    try:
        # 3 файла с разным количеством листов
        _create_multi_sheet_card(tmpdir, "001", 4)  # 4 операции
        _create_multi_sheet_card(tmpdir, "002", 2)  # 2 операции
        _create_multi_sheet_card(tmpdir, "003", 3)  # 3 операции
    except Exception:
        pass
    yield tmpdir


@pytest.fixture(scope="module")
def cards_data(card_dir: str) -> CardsData:
    """Распарсить все карты из card_dir один раз (детерминированно).

    max_workers=1 для гарантии детерминизма (без race conditions процессов).
    """
    return parse_cards(card_dir, show_progress=False, max_workers=1)


def _run_split_and_get_basenames(
    cards_data: CardsData,
    output_dir: str,
) -> List[str]:
    """Запустить split_cards_to_files и вернуть отсортированные имена файлов."""
    created = split_cards_to_files(cards_data, output_dir, max_workers=1)
    return sorted(os.path.basename(f) for f in created)


class TestSplitCardsDeterminism:
    """Три запуска split_cards_to_files → одинаковое количество и имена файлов."""

    def test_files_count_identical_across_runs(self, cards_data: CardsData, tmp_path):
        """Количество созданных файлов одинаково во всех 3 запусках."""
        runs: List[List[str]] = []

        for i in range(3):
            out_dir = os.path.join(str(tmp_path), f"run_{i}")
            basenames = _run_split_and_get_basenames(cards_data, out_dir)
            runs.append(basenames)

        # Все 3 запуска должны дать одинаковое количество файлов
        counts = [len(r) for r in runs]
        assert len(set(counts)) == 1, (
            f"Количество файлов различается между запусками: {counts}"
        )
        assert counts[0] > 0, "Должен быть создан хотя бы 1 файл"

    def test_file_names_identical_across_runs(self, cards_data: CardsData, tmp_path):
        """Имена созданных файлов идентичны во всех 3 запусках."""
        runs: List[List[str]] = []

        for i in range(3):
            out_dir = os.path.join(str(tmp_path), f"names_{i}")
            basenames = _run_split_and_get_basenames(cards_data, out_dir)
            runs.append(basenames)

        # Все файлы из run_0 должны присутствовать в run_1 и run_2
        for fname in runs[0]:
            assert fname in runs[1], (
                f"Файл '{fname}' из run_0 отсутствует в run_1\n"
                f"  run_0: {runs[0]}\n"
                f"  run_1: {runs[1]}"
            )
            assert fname in runs[2], (
                f"Файл '{fname}' из run_0 отсутствует в run_2\n"
                f"  run_0: {runs[0]}\n"
                f"  run_2: {runs[2]}"
            )

        # Проверка в обратную сторону: лишних файлов нет
        for fname in runs[1]:
            assert fname in runs[0], (
                f"В run_1 появился лишний файл '{fname}', "
                f"отсутствующий в run_0:\n"
                f"  run_0: {runs[0]}\n"
                f"  run_1: {runs[1]}"
            )

    def test_each_split_file_is_valid_xlsx(self, cards_data: CardsData, tmp_path):
        """Каждый созданный split-файл — валидный .xlsx, читаемый openpyxl."""
        out_dir = os.path.join(str(tmp_path), "valid_check")
        created = split_cards_to_files(cards_data, out_dir, max_workers=1)

        assert len(created) > 0, "Должен быть создан хотя бы 1 файл"

        for fpath in created:
            assert os.path.isfile(fpath), f"Файл не существует: {fpath}"
            # Проверяем, что файл открывается openpyxl
            wb = openpyxl.load_workbook(fpath, data_only=True)
            assert len(wb.sheetnames) == 1, (
                f"Split-файл должен содержать ровно 1 лист, "
                f"получено {len(wb.sheetnames)}: {fpath}"
            )
            ws = wb.active
            # Должны быть заголовки
            assert ws.cell(1, 1).value is not None, (
                f"Нет данных в ячейке A1: {fpath}"
            )
            wb.close()

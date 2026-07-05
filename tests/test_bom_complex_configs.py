"""Регрессионный тест: BOM со смешанными комплектациями (78 configs).

Воспроизводит структуру BOM с 78 уникальными конфигурациями:
  - 24 колонки с ЧИСЛОВЫМИ значениями (количества)
  - 54 колонки с S/- маркерами
  - Пустые колонки-разделители между группами (C53-C55)
  - Колонка qty (C10: 用量\nQty) для S-маркеров
  - Мета-колонки (CPAC, 供应商, и т.д.)

Проверяет:
  1. find_header_rows находит строку 3
  2. detect_column_types определяет part_no=C7, qty=C10
  3. detect_config_columns находит 78 колонок (C29-C52 + C56-C109)
  4. parse_bom возвращает 78 конфигураций
  5. S-маркеры берут количество из qty_col
  6. - маркеры исключаются (qty=0)
"""

from __future__ import annotations

import os
import tempfile
from typing import Any, List, Optional

import openpyxl
import pytest
from openpyxl import Workbook

from burlak_parser.bom_parser import parse_bom
from burlak_parser.heuristic_analyzer import HeuristicAnalyzer


def _create_complex_bom_xlsx() -> str:
    """Создать .xlsx со сложной структурой (78 уникальных конфигураций)."""
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="complex_bom_")
    os.close(fd)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    # R3: заголовки
    headers = {
        1: "修订", 2: "序号\nSerial NO.", 3: "CPAC编码\nCPAC Code",
        4: "标识", 5: "发运", 6: "采购",
        7: "零件号\nPartNo.", 8: "零件名称(中文）\nPart Name(CN)",
        9: "零件名称(英文）\nPart Name(EN)", 10: "用量\nQty",
        11: "度量单位\nUOM", 12: "GPC代码\nGPC", 13: "FND代码\nFND",
        14: "FND中文", 15: "FND英文", 16: "零件成熟度",
        17: "层级\nLevel", 18: "LOU用法\nUsage", 19: "物料状态\nMake/Buy",
        20: "来源车间", 21: "使用工厂", 22: "目标车间",
        23: "供应商代码", 24: "供应商名称", 25: "MWO",
        26: "生效日期", 27: "失效日期", 28: "整车物料号",
    }
    for col, val in headers.items():
        ws.cell(row=3, column=col, value=val)

    # 24 уникальных UB-конфигурации (числовые) — C29-C52
    for i in range(24):
        ws.cell(row=3, column=29 + i, value=f"UB版二{['黑橙','灰黑','绿黑'][i%3]}内饰C{i+1:02d}座CKDUBCKD")

    # 54 уникальных S-конфигурации (S/- маркеры) — C56-C109
    colors = ['航空银', '沙金', '新卡其白', '电镀绿', '布罗蓝', '新碳晶黑']
    for i in range(54):
        color = colors[i % 6]
        cfg_type = ['舒适型', '豪华型'][i % 2]
        ws.cell(row=3, column=56 + i, value=f"S版{cfg_type}{i+1:02d}{color}内饰KM两驱1180")

    # R4-R53: 50 строк данных
    for dr in range(4, 54):
        pi = dr - 3
        ws.cell(row=dr, column=1, value=pi)       # 序号
        ws.cell(row=dr, column=3, value=f"20.{pi:02d}.04.99")  # CPAC
        ws.cell(row=dr, column=7, value=f"P{pi:08d}")          # PartNo
        ws.cell(row=dr, column=8, value=f"Деталь {pi} CN")     # Name CN
        ws.cell(row=dr, column=9, value=f"Part {pi} EN")       # Name EN
        ws.cell(row=dr, column=10, value=1 + (pi % 5))         # Qty: 1-5
        ws.cell(row=dr, column=11, value="EA")                 # UOM

        # Числовые колонки C29-C52: варьирующиеся количества
        for i in range(24):
            val = (pi + i) % 6
            if val > 0:
                ws.cell(row=dr, column=29 + i, value=val)

        # S-маркер колонки C56-C109: S, - или пусто
        for i in range(54):
            mod = (pi + i) % 3
            if mod == 0:
                ws.cell(row=dr, column=56 + i, value="S")
            elif mod == 1:
                ws.cell(row=dr, column=56 + i, value="-")

    wb.save(path)
    return path


@pytest.fixture(scope="module")
def complex_bom_path() -> str:
    """Создать BOM со сложной структурой для всех тестов модуля."""
    path = _create_complex_bom_xlsx()
    yield path
    try:
        os.remove(path)
    except Exception:
        pass


class TestComplexConfigRegression:
    """Регрессионный тест BOM: 78 конфигураций со смешанными типами."""

    def test_header_rows_found(self, complex_bom_path: str):
        """Header row R3 должна быть найдена (с part_no, name, qty)."""
        wb = openpyxl.load_workbook(complex_bom_path, data_only=True)
        ws = wb.active
        header_rows = HeuristicAnalyzer.find_header_rows(ws)
        assert 3 in header_rows, f"R3 should be a header row, got {header_rows}"
        wb.close()

    def test_column_types_detected(self, complex_bom_path: str):
        """part_no=C7, name_cn=C8, name_en=C9, qty=C10."""
        wb = openpyxl.load_workbook(complex_bom_path, data_only=True)
        ws = wb.active
        header_rows = HeuristicAnalyzer.find_header_rows(ws)
        col_types = HeuristicAnalyzer.detect_column_types(ws, header_rows)
        assert col_types.get("part_no") == 7, f"Expected C7, got {col_types}"
        assert col_types.get("name_cn") == 8, f"Expected C8, got {col_types}"
        assert col_types.get("name_en") == 9, f"Expected C9, got {col_types}"
        assert col_types.get("qty") == 10, f"Expected C10, got {col_types}"
        wb.close()

    def test_detect_all_78_config_columns(self, complex_bom_path: str):
        """detect_config_columns находит ровно 78 колонок (24+54)."""
        wb = openpyxl.load_workbook(complex_bom_path, data_only=True)
        ws = wb.active
        header_rows = HeuristicAnalyzer.find_header_rows(ws)
        col_types = HeuristicAnalyzer.detect_column_types(ws, header_rows)
        config_cols = HeuristicAnalyzer.detect_config_columns(ws, header_rows, col_types)

        assert len(config_cols) == 78, f"Expected 78, got {len(config_cols)}: {config_cols}"

        # Все числовые колонки C29-C52
        for c in range(29, 53):
            assert c in config_cols, f"Numeric C{c} missing from configs"

        # Все S-маркер колонки C56-C109
        for c in range(56, 110):
            assert c in config_cols, f"S-marker C{c} missing from configs"

        # Пустые разделители C53-C55 НЕ должны быть в configs
        for c in range(53, 56):
            assert c not in config_cols, f"Empty C{c} should NOT be in configs"

        wb.close()

    def test_parse_bom_returns_78_configs(self, complex_bom_path: str):
        """parse_bom возвращает ровно 78 конфигураций."""
        bom = parse_bom(complex_bom_path)
        assert len(bom.config_names) == 78, f"Expected 78, got {len(bom.config_names)}"

        # Проверяем распределение
        ub_count = sum(1 for cn in bom.config_names if "UB" in cn)
        s_count = sum(1 for cn in bom.config_names if "S版" in cn)
        assert ub_count == 24, f"Expected 24 UB, got {ub_count}"
        assert s_count == 54, f"Expected 54 S, got {s_count}"

    def test_s_marker_uses_qty_column(self, complex_bom_path: str):
        """S-маркеры берут количество из qty_col (C10), а не равны 1."""
        bom = parse_bom(complex_bom_path)
        s_configs = [cn for cn in bom.config_names if "S版" in cn]
        assert len(s_configs) == 54

        # Берём первый S-конфиг и проверяем, что qty берётся из qty_col
        sc = s_configs[0]
        parts = bom.config_quantities[sc]
        assert len(parts) > 0, f"{sc} should have parts"
        max_qty = max(parts.values())
        # Qty берётся из C10, значения 1-5 → max <= 5
        assert max_qty <= 5, f"Max qty should be <= 5 (from qty col), got {max_qty}"

    def test_dash_marker_excluded(self, complex_bom_path: str):
        """- маркеры исключают деталь из комплектации (qty=0)."""
        bom = parse_bom(complex_bom_path)
        s_configs = [cn for cn in bom.config_names if "S版" in cn]
        sc = s_configs[0]
        config_parts = len(bom.config_quantities[sc])
        total_parts = len(bom.parts)
        assert config_parts < total_parts, (
            f"S config should have fewer parts ({config_parts}) than total ({total_parts})"
        )

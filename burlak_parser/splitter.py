"""Модуль разделения многолистовых Excel-файлов на отдельные одностраничные файлы.

Использует метод «удаления лишнего» (а не «копирования нужного»):
  1. Загружает исходный .xlsx как ZIP-архив XML.
  2. Для каждого листа создаёт копию всего Workbook.
  3. Удаляет из копии все листы, кроме целевого.
  4. Очищает глобальные именованные диапазоны (defined names / named ranges),
     ссылающиеся на удалённые листы — это устраняет ошибку Excel
     "Removed Feature: Named range from /xl/workbook.xml part (Workbook)".

Преимущества метода:
  - 100% сохранение форматирования, стилей, картинок, шрифтов.
  - Сохраняется ширина колонок, высота строк, объединённые ячейки.
  - Сохраняются изображения, диаграммы, заморозка панелей.
  - Нет ошибки "Named range" при открытии.

Поддерживает параллелизацию через ProcessPoolExecutor.
"""

from __future__ import annotations

import io
import logging
import os
import re
import shutil
import warnings
import xml.etree.ElementTree as ET

try:
    from lxml import etree as _lxml_etree
    _HAS_LXML = True
except ImportError:
    _HAS_LXML = False
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl.utils.cell import range_boundaries, get_column_letter
except ImportError:
    # Fallback implementations if openpyxl is not installed
    import string

    def get_column_letter(col_idx: int) -> str:
        """Convert column index to Excel column letter (A=1, B=2, ...)."""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = string.ascii_uppercase[remainder] + result
        return result

    def range_boundaries(range_string: str) -> tuple:
        """Parse Excel range string like 'A1:B10' into (min_col, min_row, max_col, max_row)."""
        if ':' not in range_string:
            range_string = f"{range_string}:{range_string}"
        start, end = range_string.split(':', 1)
        start_col, start_row = _split_cell_ref(start.strip())
        end_col, end_row = _split_cell_ref(end.strip())
        return (start_col, start_row, end_col, end_row)

    def _split_cell_ref(ref: str) -> tuple:
        """Split cell reference like 'A1' into (col_index, row_number)."""
        match = re.match(r'^([A-Za-z]+)(\d+)$', ref.strip())
        if not match:
            raise ValueError(f"Invalid cell reference: {ref}")
        col_str, row_str = match.groups()
        col = 0
        for ch in col_str.upper():
            col = col * 26 + (ord(ch) - ord('A') + 1)
        return (col, int(row_str))

# Подавляем предупреждения openpyxl о DrawingML (неполная поддержка)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

logger = logging.getLogger(__name__)

# Символы, запрещённые в именах файлов Windows/Linux
_ILLEGAL_FS_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# Декоративные Unicode-символы, которые нужно удалять из имён файлов
# (звёздочки, ромбы, кружки, стрелки и т.д.)
_DECORATIVE_CHARS_RE = re.compile(r'[☆★●○◆◇■□▲△▼▽♠♣♥♦↗→←↑↓«»""''„]')

# Множественные подчёркивания/точки/пробелы → одинарные
_MULTI_SEP_RE = re.compile(r'[_ .]{2,}')

# Регулярка для cell reference: "A3390" → groups ("A", "3390")
_CELL_REF_RE = re.compile(r'^([A-Z]+)(\d+)$')

# Пространства имён Excel OOXML
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_PKG_RELS = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_DRAWING = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_DRAWINGML = "http://schemas.openxmlformats.org/drawingml/2006/main"
VML_NS = "urn:schemas-microsoft-com:vml"
OFFICE_NS = "urn:schemas-microsoft-com:office:office"

# Регистрируем пространства имён глобально
# Нужно зарегистрировать ВСЕ namespace-ы, которые могут встречаться
# в OOXML-файлах, чтобы избежать появления ns0:/ns1: префиксов.
# _serialize_xml() динамически переключает default namespace при каждом вызове.
ET.register_namespace('', NS_MAIN)
ET.register_namespace('r', NS_R)
ET.register_namespace('xdr', NS_DRAWING)
ET.register_namespace('a', NS_DRAWINGML)
ET.register_namespace('ct', NS_CT)

# lxml namespace map for proper OOXML serialization
_LXML_NS = {
    'xdr': NS_DRAWING,
    'a': NS_DRAWINGML,
    'r': NS_R,
    'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
    'x14': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/main',
    'x14ac': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac',
    'x15': 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main',
    'xm': 'http://schemas.microsoft.com/office/excel/2006/main',
    'rel': NS_PKG_RELS,
    'ct': NS_CT,
    'v': VML_NS,
    'o': OFFICE_NS,
    'x': 'urn:schemas-microsoft-com:office:excel',
    'pr': 'http://schemas.microsoft.com/office/2006/relationships',
}


def _lxml_to_bytes(root, xml_declaration: bool = True) -> bytes:
    """Serialize lxml element to bytes with MS Excel-compatible declaration.

    Uses lxml.etree.tostring with proper encoding and standalone declaration.
    Falls back to ElementTree if lxml is unavailable.
    """
    if _HAS_LXML:
        return _lxml_etree.tostring(
            root,
            xml_declaration=xml_declaration,
            encoding='UTF-8',
            standalone=True,
        )
    return _serialize_xml(root, NS_MAIN)


def _serialize_xml(root: ET.Element, default_ns_uri: str,
                   extra_ns: Optional[Dict[str, str]] = None) -> bytes:
    """Serialize an ET.Element with a specific default namespace.

    Saves and restores the global ET._namespace_map to avoid corruption.
    Used because different OOXML files require different default namespaces:
      - sheet XML:  NS_MAIN as default
      - Content_Types:  NS_CT as default
      - rels files:  NS_PKG_RELS as default

    CRITICAL: Uses custom XML declaration with standalone="yes", double quotes,
    and Windows-style \r\n line endings. MS Excel requires these for compatibility.
    Python 3.14: ET.tostring(standalone=True) raises TypeError, so we work around it.
    """
    old_default_uri = None
    for uri_key, prefix_val in ET._namespace_map.items():
        if prefix_val == '':
            old_default_uri = uri_key
            break
    old_extras: Dict[str, Optional[str]] = {}
    try:
        ET.register_namespace('', default_ns_uri)
        if extra_ns:
            for p, uri in extra_ns.items():
                old_extras[p] = ET._namespace_map.get(uri)
                ET.register_namespace(p, uri)
        # Serialize without declaration, then prepend MS Excel-compatible declaration
        body = ET.tostring(root, xml_declaration=False, encoding='UTF-8')
        declaration = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        return declaration + body
    finally:
        try:
            ET.register_namespace('', old_default_uri)
        except TypeError:
            pass
        if extra_ns:
            for p, uri in extra_ns.items():
                old_prefix = old_extras.get(p)
                if old_prefix is not None:
                    ET._namespace_map[uri] = old_prefix
                else:
                    ET._namespace_map.pop(uri, None)


# ─── Типы для вертикального split ───────────────────────────────────


@dataclass
class TableBoundary:
    """Границы одной таблицы (операции) внутри листа."""
    header_row: int         # Строка заголовка таблицы
    data_start: int         # Первая строка данных (header_row + 1)
    data_end: int           # Последняя строка данных
    operation_name: str = ""  # Название операции
    source_path: str = ""
    sheet_name: str = ""
    card_label: str = ""


class CardSplitter:
    """Сервис разделения многолистовых операционных карт на отдельные файлы."""

    def __init__(self, max_workers: Optional[int] = None):
        """Инициализировать сплиттер.

        Args:
            max_workers: Максимальное количество процессов для параллельного разделения.
                         По умолчанию: количество CPU.
        """
        self.max_workers = max_workers or os.cpu_count() or 4
        self.openpyxl_fallback_count = 0
        self.openpyxl_fallback_files: Set[str] = set()
        self.copy_fallback_count = 0
        self.copy_fallback_files: Set[str] = set()
        self.manifest: Dict[str, List[str]] = {}

    def split_file(
        self,
        source_path: str,
        output_dir: str,
        sheet_names: List[str],
        file_label: str = "",
    ) -> List[str]:
        """Разделить один .xlsx файл на несколько однолистовых файлов.

        Для .xls файлов (legacy) — копирует как есть без разделения.

        Args:
            source_path: Путь к исходному .xlsx/.xls файлу.
            output_dir: Директория для сохранения результатов.
            sheet_names: Имена листов, которые нужно выделить.
            file_label: Метка файла для именования выходных файлов.

        Returns:
            Список путей к созданным файлам.
        """
        os.makedirs(output_dir, exist_ok=True)
        created: List[str] = []
        original_name = os.path.basename(source_path)

        ext_lower = os.path.splitext(source_path)[1].lower()

        if ext_lower == ".xls":
            safe_label = _safe_filename(file_label)[:50] if file_label else ""
            basename = os.path.splitext(os.path.basename(source_path))[0]
            out_name = f"{safe_label}_{basename}.xls" if safe_label else f"{basename}.xls"
            output_path = os.path.join(output_dir, out_name)
            counter = 1
            while True:
                try:
                    fd = os.open(output_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                    os.close(fd)
                    break
                except FileExistsError:
                    base, ext = os.path.splitext(out_name)
                    output_path = os.path.join(output_dir, f"{base}_{counter}{ext}")
                    counter += 1
            shutil.copy2(source_path, output_path)
            created.append(output_path)
            self.manifest.setdefault(original_name, []).append(os.path.basename(output_path))
            logger.info("Скопирован .xls файл (без разделения): %s", os.path.basename(source_path))
            return created

        if ext_lower != ".xlsx":
            logger.debug("Пропуск не-.xlsx/.xls файла: %s", source_path)
            return created

        for sheet_name in sheet_names:
            safe_label = _safe_filename(file_label)[:50] if file_label else ""
            safe_sheet = _safe_filename(sheet_name)[:50]
            if safe_label:
                output_filename = f"{safe_label}_{safe_sheet}.xlsx"
            else:
                output_filename = f"{safe_sheet}.xlsx"

            output_path = os.path.join(output_dir, output_filename)

            # Skip if path already exists (pre-allocated by main thread)
            if os.path.exists(output_path):
                continue

            try:
                self._extract_sheet(source_path, output_path, sheet_name)
                created.append(output_path)
                self.manifest.setdefault(original_name, []).append(os.path.basename(output_path))
                logger.debug("Создан: %s", os.path.basename(output_path))
            except Exception as e:
                logger.warning(
                    "Ошибка разделения листа '%s' из %s: %s",
                    sheet_name, os.path.basename(source_path), e,
                )

        return created

    def split_many_parallel(
        self,
        tasks: List[Tuple[str, str, List[str], str]],
    ) -> Tuple[List[str], List[Tuple[str, str]], int, List[str], Dict[str, List[str]]]:
        """Разделить множество файлов параллельно.

        Гарантирует детерминированный порядок: результаты сортируются
        по полному пути для воспроизводимости.

        Args:
            tasks: Список кортежей (source_path, output_dir, sheet_names, file_label).

        Returns:
            Кортеж (all_created_files, errors, openpyxl_fallback_count,
                    openpyxl_fallback_files, manifest).
        """
        all_created: List[str] = []
        errors: List[Tuple[str, str]] = []
        all_openpyxl_count = 0
        all_openpyxl_files: List[str] = []
        merged_manifest: Dict[str, List[str]] = {}

        # Предвычисляем все пути детерминированно (синхронно, главный поток)
        path_map = preallocate_split_paths(tasks, tasks[0][1] if tasks else "")

        # Собираем плоские задачи (source_path, output_path, sheet_name)
        sheet_tasks: List[Tuple[str, str, str]] = []
        for source_path, out_dir, sheet_names, file_label in tasks:
            for sheet_name in sheet_names:
                output_path = path_map.get((source_path, sheet_name))
                if output_path:
                    sheet_tasks.append((source_path, output_path, sheet_name))

        with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {}
            for src, out, sheet in sheet_tasks:
                future = executor.submit(
                    _extract_to_path_worker,
                    src, out, sheet,
                )
                futures[future] = (src, out, sheet)

            for future in as_completed(futures):
                src, out, sheet = futures[future]
                try:
                    worker_result = future.result()
                    result_path = worker_result.get("path")
                    err_msg = worker_result.get("error")
                    source_basename = worker_result.get("source_basename", "")

                    if result_path:
                        all_created.append(result_path)
                        if worker_result.get("used_fallback"):
                            all_openpyxl_count += 1
                            if source_basename:
                                all_openpyxl_files.append(source_basename)
                        if source_basename:
                            merged_manifest.setdefault(source_basename, []).append(
                                os.path.basename(result_path),
                            )
                    else:
                        logger.error(
                            "Ошибка разделения %s: %s",
                            os.path.basename(src), err_msg,
                        )
                        errors.append((src, err_msg or "Unknown error"))
                except Exception as e:
                    err_msg = str(e)
                    logger.error(
                        "Критическая ошибка параллельного разделения %s: %s",
                        os.path.basename(src), err_msg,
                    )
                    errors.append((src, err_msg))

        # Сортируем для детерминированного порядка
        all_created.sort()
        errors.sort(key=lambda x: x[0])
        all_openpyxl_files.sort()

        # Сортируем значения в манифесте
        for orig in merged_manifest:
            merged_manifest[orig].sort()

        return all_created, errors, all_openpyxl_count, all_openpyxl_files, merged_manifest

    def _extract_sheet(
        self, source_path: str, output_path: str, keep_sheet_name: str,
    ) -> None:
        """Выделить один лист из .xlsx файла.

        Стратегия ЛИНЕЙНАЯ (без рекурсии), приоритет производительности:
          1. ZIP-метод: быстрый (миллисекунды), обрабатывает 90%+ файлов.
          2. При ошибке ZIP → ОДНА попытка openpyxl (медленный, для WPS/битых).
          3. При ошибке openpyxl → исключение.

        Args:
            source_path: Путь к исходному .xlsx файлу.
            output_path: Путь для сохранения нового .xlsx файла.
            keep_sheet_name: Имя листа, который нужно оставить.

        Raises:
            ValueError: Если целевой лист не найден в файле.
            Exception: Если оба метода завершились ошибкой.
        """
        # Попытка 1: ZIP (быстро — миллисекунды на файл)
        try:
            self._extract_sheet_via_zip(source_path, output_path, keep_sheet_name)
            if not _validate_split_file(output_path):
                try:
                    os.remove(output_path)
                except OSError:
                    pass
                raise ValueError(f"Invalid split output: {output_path}")
            return
        except Exception as e:
            logger.warning(
                "ZIP-метод не смог разделить %s: %s. Пробуем openpyxl...",
                os.path.basename(source_path), e,
            )

        # Попытка 2: openpyxl (медленно — для WPS/битых файлов, одна попытка)
        try:
            self._extract_sheet_via_openpyxl(source_path, output_path, keep_sheet_name)
            self.openpyxl_fallback_count += 1
            self.openpyxl_fallback_files.add(os.path.basename(source_path))
            logger.info(
                "openpyxl успешно разделил лист: %s в файле %s",
                keep_sheet_name, os.path.basename(source_path),
            )
            return
        except Exception as openpyxl_e:
            logger.warning(
                "openpyxl не смог разделить лист '%s' из %s: %s. "
                "Пробуем скопировать исходный файл...",
                keep_sheet_name, os.path.basename(source_path), openpyxl_e,
            )

        # Попытка 3: Копирование исходного файла (последний шанс)
        # Если файл валидный и открывается, но не поддаётся разделению —
        # копируем его как есть. Лучше получить неразделённый файл,
        # чем потерять данные из-за отправки в corrupted_cards.
        try:
            _copy_source_as_fallback(source_path, output_path)
            self.copy_fallback_count += 1
            self.copy_fallback_files.add(os.path.basename(source_path))
            logger.info(
                "Исходный файл скопирован (copy fallback) для листа '%s' из %s "
                "(возможно несколько листов в выходном файле)",
                keep_sheet_name, os.path.basename(source_path),
            )
            return
        except Exception as copy_e:
            logger.error(
                "Все три метода разделения листа '%s' из %s завершились ошибкой. "
                "ZIP: см. выше. openpyxl: %s. Copy: %s",
                keep_sheet_name, os.path.basename(source_path), openpyxl_e, copy_e,
            )
            raise

    def _extract_sheet_via_openpyxl(
        self, source_path: str, output_path: str, keep_sheet_name: str,
    ) -> None:
        """Выделить один лист через openpyxl (load → remove sheets → save).

        Этот метод корректно обрабатывает файлы, созданные WPS Office
        и другими генераторами OOXML, которые могут содержать
        нестандартные CRC-суммы или повреждённые записи ZIP.

        Включает обход бага WPS: DefinedNameDict без атрибута definedName.

        Args:
            source_path: Путь к исходному .xlsx файлу.
            output_path: Путь для сохранения нового .xlsx файла.
            keep_sheet_name: Имя листа, который нужно оставить.

        Raises:
            ValueError: Если целевой лист не найден.
            Exception: При ошибке загрузки/сохранения openpyxl.
        """
        import openpyxl

        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            wb = openpyxl.load_workbook(source_path)
        sheet_names = wb.sheetnames

        if keep_sheet_name not in sheet_names:
            wb.close()
            raise ValueError(f"Лист '{keep_sheet_name}' не найден в файле")

        if len(sheet_names) <= 1:
            wb.save(output_path)
            wb.close()
            return

        sheets_to_remove = [n for n in sheet_names if n != keep_sheet_name]

        # ── WPS BUG FIX: Патчим DefinedNameDict перед удалением листов ──
        # WPS Office создаёт повреждённые OOXML, где wb.defined_names
        # не имеет атрибута definedName. openpyxl падает при del wb[sheet]
        # с AttributeError: 'DefinedNameDict' object has no attribute 'definedName'.
        # Решение: принудительно создаём пустые атрибуты.
        dn = getattr(wb, 'defined_names', None)
        if dn is not None:
            if not hasattr(dn, 'definedName'):
                dn.definedName = []
            if not hasattr(dn, 'elements'):
                dn.elements = []

        # Удаляем named ranges, ссылающиеся на удаляемые листы
        try:
            if dn is not None and dn.definedName:
                to_delete = []
                for defined_name in dn.definedName:
                    attr_text = getattr(defined_name, 'attr_text', None) or str(defined_name)
                    for deleted in sheets_to_remove:
                        if deleted in attr_text or f"'{deleted}'" in attr_text:
                            to_delete.append(defined_name)
                            break
                for defined_name in to_delete:
                    try:
                        dn.definedName.remove(defined_name)
                    except Exception as e:
                        logger.debug("Failed to remove defined name: %s", e)
        except Exception as e:
            logger.debug("Defined name cleanup failed (non-critical): %s", e)

        for name in sheets_to_remove:
            try:
                del wb[name]
            except Exception as e:
                logger.debug("Failed to remove sheet %s: %s", name, e)

        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            wb.save(output_path)
        wb.close()

    def _extract_sheet_via_zip(
        self, source_path: str, output_path: str, keep_sheet_name: str,
    ) -> None:
        """Выделить один лист через чистую ZIP-манипуляцию.

        АЛГОРИТМ (КЛЮЧЕВОЙ):
        Стратегия «сохраняем только нужное»:
          1. Читаем оригинальный ZIP в память.
          2. Находим rId и путь к сохранённому листу.
          3. Рекурсивно трассируем все .rels от листа → находим все нужные файлы
             (drawing XML, VML, OLE, изображения, printerSettings, их .rels).
          4. Добавляем обязательные: workbook, styles, theme, sharedStrings, docProps.
          5. Создаём НОВЫЙ workbook.xml: только 1 лист + очищенные definedNames.
             ВАЖНО: все остальные элементы (fileVersion, workbookPr, bookViews,
             calcPr, AlternateContent) копируются из оригинала AS-IS.
          6. Создаём НОВЫЙ workbook.xml.rels: только лист + shared items.
          7. Фильтруем Content_Types.xml: только Override для существующих файлов.
          8. Записываем новый ZIP.

        ВАЖНО: НИКАКОГО openpyxl, НИКАКОГО переименования файлов!
        Все оригинальные XML-файлы и бинарные данные копируются AS-IS
        с их оригинальными именами (sheet8.xml остаётся sheet8.xml).
        Это гарантирует 100% сохранение всех ссылок внутри drawing,
        VML, OLE и других файлов.

        Args:
            source_path: Путь к исходному .xlsx файлу.
            output_path: Путь для сохранения нового .xlsx файла.
            keep_sheet_name: Имя листа, который нужно оставить.

        Raises:
            ValueError: Если целевой лист не найден в файле.
        """
        # ── ФАЗА 1: Прочитать оригинальный ZIP ──
        # Пробуем несколько кодировок для имён файлов в ZIP.
        # Китайские Windows системы создают ZIP с GBK-кодировкой имён,
        # но Python zipfile по умолчанию использует CP437, что ломает
        # пути к файлам (mojibake) и делает невозможным поиск по имени.
        with open(source_path, 'rb') as f:
            zip_data = f.read()

        orig_entries: Dict[str, bytes] = {}
        _zip_loaded = False
        for _enc in (None, 'gbk', 'utf-8', 'cp1251', 'latin-1'):
            try:
                kwargs = {'metadata_encoding': _enc} if _enc else {}
                with zipfile.ZipFile(io.BytesIO(zip_data), 'r', **kwargs) as zf:
                    for name in zf.namelist():
                        try:
                            orig_entries[name] = zf.read(name)
                        except (zipfile.BadZipFile, Exception):
                            pass
                _zip_loaded = True
                break
            except (zipfile.BadZipFile, UnicodeDecodeError):
                continue
        if not _zip_loaded:
            raise ValueError(f"Cannot read source ZIP with any encoding: {source_path}")

        # ── ФАЗА 2: Найти лист в workbook.xml ──
        wb_xml = orig_entries.get('xl/workbook.xml')
        if wb_xml is None:
            raise ValueError("xl/workbook.xml not found")

        wb_root = ET.fromstring(wb_xml)
        sheets_elem = wb_root.find(f'{{{NS_MAIN}}}sheets')
        if sheets_elem is None:
            raise ValueError("<sheets> not found in original workbook.xml")

        # Находим rId сохранённого листа
        target_r_id: Optional[str] = None
        for sheet_el in sheets_elem.findall(f'{{{NS_MAIN}}}sheet'):
            if sheet_el.get('name') == keep_sheet_name:
                target_r_id = sheet_el.get(f'{{{NS_R}}}id') or sheet_el.get('r:id')
                break

        if target_r_id is None:
            raise ValueError(f"Sheet '{keep_sheet_name}' not found")

        # ── ФАЗА 3a: Найти путь к листу из workbook.xml.rels ──
        rels_xml = orig_entries.get('xl/_rels/workbook.xml.rels')
        if rels_xml is None:
            raise ValueError("xl/_rels/workbook.xml.rels not found")

        rels_root = ET.fromstring(rels_xml)
        orig_sheet_path = ''
        for rel_el in rels_root:
            if rel_el.get('Id') == target_r_id:
                orig_sheet_path = rel_el.get('Target', '')
                break

        if not orig_sheet_path:
            raise ValueError(f"No target for rId {target_r_id}")

        # Нормализуем путь
        orig_sheet_path = orig_sheet_path.lstrip('/')
        if not orig_sheet_path.startswith('xl/'):
            orig_sheet_path = 'xl/' + orig_sheet_path

        # ── ФАЗА 3b: Рекурсивно трассировать все .rels ──
        needed: Set[str] = set()

        def _trace_rels(rels_path: str, base_dir: str) -> None:
            """Рекурсивно трассировать .rels, добавляя все найденные файлы."""
            if rels_path not in orig_entries:
                return
            try:
                tr_root = ET.fromstring(orig_entries[rels_path])
                for tr_el in tr_root:
                    target = tr_el.get('Target', '')
                    if not target:
                        continue
                    # Ресолвим относительный путь от base_dir
                    resolved = os.path.normpath(
                        os.path.join(base_dir, target)
                    ).replace(os.sep, '/')
                    if resolved in orig_entries and resolved not in needed:
                        needed.add(resolved)
                        # Ищем под-rels (drawing.rels, vml.rels)
                        res_dir = os.path.dirname(resolved)
                        res_base = os.path.basename(resolved)
                        sub_rels = f"{res_dir}/_rels/{res_base}.rels"
                        if sub_rels in orig_entries:
                            needed.add(sub_rels)
                            _trace_rels(sub_rels, res_dir)
            except Exception as e:
                logger.warning("Trace rels failed for %s: %s", rels_path, e)

        # Всегда нужны базовые файлы
        needed.add('[Content_Types].xml')
        needed.add('_rels/.rels')
        needed.add('xl/workbook.xml')
        needed.add('xl/_rels/workbook.xml.rels')

        # Сам лист
        needed.add(orig_sheet_path)

        # .rels файл листа и его рекурсивные зависимости
        sheet_dir = os.path.dirname(orig_sheet_path)
        sheet_base = os.path.basename(orig_sheet_path)
        sheet_rels_path = f"{sheet_dir}/_rels/{sheet_base}.rels"
        if sheet_rels_path in orig_entries:
            needed.add(sheet_rels_path)
            _trace_rels(sheet_rels_path, sheet_dir)

        # Добавляем зависимости из workbook.xml.rels.
        # Оставляем shared items, customXml, datastore, VBA и т.д.,
        # но исключаем другие листы (worksheet/chartsheet/dialogsheet)
        # и calcChain (цепь вычислений, невалидна после удаления листов).
        for rel_el in rels_root:
            rel_id = rel_el.get('Id', '')
            rel_type = rel_el.get('Type', '').lower()
            rel_target = rel_el.get('Target', '')
            if rel_id == target_r_id:
                continue  # Пропускаем сам лист (уже добавлен)
            # Исключаем связи на другие листы и невалидный calcChain
            if any(t in rel_type for t in [
                'worksheet', 'chartsheet', 'dialogsheet', 'calcchain',
            ]):
                continue
            # Ресолвим путь
            if rel_target.startswith('/'):
                resolved = rel_target.lstrip('/')
            else:
                resolved = os.path.normpath(
                    os.path.join('xl', rel_target)
                ).replace(os.sep, '/')
            if resolved in orig_entries:
                needed.add(resolved)
                # Трассируем под-связи (например customXml/_rels/item1.xml.rels)
                res_dir = os.path.dirname(resolved)
                res_base = os.path.basename(resolved)
                sub_rels = f"{res_dir}/_rels/{res_base}.rels"
                if sub_rels in orig_entries:
                    needed.add(sub_rels)
                    _trace_rels(sub_rels, res_dir)

        # Добавляем docProps (core, app, custom) — не влияют на загрузку листа
        doc_props = [n for n in orig_entries if n.startswith('docProps/')]
        needed.update(doc_props)

        # Добавляем customXml (если есть)
        custom_xml = [n for n in orig_entries if n.startswith('customXml/')]
        needed.update(custom_xml)

        # ── ФАЗА 4: Собрать имена удалённых листов ──
        other_sheet_names: Set[str] = set()
        for sheet_el in sheets_elem.findall(f'{{{NS_MAIN}}}sheet'):
            sn = sheet_el.get('name', '')
            if sn != keep_sheet_name:
                other_sheet_names.add(sn)

        # ── ФАЗА 5: Модифицировать workbook.xml через строковые операции ──
        # ВАЖНО: используем строковые операции, а НЕ XML парсинг,
        # чтобы сохранить оригинальные namespace declarations, XML declaration,
        # line endings и все остальные детали исходного файла AS-IS.
        new_wb_text = _modify_workbook_xml_text(
            orig_entries['xl/workbook.xml'].decode('utf-8'),
            keep_sheet_name,
            target_r_id,
            other_sheet_names,
        )

        # ── ФАЗА 6: Модифицировать workbook.xml.rels — удалить лишние Relationship ──
        new_rels_text = _modify_workbook_rels_text(
            orig_entries['xl/_rels/workbook.xml.rels'].decode('utf-8'),
            target_r_id,
        )

        # ── ФАЗА 7: Собрать выходной словарь ──
        output_entries: Dict[str, bytes] = {}

        for name in needed:
            if name == 'xl/workbook.xml':
                output_entries[name] = new_wb_text.encode('utf-8')
            elif name == 'xl/_rels/workbook.xml.rels':
                output_entries[name] = new_rels_text.encode('utf-8')
            elif name.startswith('xl/printerSettings/'):
                # Skip printerSettings — binary files that reference removed sheets
                # cause "Removed Part" errors in Excel
                continue
            else:
                output_entries[name] = orig_entries[name]

        # ── ФАЗА 7b: Очистить .rels файлы от ссылок на удалённые printerSettings ──
        for rels_name in list(output_entries.keys()):
            if rels_name.endswith('.rels') and 'printerSettings' not in rels_name:
                try:
                    rels_text = output_entries[rels_name].decode('utf-8')
                    if 'printerSettings' in rels_text:
                        # Remove Relationship entries pointing to printerSettings
                        cleaned = re.sub(
                            r'<Relationship[^>]*Target="[^"]*printerSettings[^"]*"[^>]*/>\s*',
                            '', rels_text)
                        output_entries[rels_name] = cleaned.encode('utf-8')
                except Exception:
                    pass

        # ── ФАЗА 8: Фильтровать Content_Types.xml — удалить Override для отсутствующих файлов ──
        if '[Content_Types].xml' in needed:
            ct_text = orig_entries['[Content_Types].xml'].decode('utf-8')
            new_ct_text = _filter_content_types_text(ct_text, set(output_entries.keys()))
            output_entries['[Content_Types].xml'] = new_ct_text.encode('utf-8')

        # ── ФАЗА 8: Записать новый ZIP с сохранением оригинального сжатия ──
        # ВАЖНО: MS Excel требует, чтобы изображения (PNG, EMF, JPEG) были
        # STORED (без сжатия), а XML/DATA файлы — DEFLATED.
        # Используем оригинальный compression_type если известен.
        if os.path.exists(output_path):
            os.remove(output_path)

        def _get_compress_type(name: str) -> int:
            """Определить метод сжатия: STORED для изображений, DEFLATED для всего остального.

            MS Office хранит изображения в исходном виде (STORED), так как они
            уже сжаты. XML и другие текстовые данные — DEFLATED.
            """
            name_lower = name.lower()
            # Изображения — без сжатия (уже сжаты, DEFLATE не помогает)
            if any(name_lower.endswith(ext) for ext in ['.png', '.emf', '.wmf', '.jpeg', '.jpg',
                                                         '.gif', '.tiff', '.tif', '.bmp', '.svg']):
                return zipfile.ZIP_STORED
            return zipfile.ZIP_DEFLATED

        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name in sorted(output_entries.keys()):
                compress_type = _get_compress_type(name)
                zout.writestr(name, output_entries[name], compress_type=compress_type)


_CT_CACHE: Dict[str, Optional[str]] = {}

def _modify_workbook_xml_text(
    xml_text: str,
    keep_sheet_name: str,
    target_r_id: str,
    other_sheet_names: Set[str],
) -> str:
    """Модифицировать workbook.xml строковыми операциями.

    1. Удалить лишние <sheet> из <sheets>.
    2. Удалить <definedName>, ссылающиеся на удалённые листы.

    ВСЁ остальное сохраняется AS-IS (XML declaration, namespace, line endings).
    """
    # ── 1. Замена <sheets> — оставляем только 1 лист ──
    def _replace_sheets(m: re.Match) -> str:
        """Callback для замены содержимого <sheets>."""
        open_tag = m.group(1)
        close_tag = m.group(3)
        content = m.group(2)
        # Ищем сохранённый лист по name или r:id
        kept = None
        for sh in re.finditer(r'<sheet[^>]*/>', content):
            sh_tag = sh.group(0)
            name_m = re.search(r'name="([^"]+)"', sh_tag)
            if name_m and name_m.group(1) == keep_sheet_name:
                kept = sh_tag
                break
        # Fallback: по r:id
        if kept is None:
            for sh in re.finditer(r'<sheet[^>]*/>', content):
                sh_tag = sh.group(0)
                rid_m = re.search(r'r:id="([^"]+)"', sh_tag)
                if rid_m and rid_m.group(1) == target_r_id:
                    kept = sh_tag
                    break
        if kept:
            return f'{open_tag}\n{kept}\n{close_tag}'
        return m.group(0)  # fallback: без изменений

    xml_text = re.sub(r'(<sheets[^>]*>)(.*?)(</sheets>)', _replace_sheets, xml_text, count=1, flags=re.DOTALL)

    # ── 2. Очистка definedNames ──
    def _filter_defined_names(m: re.Match) -> str:
        """Callback: удалить definedName, ссылающиеся на other_sheet_names."""
        dn_block = m.group(0)
        # Находим границы тега
        dn_open_m = re.match(r'(<definedNames[^>]*>)', dn_block)
        if not dn_open_m:
            return dn_block
        dn_open = dn_open_m.group(1)
        # Находим закрывающий тег
        close_idx = dn_block.rfind('</definedNames>')
        if close_idx == -1:
            return dn_block
        content = dn_block[len(dn_open):close_idx]

        kept_lines = []
        for dn_match_inner in re.finditer(r'<definedName[^>]*>.*?</definedName>', content, re.DOTALL):
            dn_xml = dn_match_inner.group(0)
            dn_text = re.sub(r'<[^>]+>', '', dn_xml).strip()  # extract text content
            formula = dn_text
            should_remove = False
            for deleted_name in other_sheet_names:
                if f"'{deleted_name}'!" in formula or formula.startswith(f"{deleted_name}!"):
                    should_remove = True
                    break
            if not should_remove:
                # Обновляем localSheetId на 0 (сохранённый лист теперь единственный)
                dn_xml = re.sub(r'localSheetId="[^"]+"', 'localSheetId="0"', dn_xml)
                kept_lines.append(dn_xml)

        if not kept_lines:
            # definedNames пуст — удаляем весь блок
            return ''
        return dn_open + ''.join(kept_lines) + '</definedNames>'

    xml_text = re.sub(r'<definedNames[^>]*>.*?</definedNames>', _filter_defined_names, xml_text, count=1, flags=re.DOTALL)

    # ── 3. Очистка bookViews/workbookView — сбросить activeTab/firstSheet ──
    def _fix_workbook_view(m: re.Match) -> str:
        tag = m.group(0)
        tag = re.sub(r'activeTab="[^"]*"', 'activeTab="0"', tag)
        tag = re.sub(r'firstSheet="[^"]*"', 'firstSheet="0"', tag)
        return tag

    xml_text = re.sub(r'<(?:[\w\-]+:)?workbookView\b[^>]*/>', _fix_workbook_view, xml_text)

    # ── 4. Удалить customWorkbookViews ──
    xml_text = re.sub(r'<customWorkbookViews[^>]*>.*?</customWorkbookViews>', '', xml_text, flags=re.DOTALL)

    return xml_text


def _modify_workbook_rels_text(
    rels_text: str,
    target_r_id: str,
) -> str:
    """Модифицировать workbook.xml.rels — удалить Relationship для других листов.

    Оставляет ТОЛЬКО:
      - worksheet (сохранённый лист)
      - styles
      - theme
      - sharedStrings

    ВСЁ остальное (XML declaration, форматирование) сохраняется AS-IS.
    Используется re.sub для удаления отдельных <Relationship .../> строк.
    """
    def _keep_relevant_rels(m: re.Match) -> str:
        """Callback: вернуть Relationship строку только если она нужна."""
        rel = m.group(0)
        rid_m = re.search(r'Id="([^"]+)"', rel)
        rtype_m = re.search(r'Type="([^"]+)"', rel)
        rid = rid_m.group(1) if rid_m else ''
        rtype = rtype_m.group(1).lower() if rtype_m else ''

        # Всегда оставляем сохранённый лист
        if rid == target_r_id:
            return rel
        # Удаляем связи других листов и невалидный calcChain
        if any(t in rtype for t in [
            'worksheet', 'chartsheet', 'dialogsheet', 'calcchain',
        ]):
            return ''
        # Остальные связи (styles, theme, sharedStrings, customXml, VBA и др.) сохраняем
        return rel

    return re.sub(r'<Relationship[^>]*/>', _keep_relevant_rels, rels_text)


def _filter_content_types_text(
    ct_text: str,
    existing_files: Set[str],
) -> str:
    """Фильтровать [Content_Types].xml — удалить Override для несуществующих файлов.

    Args:
        ct_text: Оригинальный текст [Content_Types].xml.
        existing_files: Множество путей файлов в выходном ZIP.

    Returns:
        Отфильтрованный XML текст (ВСЁ остальное AS-IS).
    """
    def _filter_override(m: re.Match) -> str:
        """Callback: вернуть Override только если файл существует."""
        override_line = m.group(0)
        pn_m = re.search(r'PartName="([^"]+)"', override_line)
        if pn_m:
            part_name = pn_m.group(1)
            if part_name.startswith('/'):
                clean_name = part_name[1:]
            else:
                clean_name = part_name
            if clean_name not in existing_files:
                return ''  # Удаляем
        return override_line

    return re.sub(r'<Override[^>]*>(?:</Override>)?', _filter_override, ct_text)


def _infer_content_type(path: str) -> Optional[str]:
    """Определить OOXML ContentType по пути файла."""
    if path in _CT_CACHE:
        return _CT_CACHE[path]

    result: Optional[str] = None
    path_lower = path.lower()

    if path_lower.endswith('.xml'):
        if 'drawing' in path_lower and 'rels' not in path_lower:
            result = 'application/vnd.openxmlformats-officedocument.drawing+xml'
        elif 'vml' in path_lower:
            result = 'application/vnd.openxmlformats-officedocument.vmlDrawing'
    elif path_lower.endswith('.bin'):
        result = 'application/vnd.openxmlformats-officedocument.oleObject'
    elif path_lower.endswith('.rels'):
        result = 'application/vnd.openxmlformats-package.relationships+xml'
    elif path_lower.endswith('.png'):
        result = 'image/png'
    elif path_lower.endswith('.jpeg') or path_lower.endswith('.jpg'):
        result = 'image/jpeg'
    elif path_lower.endswith('.emf'):
        result = 'image/x-emf'
    elif path_lower.endswith('.wmf'):
        result = 'image/x-wmf'
    elif path_lower.endswith('.gif'):
        result = 'image/gif'
    elif path_lower.endswith('.tiff') or path_lower.endswith('.tif'):
        result = 'image/tiff'
    elif path_lower.endswith('.bmp'):
        result = 'image/bmp'
    elif path_lower.endswith('.svg'):
        result = 'image/svg+xml'

    _CT_CACHE[path] = result
    return result


def _validate_split_file(path: str) -> bool:
    """Verify a split .xlsx file has valid sheet XML and can be opened.

    Использует ValidationPipeline для comprehensive проверки
    (structural + schema + split-quality levels).

    Returns True if the file is valid, False if it should be deleted.
    """
    try:
        from burlak_parser.validator import validate_split_file
        result = validate_split_file(
            path,
            has_images_in_original=True,  #保守но: предполагаем что были изображения
        )
        if not result.is_valid:
            for issue in result.errors:
                logger.warning(
                    "Invalid split file %s: [%s] %s",
                    os.path.basename(path), issue.level, issue.message,
                )
        return result.is_valid
    except ImportError:
        # Fallback: если validator недоступен, используем простую проверку
        try:
            with zipfile.ZipFile(path, 'r') as zf:
                has_sheet = False
                for name in zf.namelist():
                    if (name.endswith('.xml')
                            and 'sheet' in name.lower()
                            and '_rels' not in name):
                        data = zf.read(name)
                        root = ET.fromstring(data)
                        ns = f'{{{NS_MAIN}}}sheetData'
                        if root.find(ns) is None:
                            logger.warning(
                                "Invalid split file %s: missing sheetData in %s",
                                os.path.basename(path), name)
                            return False
                        has_sheet = True
                        break
                return has_sheet
        except (zipfile.BadZipFile, ET.ParseError, OSError) as e:
            logger.warning("Invalid split file %s: %s", os.path.basename(path), e)
            return False


def _safe_filename(name: str) -> str:
    """Очистить имя файла, сохранив Unicode-символы.

    Удаляет:
      - Символы, запрещённые в именах файлов ОС: < > : " / \\ | ? *
      - Управляющие символы (0x00-0x1f)
      - Декоративные Unicode: ☆ ★ ● ○ ◆ ◇ ■ □ и т.д.
      - Суррогатные пары (некорректный Unicode)
      - Множественные подчёркивания/точки/пробелы → одинарные

    Сохраняет:
      - Китайские иероглифы (CJK)
      - Кириллицу
      - Латиницу и цифры

    Args:
        name: Исходное имя файла.

    Returns:
        Безопасное имя файла с сохранёнными кириллицей/иероглифами.
    """
    # Удаляем суррогатные пары (некорректный Unicode из битых кодировок)
    result = name.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')
    result = _ILLEGAL_FS_CHARS_RE.sub("_", result)
    result = _DECORATIVE_CHARS_RE.sub("", result)
    result = _MULTI_SEP_RE.sub("_", result)
    return result.strip("_ .")


def _collect_related_files(
    zip_entries: Dict[str, bytes],
    removed_sheet: str,
    files_to_remove: Set[str],
) -> None:
    """Собрать все файлы, связанные с удаляемым листом (.rels, drawings, VML, charts).

    Args:
        zip_entries: Словарь {имя_в_zip: содержимое}.
        removed_sheet: Путь к удаляемому листу (напр. 'xl/worksheets/sheet2.xml').
        files_to_remove: Множество для добавления найденных файлов.
    """
    # .rels файл для листа
    base = os.path.basename(removed_sheet)
    removed_rels = f"xl/worksheets/_rels/{base}.rels"
    if removed_rels in zip_entries:
        files_to_remove.add(removed_rels)

        # Находим связанные drawings, VML, charts
        try:
            sr_root = ET.fromstring(zip_entries[removed_rels])
            # OOXML relationship targets resolve relative to the package part
            # (e.g. xl/worksheets/), NOT relative to the .rels directory
            # (e.g. xl/worksheets/_rels/). Go up one extra level.
            sr_dir = os.path.dirname(os.path.dirname(removed_rels))
            for sr_el in sr_root:
                sr_target = sr_el.get('Target', '')
                # Резолвим относительный путь (../drawings/drawing1.xml)
                resolved = os.path.normpath(os.path.join(sr_dir, sr_target))
                resolved = resolved.replace(os.sep, '/')
                files_to_remove.add(resolved)

                # Рекурсивно: .rels для drawing, VML
                resolved_base = os.path.basename(resolved)
                resolved_dir = os.path.dirname(resolved)
                resolved_rels = f"{resolved_dir}/_rels/{resolved_base}.rels"
                if resolved_rels in zip_entries:
                    files_to_remove.add(resolved_rels)
        except Exception as e:
            logger.debug("Relationship resolution failed: %s", e)


def _clean_named_ranges(
    wb_root: ET.Element,
    deleted_sheet_names: Set[str],
    keep_sheet_name: str,
) -> None:
    """Очистить definedNames (named ranges), ссылающиеся на удалённые листы.

    Это КЛЮЧЕВОЙ шаг для устранения ошибки Excel:
    "Removed Feature: Named range from /xl/workbook.xml part (Workbook)"

    Алгоритм:
      1. Найти элемент <definedNames> в workbook.xml.
      2. Для каждого <definedName> проверить, ссылается ли он на удалённый лист.
      3. Ссылка на лист в definedName обычно выглядит как: SheetName!$A$1
         или заключена в одиночные кавычки если имя с пробелами: 'Sheet Name'!$A$1.
      4. Удалить все definedNames, ссылающиеся на удалённые листы.

    Args:
        wb_root: Корневой элемент workbook.xml.
        deleted_sheet_names: Множество имён удалённых листов.
        keep_sheet_name: Имя оставленного листа.
    """
    defined_names_elem = wb_root.find(f'{{{NS_MAIN}}}definedNames')
    if defined_names_elem is None:
        return  # Нет именованных диапазонов — нечего чистить

    names_to_remove: List[ET.Element] = []

    for dn in defined_names_elem.findall(f'{{{NS_MAIN}}}definedName'):
        # Текст definedName — это формула со ссылкой на лист
        formula = (dn.text or '').strip()
        name_attr = dn.get('name', '')

        # Проверяем, ссылается ли definedName на удалённый лист
        # Шаблоны ссылок:
        #   'Sheet Name'!$A$1:$B$2
        #   SheetName!$A$1
        #   SheetName!$A$1:$B$2
        should_remove = False

        for deleted_name in deleted_sheet_names:
            # Проверка с кавычками (для имён с пробелами/спецсимволами)
            if f"'{deleted_name}'!" in formula:
                should_remove = True
                break
            # Проверка без кавычек
            if formula.startswith(f"{deleted_name}!"):
                should_remove = True
                break
            # Проверка на вхождение (менее точная, но покрывает edge cases)
            # Ищем паттерн: граница слова + имя листа + !
            if re.search(rf"\b{re.escape(deleted_name)}!", formula):
                should_remove = True
                break

        # Также проверяем локальные имена (localSheetId атрибут)
        if not should_remove:
            local_sheet_id = dn.get('localSheetId')
            if local_sheet_id is not None and local_sheet_id != '0':
                # После удаления всех остальных листов, оставшийся лист
                # становится единственным с индексом 0.
                # Обновляем localSheetId на 0.
                dn.set('localSheetId', '0')

        if should_remove:
            names_to_remove.append(dn)
            logger.debug(
                "Удалён definedName '%s' (ссылка на удалённый лист)",
                name_attr,
            )

    for dn in names_to_remove:
        defined_names_elem.remove(dn)

    # Если после очистки definedNames пуст — удаляем элемент целиком
    if len(defined_names_elem) == 0:
        wb_root.remove(defined_names_elem)


def preallocate_split_paths(
    tasks: List[Tuple[str, str, List[str], str]],
    output_dir: str,
) -> Dict[Tuple[str, str], str]:
    """Детерминированная предварительная разметка путей для всех листов.

    ВЫПОЛНЯЕТСЯ В ГЛАВНОМ ПОТОКЕ (один поток, детерминированно).

    Алгоритм:
      1. Собрать все задачи (источник + лист + метка) в плоский список.
      2. Отсортировать по (source_path, sheet_name) — детерминированный порядок.
      3. Для каждой задачи вычислить целевой путь.
      4. При коллизии имён — разрешить последовательно (_1, _2, ...).
         Так как список отсортирован, разрешение 100% детерминированно.
      5. Вернуть словарь {(source_path, sheet_name) -> abs_output_path}.

    Args:
        tasks: Список кортежей (source_path, output_dir, sheet_names, file_label)
               — такой же формат, как в split_many_parallel.
        output_dir: Директория для сохранения результатов.

    Returns:
        Словарь, отображающий (source_path, sheet_name) в уникальный
        абсолютный путь выходного файла.
    """
    path_registry: Set[str] = set()
    path_map: Dict[Tuple[str, str], str] = {}

    # 1. Собираем плоский список (source_path, sheet_name, file_label)
    sheet_tasks: List[Tuple[str, str, str]] = []
    for source_path, _out_dir, sheet_names, file_label in tasks:
        for sheet_name in sheet_names:
            sheet_tasks.append((source_path, sheet_name, file_label or ""))

    # 2. Детерминированная сортировка
    sheet_tasks.sort(key=lambda t: (t[0], t[1], t[2]))

    # 3-4. Предвычисляем пути с детерминированным разрешением коллизий
    for source_path, sheet_name, file_label in sheet_tasks:
        safe_label = _safe_filename(file_label)[:50] if file_label else ""
        safe_sheet = _safe_filename(sheet_name)[:50]
        if safe_label:
            output_filename = f"{safe_label}_{safe_sheet}.xlsx"
        else:
            output_filename = f"{safe_sheet}.xlsx"

        output_path = os.path.join(output_dir, output_filename)
        base_no_ext = os.path.splitext(output_filename)[0]
        ext = ".xlsx"

        # Детерминированное разрешение коллизий (проверка по set, не по файловой системе)
        counter = 1
        while output_path in path_registry:
            output_path = os.path.join(output_dir, f"{base_no_ext}_{counter}{ext}")
            counter += 1

        path_registry.add(output_path)
        path_map[(source_path, sheet_name)] = output_path

    return path_map


def _verify_xlsx_integrity(file_path: str) -> Tuple[bool, str]:
    """Проверить целостность .xlsx файла (ЛЕНЬЯНАЯ проверка, как Microsoft Excel).

    Файл считается ПОВРЕЖДЁННЫМ (вернёт False) только если openpyxl не может
    загрузить его даже в read_only режиме — т.е. при фатальных исключениях:
      - zipfile.BadZipFile (архив не является ZIP)
      - InvalidFileException (битая OOXML-структура)

    Args:
        file_path: Путь к .xlsx файлу.

    Returns:
        Кортеж (is_valid: bool, error_message: str).
        error_message пуст если файл корректен.
    """
    import warnings

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            warnings.filterwarnings('ignore', category=DeprecationWarning)
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            _ = wb.sheetnames
            wb.close()
    except zipfile.BadZipFile as e:
        return False, f"BadZipFile: {e}"
    except Exception as e:
        exc_name = type(e).__name__
        if exc_name in ("InvalidFileException", "InvalidFormatException", "LoadWorkbookException"):
            return False, f"{exc_name}: {e}"
        # All other exceptions are non-fatal (warnings, DrawingML, etc.)
        return True, ""

    return True, ""





def _copy_source_as_fallback(
    source_path: str,
    output_path: str,
) -> None:
    """Скопировать исходный файл как есть — последний шанс перед corrupted.

    Используется когда и ZIP-метод, и openpyxl не смогли выделить лист.
    Проверяет, что исходный файл открывается через openpyxl (lenient check).
    Если файл валидный — копирует его в выходной путь.

    Это предотвращает попадание в corrupted_cards файлов, которые
    являются функционально валидными, но не поддаются разделению
    из-за нестандартной структуры OOXML (WPS Office и т.д.).

    Args:
        source_path: Путь к исходному .xlsx файлу.
        output_path: Путь для сохранения.

    Raises:
        ValueError: Если исходный файл не открывается.
    """
    try:
        from burlak_parser.validator import validate_split_file_lenient
        result = validate_split_file_lenient(source_path)
        if not result.is_valid:
            raise ValueError(
                f"Source file cannot be opened: {result.errors[0].message if result.errors else 'unknown'}"
            )
        shutil.copy2(source_path, output_path)
    except ImportError:
        # Fallback: just try to copy if validator not available
        shutil.copy2(source_path, output_path)


def _extract_to_path_worker(
    source_path: str,
    output_path: str,
    sheet_name: str,
) -> Dict[str, Any]:
    """Рабочая функция: выделить один лист в предварительно размеченный путь.

    Выполняется в отдельном процессе. НЕ проверяет существование файла —
    уникальность пути гарантирована главным потоком через preallocate_split_paths.

    Вызывает _extract_sheet(), который пробует ZIP-метод, затем openpyxl fallback.
    Если оба метода завершаются ошибкой — файл считается повреждённым.

    Args:
        source_path: Путь к исходному .xlsx файлу.
        output_path: Абсолютный путь для сохранения (уже гарантированно уникальный).
        sheet_name: Имя листа для выделения.

    Returns:
        Словарь с результатами:
          - "path": output_path при успехе, None при ошибке
          - "error": сообщение об ошибке или None
          - "used_fallback": True если использован openpyxl fallback
          - "source_basename": os.path.basename(source_path)
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    splitter = CardSplitter(max_workers=1)
    result: Dict[str, Any] = {
        "path": None,
        "error": None,
        "used_fallback": False,
        "source_basename": os.path.basename(source_path),
        "source_path": source_path,
        "sheet_name": sheet_name,
    }
    try:
        splitter._extract_sheet(source_path, output_path, sheet_name)
        if splitter.openpyxl_fallback_count > 0:
            result["used_fallback"] = True
        result["path"] = output_path
    except Exception as e:
        err_msg = f"Лист '{sheet_name}' из {os.path.basename(source_path)}: {e}"
        logger.error("Ошибка разделения: %s", err_msg)
        result["error"] = err_msg
    return result




def _detect_xinyuan_boundaries(
    source_path: str,
    sheet_name: str,
) -> List[TableBoundary]:
    """Обнаружить границы операций по маркеру '鑫源汽车'.

    Специальный детектор для SWM мега-файлов (4_G01P作业指导书, G01P后备箱 и т.д.),
    где каждая операционная карта начинается с '鑫源汽车'.
    Карты расположены вертикально с фиксированным шагом (36-37 строк).

    Args:
        source_path: Путь к .xlsx файлу.
        sheet_name: Имя листа.

    Returns:
        Список TableBoundary для каждой найденной операции.
    """
    from burlak_parser.card_parser import ExcelReader

    boundaries: List[TableBoundary] = []
    reader = ExcelReader(source_path)
    try:
        if sheet_name not in reader.sheet_names:
            return boundaries

        ws = reader.get_sheet(sheet_name)
        max_row = ws.max_row or 0
        if max_row < 10:
            return boundaries

        max_col = min((ws.max_column or 10) + 1, 20)

        # Ищем все строки с '鑫源汽车' в первых 10 колонках
        marker_rows: List[int] = []
        for r in range(1, max_row + 1):
            for c in range(1, min(max_col, 10)):
                val = ws.cell_value(r, c)
                if val is not None and '鑫源汽车' in str(val):
                    marker_rows.append(r)
                    break

        if len(marker_rows) < 2:
            return boundaries

        # Вычисляем шаг (медиана интервалов)
        spacings = [marker_rows[i + 1] - marker_rows[i]
                    for i in range(len(marker_rows) - 1)]
        step = sorted(spacings)[len(spacings) // 2]  # медиана

        # Проверяем стабильность шага (>60% интервалов в пределах ±5 от медианы)
        consistent = sum(1 for s in spacings if abs(s - step) <= 5)
        if consistent < len(spacings) * 0.6:
            logger.debug(
                'Xinyuan boundaries: inconsistent spacing (step=%d, consistent=%d/%d)',
                step, consistent, len(spacings),
            )
            return boundaries

        # Строим границы: каждая '鑫源汽车' — начало новой карты
        for idx, marker_row in enumerate(marker_rows):
            # Граница данных: от текущего маркера до следующего (или конца файла)
            if idx + 1 < len(marker_rows):
                data_end = marker_rows[idx + 1] - 1
            else:
                data_end = max_row

            # Извлекаем имя операции: ищем CJK текст в строке маркера
            # (колонки B-J, пропуская колонку A где сам маркер)
            op_name = ''
            for c in range(2, min(max_col, 10)):
                val = ws.cell_value(marker_row, c)
                if val is not None:
                    val_str = str(val).strip()
                    if len(val_str) > 2 and re.search(r'[\u4e00-\u9fff]', val_str):
                        op_name = val_str
                        break

            boundaries.append(TableBoundary(
                header_row=marker_row,
                data_start=marker_row + 1,
                data_end=data_end,
                operation_name=op_name,
                source_path=source_path,
                sheet_name=sheet_name,
                card_label=(
                    f"{idx + 1:03d}_{_safe_filename(op_name)[:30]}"
                    if op_name else f"Op{idx + 1:03d}"
                ),
            ))

        logger.info(
            'Xinyuan boundaries: found %d cards (step=%d rows) in %s',
            len(boundaries), step, os.path.basename(source_path),
        )

    finally:
        reader.close()

    return boundaries


def find_table_boundaries(
    source_path: str,
    sheet_name: str,
    min_confidence: float = 0.35,
) -> List[TableBoundary]:
    """Обнаружить границы таблиц (операций) внутри одного листа.

    Универсальный детектор для любых брендов и структур файлов.
    Жёсткий лимит max_row <= 500 УДАЛЁН — анализируются все листы.
    Каждая найденная граница получает confidence score для фильтрации
    ложных срабатываний (меньше false positives на маленьких файлах).

    SWM-формат: таблицы могут не иметь явной qty-колонки (qty_col=0).
    Для таких случаев confidence score вычисляется без учёта qty.

    Args:
        source_path: Путь к .xlsx файлу.
        sheet_name: Имя листа для анализа.
        min_confidence: Минимальный порог уверенности (0.0-1.0).
                        Понижен с 0.3 до 0.2 для поддержки SWM-формата.

    Returns:
        Список TableBoundary с границами каждой таблицы.
    """
    from burlak_parser.heuristic_analyzer import HeuristicAnalyzer
    from burlak_parser.card_parser import ExcelReader

    boundaries: List[TableBoundary] = []

    reader = ExcelReader(source_path)
    try:
        if sheet_name not in reader.sheet_names:
            return boundaries

        ws = reader.get_sheet(sheet_name)
        start_search = 1
        max_row = ws.max_row or 0

        if max_row < 3:
            return boundaries

        # ── High-priority: SWM '鑫源汽车' marker detection ──
        # Проверяем FIRST, до HeuristicAnalyzer, потому что
        # find_part_table находит границы НЕ совпадающие с 鑫源汽车
        # (смещены на ~20 строк), что приводит к 2 картам в одном файле.
        xinyuan_first = _detect_xinyuan_boundaries(source_path, sheet_name)
        if xinyuan_first:
            logger.info(
                'Xinyuan (primary): found %d cards in %s',
                len(xinyuan_first), os.path.basename(source_path),
            )
            reader.close()
            return xinyuan_first

        max_tables = 500

        for table_idx in range(max_tables):
            if start_search >= max_row:
                break

            table_info = HeuristicAnalyzer.find_part_table(ws, start_row=start_search)
            if table_info is None:
                break

            header_row, part_no_col, qty_col, name_col = table_info

            if header_row < start_search:
                break

            # Определяем operation_name
            operation_name = HeuristicAnalyzer.extract_operation_name(ws, header_row)

            # Определяем последнюю строку данных (data_end)
            data_end = _find_table_data_end(ws, header_row, max_row, part_no_col)

            # Confidence scoring: отсеиваем false positives
            # ВАЖНО: qty_col может быть 0 в SWM-формате — confidence
            # вычисляется и без qty (с пониженным порогом)
            confidence = _compute_boundary_confidence(
                ws, header_row, data_end, part_no_col,
                qty_col if qty_col and qty_col > 0 else 0,
                name_col if name_col and name_col > 0 else 0,
            )
            if confidence < min_confidence:
                logger.debug(
                    "Boundary at row %d rejected: confidence %.2f < %.2f",
                    header_row, confidence, min_confidence,
                )
                start_search = header_row + 1
                continue

            boundaries.append(TableBoundary(
                header_row=header_row,
                data_start=header_row + 1,
                data_end=data_end,
                operation_name=operation_name,
                source_path=source_path,
                sheet_name=sheet_name,
                card_label=f"{table_idx + 1:03d}_{_safe_filename(operation_name)[:30]}" if operation_name else f"Op{table_idx + 1:03d}",
            ))

            start_search = data_end + 1

    finally:
        reader.close()

    # High-priority fallback: SWM-формат '鑫源汽车' (mega-files с 83+ картами)
    if not boundaries and max_row > 50:
        boundaries = _detect_xinyuan_boundaries(source_path, sheet_name)

    # Fallback: обнаружение таблиц проверки качества (检验项目 pattern)
    if not boundaries:
        boundaries = _detect_inspection_boundaries(source_path, sheet_name)

    # Fallback: универсальный детектор по повторяющимся шаблонам строк
    # (для SWM-формата, где find_part_table может пропускать таблицы)
    if not boundaries and max_row > 100:
        boundaries = _detect_repeating_pattern_boundaries(source_path, sheet_name)

    # Mega-sheet force: если лист очень большой (500+ строк), а найдено
    # слишком мало границ (менее 5% строк покрыто) — эвристика могла
    # пропустить большинство таблиц. Принудительно запускаем
    # универсальный детектор повторяющихся шаблонов.
    if boundaries and max_row > 200:
        covered_rows = sum(b.data_end - b.header_row for b in boundaries)
        coverage_ratio = covered_rows / max(max_row, 1)
        if coverage_ratio < 0.3 or len(boundaries) < 3:
            logger.info(
                "Mega-sheet (%d rows, %d boundaries, %.1f%% coverage) "
                "— forcing re-detection",
                max_row, len(boundaries), coverage_ratio * 100,
            )
            # Prefer xinyuan detection for SWM files
            xinyuan_boundaries = _detect_xinyuan_boundaries(
                source_path, sheet_name,
            )
            if xinyuan_boundaries and len(xinyuan_boundaries) > len(boundaries):
                boundaries = xinyuan_boundaries
                logger.info(
                    "Xinyuan detection found %d boundaries",
                    len(boundaries),
                )
            else:
                pattern_boundaries = _detect_repeating_pattern_boundaries(
                    source_path, sheet_name,
                )
                if pattern_boundaries and len(pattern_boundaries) > len(boundaries):
                    boundaries = pattern_boundaries
                    logger.info(
                        "Repeating pattern detection found %d boundaries",
                        len(boundaries),
                    )

    # ── Boundary snapping: close gaps between operations ──
    # Ensures every row between consecutive operations is assigned.
    # Without this, rows between data_end and next header_row are lost.
    if boundaries:
        boundaries.sort(key=lambda b: b.header_row)
        for i in range(len(boundaries) - 1):
            if boundaries[i].data_end < boundaries[i + 1].header_row - 1:
                boundaries[i].data_end = boundaries[i + 1].header_row - 1

    # ── Validation: reject boundary sets with inconsistent spacing ──
    # If intervals between boundaries vary wildly (>50% from median),
    # it's likely a false positive (not real operations but random data patterns).
    if len(boundaries) >= 3:
        spacings = [
            boundaries[i + 1].header_row - boundaries[i].header_row
            for i in range(len(boundaries) - 1)
        ]
        if spacings:
            median_spacing = sorted(spacings)[len(spacings) // 2]
            if median_spacing > 0:
                consistent = sum(
                    1 for s in spacings
                    if abs(s - median_spacing) <= median_spacing * 0.5
                )
                if consistent < len(spacings) * 0.5:
                    logger.warning(
                        "Boundary set rejected: inconsistent spacing "
                        "(median=%d, consistent=%d/%d)",
                        median_spacing, consistent, len(spacings),
                    )
                    boundaries = []

    return boundaries


def _compute_boundary_confidence(
    ws: Any,
    header_row: int,
    data_end: int,
    part_no_col: int,
    qty_col: int,
    name_col: int,
) -> float:
    """Вычислить уверенность в границах таблицы (0.0-1.0).

    Оценка на основе:
      - Количество ключевых слов в заголовке (до 0.4)
      - Плотность данных: непустые строки / общие строки (до 0.3)
      - Количество валидных part-номеров в данных (до 0.3)

    ВАЖНО: qty_col может быть 0 (SWM-формат) — в этом случае
    проверка qty пропускается, confidence снижается через более
    низкий min_confidence порог.
    """
    from burlak_parser.heuristic_analyzer import (
        HeuristicAnalyzer, PART_NO_KEYWORDS, QTY_KEYWORDS, NAME_KEYWORDS,
    )
    from burlak_parser.normalizer import is_valid_part_number

    score = 0.0

    # 1. Header quality (0.0-0.4)
    max_check_col = min((ws.max_column or 10) + 1, 50)
    header_non_empty = 0
    header_keywords = 0
    for hc in range(1, max_check_col):
        hv = HeuristicAnalyzer.get_cell_value(ws, header_row, hc)
        if hv is not None and str(hv).strip():
            header_non_empty += 1
            hv_lower = str(hv).strip().lower()
            if any(kw in hv_lower for kw in PART_NO_KEYWORDS):
                header_keywords += 1
            if qty_col > 0 and any(kw in hv_lower for kw in QTY_KEYWORDS):
                header_keywords += 1
            if name_col > 0 and any(kw in hv_lower for kw in NAME_KEYWORDS):
                header_keywords += 1

    if header_non_empty >= 3:
        score += 0.2
    elif header_non_empty >= 2:
        score += 0.1
    score += min(header_keywords * 0.1, 0.2)

    # 2. Data density (0.0-0.3)
    data_rows = data_end - header_row
    if data_rows > 0:
        non_empty_data = 0
        sample_start = header_row + 1
        sample_end = min(data_end + 1, header_row + 50)
        sample_count = sample_end - sample_start
        for r in range(sample_start, sample_end):
            v = HeuristicAnalyzer.get_cell_value(ws, r, part_no_col)
            if v is not None and str(v).strip():
                non_empty_data += 1
        if sample_count > 0:
            density = non_empty_data / sample_count
            score += density * 0.3

    # 3. Valid part numbers (0.0-0.3)
    valid_pn = 0
    total_pn = 0
    check_end = min(data_end + 1, header_row + 30)
    for r in range(header_row + 1, check_end):
        v = HeuristicAnalyzer.get_cell_value(ws, r, part_no_col)
        if v is not None:
            total_pn += 1
            pn_str = str(v).strip()
            if is_valid_part_number(pn_str):
                valid_pn += 1
    if total_pn > 0:
        pn_ratio = valid_pn / total_pn
        score += pn_ratio * 0.3

    return min(score, 1.0)


# Ключевые слова для обнаружения таблиц проверки качества
_INSPECTION_HEADER_KW = '检验项目'
_INSPECTION_SUBHEADER_KW = '作业内容图示'


def _detect_inspection_boundaries(
    source_path: str,
    sheet_name: str,
) -> List[TableBoundary]:
    """Обнаружить границы таблиц проверки качества (检验作业指导书).

    Ищет повторяющиеся блоки с заголовком "检验项目" в колонке B.
    Каждый блок содержит операцию проверки качества.

    Args:
        source_path: Путь к .xlsx файлу.
        sheet_name: Имя листа.

    Returns:
        Список TableBoundary для каждой операции проверки.
    """
    from burlak_parser.card_parser import ExcelReader

    boundaries: List[TableBoundary] = []
    reader = ExcelReader(source_path)
    try:
        if sheet_name not in reader.sheet_names:
            return boundaries

        ws = reader.get_sheet(sheet_name)
        max_row = ws.max_row or 0
        if max_row < 3:
            return boundaries

        # Находим все строки с "检验项目" в колонке B (col 2)
        header_rows: List[int] = []
        for r in range(1, max_row + 1):
            val = ws.cell_value(r, 2)
            if val is not None and _INSPECTION_HEADER_KW in str(val):
                header_rows.append(r)

        if len(header_rows) < 2:
            return boundaries

        # Определяем шаг между заголовками (медиана интервалов)
        spacings = [header_rows[i + 1] - header_rows[i]
                    for i in range(len(header_rows) - 1)]
        if not spacings:
            return boundaries
        step = sorted(spacings)[len(spacings) // 2]  # медиана

        # Проверяем что шаг стабилен (>50% интервалов в пределах ±3 от медианы)
        consistent = sum(1 for s in spacings if abs(s - step) <= 3)
        if consistent < len(spacings) * 0.5:
            return boundaries

        # Группируем заголовки: каждый заголовок — отдельная операция,
        # данные идут до следующего заголовка

        # Pre-compute title rows for all header rows
        title_rows: List[int] = []
        for header_row in header_rows:
            title_row = header_row
            for tr in range(max(1, header_row - 5), header_row):
                tr_val = ws.cell_value(tr, 4)  # колонка D
                if tr_val is not None:
                    tr_str = str(tr_val).strip()
                    if '检验' in tr_str or '作业指导' in tr_str or '指导书' in tr_str:
                        title_row = tr
                        break
            title_rows.append(title_row)

        for group_idx, header_row in enumerate(header_rows):
            cur_title = title_rows[group_idx]
            # data_end: до следующего title_row (не header_row!), чтобы не было overlap
            if group_idx + 1 < len(header_rows):
                next_title = title_rows[group_idx + 1]
                data_end = next_title - 1
            else:
                data_end = max_row

            # Извлекаем имя операции из колонки D строки header_row
            op_name = ""
            op_val = ws.cell_value(header_row, 4)
            if op_val is not None and str(op_val).strip():
                op_name = str(op_val).strip()

            boundaries.append(TableBoundary(
                header_row=cur_title,
                data_start=header_row,
                data_end=data_end,
                operation_name=op_name,
                source_path=source_path,
                sheet_name=sheet_name,
                card_label=(
                    f"{group_idx + 1:03d}_{_safe_filename(op_name)[:30]}"
                    if op_name else f"Op{group_idx + 1:03d}"
                ),
            ))

    finally:
        reader.close()

    return boundaries


def _find_table_data_end(
    ws: Any,
    header_row: int,
    max_row: int,
    part_no_col: int,
) -> int:
    """Найти последнюю строку данных таблицы.

    Определяет границу между текущей таблицей и следующей операцией.
    Срабатывает на:
      1. Строку-заголовок следующей таблицы (содержит PART_NO_KEYWORDS)
      2. Название следующей операции (строка с CJK текстом, где part_no_col пуст)
      3. 5+ полностью пустых строк подряд (ВСЕ колонки пусты)
      4. Резкое изменение формата строки (мерджи, пустые колонки)

    ВАЖНО: empty-run проверяет ВСЮ строку на пустоту, а не только part_no колонку.
    В SWM мега-файлах part-номера занимают первые несколько строк операции,
    а затем идут строки инструкций и картинок где part_no_col пуст.
    Проверка только part_no_col обрезала бы операцию после 2 строк.
    """
    from burlak_parser.heuristic_analyzer import HeuristicAnalyzer, PART_NO_KEYWORDS
    CJK_RE = re.compile(r'[一-鿿㐀-䶿]')

    empty_run = 0
    max_scan = min(max_row - header_row, 500)
    for r in range(header_row + 1, header_row + max_scan + 1):
        if r > max_row:
            break

        # Count non-empty cells across ALL scanned columns to determine
        # if the ENTIRE row is empty (not just part_no column)
        non_empty = 0
        row_values_check: List[str] = []
        max_check_col = min((ws.max_column or 10) + 1, 25)
        for c in range(1, max_check_col):
            v = ws.cell_value(r, c)
            if v is not None:
                non_empty += 1
                rv = str(v).strip().lower()
                if len(rv) < 50:
                    row_values_check.append(rv)

        if non_empty >= 2:
            has_part_no_keyword = any(
                any(kw in rv for kw in PART_NO_KEYWORDS)
                for rv in row_values_check
            )
            if has_part_no_keyword:
                return r - 1

        # ── Operation title detection ──
        # Если part_no_col пуст, но в строке есть CJK текст (название операции),
        # это граница следующей операции
        pn_val = ws.cell_value(r, part_no_col)
        pn_is_empty = pn_val is None or (isinstance(pn_val, str) and not pn_val.strip())

        if pn_is_empty and non_empty >= 1:
            # Проверяем: есть ли CJK текст в строке (признак названия операции)
            has_cjk = False
            for c in range(1, max_check_col):
                v = ws.cell_value(r, c)
                if v is not None and CJK_RE.search(str(v)):
                    has_cjk = True
                    break
            if has_cjk:
                # Проверяем: не является ли это просто пустой строкой с одним значением
                # Если CJK текст и part_no_col пуст — вероятно, это название операции
                # Проверяем что следующая строка тоже пуста или содержит заголовок
                if r + 1 <= max_row:
                    next_pn = ws.cell_value(r + 1, part_no_col)
                    next_is_empty = next_pn is None or (isinstance(next_pn, str) and not next_pn.strip())
                    if next_is_empty:
                        # Два пустых part_no подряд с CJK текстом — граница
                        return r - 1
                # Одинокая CJK-строка с пустым part_no — тоже граница
                # (для SWM где заголовки идут вплотную)
                if r - header_row > 3:
                    return r - 1

        # ── Empty-run detection: check FULL row emptiness ──
        # Only count as "empty" when ALL scanned columns are empty.
        # Part-number column being empty is normal for instruction/image rows.
        if non_empty == 0:
            empty_run += 1
            if empty_run >= 5:
                return r - 5
        else:
            empty_run = 0

    return min(max_row, header_row + max_scan)


# ═══════════════════════════════════════════════════════════════════════
# УНИВЕРСАЛЬНЫЙ ДЕТЕКТОР ПОВТОРЯЮЩИХСЯ ШАБЛОНОВ (SWM-style)
# ═══════════════════════════════════════════════════════════════════════

def _detect_repeating_pattern_boundaries(
    source_path: str,
    sheet_name: str,
) -> List[TableBoundary]:
    """Обнаружить границы таблиц через поиск повторяющихся шаблонов строк.

    Используется как универсальный fallback для SWM-формата, где
    find_part_table() может пропускать таблицы из-за нестандартных
    заголовков или отсутствия явных qty/name колонок.

    Алгоритм:
      1. Находит строки, где колонка A содержит числа (признак part-number)
      2. Группирует последовательные блоки данных
      3. Разделяет блоки по пустым строкам или строкам-заголовкам

    Returns:
        Список TableBoundary.
    """
    from burlak_parser.heuristic_analyzer import HeuristicAnalyzer
    from burlak_parser.card_parser import ExcelReader

    boundaries: List[TableBoundary] = []
    reader = ExcelReader(source_path)
    try:
        if sheet_name not in reader.sheet_names:
            return boundaries

        ws = reader.get_sheet(sheet_name)
        max_row = ws.max_row or 0
        if max_row < 20:
            return boundaries

        CJK_RE = re.compile(r'[一-鿿㐀-䶿]')

        # Сканируем все строки: ищем блоки данных (part-number в колонках A-D)
        # Универсальный поиск: part-numbers могут быть в любой из первых 4 колонок
        data_blocks: List[Tuple[int, int]] = []  # (start_row, end_row)
        in_block = False
        block_start = 0
        empty_count = 0
        SCAN_COLS = 8  # Columns A-H (wider scan for SWM where data may be right-aligned)

        for r in range(1, max_row + 1):
            # Check data in columns A-H
            has_data = False
            for c in range(1, SCAN_COLS + 1):
                val = ws.cell_value(r, c)
                if val is not None and str(val).strip():
                    has_data = True
                    break

            # Check for CJK header row (operation title) in columns A-C
            is_cjk_header = False
            for c in range(1, 4):
                val = ws.cell_value(r, c)
                if val is not None and CJK_RE.search(str(val)):
                    is_cjk_header = True
                    break

            if has_data and not is_cjk_header:
                if not in_block:
                    in_block = True
                    block_start = r
                    empty_count = 0
                empty_count = 0
            else:
                if in_block:
                    empty_count += 1
                    # Use threshold 5 (consistent with _find_table_data_end)
                    # to avoid premature block splitting on instruction/image rows
                    if empty_count >= 5 or is_cjk_header:
                        # End of block
                        data_blocks.append((block_start, r - empty_count))
                        in_block = False
                        empty_count = 0

        # Закрываем последний блок
        if in_block:
            data_blocks.append((block_start, max_row))

        # Фильтруем: минимум 3 строки данных в блоке
        for idx, (start, end) in enumerate(data_blocks):
            if end - start < 2:
                continue

            # Ищем заголовок над блоком
            header_row = start
            for r in range(max(1, start - 5), start):
                row_vals = []
                for c in range(1, 12):
                    v = ws.cell_value(r, c)
                    if v is not None:
                        row_vals.append(str(v).strip().lower())
                if any(
                    any(kw in rv for kw in HeuristicAnalyzer._get_part_no_keywords())
                    for rv in row_vals
                ):
                    header_row = r
                    break

            # Extract operation name
            op_name = ""
            for r in range(max(1, header_row - 3), header_row):
                for c in range(1, 8):
                    v = ws.cell_value(r, c)
                    if v is not None and CJK_RE.search(str(v)):
                        op_name = str(v).strip()
                        if len(op_name) > 3:
                            break
                if op_name:
                    break

            boundaries.append(TableBoundary(
                header_row=header_row,
                data_start=start,
                data_end=end,
                operation_name=op_name,
                source_path=source_path,
                sheet_name=sheet_name,
                card_label=f"{idx + 1:03d}_{_safe_filename(op_name)[:30]}" if op_name else f"Op{idx + 1:03d}",
            ))

    finally:
        reader.close()

    return boundaries


def _cleanup_workbook_for_single_sheet(
    all_entries: Dict[str, bytes],
    target_sheet_name: str,
) -> None:
    """Clean up workbook.xml, workbook.xml.rels for a single-sheet file.

    After vertical split, the output ZIP contains a workbook with only one sheet,
    but the workbook.xml may still reference other sheets, have wrong activeTab,
    high sheetId, etc. This function fixes those issues.

    Modifies all_entries in-place.
    """
    # First, find the target rId from the existing workbook.xml.rels
    target_rid = None
    wb_rels_bytes = all_entries.get('xl/_rels/workbook.xml.rels')
    if wb_rels_bytes is not None:
        if _HAS_LXML:
            try:
                rels_root = _lxml_etree.fromstring(wb_rels_bytes)
                # Find which rId maps to the worksheet that has our target sheet
                wb_bytes = all_entries.get('xl/workbook.xml')
                if wb_bytes is not None:
                    wb_root = _lxml_etree.fromstring(wb_bytes)
                    sheets = wb_root.find(f'{{{NS_MAIN}}}sheets')
                    if sheets is not None:
                        for sheet_el in sheets.findall(f'{{{NS_MAIN}}}sheet'):
                            if sheet_el.get('name') == target_sheet_name:
                                target_rid = sheet_el.get(f'{{{NS_R}}}id')
                                break
            except Exception:
                pass

    # Fix workbook.xml
    wb_bytes = all_entries.get('xl/workbook.xml')
    if wb_bytes is None:
        return

    if _HAS_LXML:
        wb_root = _lxml_etree.fromstring(wb_bytes)
        ns = NS_MAIN
        ns_r = NS_R

        # Fix bookViews: activeTab=0, firstSheet=0
        book_views = wb_root.find(f'{{{ns}}}bookViews')
        if book_views is not None:
            wb_view = book_views.find(f'{{{ns}}}workbookView')
            if wb_view is not None:
                wb_view.set('activeTab', '0')
                wb_view.set('firstSheet', '0')

        # Fix sheets: keep only the target sheet, set sheetId=1
        sheets = wb_root.find(f'{{{ns}}}sheets')
        if sheets is not None:
            to_remove = []
            found_target = False
            for sheet_el in sheets.findall(f'{{{ns}}}sheet'):
                name = sheet_el.get('name', '')
                if name == target_sheet_name and not found_target:
                    sheet_el.set('sheetId', '1')
                    found_target = True
                else:
                    to_remove.append(sheet_el)
            for el in to_remove:
                sheets.remove(el)

        # Remove all definedName elements (they reference other sheets)
        defined_names = wb_root.find(f'{{{ns}}}definedNames')
        if defined_names is not None:
            wb_root.remove(defined_names)

        # Remove customWorkbookViews (can cause openpyxl parse errors)
        custom_views = wb_root.find(f'{{{ns}}}customWorkbookViews')
        if custom_views is not None:
            wb_root.remove(custom_views)

        all_entries['xl/workbook.xml'] = _lxml_etree.tostring(
            wb_root, xml_declaration=True, encoding='UTF-8', standalone=True)
    else:
        # Regex fallback
        wb_text = wb_bytes.decode('utf-8')
        wb_text = re.sub(r'activeTab="\d+"', 'activeTab="0"', wb_text)
        wb_text = re.sub(r'firstSheet="\d+"', 'firstSheet="0"', wb_text)
        def _replace_sheets(m: re.Match) -> str:
            content = m.group(2)
            kept = None
            for sh in re.finditer(r'<sheet[^>]*/>', content):
                name_m = re.search(r'name="([^"]+)"', sh.group(0))
                if name_m and name_m.group(1) == target_sheet_name:
                    kept = sh.group(0)
                    break
            if kept:
                return f'{m.group(1)}\n{kept}\n{m.group(3)}'
            return m.group(0)
        wb_text = re.sub(
            r'(<sheets[^>]*>)(.*?)(</sheets>)', _replace_sheets, wb_text,
            count=1, flags=re.DOTALL)
        wb_text = re.sub(r'<definedNames[^>]*>.*?</definedNames>', '', wb_text,
                         flags=re.DOTALL)
        wb_text = re.sub(r'<customWorkbookViews[^>]*>.*?</customWorkbookViews>', '', wb_text,
                         flags=re.DOTALL)
        all_entries['xl/workbook.xml'] = wb_text.encode('utf-8')

    # Fix workbook.xml.rels — keep only the target sheet + shared resources
    if wb_rels_bytes is not None:
        if _HAS_LXML:
            rels_root = _lxml_etree.fromstring(wb_rels_bytes)
            to_remove = []
            for rel in list(rels_root):
                rel_type = rel.get('Type', '')
                rid = rel.get('Id', '')
                if 'worksheet' in rel_type and rid != target_rid:
                    to_remove.append(rel)
            for el in to_remove:
                rels_root.remove(el)
            all_entries['xl/_rels/workbook.xml.rels'] = _lxml_etree.tostring(
                rels_root, xml_declaration=True, encoding='UTF-8', standalone=True)
        else:
            rels_text = wb_rels_bytes.decode('utf-8')
            if target_rid:
                # Remove all worksheet relationships except the target
                def _filter_rels(m: re.Match) -> str:
                    rel_xml = m.group(0)
                    rid_m = re.search(r'Id="([^"]+)"', rel_xml)
                    if rid_m and rid_m.group(1) != target_rid:
                        if 'worksheet' in rel_xml:
                            return ''
                    return rel_xml
                rels_text = re.sub(r'<Relationship\b[^>]*/>', _filter_rels, rels_text)
            all_entries['xl/_rels/workbook.xml.rels'] = rels_text.encode('utf-8')


def _cleanup_app_xml(all_entries: Dict[str, bytes]) -> None:
    """Update docProps/app.xml to reflect 1 sheet."""
    app_bytes = all_entries.get('docProps/app.xml')
    if app_bytes is None:
        return

    if _HAS_LXML:
        try:
            root = _lxml_etree.fromstring(app_bytes)
            # Fix HeadingPairs: <vt:i4>4</vt:i4> → <vt:i4>1</vt:i4>
            for elem in root.iter():
                if elem.tag == f'{{{NS_R}}}i4' or elem.tag.endswith('}i4'):
                    if elem.text and elem.text.strip() == '4':
                        parent = elem.getparent()
                        if parent is not None:
                            gparent = parent.getparent()
                            if gparent is not None:
                                # Check if this is inside HeadingPairs
                                gparent_tag = _lxml_etree.QName(gparent.tag).localname
                                if gparent_tag == 'vector':
                                    elem.text = '1'
            # Fix TitlesOfParts: keep only the first sheet name
            for vector in root.iter():
                if _lxml_etree.QName(vector.tag).localname == 'vector':
                    children = list(vector)
                    # Check if this is TitlesOfParts (contains lpstr children)
                    if (len(children) > 1 and
                            _lxml_etree.QName(children[0].tag).localname == 'lpstr'):
                        # This is TitlesOfParts — keep only first entry
                        for child in children[1:]:
                            vector.remove(child)
                        vector.set('size', '1')
            all_entries['docProps/app.xml'] = _lxml_etree.tostring(
                root, xml_declaration=True, encoding='UTF-8', standalone=True)
        except Exception:
            pass
    else:
        # Regex fallback
        app_text = app_bytes.decode('utf-8')
        # Replace the i4 in HeadingPairs
        app_text = re.sub(
            r'(<HeadingPairs>.*?<vt:i4>)\d+(</vt:i4>.*?</HeadingPairs>)',
            r'\g<1>1\2', app_text, count=1, flags=re.DOTALL)
        # Fix TitlesOfParts vector size and remove extra entries
        app_text = re.sub(
            r'(<TitlesOfParts>.*?<vt:vector\s+)size="\d+"',
            r'\g<1>size="1"', app_text, count=1, flags=re.DOTALL)
        # Remove all but first <vt:lpstr> in TitlesOfParts
        tp_match = re.search(
            r'(<TitlesOfParts>.*?<vt:vector[^>]*>)(.*?)(</vt:vector>)',
            app_text, re.DOTALL)
        if tp_match:
            inner = tp_match.group(2)
            first_lpstr = re.search(r'<vt:lpstr>.*?</vt:lpstr>', inner, re.DOTALL)
            if first_lpstr:
                new_inner = first_lpstr.group(0)
                app_text = app_text[:tp_match.start(2)] + new_inner + app_text[tp_match.end(2):]
        all_entries['docProps/app.xml'] = app_text.encode('utf-8')


def _cleanup_custom_xml(all_entries: Dict[str, bytes]) -> None:
    """Remove WPS-specific custom.xml metadata (large, causes issues)."""
    # Replace with minimal valid custom.xml
    minimal = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"'
        ' xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '</Properties>'
    )
    all_entries['docProps/custom.xml'] = minimal.encode('utf-8')


def _vertical_split_worker(
    source_path: str,
    output_dir: str,
    sheet_name: str,
    boundaries: List[TableBoundary],
    card_label: str,
    preloaded_zip: Optional[bytes] = None,
) -> List[str]:
    """Разделить один лист по вертикальным границам через ZIP-манипуляцию.

    Сохраняет изображения, форматирование и стили, работая напрямую
    с ZIP-структурой .xlsx файла (а не через openpyxl Workbook).

    Алгоритм для каждой операции:
      1. Скопировать исходный .xlsx (один лист, все изображения).
      2. Отфильтровать sheet XML: оставить только <row> нужного диапазона.
      3. Отфильтровать drawing XML: оставить только anchors нужного диапазона.
      4. Скорректировать row-позиции в anchors.
      5. Записать изменённый ZIP.

    Args:
        source_path: Путь к исходному .xlsx файлу (уже один лист).
        output_dir: Директория для сохранения результатов.
        sheet_name: Имя листа.
        boundaries: Список границ таблиц (операций).
        card_label: Метка для именования файлов.

    Returns:
        Список путей к созданным файлам.
    """
    os.makedirs(output_path_dir := output_dir, exist_ok=True)
    created: List[str] = []
    safe_label = _safe_filename(card_label)[:50] if card_label else ""

    # Читаем ZIP в память
    if preloaded_zip is not None:
        zip_data = preloaded_zip
    else:
        try:
            with open(source_path, 'rb') as f:
                zip_data = f.read()
        except OSError as e:
            logger.error("Cannot read source for vertical split: %s", e)
            return created

    # Читаем все ZIP-entries один раз
    try:
        with zipfile.ZipFile(io.BytesIO(zip_data), 'r') as zf:
            all_entries: Dict[str, bytes] = {}
            for name in zf.namelist():
                try:
                    all_entries[name] = zf.read(name)
                except zipfile.BadZipFile as e:
                    logger.warning("Skipping corrupt entry %s: %s", name, e)
    except zipfile.BadZipFile as e:
        logger.error("Cannot open source ZIP for vertical split: %s", e)
        return created

    all_names = set(all_entries.keys())

    # Находим имя листа в workbook.xml для определения rId
    wb_xml_bytes = all_entries.get('xl/workbook.xml')
    if wb_xml_bytes is None:
        return created

    wb_root = ET.fromstring(wb_xml_bytes)
    sheets_elem = wb_root.find(f'{{{NS_MAIN}}}sheets')
    if sheets_elem is None:
        return created

    # rId → sheet name mapping
    target_r_id: Optional[str] = None
    for sheet_el in sheets_elem.findall(f'{{{NS_MAIN}}}sheet'):
        if sheet_el.get('name') == sheet_name:
            target_r_id = (sheet_el.get(f'{{{NS_R}}}id')
                           or sheet_el.get('r:id'))
            break

    if target_r_id is None:
        return created

    # rId → target path из workbook.xml.rels
    sheet_target: Optional[str] = None
    rels_bytes = all_entries.get('xl/_rels/workbook.xml.rels')
    if rels_bytes is not None:
        rels_root = ET.fromstring(rels_bytes)
        for rel_el in rels_root:
            if (rel_el.get('Id') == target_r_id
                    and 'worksheet' in rel_el.get('Type', '')):
                sheet_target = rel_el.get('Target', '').lstrip('/')
                if not sheet_target.startswith('xl/'):
                    sheet_target = 'xl/' + sheet_target
                break

    if sheet_target is None:
        return created

    # Определяем drawing XML для этого листа
    sheet_dir = os.path.dirname(sheet_target)
    sheet_base = os.path.basename(sheet_target)
    sheet_rels_path = f"{sheet_dir}/_rels/{sheet_base}.rels"
    drawing_path: Optional[str] = None
    vml_path: Optional[str] = None
    comments_path: Optional[str] = None

    sr_bytes = all_entries.get(sheet_rels_path)
    if sr_bytes is not None:
        sr_root = ET.fromstring(sr_bytes)
        sr_base_dir = os.path.dirname(sheet_target)
        for sr_el in sr_root:
            target = sr_el.get('Target', '')
            rtype = sr_el.get('Type', '')
            resolved = os.path.normpath(
                os.path.join(sr_base_dir, target)).replace(os.sep, '/')
            if 'drawing' in rtype.lower() and 'vml' not in rtype.lower():
                drawing_path = resolved
            elif 'vml' in rtype.lower():
                vml_path = resolved
            elif 'comment' in rtype.lower():
                comments_path = resolved

    # Читаем drawing XML как bytes (НЕ через ET — сохраняем оригинальные namespaces)
    drawing_xml_bytes: Optional[bytes] = None
    if drawing_path and drawing_path in all_entries:
        drawing_xml_bytes = all_entries[drawing_path]

    safe_label_prefix = _safe_filename(card_label)[:50] if card_label else ""

    # ── Находим ВСЕ drawing файлы и их rels (WPS может привязывать image к не тому sheet) ──
    all_drawings: Dict[str, bytes] = {}        # drawing_path -> raw bytes
    all_drawing_rels: Dict[str, str] = {}      # drawing_rels_path -> drawing_path
    all_drawing_rels_map: Dict[str, Dict[str, str]] = {}  # drawing_rels_path -> {rId -> media_path}

    for entry_name in list(all_entries.keys()):
        if (entry_name.startswith('xl/drawings/drawing')
                and entry_name.endswith('.xml')
                and '_rels' not in entry_name):
            all_drawings[entry_name] = all_entries[entry_name]
            dr_path = f"{os.path.dirname(entry_name)}/_rels/{os.path.basename(entry_name)}.rels"
            all_drawing_rels[dr_path] = entry_name
            if dr_path in all_entries:
                rid_map: Dict[str, str] = {}
                try:
                    dr_root = ET.fromstring(all_entries[dr_path])
                    for dr_el in dr_root:
                        rid = dr_el.get('Id', '')
                        target = dr_el.get('Target', '')
                        if rid and target:
                            resolved = os.path.normpath(
                                os.path.join(os.path.dirname(entry_name), target)
                            ).replace(os.sep, '/')
                            rid_map[rid] = resolved
                except Exception as e:
                    logger.debug("Failed to parse drawing rels %s: %s", dr_path, e)
                all_drawing_rels_map[dr_path] = rid_map

    for i, boundary in enumerate(boundaries):
        op_label = boundary.card_label or f"Op{i + 1:03d}"
        if safe_label_prefix:
            output_filename = (
                f"{safe_label_prefix}_{_safe_filename(op_label)[:40]}.xlsx")
        else:
            output_filename = f"{_safe_filename(op_label)[:50]}.xlsx"

        output_path = os.path.join(output_dir, output_filename)

        # Post-split validation: skip empty operations
        if boundary.data_end <= boundary.header_row:
            logger.warning(
                "Skipping empty operation %d in %s",
                i + 1, os.path.basename(source_path))
            continue

        # ── Фильтруем ВСЕ drawing XML и собираем retained rIds ──
        filtered_drawings: Dict[str, Optional[bytes]] = {}  # drawing_path -> filtered bytes (None = skip)
        retained_image_paths: Set[str] = set()
        current_retained_rids: Set[str] = set()
        comments_fully_removed = False  # True если все комментарии вне диапазона
        if comments_path and comments_path in all_entries:
            test_filtered = _filter_comments_xml(
                all_entries[comments_path], boundary.header_row, boundary.data_end)
            if test_filtered is None:
                comments_fully_removed = True

        for dr_path, dr_bytes in all_drawings.items():
            # Drawing XML uses 0-indexed rows (row 0 = Excel row 1)
            # Boundary header_row/data_end are 1-indexed (Excel rows)
            # Convert: subtract 1 for drawing filter
            filtered, rids = _filter_drawing_xml_with_rids(
                dr_bytes, boundary.header_row - 1, boundary.data_end - 1)
            filtered_drawings[dr_path] = filtered
            if filtered is not None:
                current_retained_rids.update(rids)
                # Map retained rIds → media paths
                # all_drawing_rels_map is keyed by rels path, not drawing path
                dr_rels_path = f"{os.path.dirname(dr_path)}/_rels/{os.path.basename(dr_path)}.rels"
                rid_map = all_drawing_rels_map.get(dr_rels_path, {})
                for rid in rids:
                    media_path = rid_map.get(rid, '')
                    if media_path:
                        retained_image_paths.add(media_path)
                logger.debug("Drawing %s: filtered %d bytes -> %d bytes, rids=%s",
                    dr_path, len(dr_bytes), len(filtered), rids)
            else:
                logger.debug("Drawing %s: fully outside range, skipping", dr_path)

        try:
            # Apply workbook/docProps cleanup in-place before writing
            _cleanup_workbook_for_single_sheet(all_entries, sheet_name)
            _cleanup_app_xml(all_entries)
            _cleanup_custom_xml(all_entries)

            with zipfile.ZipFile(output_path, 'w',
                                 zipfile.ZIP_DEFLATED) as zf_write:
                for name, data in all_entries.items():
                    # Skip unused sheet XML files (keep only the target sheet)
                    if (name.startswith('xl/worksheets/sheet')
                            and name.endswith('.xml')
                            and '_rels' not in name
                            and name != sheet_target):
                        continue
                    # Skip unused sheet .rels files
                    if (name.startswith('xl/worksheets/_rels/sheet')
                            and name.endswith('.rels')
                            and name != sheet_rels_path):
                        continue
                    # Skip unused sheet comments
                    if (name.startswith('xl/comments')
                            and name.endswith('.xml')
                            and name != comments_path):
                        continue
                    if name == sheet_target:
                        data = _filter_sheet_xml(
                            data, boundary.header_row, boundary.data_end)
                    elif name in filtered_drawings and filtered_drawings[name] is not None:
                        data = filtered_drawings[name]
                    elif name in filtered_drawings and filtered_drawings[name] is None:
                        continue  # Drawing fully outside keep range — skip
                    elif (name in all_drawing_rels and filtered_drawings.get(all_drawing_rels[name]) is not None):
                        parent_drawing = all_drawing_rels[name]
                        rids_for_dr = set()
                        # all_drawing_rels_map is keyed by rels path (name), not drawing path
                        for r, p in all_drawing_rels_map.get(name, {}).items():
                            if p in retained_image_paths:
                                rids_for_dr.add(r)
                        data = _filter_drawing_rels(data, rids_for_dr)
                    elif (name in all_drawing_rels and filtered_drawings.get(all_drawing_rels[name]) is None):
                        continue  # Parent drawing fully outside — skip rels too
                    elif name == vml_path and vml_path is not None:
                        data = _filter_vml_xml(
                            data, boundary.header_row,
                            boundary.data_end)
                    elif (comments_path and name == comments_path
                          and comments_path is not None):
                        filtered_comments = _filter_comments_xml(
                            data, boundary.header_row, boundary.data_end)
                        if filtered_comments is None:
                            continue  # Все комментарии вне диапазона — пропускаем
                        data = filtered_comments

                    # ── Удаляем ссылку на comments из sheet .rels ──
                    if (comments_fully_removed
                            and name == sheet_rels_path
                            and comments_path):
                        try:
                            rels_root_cleanup = ET.fromstring(data)
                            to_drop = []
                            for rel_el in rels_root_cleanup:
                                target = rel_el.get('Target', '')
                                resolved = os.path.normpath(
                                    os.path.join(
                                        os.path.dirname(sheet_target), target)
                                ).replace(os.sep, '/')
                                if resolved == comments_path:
                                    to_drop.append(rel_el)
                            for el in to_drop:
                                rels_root_cleanup.remove(el)
                            data = _serialize_xml(
                                rels_root_cleanup, NS_PKG_RELS)
                        except ET.ParseError:
                            pass

                    # ── Удаляем Override для comments из Content_Types ──
                    if (comments_fully_removed
                            and name == '[Content_Types].xml'
                            and comments_path):
                        ct_str = data.decode('utf-8', errors='replace')
                        ct_str = re.sub(
                            r'<Override[^>]*PartName="[^"]*comment[^"]*"[^>]*/>',
                            '', ct_str)
                        ct_str = re.sub(
                            r'<Override[^>]*PartName="[^"]*Comment[^"]*"[^>]*/>',
                            '', ct_str)
                        data = ct_str.encode('utf-8')

                    # ── Фильтрация медиа: пропускаем неиспользуемые изображения ──
                    # БЕЗОПАСНАЯ СТРАТЕГИЯ: копируем все медиа по умолчанию.
                    # Фильтруем ТОЛЬКО если:
                    #   1. all_drawing_rels_map непустой (успешно распарсили .rels)
                    #   2. retained_image_paths непустой (есть anchors в диапазоне)
                    # Если хотя бы одно условие не выполнено — копируем все медиа.
                    # Это предотвращает потерю изображений при неполных данных.
                    any_rels_parsed = any(
                        rid_map for rid_map in all_drawing_rels_map.values()
                    )
                    should_filter_media = (
                        any_rels_parsed
                        and retained_image_paths
                    )
                    if name.startswith('xl/media/'):
                        if should_filter_media:
                            if name not in retained_image_paths:
                                logger.debug(
                                    "Filtered media: %s (not in retained set)",
                                    name,
                                )
                                continue  # Не копируем неиспользуемое изображение
                        # else: копируем все медиа (safe default)

                    zf_write.writestr(name, data)

        except zipfile.BadZipFile as e:
            logger.error("Failed to create vertical split %s: %s",
                         output_filename, e)
            continue

        # Validate the output file
        if not _validate_split_file(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass
            logger.warning("Deleted invalid split file: %s", output_filename)
            continue

        created.append(output_path)
        logger.info(
            "Вертикальный split (ZIP): операция %d '%s' [%d-%d] → %s",
            i + 1, boundary.operation_name[:30] or "",
            boundary.header_row, boundary.data_end,
            os.path.basename(output_path))

    return created


def _filter_sheet_xml(
    sheet_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> bytes:
    """Filter sheet XML — keep only rows in range with reindexed references.

    Uses lxml for proper namespace handling and valid XML output.
    Falls back to regex when lxml is unavailable.
    """
    if _HAS_LXML:
        return _filter_sheet_xml_lxml(sheet_data, keep_from_row, keep_to_row)
    return _filter_sheet_xml_regex(sheet_data, keep_from_row, keep_to_row)


def _reindex_cell_ref(col_letters: str, old_row: int, keep_from_row: int) -> str:
    """Reindex a cell reference: B4 with keep_from_row=4 → B1."""
    new_row = max(1, old_row - keep_from_row + 1)
    return f'{col_letters}{new_row}'


def _reindex_range_ref(ref: str, keep_from_row: int, keep_to_row: int) -> Optional[str]:
    """Reindex a range reference like B4:D10, keeping only the intersection with [keep_from_row, keep_to_row].

    Returns None if the range is completely outside the keep range.
    """
    if ':' not in ref:
        # Single cell
        m = re.match(r'^([A-Z]+)(\d+)$', ref)
        if not m:
            return ref
        row = int(m.group(2))
        if row < keep_from_row or row > keep_to_row:
            return None
        return _reindex_cell_ref(m.group(1), row, keep_from_row)

    parts = ref.split(':')
    m1 = re.match(r'^([A-Z]+)(\d+)$', parts[0])
    m2 = re.match(r'^([A-Z]+)(\d+)$', parts[1])
    if not m1 or not m2:
        return ref

    r1 = int(m1.group(2))
    r2 = int(m2.group(2))
    c1 = m1.group(1)
    c2 = m2.group(1)

    if r2 < keep_from_row or r1 > keep_to_row:
        return None

    new_r1 = max(1, max(r1, keep_from_row) - keep_from_row + 1)
    new_r2 = max(1, min(r2, keep_to_row) - keep_from_row + 1)
    return f'{c1}{new_r1}:{c2}{new_r2}'


def _filter_sheet_xml_lxml(
    sheet_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> bytes:
    """lxml-based sheet XML filter."""
    root = _lxml_etree.fromstring(sheet_data)
    ns = NS_MAIN
    _cell_ref_re = re.compile(r'^([A-Z]+)(\d+)$')

    # Register namespaces for proper output (skip empty prefix — lxml rejects it)
    for prefix, uri in _LXML_NS.items():
        if prefix:
            _lxml_etree.register_namespace(prefix, uri)

    # 1. Update <dimension>
    dim = root.find(f'{{{ns}}}dimension')
    if dim is not None:
        ref = dim.get('ref', '')
        try:
            _, _, max_col, _ = range_boundaries(ref)
        except (ValueError, IndexError):
            max_col = 10
        new_count = keep_to_row - keep_from_row + 1
        dim.set('ref', f'A1:{get_column_letter(max_col)}{new_count}')

    # 2. Filter <sheetData>/<row> and reindex cell references
    sheet_data_elem = root.find(f'{{{ns}}}sheetData')
    if sheet_data_elem is not None:
        rows_to_remove = []
        for row_elem in sheet_data_elem.findall(f'{{{ns}}}row'):
            r_attr = row_elem.get('r')
            if r_attr is None:
                rows_to_remove.append(row_elem)
                continue
            try:
                r = int(r_attr)
            except ValueError:
                rows_to_remove.append(row_elem)
                continue

            if r < keep_from_row or r > keep_to_row:
                rows_to_remove.append(row_elem)
                continue

            new_r = r - keep_from_row + 1
            row_elem.set('r', str(new_r))

            # Reindex cell r attributes
            for c_elem in row_elem.findall(f'{{{ns}}}c'):
                cell_ref = c_elem.get('r', '')
                m = _cell_ref_re.match(cell_ref)
                if m:
                    c_elem.set('r', _reindex_cell_ref(m.group(1), int(m.group(2)), keep_from_row))

        for elem in rows_to_remove:
            sheet_data_elem.remove(elem)

    # 3. Filter <mergeCells>
    merge_cells = root.find(f'{{{ns}}}mergeCells')
    if merge_cells is not None:
        to_remove = []
        kept_count = 0
        for mc in merge_cells.findall(f'{{{ns}}}mergeCell'):
            ref = mc.get('ref', '')
            try:
                min_col, min_r, max_col, max_r = range_boundaries(ref)
            except (ValueError, IndexError):
                to_remove.append(mc)
                continue
            if max_r < keep_from_row or min_r > keep_to_row:
                to_remove.append(mc)
                continue
            nr1 = max(min_r, keep_from_row) - keep_from_row + 1
            nr2 = min(max_r, keep_to_row) - keep_from_row + 1
            mc.set('ref', f'{get_column_letter(min_col)}{nr1}:{get_column_letter(max_col)}{nr2}')
            kept_count += 1
        for elem in to_remove:
            merge_cells.remove(elem)
        merge_cells.set('count', str(kept_count))
        if kept_count == 0:
            root.remove(merge_cells)

    # 4. Remove autoFilter
    for af in root.findall(f'{{{ns}}}autoFilter'):
        root.remove(af)

    # Remove filterMode from sheetPr
    sheet_pr = root.find(f'{{{ns}}}sheetPr')
    if sheet_pr is not None:
        if 'filterMode' in sheet_pr.attrib:
            del sheet_pr.attrib['filterMode']

    # Remove extLst (references features that become invalid after reindex)
    for ext_lst in root.findall(f'{{{ns}}}extLst'):
        root.remove(ext_lst)

    # Also remove extLst inside sheetPr
    if sheet_pr is not None:
        for ext_lst in sheet_pr.findall(f'{{{ns}}}extLst'):
            sheet_pr.remove(ext_lst)

    # 5. Remove dataValidations
    for dv in root.findall(f'{{{ns}}}dataValidations'):
        root.remove(dv)

    # 5b. Filter conditionalFormatting
    for cf in list(root.findall(f'{{{ns}}}conditionalFormatting')):
        sqref = cf.get('sqref', '')
        parts = re.split(r'\s+', sqref)
        kept_parts = []
        for part in parts:
            reindexed = _reindex_range_ref(part, keep_from_row, keep_to_row)
            if reindexed is not None:
                kept_parts.append(reindexed)
        if not kept_parts:
            root.remove(cf)
        else:
            cf.set('sqref', ' '.join(kept_parts))

    # 6. Reset sheetView
    for sv in root.findall(f'{{{ns}}}sheetView'):
        sv.set('topLeftCell', 'A1')
        if sv.get('view') == 'pageBreakPreview':
            del sv.attrib['view']
        # Reset selection
        for sel in sv.findall(f'{{{ns}}}selection'):
            sel.set('activeCell', 'A1')
            sel.set('sqref', 'A1')

    # 7. Filter rowBreaks
    for rb in list(root.findall(f'{{{ns}}}rowBreaks')):
        to_remove = []
        for brk in rb.findall(f'{{{ns}}}brk'):
            brk_id = brk.get('id')
            if brk_id is None:
                to_remove.append(brk)
                continue
            try:
                br = int(brk_id)
            except ValueError:
                to_remove.append(brk)
                continue
            if br < keep_from_row or br > keep_to_row:
                to_remove.append(brk)
            else:
                brk.set('id', str(max(1, br - keep_from_row + 1)))
        for elem in to_remove:
            rb.remove(elem)
        if len(rb) == 0:
            root.remove(rb)

    return _lxml_etree.tostring(
        root, xml_declaration=True, encoding='UTF-8', standalone=True)


def _filter_sheet_xml_regex(
    sheet_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> bytes:
    """Regex fallback for _filter_sheet_xml when lxml is unavailable."""
    xml_text = sheet_data.decode("utf-8")

    def _update_dim(m: re.Match) -> str:
        full = m.group(0)
        ref = m.group(1)
        try:
            _, _, max_col, _ = range_boundaries(ref)
        except (ValueError, IndexError):
            max_col = 10
        new_count = keep_to_row - keep_from_row + 1
        return re.sub(
            r'ref="[^"]*"',
            f'ref="A1:{get_column_letter(max_col)}{new_count}"',
            full,
        )
    xml_text = re.sub(
        r'<[^>]*dimension[^>]*ref="([^"]+)"[^>]*/?\s*>',
        _update_dim, xml_text, count=1,
    )

    _CELL_REF_RE = re.compile(r'(r=")([A-Za-z]+)(\d+)(")')

    def _filter_sd(m: re.Match) -> str:
        sd_open = m.group(1)
        sd_content = m.group(2)
        sd_close = m.group(3)

        kept_rows = []
        for row_m in re.finditer(
            r'(<(?:[\w\-]+:)?row\b[^>]*>.*?</(?:[\w\-]+:)?row>)',
            sd_content, re.DOTALL,
        ):
            row_xml = row_m.group(0)
            r_match = re.search(r'\br="(\d+)"', row_xml)
            if not r_match:
                continue
            r = int(r_match.group(1))
            if r < keep_from_row or r > keep_to_row:
                continue

            new_r = r - keep_from_row + 1
            row_xml = re.sub(
                r'(\br=")\d+(")',
                lambda m, nr=new_r: f'{m.group(1)}{nr}{m.group(2)}',
                row_xml,
            )
            def _upd_cref(cm: re.Match, nr=new_r) -> str:
                return f'{cm.group(1)}{cm.group(2)}{nr}{cm.group(4)}'
            row_xml = _CELL_REF_RE.sub(_upd_cref, row_xml)
            kept_rows.append(row_xml)

        return sd_open + "".join(kept_rows) + sd_close

    xml_text = re.sub(
        r'(<(?:[\w\-]+:)?sheetData[^>]*>)(.*?)(</(?:[\w\-]+:)?sheetData>)',
        _filter_sd, xml_text, flags=re.DOTALL,
    )

    def _filter_mc(m: re.Match) -> str:
        mc_open = m.group(1)
        mc_content = m.group(2)
        mc_close = m.group(3)

        kept_mcs = []
        for cell_m in re.finditer(r'<(?:[\w\-]+:)?mergeCell[^/]*/>', mc_content):
            cx = cell_m.group(0)
            ref_m = re.search(r'ref="([^"]+)"', cx)
            if not ref_m:
                continue
            ref = ref_m.group(1)
            parts = ref.split(":")
            if len(parts) != 2:
                continue
            try:
                min_col, min_r, max_col, max_r = range_boundaries(ref)
            except (ValueError, IndexError):
                continue
            if max_r < keep_from_row or min_r > keep_to_row:
                continue
            nr1 = max(min_r, keep_from_row) - keep_from_row + 1
            nr2 = min(max_r, keep_to_row) - keep_from_row + 1
            new_ref = f"{get_column_letter(min_col)}{nr1}:{get_column_letter(max_col)}{nr2}"
            cx = re.sub(r'ref="[^"]*"', f'ref="{new_ref}"', cx)
            kept_mcs.append(cx)

        if not kept_mcs:
            return ""
        mc_open = re.sub(r'count="\d+"', f'count="{len(kept_mcs)}"', mc_open)
        return mc_open + "".join(kept_mcs) + mc_close

    xml_text = re.sub(
        r'(<(?:[\w\-]+:)?mergeCells[^>]*>)(.*?)(</(?:[\w\-]+:)?mergeCells>)',
        _filter_mc, xml_text, flags=re.DOTALL,
    )

    xml_text = re.sub(r'<(?:[\w\-]+:)?autoFilter[^>]*/>\s*', "", xml_text)
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?autoFilter[^>]*>.*?</(?:[\w\-]+:)?autoFilter>\s*',
        "", xml_text, flags=re.DOTALL,
    )
    xml_text = re.sub(r'(\bsheetPr[^>]*?)\s+filterMode="[^"]*"', r'\1', xml_text)
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?extLst[^>]*>.*?</(?:[\w\-]+:)?extLst>\s*',
        "", xml_text, flags=re.DOTALL,
    )

    xml_text = re.sub(r'<(?:[\w\-]+:)?dataValidations[^>]*/>\s*', "", xml_text)
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?dataValidations[^>]*>.*?</(?:[\w\-]+:)?dataValidations>\s*',
        "", xml_text, flags=re.DOTALL,
    )

    def _filter_cond_fmt(m: re.Match) -> str:
        cf_text = m.group(0)
        sqref_m = re.search(r'sqref="([^"]+)"', cf_text)
        if not sqref_m:
            return cf_text
        sqref = sqref_m.group(1)
        parts = re.split(r'\s+', sqref)
        kept_parts = []
        for part in parts:
            if ':' in part:
                try:
                    _, min_r, _, max_r = range_boundaries(part)
                    if min_r <= keep_to_row and max_r >= keep_from_row:
                        kept_parts.append(part)
                except (ValueError, IndexError):
                    kept_parts.append(part)
            else:
                cell_match = re.match(r'^([A-Z]+)(\d+)$', part)
                if cell_match:
                    row_num = int(cell_match.group(2))
                    if keep_from_row <= row_num <= keep_to_row:
                        kept_parts.append(part)
                else:
                    kept_parts.append(part)
        if not kept_parts:
            return ''
        new_sqref = ' '.join(kept_parts)
        return re.sub(r'sqref="[^"]*"', f'sqref="{new_sqref}"', cf_text)

    xml_text = re.sub(
        r'<(?:[\w\-]+:)?conditionalFormatting[^>]*>.*?</(?:[\w\-]+:)?conditionalFormatting>',
        _filter_cond_fmt, xml_text, flags=re.DOTALL,
    )
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?conditionalFormatting[^/]*/>\s*',
        _filter_cond_fmt, xml_text,
    )

    def _reset_sheetview(m: re.Match) -> str:
        sv = m.group(0)
        sv = re.sub(r'topLeftCell="[A-Z]+\d+"', 'topLeftCell="A1"', sv)
        sv = re.sub(r' ?view="pageBreakPreview"', '', sv)
        return sv
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?sheetView\b[^>]*/>', _reset_sheetview, xml_text,
    )
    xml_text = re.sub(
        r'<(?:[\w\-]+:)?sheetView\b[^>]*>.*?</(?:[\w\-]+:)?sheetView>',
        _reset_sheetview, xml_text, flags=re.DOTALL,
    )

    def _reset_selection(m: re.Match) -> str:
        sel = m.group(0)
        sel = re.sub(r'activeCell="[A-Z]+\d+"', 'activeCell="A1"', sel)
        sel = re.sub(r'sqref="[A-Z]+\d+(?::[A-Z]+\d+)?"', 'sqref="A1"', sel)
        return sel
    xml_text = re.sub(r'<(?:[\w\-]+:)?selection\b[^>]*/>', _reset_selection, xml_text)

    def _filter_breaks(m: re.Match) -> str:
        bk_open = m.group(1)
        bk_content = m.group(2)
        bk_close = m.group(3)
        kept_brs = []
        for br_m in re.finditer(r'<(?:[\w\-]+:)?brk[^>]*/>', bk_content):
            bx = br_m.group(0)
            id_m = re.search(r'id="(\d+)"', bx)
            if not id_m:
                continue
            br_r = int(id_m.group(1))
            if br_r < keep_from_row or br_r > keep_to_row:
                continue
            new_br_r = br_r - keep_from_row + 1
            bx = re.sub(r'id="\d+"', f'id="{new_br_r}"', bx)
            kept_brs.append(bx)
        if not kept_brs:
            return ""
        return bk_open + "".join(kept_brs) + bk_close

    xml_text = re.sub(
        r'(<(?:[\w\-]+:)?rowBreaks[^>]*>)(.*?)(</(?:[\w\-]+:)?rowBreaks>)',
        _filter_breaks, xml_text, flags=re.DOTALL,
    )

    return xml_text.encode("utf-8")

def _filter_drawing_xml(
    drawing_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> bytes:
    """Отфильтровать drawing XML, оставляя только anchors нужного диапазона.

    Использует regex-based строковые операции вместо ET.fromstring/ET.tostring
    для сохранения оригинальных namespace declarations (xmlns:ns2, xmlns:ns4
    и т.д.), которые Python ET может переименовать при сериализации,
    вызывая ошибку Excel "Repaired Records: Drawing shape".

    Поддерживает twoCellAnchor, oneCellAnchor и absoluteAnchor.
    Корректирует row-позиции anchors.
    """
    result, _ = _filter_drawing_xml_with_rids(
        drawing_data, keep_from_row, keep_to_row,
    )
    return result


def _filter_drawing_xml_with_rids(
    drawing_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> Tuple[bytes, Set[str]]:
    """Filter drawing XML using lxml and return retained image rIds.

    Uses proper XML DOM parsing instead of regex to produce valid OOXML.
    Handles twoCellAnchor, oneCellAnchor, and absoluteAnchor elements.
    Collects r:embed rIds from retained <a:blip> elements.

    Returns:
        Tuple of (filtered_xml_bytes, retained_rids).
    """
    retained_rids: Set[str] = set()

    if _HAS_LXML:
        return _filter_drawing_xml_lxml(
            drawing_data, keep_from_row, keep_to_row, retained_rids)

    # Fallback: regex-based (original code)
    return _filter_drawing_xml_regex(
        drawing_data, keep_from_row, keep_to_row, retained_rids)


def _filter_drawing_xml_lxml(
    drawing_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
    retained_rids: Set[str],
) -> Tuple[bytes, Set[str]]:
    """lxml-based drawing XML filter — produces valid OOXML output."""
    root = _lxml_etree.fromstring(drawing_data)

    # Register namespaces so lxml uses proper prefixes when serializing
    for prefix, uri in _LXML_NS.items():
        _lxml_etree.register_namespace(prefix, uri)

    ns_xdr = NS_DRAWING
    ns_a = NS_DRAWINGML
    ns_r = NS_R

    to_remove = []

    for anchor in root:
        tag = _lxml_etree.QName(anchor.tag).localname

        if tag in ('twoCellAnchor', 'oneCellAnchor'):
            from_elem = anchor.find(f'{{{ns_xdr}}}from')
            to_elem = anchor.find(f'{{{ns_xdr}}}to')

            from_row = 0
            to_row = 0

            if from_elem is not None:
                row_elem = from_elem.find(f'{{{ns_xdr}}}row')
                if row_elem is not None and row_elem.text:
                    from_row = int(row_elem.text)

            if to_elem is not None:
                row_elem = to_elem.find(f'{{{ns_xdr}}}row')
                if row_elem is not None and row_elem.text:
                    to_row = int(row_elem.text)

            # Remove anchors completely outside the range
            if to_row < keep_from_row or from_row > keep_to_row:
                to_remove.append(anchor)
                continue

            # Remove anchors whose CENTER is outside the range
            # (prevents bleeding from adjacent operations)
            center_row = (from_row + to_row) / 2.0
            if center_row < keep_from_row or center_row > keep_to_row:
                to_remove.append(anchor)
                continue

            # Clamp from/to to the keep range, then reindex
            clamped_from = max(from_row, keep_from_row)
            clamped_to = min(to_row, keep_to_row)

            if from_elem is not None:
                row_elem = from_elem.find(f'{{{ns_xdr}}}row')
                if row_elem is not None:
                    row_elem.text = str(clamped_from - keep_from_row)
                # Reset rowOff to 0 when clamping from_row
                if from_row < keep_from_row:
                    row_off = from_elem.find(f'{{{ns_xdr}}}rowOff')
                    if row_off is not None:
                        row_off.text = '0'

            if to_elem is not None:
                row_elem = to_elem.find(f'{{{ns_xdr}}}row')
                if row_elem is not None:
                    row_elem.text = str(clamped_to - keep_from_row)

            # Collect rIds from <a:blip r:embed="..."> elements
            for blip in anchor.iter(f'{{{ns_a}}}blip'):
                embed = blip.get(f'{{{ns_r}}}embed')
                if embed:
                    retained_rids.add(embed)

        elif tag == 'absoluteAnchor':
            to_remove.append(anchor)

    for elem in to_remove:
        root.remove(elem)

    # Renumber all cNvPr ids sequentially (WPS generates huge ids like 101820)
    id_counter = 0
    for cnvpr in root.iter():
        if _lxml_etree.QName(cnvpr.tag).localname == 'cNvPr':
            id_counter += 1
            cnvpr.set('id', str(id_counter))

    result_bytes = _lxml_etree.tostring(
        root, xml_declaration=True, encoding='UTF-8', standalone=True)

    return result_bytes, retained_rids


def _filter_drawing_xml_regex(
    drawing_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
    retained_rids: Set[str],
) -> Tuple[bytes, Set[str]]:
    """Regex-based fallback for when lxml is unavailable."""
    xml_text = drawing_data.decode('utf-8')

    _ANCHOR_TAG_RE = re.compile(
        r'<(?:[\w\-]+:)?(?:two|one)CellAnchor\b[^>]*>'
        r'.*?'
        r'</(?:[\w\-]+:)?(?:two|one)CellAnchor>',
        re.DOTALL,
    )
    _ABS_ANCHOR_RE = re.compile(
        r'<(?:[\w\-]+:)?absoluteAnchor\b[^>]*>'
        r'.*?'
        r'</(?:[\w\-]+:)?absoluteAnchor>',
        re.DOTALL,
    )

    def _filter_anchor(m: re.Match) -> str:
        anchor_xml = m.group(0)

        from_row = 0
        from_m = re.search(
            r'<(?:[\w\-]+:)?from\b[^>]*>(.*?)</(?:[\w\-]+:)?from>',
            anchor_xml, re.DOTALL,
        )
        if from_m:
            row_m = re.search(
                r'<(?:[\w\-]+:)?row>(\d+)</(?:[\w\-]+:)?row>',
                from_m.group(1),
            )
            if row_m:
                from_row = int(row_m.group(1))

        to_row = from_row
        to_m = re.search(
            r'<(?:[\w\-]+:)?to\b[^>]*>(.*?)</(?:[\w\-]+:)?to>',
            anchor_xml, re.DOTALL,
        )
        if to_m:
            row_m = re.search(
                r'<(?:[\w\-]+:)?row>(\d+)</(?:[\w\-]+:)?row>',
                to_m.group(1),
            )
            if row_m:
                to_row = int(row_m.group(1))

        if to_row < keep_from_row or from_row > keep_to_row:
            return ''

        def _update_row(row_m_inner: re.Match) -> str:
            tag_open = row_m_inner.group(1)
            val = int(row_m_inner.group(2))
            tag_close = row_m_inner.group(3)
            new_val = max(0, val - keep_from_row)
            return f'{tag_open}{new_val}{tag_close}'

        anchor_xml = re.sub(
            r'(<(?:[\w\-]+:)?row>)(\d+)(</(?:[\w\-]+:)?row>)',
            _update_row, anchor_xml,
        )

        for blip_m in re.finditer(
            r'<(?:[\w\-]+:)?blip\b[^>]*>', anchor_xml,
        ):
            embed_m = re.search(
                r'r:embed="([^"]+)"', blip_m.group(0),
            )
            if embed_m:
                retained_rids.add(embed_m.group(1))

        return anchor_xml

    xml_text = _ANCHOR_TAG_RE.sub(_filter_anchor, xml_text)

    def _filter_absolute(m: re.Match) -> str:
        if keep_from_row > 0:
            return ''
        return m.group(0)

    xml_text = _ABS_ANCHOR_RE.sub(_filter_absolute, xml_text)

    _id_counter = [0]
    def _renumber_id(m: re.Match) -> str:
        _id_counter[0] += 1
        return f'{m.group(1)}{_id_counter[0]}{m.group(3)}'

    xml_text = re.sub(
        r'(id=")(\d+)(")',
        _renumber_id, xml_text,
    )

    return xml_text.encode('utf-8'), retained_rids


def _filter_drawing_rels(
    rels_data: bytes,
    retained_rids: Set[str],
) -> bytes:
    """Filter drawing .rels, keeping only retained rIds.

    Removes <Relationship> entries whose Id is not in retained_rids.
    Uses lxml for valid XML output.
    """
    if _HAS_LXML:
        root = _lxml_etree.fromstring(rels_data)
        ns = NS_PKG_RELS
        to_remove = []
        for rel in root:
            rid = rel.get('Id', '')
            if rid and rid not in retained_rids:
                to_remove.append(rel)
        for elem in to_remove:
            root.remove(elem)
        return _lxml_etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)

    # Regex fallback
    rels_text = rels_data.decode('utf-8')
    def _filter_rel(m: re.Match) -> str:
        rel_xml = m.group(0)
        id_m = re.search(r'Id="([^"]+)"', rel_xml)
        if id_m and id_m.group(1) not in retained_rids:
            return ''
        return rel_xml
    return re.sub(r'<Relationship\b[^>]*/>', _filter_rel, rels_text).encode('utf-8')


def _filter_vml_xml(
    vml_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> bytes:
    """Filter VML XML, removing shapes outside the row range.

    Uses lxml for proper namespace handling.
    VML shapes use two positioning modes:
      1. row attribute (direct row binding)
      2. CSS style="position:absolute;top:..." (approximate)
    """
    if _HAS_LXML:
        try:
            root = _lxml_etree.fromstring(vml_data)
        except Exception:
            return vml_data
    else:
        try:
            root = ET.fromstring(vml_data)
        except ET.ParseError:
            return vml_data

    vml_ns = 'urn:schemas-microsoft-com:vml'
    office_ns = 'urn:schemas-microsoft-com:office:office'
    ROW_HEIGHT_PT = 15

    to_remove = []
    for elem in root.iter(f'{{{vml_ns}}}shape'):
        row_attr = elem.get('row') or elem.get(f'{{{office_ns}}}row')
        if row_attr is not None:
            try:
                row_num = int(row_attr.split()[0])
                if row_num < keep_from_row or row_num > keep_to_row:
                    to_remove.append(elem)
                    continue
            except (ValueError, IndexError):
                pass

        style = elem.get('style', '')
        if 'top:' in style.lower() or 'top: ' in style.lower():
            try:
                top_match = re.search(r'top:\s*([\d.]+)\s*(?:pt|mm|cm)?', style, re.IGNORECASE)
                if top_match:
                    top_pt = float(top_match.group(1))
                    approx_row = int(top_pt / ROW_HEIGHT_PT) + 1
                    if approx_row < keep_from_row or approx_row > keep_to_row:
                        to_remove.append(elem)
                        continue
            except (ValueError, IndexError):
                pass

    for elem in to_remove:
        for parent in root.iter():
            if elem in list(parent):
                parent.remove(elem)
                break

    if _HAS_LXML:
        return _lxml_etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)
    return _serialize_xml(root, NS_MAIN, extra_ns={'v': VML_NS, 'o': OFFICE_NS})


def _filter_comments_xml(
    comments_data: bytes,
    keep_from_row: int,
    keep_to_row: int,
) -> Optional[bytes]:
    """Filter xl/commentsN.xml — keep and reindex comments in row range.

    Comments whose ref falls within [keep_from_row, keep_to_row] are kept
    with reindexed row numbers.

    Returns:
        Filtered XML bytes, or None if no comments in range.
    """
    if _HAS_LXML:
        try:
            root = _lxml_etree.fromstring(comments_data)
        except Exception:
            return comments_data
    else:
        try:
            root = ET.fromstring(comments_data)
        except ET.ParseError:
            return comments_data

    ns = NS_MAIN
    comment_list = root.find(f'{{{ns}}}commentList')
    if comment_list is None:
        return comments_data

    to_remove = []
    for comment_el in comment_list.findall(f'{{{ns}}}comment'):
        ref = comment_el.get('ref', '')
        ref_match = re.match(r'R(\d+)(.*)', ref)
        if ref_match:
            row_num = int(ref_match.group(1))
            suffix = ref_match.group(2)
            if row_num < keep_from_row or row_num > keep_to_row:
                to_remove.append(comment_el)
            else:
                new_row = row_num - keep_from_row + 1
                comment_el.set('ref', f'R{new_row}{suffix}')

    for elem in to_remove:
        comment_list.remove(elem)

    if len(comment_list) == 0:
        return None

    if _HAS_LXML:
        return _lxml_etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)
    return _serialize_xml(root, NS_MAIN)

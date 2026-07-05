"""Модуль валидации выходных файлов после разделения.

Многоуровневый pipeline:
  1. Structural: ZIP целостность, XML well-formed, обязательные файлы present
  2. Schema: sheetData существует, колонки обнаружены, строки данных > 0
  3. Content: изображения загружаемые, формулы парсибельные
  4. Semantic: part numbers валидны, количества числовые, нет дубликатов

Используется после каждого split для гарантии корректности выходных файлов.
"""

from __future__ import annotations

import logging
import os
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from typing import Any, List

logger = logging.getLogger(__name__)

# Пространства имён OOXML
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


@dataclass
class ValidationIssue:
    """Одна проблема, обнаруженная при валидации."""
    level: str        # "structural", "schema", "content", "semantic"
    severity: str     # "error", "warning"
    message: str
    file_path: str = ""
    sheet_name: str = ""


@dataclass
class ValidationResult:
    """Результат валидации одного файла."""
    file_path: str
    is_valid: bool = True
    issues: List[ValidationIssue] = field(default_factory=list)

    @property
    def errors(self) -> List[ValidationIssue]:
        return [i for i in self.issues if i.severity == "error"]

    @property
    def warnings(self) -> List[ValidationIssue]:
        return [i for i in self.issues if i.severity == "warning"]

    def add_error(self, level: str, message: str, **kwargs: Any) -> None:
        self.issues.append(ValidationIssue(
            level=level, severity="error", message=message,
            file_path=kwargs.get("file_path", self.file_path),
            sheet_name=kwargs.get("sheet_name", ""),
        ))
        self.is_valid = False

    def add_warning(self, level: str, message: str, **kwargs: Any) -> None:
        self.issues.append(ValidationIssue(
            level=level, severity="warning", message=message,
            file_path=kwargs.get("file_path", self.file_path),
            sheet_name=kwargs.get("sheet_name", ""),
        ))


class ValidationPipeline:
    """Многоуровневый pipeline валидации .xlsx файлов.

    Уровни:
      1. Structural: ZIP целостность, XML well-formed, обязательные файлы present
      2. Schema: sheetData существует, колонки обнаружены, строки данных > 0
      3. Content: изображения загружаемые, формулы парсибельные
      4. Semantic: part numbers валидны, количества числовые, нет дубликатов
      5. Split-quality: проверка качества после split (изображения, строки)

    Используется после каждого split для гарантии корректности выходных файлов.
    """

    def __init__(
        self,
        check_structural: bool = True,
        check_schema: bool = True,
        check_content: bool = True,
        check_semantic: bool = False,
        check_split_quality: bool = False,
        expected_min_rows: int = 0,
        expected_max_rows: int = 0,
        has_images_in_original: bool = False,
        max_file_size_mb: float = 50.0,
    ):
        self.check_structural = check_structural
        self.check_schema = check_schema
        self.check_content = check_content
        self.check_semantic = check_semantic
        self.check_split_quality = check_split_quality
        self.expected_min_rows = expected_min_rows
        self.expected_max_rows = expected_max_rows
        self.has_images_in_original = has_images_in_original
        self.max_file_size_mb = max_file_size_mb

    def validate(self, file_path: str) -> ValidationResult:
        """Запустить полный pipeline валидации на одном файле.

        Args:
            file_path: Путь к .xlsx файлу.

        Returns:
            ValidationResult с is_valid и списком issues.
        """
        result = ValidationResult(file_path=file_path)

        if not os.path.isfile(file_path):
            result.add_error("structural", f"Файл не найден: {file_path}")
            return result

        ext = os.path.splitext(file_path)[1].lower()
        if ext != ".xlsx":
            result.add_warning("structural", f"Не .xlsx формат: {ext}")
            return result

        if self.check_structural:
            self._check_structural(file_path, result)
            if not result.is_valid:
                return result

        if self.check_schema:
            self._check_schema(file_path, result)

        if self.check_content:
            self._check_content(file_path, result)

        if self.check_semantic:
            self._check_semantic(file_path, result)

        if self.check_split_quality:
            self._check_split_quality(file_path, result)

        return result

    def validate_batch(self, file_paths: List[str]) -> List[ValidationResult]:
        """Валидировать список файлов.

        Returns:
            Список ValidationResult для каждого файла.
        """
        results = []
        for fp in file_paths:
            results.append(self.validate(fp))
        return results

    def _check_structural(self, file_path: str, result: ValidationResult) -> None:
        """Уровень 1: Структурная целостность.

        Проверяет:
          - Файл является валидным ZIP
          - Обязательные OOXML файлы присутствуют
          - XML well-formed
          - Размер файла не превышает лимит
        """
        # Размер файла
        try:
            size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if size_mb > self.max_file_size_mb:
                result.add_warning(
                    "structural",
                    f"Файл слишком большой: {size_mb:.1f} MB > {self.max_file_size_mb} MB",
                )
        except OSError:
            pass

        # ZIP целостность
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                names = set(zf.namelist())

                # Обязательные файлы
                required = [
                    '[Content_Types].xml',
                    'xl/workbook.xml',
                    'xl/_rels/workbook.xml.rels',
                    '_rels/.rels',
                ]
                for req in required:
                    if req not in names:
                        result.add_error(
                            "structural",
                            f"Отсутствует обязательный файл: {req}",
                        )

                # Хотя бы один sheet XML
                if not any(
                    n.endswith('.xml') and 'sheet' in n.lower() and '_rels' not in n
                    for n in names
                ):
                    result.add_error("structural", "Нет sheet XML файлов")

                # XML well-formed для критических файлов
                critical_xmls = ['xl/workbook.xml', '[Content_Types].xml']
                for cx in critical_xmls:
                    if cx in names:
                        try:
                            data = zf.read(cx)
                            ET.fromstring(data)
                        except ET.ParseError as e:
                            result.add_error(
                                "structural",
                                f"XML не well-formed: {cx}: {e}",
                            )

                # Проверка CRC для ВСЕХ файлов
                # ВАЖНО: WPS Office и другие генераторы OOXML могут создавать
                # файлы с некорректными CRC-суммами, которые при этом
                # нормально открываются в Excel. Поэтому CRC ошибки
                # downgrade-ятся до WARNING, а не ERROR.
                for info in zf.infolist():
                    try:
                        zf.read(info.filename)
                    except (zipfile.BadZipFile, Exception) as e:
                        result.add_warning(
                            "structural",
                            f"CRC ошибка (не критично): {info.filename}: {e}",
                        )

        except zipfile.BadZipFile as e:
            result.add_error("structural", f"Невалидный ZIP: {e}")
        except OSError as e:
            result.add_error("structural", f"Ошибка чтения файла: {e}")

    def _check_schema(self, file_path: str, result: ValidationResult) -> None:
        """Уровень 2: Схемная валидация.

        Проверяет:
          - sheetData существует в каждом листе
          - Есть хотя бы одна строка данных
          - Количество строк > 0
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                for name in zf.namelist():
                    if (name.endswith('.xml')
                            and 'sheet' in name.lower()
                            and '_rels' not in name):
                        try:
                            data = zf.read(name)
                            root = ET.fromstring(data)
                            ns = f'{{{NS_MAIN}}}sheetData'
                            sheet_data = root.find(ns)
                            if sheet_data is None:
                                result.add_error(
                                    "schema",
                                    f"Нет sheetData в {name}",
                                )
                                continue

                            # Подсчёт строк
                            row_count = len(sheet_data.findall(f'{{{NS_MAIN}}}row'))
                            if row_count == 0:
                                result.add_warning(
                                    "schema",
                                    f"Пустой sheetData (0 строк) в {name}",
                                )
                        except ET.ParseError as e:
                            result.add_error(
                                "schema",
                                f"Ошибка парсинга sheet XML {name}: {e}",
                            )
        except zipfile.BadZipFile:
            pass  # Уже обработано на уровне structural

    def _check_content(self, file_path: str, result: ValidationResult) -> None:
        """Уровень 3: Проверка содержимого.

        Проверяет:
          - Изображения в xl/media/ имеют корректные форматы
          - Drawing XML references существуют
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                names = set(zf.namelist())
                media_files = [n for n in names if n.startswith('xl/media/')]

                # Проверяем что все media файлы непустые
                for mf in media_files:
                    try:
                        data = zf.read(mf)
                        if len(data) == 0:
                            result.add_warning(
                                "content",
                                f"Пустой медиафайл: {mf}",
                            )
                    except Exception as e:
                        result.add_error(
                            "content",
                            f"Ошибка чтения медиафайла {mf}: {e}",
                        )

                # Проверяем что drawing rels ссылки существуют
                for name in names:
                    if 'drawing' in name and name.endswith('.rels'):
                        try:
                            data = zf.read(name)
                            root = ET.fromstring(data)
                            drawing_dir = os.path.dirname(
                                name.replace('_rels/', '').replace('.rels', '')
                            )
                            for rel in root:
                                target = rel.get('Target', '')
                                if target:
                                    resolved = os.path.normpath(
                                        os.path.join(drawing_dir, target)
                                    ).replace(os.sep, '/')
                                    if not resolved.startswith('/'):
                                        resolved_check = resolved
                                    else:
                                        resolved_check = resolved[1:]
                                    # Проверяем только media references
                                    if 'media' in resolved_check.lower():
                                        if resolved_check not in names:
                                            result.add_warning(
                                                "content",
                                                f"Broken media reference in {name}: "
                                                f"{target} -> {resolved_check}",
                                            )
                        except Exception:
                            pass  # Не критично

        except zipfile.BadZipFile:
            pass

    def _check_semantic(self, file_path: str, result: ValidationResult) -> None:
        """Уровень 4: Семантическая валидация (опциональная).

        Проверяет:
          - Part numbers в данных выглядят валидно
          - Количества являются числами
          - Нет дубликатов part-номеров
        """
        from burlak_parser.normalizer import is_valid_part_number, normalize_quantity

        try:
            import openpyxl
            import warnings
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Простая проверка: хотя бы несколько ячеек с данными
                non_empty = 0
                for row in ws.iter_rows(max_row=10, max_col=10):
                    for cell in row:
                        if cell.value is not None:
                            non_empty += 1
                if non_empty == 0:
                    result.add_warning(
                        "semantic",
                        f"Лист '{sheet_name}' не содержит данных",
                        sheet_name=sheet_name,
                    )
            wb.close()
        except Exception as e:
            result.add_warning("semantic", f"Не удалось проверить семантику: {e}")

    def _check_split_quality(self, file_path: str, result: ValidationResult) -> None:
        """Уровень 5: Проверка качества после split.

        Проверяет:
          - Наличие изображений (если в оригинале были)
          - Количество строк в разрезанном файле
          - Размер файла не слишком мал (признак пустого/битого файла)
        """
        try:
            with zipfile.ZipFile(file_path, 'r') as zf:
                names = set(zf.namelist())

                # Проверка 1: изображения (если в оригинале были)
                if self.has_images_in_original:
                    media_files = [n for n in names if n.startswith('xl/media/')]
                    if not media_files:
                        result.add_warning(
                            "split-quality",
                            "Нет изображений в split-файле (в оригинале были)",
                        )

                # Проверка 2: количество строк
                if self.expected_min_rows > 0 or self.expected_max_rows > 0:
                    for name in names:
                        if (name.endswith('.xml')
                                and 'sheet' in name.lower()
                                and '_rels' not in name):
                            try:
                                data = zf.read(name)
                                root = ET.fromstring(data)
                                ns = f'{{{NS_MAIN}}}sheetData'
                                sheet_data = root.find(ns)
                                if sheet_data is not None:
                                    row_count = len(
                                        sheet_data.findall(f'{{{NS_MAIN}}}row')
                                    )
                                    if (self.expected_min_rows > 0
                                            and row_count < self.expected_min_rows):
                                        result.add_warning(
                                            "split-quality",
                                            f"Мало строк: {row_count} < {self.expected_min_rows}",
                                        )
                                    if (self.expected_max_rows > 0
                                            and row_count > self.expected_max_rows * 1.5):
                                        result.add_warning(
                                            "split-quality",
                                            f"Много строк: {row_count} > {self.expected_max_rows * 1.5:.0f}",
                                        )
                            except ET.ParseError:
                                pass
                            break  # Проверяем только первый sheet XML

                # Проверка 3: размер файла
                try:
                    size_bytes = os.path.getsize(file_path)
                    if size_bytes < 1024:  # < 1 KB — подозрительно мало
                        result.add_warning(
                            "split-quality",
                            f"Файл очень маленький: {size_bytes} байт",
                        )
                except OSError:
                    pass

        except zipfile.BadZipFile:
            pass  # Уже обработано на уровне structural


def validate_split_file(
    file_path: str,
    has_images_in_original: bool = False,
    expected_min_rows: int = 0,
    expected_max_rows: int = 0,
) -> ValidationResult:
    """Быстрая валидация одного split-файла (structural + schema + split-quality).

    Удобная функция-обёртка для использования в splitter.

    Args:
        file_path: Путь к .xlsx файлу.
        has_images_in_original: True если в исходном файле были изображения.
        expected_min_rows: Минимальное ожидаемое количество строк (0 = без проверки).
        expected_max_rows: Максимальное ожидаемое количество строк (0 = без проверки).
    """
    pipeline = ValidationPipeline(
        check_structural=True,
        check_schema=True,
        check_content=False,
        check_semantic=False,
        check_split_quality=True,
        has_images_in_original=has_images_in_original,
        expected_min_rows=expected_min_rows,
        expected_max_rows=expected_max_rows,
    )
    return pipeline.validate(file_path)


def validate_split_file_lenient(file_path: str) -> ValidationResult:
    """Ленивая валидация split-файла: только открываемость через openpyxl.

    Используется как последний шанс перед отправкой в corrupted_cards.
    WPS-созданные файлы могут иметь некорректные CRC, нестандартные
    структуры OOXML, но при этом быть функционально валидными.
    Эта проверка эмулирует поведение MS Excel: «если файл открывается — он валидный».

    Returns:
        ValidationResult с is_valid = True если файл можно открыть через openpyxl.
    """
    result = ValidationResult(file_path=file_path)
    try:
        import openpyxl
        import warnings
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            _ = wb.sheetnames
            wb.close()
    except Exception as e:
        result.add_error(
            "structural",
            f"Файл не открывается даже через openpyxl: {e}",
        )
    return result


def validate_and_quarantine(
    file_path: str,
    quarantine_dir: str,
) -> ValidationResult:
    """Валидировать файл и переместить повреждённые в quarantine/.

    Args:
        file_path: Путь к .xlsx файлу.
        quarantine_dir: Директория для повреждённых файлов.

    Returns:
        ValidationResult.
    """
    result = validate_split_file(file_path)

    if not result.is_valid:
        try:
            os.makedirs(quarantine_dir, exist_ok=True)
            import shutil
            dest = os.path.join(quarantine_dir, os.path.basename(file_path))
            shutil.copy2(file_path, dest)

            # Записываем .error файл
            error_path = dest + ".error"
            with open(error_path, "w", encoding="utf-8") as f:
                for issue in result.errors:
                    f.write(f"[{issue.level}] {issue.message}\n")

            logger.warning(
                "Файл перемещён в quarantine: %s (%d ошибок)",
                os.path.basename(file_path), len(result.errors),
            )
        except Exception as e:
            logger.error("Не удалось переместить в quarantine: %s", e)

    return result

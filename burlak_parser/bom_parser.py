"""Модуль чтения BOM-файла (Bill of Materials / Ведомость материалов).

Формат: .xlsx (таблица на китайском/английском/русском языках).

Алгоритм работы:
  1. Загружает .xlsx и обходит ВСЕ листы.
  2. Для каждого листа использует эвристический анализатор для поиска:
     - Строки заголовков
     - Колонок с парт-номерами, названиями и количествами
     - Колонок комплектаций
  3. Строит ГЛОБАЛЬНЫЙ словарь парт-номеров и названий (сканирует ВСЕ строки,
     а не только для конкретной комплектации).
  4. Извлекает количества по каждой комплектации.
  5. Агрегирует данные по всем листам.

Универсален — не привязан к конкретным моделям автомобилей, брендам или
форматам. Использует эвристический анализатор из heuristic_analyzer.py.

Класс BOMService — обёртка для использования в FastAPI/серверной архитектуре.
"""

from __future__ import annotations

import logging
import os
import tempfile
import zipfile
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

from burlak_parser.heuristic_analyzer import (
    HeuristicAnalyzer,
    clean_cell_text,
    NAME_KEYWORDS,
    QTY_KEYWORDS,
)
from burlak_parser.normalizer import (
    normalize_quantity,
    normalize_part_number,
    clean_part_number,
    is_valid_part_number,
)

# Листы с мета-именами, которые НЕ создают конфигурации (单车用量, 发动机附件 и т.д.)
_NON_CONFIG_SHEET_KEYWORDS = (
    "单车用量", "组件数量", "发动机附件",
    "Расход на один автомобиль", "количество компонентов",
)

# Ключевые слова для идентификации SWM-стиля листов (总装BOM/涂装BOM/焊装BOM)
# Эти листы не имеют отдельных config-колонок, но имеют qty-колонку;
# все три листа агрегируются в одну общую конфигурацию.
_SWM_MULTISHEET_BOM_KEYWORDS = (
    "总装bom", "涂装bom", "焊装bom",
    "总装", "涂装", "焊装",
)

logger = logging.getLogger(__name__)


@dataclass
class PartInfo:
    """Информация о детали из BOM."""
    part_number: str
    name_cn: str = ""
    name_en: str = ""
    # Количество для конкретной комплектации (будет заполнено после выбора)
    quantity: float = 0.0
    # Номера/коды комплектаций, для которых указана деталь
    applicable_configs: List[str] = field(default_factory=list)


@dataclass
class BOMData:
    """Результат парсинга BOM-файла."""
    parts: Dict[str, PartInfo]  # part_number -> PartInfo
    config_names: List[str]  # названия колонок комплектаций
    config_quantities: Dict[str, Dict[str, float]]  # config_name -> {part_number -> qty}
    source_file: str = ""
    # Глобальный словарь названий (составлен из ВСЕХ строк, а не только для комплектации)
    global_names: Dict[str, Tuple[str, str]] = field(default_factory=dict)  # part_number -> (name_cn, name_en)


def _detect_multi_block_layout(
    ws: Any,
    header_row: int,
    part_no_col: int,
    name_cn_col: int,
    qty_col: int,
) -> List[Tuple[int, int, int]]:
    """Detect multi-block horizontal layout (side-by-side tables).

    Some BOM sheets (e.g., SWM 舒享版焊装合件) have multiple blocks
    of columns laid out horizontally, separated by empty columns.
    Each block has the same structure: part_no, name, qty.

    Returns:
        List of (part_no_col, name_col, qty_col) tuples for each block.
        Always includes the primary block first.
    """
    from burlak_parser.heuristic_analyzer import PART_NO_KEYWORDS

    max_col = ws.max_column or 20
    blocks: List[Tuple[int, int, int]] = [(part_no_col, name_cn_col, qty_col)]

    if header_row < 1 or max_col < 5:
        return blocks

    # Scan header row for additional part_no columns
    # Pattern: part_no, name, qty, [empty separator], part_no, name, qty, ...
    row_vals: List[str] = []
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        row_vals.append(str(v).strip().lower() if v is not None else "")

    # Find all columns that contain part_no keywords
    pn_cols: List[int] = []
    for idx, val in enumerate(row_vals):
        if val and any(kw in val for kw in PART_NO_KEYWORDS):
            pn_cols.append(idx + 1)  # 1-indexed

    if len(pn_cols) <= 1:
        return blocks

    # For each additional part_no column, find matching name and qty columns
    for additional_pn_col in pn_cols:
        if additional_pn_col == part_no_col:
            continue

        # Find name and qty columns near this part_no column
        # They should be within +1 to +3 columns
        found_name = None
        found_qty = None
        for offset in range(1, 5):
            check_col = additional_pn_col + offset
            if check_col > max_col:
                break
            val = row_vals[check_col - 1] if check_col - 1 < len(row_vals) else ""
            if val and any(kw in val for kw in NAME_KEYWORDS):
                found_name = check_col
            if val and any(kw in val for kw in QTY_KEYWORDS):
                found_qty = check_col

        if found_name and found_qty:
            blocks.append((additional_pn_col, found_name, found_qty))

    if len(blocks) > 1:
        logger.debug(
            "Multi-block layout detected: %d blocks in %s (header row %d)",
            len(blocks), ws.title, header_row,
        )

    return blocks


def parse_bom(file_path: str) -> BOMData:
    """Разобрать BOM-файл и вернуть структурированные данные.

    Алгоритм SWM multi-sheet BOM:
      - Если файл содержит листы (总装BOM/涂装BOM/焊装BOM) без config-колонок,
        они агрегируются в ЕДИНЫЙ конфиг "SWM_COMBINED".
      - Зачёркнутые ячейки (strike) пропускаются для парт-номеров И кол-ва.

    Args:
        file_path: Путь к .xlsx файлу BOM.

    Returns:
        BOMData со всеми извлечёнными данными.
    """
    logger.info("Загрузка BOM-файла: %s", file_path)

    wb = None
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.warning("Не удалось загрузить обычным режимом (%s), пробуем read_only", e)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # Check if data_only=True produced empty results (formulas never cached)
    # If first data sheet has 0 non-None values in non-header rows, reload without data_only
    _needs_data_only_reload = False
    try:
        for sn in wb.sheetnames[:3]:  # Check first 3 sheets only
            ws_check = wb[sn]
            if (ws_check.max_row or 0) < 2:
                continue
            non_none_count = 0
            check_rows = min(10, ws_check.max_row or 1)
            check_cols = min(20, ws_check.max_column or 10)
            for r in range(2, check_rows + 1):
                for c in range(1, check_cols + 1):
                    if ws_check.cell(r, c).value is not None:
                        non_none_count += 1
            if non_none_count == 0:
                _needs_data_only_reload = True
                break
    except (AttributeError, TypeError, IndexError):
        pass

    if _needs_data_only_reload:
        logger.warning(
            "data_only=True produced empty data — formulas may not be cached. "
            "Reloading without data_only."
        )
        if wb is not None:
            wb.close()
        try:
            wb = openpyxl.load_workbook(file_path)
        except (OSError, zipfile.BadZipFile):
            wb = openpyxl.load_workbook(file_path, read_only=True)

    # ── SWM multi-sheet BOM aggregate config name ──
    # Все sub-листы SWM (总装/涂装/焊装) вносят данные в ОДНУ конфигурацию
    # Инициализируются ДО try-блока для гарантии доступности в finally
    _SWM_COMBINED_CONFIG = "SWM_COMBINED"
    _swm_multisheet_sheets: List[str] = []  # собираем имена для логирования

    try:
        sheet_names = wb.sheetnames

        # ── Результаты, агрегированные по всем листам ──
        all_parts: Dict[str, PartInfo] = {}
        all_config_quantities: Dict[str, Dict[str, float]] = {}
        all_config_names: List[str] = []
        all_global_names: Dict[str, Tuple[str, str]] = {}
        seen_config_names: Dict[str, str] = {}  # config_name -> нормализованный оригинал

        for sheet_name in sheet_names:
            try:
                ws = wb[sheet_name]
                logger.info(
                    "Анализ листа: %s (строк: %s, колонок: %s)",
                    sheet_name, ws.max_row, ws.max_column,
                )

                # Анализируем лист одним вызовом (без дублирования)
                # min_configs=1: поддержка SWM-стиля листов (总装/涂装/焊装) с 0-1 конфиг-колонками
                analysis = HeuristicAnalyzer.analyze_bom_sheet(
                    ws, min_configs=1, sheet_name=sheet_name,
                )
                if analysis is None:
                    logger.info("Лист не является BOM-кандидатом, пропуск: %s", sheet_name)
                    continue

                header_rows, col_types, config_cols = analysis
                part_no_col = col_types.get("part_no", 0)
                name_cn_col = col_types.get("name_cn", 0)
                name_en_col = col_types.get("name_en", 0)

                if part_no_col == 0:
                    logger.warning("Не найдена колонка парт-номеров в листе: %s", sheet_name)
                    continue

                header_row = header_rows[0]

                # ── 3. Строим ГЛОБАЛЬНЫЙ словарь названий (ВСЕ строки, ВСЕ листы) ──
                sheet_names_dict = HeuristicAnalyzer.build_global_name_dict(
                    ws, part_no_col, name_cn_col, name_en_col, header_row,
                )
                for pn, (nc, ne) in sheet_names_dict.items():
                    if pn not in all_global_names:
                        all_global_names[pn] = (nc, ne)
                    else:
                        existing_cn, existing_en = all_global_names[pn]
                        if not existing_cn and nc:
                            existing_cn = nc
                        if not existing_en and ne:
                            existing_en = ne
                        all_global_names[pn] = (existing_cn, existing_en)

                # ── 4. Определяем колонки комплектаций (уже из analysis) ──
                qty_col = col_types.get("qty", 0)

                # ── 5. Если есть отдельная qty-колонка (спец-листы 附件 или SWM multi-sheet) ──
                is_non_config_sheet = any(kw in sheet_name for kw in _NON_CONFIG_SHEET_KEYWORDS)

                # Определяем, является ли это SWM-стилем листа (总装/涂装/焊装)
                # Такие листы агрегируются в ЕДИНЫЙ конфиг, а не создают отдельные конфиги
                sheet_name_lower = sheet_name.lower()
                is_swm_multisheet = any(
                    kw in sheet_name_lower for kw in _SWM_MULTISHEET_BOM_KEYWORDS
                )

                if (not config_cols or len(config_cols) == 0) and qty_col > 0 and not is_non_config_sheet:
                    data_start = header_row + 1

                    # SWM-стиль: все листы идут в один агрегированный конфиг
                    if is_swm_multisheet:
                        config_name = _SWM_COMBINED_CONFIG
                        _swm_multisheet_sheets.append(sheet_name)
                    else:
                        config_name = sheet_name

                    # Создаём конфиг если ещё не существует
                    if config_name not in seen_config_names:
                        seen_config_names[config_name] = config_name
                        all_config_names.append(config_name)
                        all_config_quantities[config_name] = {}

                    sheet_parts_added = 0
                    for row_idx in range(data_start, (ws.max_row or data_start) + 1):
                        # Пропускаем зачёркнутые строки (отменённые позиции)
                        if HeuristicAnalyzer.is_cell_strike(ws, row_idx, part_no_col):
                            continue
                        if qty_col > 0 and HeuristicAnalyzer.is_cell_strike(ws, row_idx, qty_col):
                            continue

                        pn = HeuristicAnalyzer.get_cell_value(ws, row_idx, part_no_col)
                        if pn is None:
                            continue
                        pn_str = clean_cell_text(pn)
                        if not pn_str or pn_str.startswith("~$"):
                            continue
                        if not is_valid_part_number(pn_str):
                            continue

                        qty_val = HeuristicAnalyzer.get_cell_value(ws, row_idx, qty_col)
                        qty = normalize_quantity(qty_val)

                        if qty > 0:
                            pn_normalized = clean_part_number(pn_str)
                            current_qty = all_config_quantities[config_name].get(pn_normalized, 0.0)
                            all_config_quantities[config_name][pn_normalized] = current_qty + qty
                            if current_qty > 0:
                                logger.debug(
                                    "Cross-sheet aggregation: %s qty %.1f + %.1f = %.1f "
                                    "for config '%s' (sheet: %s)",
                                    pn_normalized, current_qty, qty,
                                    current_qty + qty, config_name, sheet_name,
                                )

                            if pn_normalized not in all_parts:
                                all_parts[pn_normalized] = PartInfo(part_number=pn_str)
                            if config_name not in all_parts[pn_normalized].applicable_configs:
                                all_parts[pn_normalized].applicable_configs.append(config_name)
                            sheet_parts_added += 1

                    logger.info(
                        "Лист %s → config='%s': %d деталей добавлено (итого в конфиге: %d)",
                        sheet_name, config_name, sheet_parts_added,
                        len(all_config_quantities[config_name]),
                    )
                    continue

                # ── 5b. Non-config sheets (单车用量, 发动机附件) — collect parts only ──
                if is_non_config_sheet and qty_col > 0:
                    data_start = header_row + 1
                    for row_idx in range(data_start, (ws.max_row or data_start) + 1):
                        if HeuristicAnalyzer.is_cell_strike(ws, row_idx, part_no_col):
                            continue
                        pn = HeuristicAnalyzer.get_cell_value(ws, row_idx, part_no_col)
                        if pn is None:
                            continue
                        pn_str = clean_cell_text(pn)
                        if not pn_str or pn_str.startswith("~$"):
                            continue
                        if not is_valid_part_number(pn_str):
                            continue
                        pn_normalized = clean_part_number(pn_str)
                        if pn_normalized not in all_parts:
                            all_parts[pn_normalized] = PartInfo(part_number=pn_str)
                    logger.info("Лист %s: не-конфигурационный, детали собраны в all_parts", sheet_name)
                    continue

                if not config_cols:
                    logger.info("Лист %s: не найдено колонок комплектаций, пропуск", sheet_name)
                    continue

                # ── 6. Дедупликация имён комплектаций ──
                config_names: List[str] = []
                for col_idx in config_cols:
                    name = HeuristicAnalyzer.get_cell_value(ws, header_row, col_idx)
                    name_str = str(name) if name is not None else ""
                    name_str = name_str.replace("\n", " ").replace("\r", "").strip()

                    if not name_str:
                        for look_row in range(max(1, header_row - 1), 0, -1):
                            meta_val = HeuristicAnalyzer.get_cell_value(ws, look_row, col_idx)
                            if meta_val is not None:
                                meta_str = str(meta_val).strip()
                                if meta_str and len(meta_str) < 80:
                                    name_str = meta_str
                                    break
                    if not name_str:
                        name_str = f"Config_{col_idx}"
                    config_names.append(name_str)

                deduped_indices: List[int] = []
                seen_norm: Set[str] = set()
                for i, name in enumerate(config_names):
                    norm = name.lower().replace(" ", "").replace("-", "")
                    if norm not in seen_norm:
                        seen_norm.add(norm)
                        deduped_indices.append(i)

                if len(deduped_indices) < len(config_cols):
                    logger.info(
                        "Дедупликация: %d -> %d имён комплектаций",
                        len(config_cols), len(deduped_indices),
                    )
                    config_cols = [config_cols[i] for i in deduped_indices]
                    config_names = [config_names[i] for i in deduped_indices]

                # ── 7. Парсинг данных комплектаций ──
                data_start = header_row + 1
                max_row = ws.max_row or data_start
                sheet_config_count = 0

                # Detect multi-block horizontal layout (side-by-side tables)
                multi_blocks = _detect_multi_block_layout(
                    ws, header_row, part_no_col, name_cn_col, qty_col,
                )

                # Precompute strikethrough rows for part_no + qty columns
                # (config columns not checked — too many cols × rows for font access)
                strike_cols = [part_no_col]
                if qty_col > 0:
                    strike_cols.append(qty_col)
                strike_rows_cache = HeuristicAnalyzer.get_strike_rows(
                    ws, range(data_start, max_row + 1), strike_cols,
                )

                for row_idx in range(data_start, max_row + 1):
                    if row_idx in strike_rows_cache:
                        continue
                    pn = HeuristicAnalyzer.get_cell_value(ws, row_idx, part_no_col)
                    if pn is None:
                        continue
                    pn_str = clean_cell_text(pn)
                    if not pn_str or pn_str.startswith("~$"):
                        continue
                    if not is_valid_part_number(pn_str):
                        continue

                    pn_normalized = clean_part_number(pn_str)

                    if pn_normalized not in all_parts:
                        all_parts[pn_normalized] = PartInfo(part_number=pn_str)

                    part = all_parts[pn_normalized]

                    for i, col_idx in enumerate(config_cols):
                        if HeuristicAnalyzer.is_cell_strike(ws, row_idx, col_idx):
                            continue
                        config_val = str(HeuristicAnalyzer.get_cell_value(ws, row_idx, col_idx) or '').strip()

                        if config_val.upper() == 'S' and qty_col > 0:
                            qty = normalize_quantity(HeuristicAnalyzer.get_cell_value(ws, row_idx, qty_col))
                        elif config_val in ('-', '–', '—', ''):
                            continue
                        else:
                            qty = normalize_quantity(config_val)

                        if qty > 0:
                            config_name = config_names[i]
                            if config_name not in seen_config_names:
                                seen_config_names[config_name] = config_name
                                all_config_names.append(config_name)

                            if config_name not in all_config_quantities:
                                all_config_quantities[config_name] = {}

                            current_qty = all_config_quantities[config_name].get(pn_normalized, 0.0)
                            all_config_quantities[config_name][pn_normalized] = current_qty + qty
                            if current_qty > 0:
                                logger.debug(
                                    "Cross-sheet aggregation: %s qty %.1f + %.1f = %.1f "
                                    "for config '%s' (sheet: %s)",
                                    pn_normalized, current_qty, qty,
                                    current_qty + qty, config_name, sheet_name,
                                )

                            if config_name not in part.applicable_configs:
                                part.applicable_configs.append(config_name)

                            sheet_config_count += 1

                # ── 7b. Multi-block: read additional side-by-side blocks ──
                if len(multi_blocks) > 1:
                    for blk_pn, blk_name, blk_qty in multi_blocks[1:]:
                        blk_count = 0
                        # Precompute strike rows for this block's column
                        blk_strike_rows = HeuristicAnalyzer.get_strike_rows(
                            ws, range(data_start, max_row + 1), [blk_pn],
                        )
                        for row_idx in range(data_start, max_row + 1):
                            if row_idx in blk_strike_rows:
                                continue
                            pn = HeuristicAnalyzer.get_cell_value(ws, row_idx, blk_pn)
                            if pn is None:
                                continue
                            pn_str = clean_cell_text(pn)
                            if not pn_str or pn_str.startswith("~$"):
                                continue
                            if not is_valid_part_number(pn_str):
                                continue

                            pn_normalized = clean_part_number(pn_str)
                            if pn_normalized not in all_parts:
                                all_parts[pn_normalized] = PartInfo(part_number=pn_str)

                            blk_count += 1

                        if blk_count > 0:
                            logger.debug(
                                "Multi-block: block at col %d contributed %d parts (sheet: %s)",
                                blk_pn, blk_count, sheet_name,
                            )

                logger.info(
                    "Лист %s: BOM, %d колонок комплектаций, %d строк с данными",
                    sheet_name, len(config_cols), sheet_config_count,
                )
            except Exception as e:
                logger.error(
                    "Ошибка обработки листа %s: %s — пропуск листа",
                    sheet_name, e,
                )
                continue

    finally:
        if wb is not None:
            wb.close()

    # ── Финальная агрегация ──
    if _swm_multisheet_sheets:
        logger.info(
            "SWM multi-sheet BOM: агрегированы листы %s → '%s' (%d деталей)",
            _swm_multisheet_sheets,
            _SWM_COMBINED_CONFIG,
            len(all_config_quantities.get(_SWM_COMBINED_CONFIG, {})),
        )

    # Удаляем из all_parts детали, у которых нет qty > 0 ни в одной конфигурации.
    # Это детали, где ВСЕ колонки конфигураций содержат '-', пусто или 0.
    # Они не используются ни в одной комплектации и не должны считаться.
    qty_zero_pns = [
        pn for pn in all_parts
        if not any(all_config_quantities.get(cn, {}).get(pn, 0) > 0 for cn in all_config_names)
    ]
    for pn in qty_zero_pns:
        del all_parts[pn]

    logger.info(
        "Загружено деталей с qty>0: %d (отброшено %d деталей с qty=0)",
        len(all_parts), len(qty_zero_pns),
    )
    logger.info("Найдено комплектаций: %d", len(all_config_names))
    logger.info("Глобальный словарь названий: %d записей", len(all_global_names))

    for cn in all_config_names[:10]:
        qty_count = len(all_config_quantities.get(cn, {}))
        logger.info("  %s: %d деталей", cn[:50], qty_count)
    if len(all_config_names) > 10:
        logger.info("  ... и ещё %d комплектаций", len(all_config_names) - 10)

    # Применяем глобальные названия к деталям, у которых нет названия
    for pn, part in all_parts.items():
        if (not part.name_cn and not part.name_en) and pn in all_global_names:
            gc, ge = all_global_names[pn]
            if not part.name_cn and gc:
                part.name_cn = gc
            if not part.name_en and ge:
                part.name_en = ge

    return BOMData(
        parts=all_parts,
        config_names=all_config_names,
        config_quantities=all_config_quantities,
        source_file=file_path,
        global_names=all_global_names,
    )


def get_config_quantities(bom: BOMData, config_name: str) -> Dict[str, PartInfo]:
    """Получить данные деталей для выбранной комплектации.

    Args:
        bom: Распарсенные BOM-данные.
        config_name: Название комплектации.

    Returns:
        Словарь {part_number: PartInfo} с заполненным quantity для комплектации.
    """
    if config_name not in bom.config_quantities:
        raise ValueError(
            f"Комплектация '{config_name}' не найдена. "
            f"Доступные: {bom.config_names[:10]}..."
        )

    result: Dict[str, PartInfo] = {}
    for part_no, qty in bom.config_quantities[config_name].items():
        if part_no in bom.parts:
            part = bom.parts[part_no]
            result[part_no] = PartInfo(
                part_number=part_no,
                name_cn=part.name_cn,
                name_en=part.name_en,
                quantity=qty,
            )
        else:
            # Берём из глобального словаря
            gc, ge = bom.global_names.get(part_no, ("", ""))
            result[part_no] = PartInfo(
                part_number=part_no,
                name_cn=gc,
                name_en=ge,
                quantity=qty,
            )

    return result


def get_all_config_quantities(bom: BOMData) -> Dict[str, Dict[str, PartInfo]]:
    """Получить данные деталей для ВСЕХ комплектаций одновременно.

    Args:
        bom: Распарсенные BOM-данные.

    Returns:
        Словарь {config_name: {part_number: PartInfo}}.
    """
    return {cn: get_config_quantities(bom, cn) for cn in bom.config_names}


def lookup_part_name(bom: BOMData, part_number: str) -> Tuple[str, str]:
    """Найти название детали по парт-номеру.

    Сначала ищет в parts, затем в global_names.

    Args:
        bom: BOM-данные.
        part_number: Парт-номер.

    Returns:
        (name_cn, name_en)
    """
    if part_number in bom.parts:
        part = bom.parts[part_number]
        if part.name_cn or part.name_en:
            return (part.name_cn, part.name_en)
    return bom.global_names.get(part_number, ("", ""))


class BOMService:
    """Сервис парсинга BOM-файлов.

    Готов к использованию в серверной архитектуре (FastAPI).
    Поддерживает:
      - Загрузку из файла (load)
      - Загрузку из памяти (load_from_bytes) — для HTTP upload
      - Асинхронную загрузку (load_async) — не блокирует event loop
      - Автоочистку временных файлов (cleanup / context manager)
    """

    def __init__(self):
        self._bom: Optional[BOMData] = None
        self._temp_paths: List[str] = []

    @property
    def bom(self) -> Optional[BOMData]:
        return self._bom

    @property
    def is_loaded(self) -> bool:
        return self._bom is not None

    def load(self, file_path: str) -> BOMData:
        """Загрузить и распарсить BOM-файл.

        Args:
            file_path: Путь к .xlsx файлу BOM.

        Returns:
            Распарсенные данные BOMData.
        """
        self._bom = parse_bom(file_path)
        return self._bom

    def load_from_bytes(self, data: bytes, filename: str = "bom.xlsx") -> BOMData:
        """Загрузить BOM из байтового содержимого (in-memory upload).

        Сохраняет данные во временный файл, парсит, возвращает результат.
        Временный файл будет удалён при вызове cleanup() или выходе из
        контекстного менеджера.

        Args:
            data: Байтовое содержимое .xlsx файла.
            filename: Имя файла для определения расширения.

        Returns:
            Распарсенные данные BOMData.
        """
        suffix = os.path.splitext(filename)[1] or ".xlsx"
        fd, path = tempfile.mkstemp(suffix=suffix, prefix="bom_upload_")
        os.close(fd)
        with open(path, "wb") as f:
            f.write(data)
        self._temp_paths.append(path)
        return self.load(path)

    async def load_async(self, data: bytes, filename: str = "bom.xlsx") -> BOMData:
        """Асинхронная загрузка BOM из байтов.

        Парсинг CPU-bound — выполняется в отдельном потоке,
        не блокируя event loop.

        Args:
            data: Байтовое содержимое .xlsx файла.
            filename: Имя файла для определения расширения.

        Returns:
            Распарсенные данные BOMData.
        """
        import asyncio
        return await asyncio.to_thread(self.load_from_bytes, data, filename)

    def cleanup(self) -> None:
        """Удалить все временные файлы, созданные при load_from_bytes."""
        for path in self._temp_paths:
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except Exception as e:
                logger.debug("Failed to remove temp file %s: %s", path, e)
        self._temp_paths.clear()

    def __enter__(self) -> BOMService:
        return self

    def __exit__(self, *args: object) -> None:
        self.cleanup()

    def get_config_names(self) -> List[str]:
        """Получить список названий всех найденных комплектаций."""
        if not self._bom:
            raise RuntimeError("BOM не загружен. Вызовите load() сначала.")
        return list(self._bom.config_names)

    def get_config_count(self) -> int:
        """Получить количество найденных комплектаций."""
        if not self._bom:
            return 0
        return len(self._bom.config_names)

    def get_parts_for_config(self, config_name: str) -> Dict[str, PartInfo]:
        """Получить детали для конкретной комплектации."""
        if not self._bom:
            raise RuntimeError("BOM не загружен. Вызовите load() сначала.")
        return get_config_quantities(self._bom, config_name)

    def get_all_configs(self) -> Dict[str, Dict[str, PartInfo]]:
        """Получить детали для ВСЕХ комплектаций."""
        if not self._bom:
            raise RuntimeError("BOM не загружен. Вызовите load() сначала.")
        return get_all_config_quantities(self._bom)

    def get_all_part_numbers(self) -> Set[str]:
        """Получить множество ВСЕХ уникальных парт-номеров из BOM."""
        if not self._bom:
            return set()
        return set(self._bom.parts.keys())

    def lookup_name(self, part_number: str) -> Tuple[str, str]:
        """Найти название детали по парт-номеру (с учётом глобального словаря)."""
        if not self._bom:
            return ("", "")
        return lookup_part_name(self._bom, part_number)

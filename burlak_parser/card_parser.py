"""Модуль чтения операционных карт (ОК).

Формат: Множество файлов .xlsx/.xls (распределённых по папкам или архивом).

Алгоритм обработки:
  - Автоматическая фильтрация: операционные карты vs служебные файлы.
  - Каждый файл может содержать несколько листов. Один лист = одна операция.
  - Пустые листы или листы без номера карты — игнорируются.
  - Извлекаются: [Парт-номер запчасти] и [Необходимое количество].
  - Если парт-номер переносится на следующую строку (символ «-» на конце) —
    система склеивает строки.
  - Повторяющиеся детали в одной карте или в разных картах — суммируются.
  - Поддерживаются .xlsx (openpyxl) и .xls (xlrd).
  - Многопоточный парсинг (ProcessPoolExecutor) для больших объёмов (>1500 карт).

Использует эвристический анализатор (heuristic_analyzer.py) для универсального
поиска таблиц деталей и извлечения номеров карт без привязки к брендам.

Класс CardService — обёртка для использования в FastAPI/серверной архитектуре.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import tempfile
import warnings
import zipfile
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional, Tuple

from tqdm import tqdm

from burlak_parser.heuristic_analyzer import (
    HeuristicAnalyzer,
    extract_card_number,
)
from burlak_parser.normalizer import (
    normalize_quantity,
    clean_part_number,
    is_valid_part_number,
)
from burlak_parser.xls_converter import (
    is_libreoffice_available,
    convert_xls_files_batch,
)

logger = logging.getLogger(__name__)


def _safe_extractall(zf: zipfile.ZipFile, extract_dir: str) -> None:
    """Безопасная распаковка ZIP с защитой от path traversal (zip slip).

    Проверяет, что все пути в архивах не выходят за пределы целевой директории.
    """
    for info in zf.infolist():
        target_path = os.path.normpath(os.path.join(extract_dir, info.filename))
        if not target_path.startswith(os.path.normpath(extract_dir) + os.sep) and target_path != os.path.normpath(extract_dir):
            raise ValueError(f"ZIP path traversal detected: {info.filename}")
    zf.extractall(extract_dir)


# ─── Универсальный загрузчик Excel (.xlsx + .xls) ────────────────────────────


class ExcelReader:
    """Универсальный читатель Excel-файлов.
    Поддерживает .xlsx (openpyxl) и .xls (xlrd).
    Использует openpyxl для .xlsx, xlrd для .xls.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._wb: Any = None
        self._engine: str = ""
        self._sheet_names: List[str] = []
        self._sheets: Dict[str, Any] = {}
        self._load()

    def _load(self) -> None:
        ext = os.path.splitext(self.file_path)[1].lower()

        if ext == ".xls":
            self._load_via_xlrd()
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(
                    self.file_path, data_only=True,
                )
                self._engine = "openpyxl"
                self._wb = wb
                self._sheet_names = list(wb.sheetnames)
                for sn in self._sheet_names:
                    self._sheets[sn] = wb[sn]
                logger.debug(
                    "Файл %s загружен через openpyxl",
                    os.path.basename(self.file_path),
                )
                return
            except Exception as e:
                logger.debug("openpyxl normal load failed: %s", e)
            # Fallback: read_only mode (handles WPS/slightly corrupted files)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(
                    self.file_path, data_only=True, read_only=True,
                )
                self._engine = "openpyxl"
                self._wb = wb
                self._sheet_names = list(wb.sheetnames)
                for sn in self._sheet_names:
                    self._sheets[sn] = wb[sn]
                logger.info(
                    "Файл %s загружен через openpyxl (read_only fallback)",
                    os.path.basename(self.file_path),
                )
                return
            except Exception as e:
                logger.debug("openpyxl read_only fallback failed: %s", e)
            self._load_via_xlrd()

    def _load_via_xlrd(self) -> None:
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "Для чтения .xls файлов требуется xlrd. Установите: pip install xlrd"
            )

        try:
            wb = xlrd.open_workbook(self.file_path)
            self._engine = "xlrd"
            self._wb = wb
            self._sheet_names = list(wb.sheet_names())
            for sn in self._sheet_names:
                self._sheets[sn] = wb.sheet_by_name(sn)
            logger.debug(
                "Файл %s загружен через xlrd",
                os.path.basename(self.file_path),
            )
        except Exception as e:
            raise ValueError(
                f"Не удалось открыть Excel-файл {self.file_path}: {e}"
            )

    @property
    def sheet_names(self) -> List[str]:
        return self._sheet_names

    def get_sheet(self, name: str) -> "ExcelSheet":
        if name not in self._sheets:
            raise KeyError(f"Лист '{name}' не найден")
        return ExcelSheet(self._sheets[name], self._engine)

    def close(self) -> None:
        if self._engine == "openpyxl" and self._wb is not None:
            self._wb.close()


class ExcelSheet:
    """Обёртка над листом Excel для единого API openpyxl / xlrd."""

    def __init__(self, ws: Any, engine: str):
        self._ws = ws
        self._engine = engine

    @property
    def max_row(self) -> int:
        if self._engine == "openpyxl":
            return self._ws.max_row or 0
        else:
            return self._ws.nrows

    @property
    def max_column(self) -> int:
        if self._engine == "openpyxl":
            return self._ws.max_column or 0
        else:
            return self._ws.ncols

    def cell_value(self, row: int, column: int) -> Any:
        try:
            if self._engine == "openpyxl":
                return self._ws.cell(row=row, column=column).value
            else:
                val = self._ws.cell_value(row - 1, column - 1)
                if val == "" or val is None:
                    return None
                if isinstance(val, float) and val == int(val):
                    return int(val)
                return val
        except Exception as e:
            logger.debug("cell_value error at row=%d, col=%d: %s", row, column, e)
            return None


# ─── Структуры данных ────────────────────────────────────────────────────────


@dataclass
class CardSheetInfo:
    """Информация об одном листе операционной карты."""
    card_number: str
    sheet_name: str
    operation_name: str = ""
    is_valid: bool = False
    has_data: bool = False
    max_data_row: int = 0  # Максимальное количество строк данных на листе (для защиты от ложного вертикального split)


@dataclass
class CardPart:
    """Деталь, найденная в операционной карте."""
    part_number: str
    quantity: float
    source_card: str
    source_sheet: str


@dataclass
class CardParseResult:
    """Результат парсинга одной операционной карты."""
    card_number: str
    file_path: str
    sheets: List[CardSheetInfo]
    parts: List[CardPart]
    aggregated_parts: Dict[str, float]  # part_number -> total_qty
    is_service_file: bool = False  # True если файл был определён как служебный
    tables_extracted: int = 0  # Количество таблиц (операций) найденных во всех листах


@dataclass
class CardsData:
    """Результат парсинга всех операционных карт."""
    all_parts: Dict[str, float]  # part_number -> суммарное количество
    original_part_numbers: Dict[str, str] = field(default_factory=dict)  # cleaned_part_no -> оригинальный (с тире и т.д.)
    part_sources: Dict[str, List[Tuple[str, str, float]]] = field(default_factory=dict)
    card_results: List[CardParseResult] = field(default_factory=list)
    total_cards_processed: int = 0
    total_sheets_processed: int = 0
    total_sheets_skipped: int = 0
    service_files_skipped: int = 0
    corrupted_files: List[str] = field(default_factory=list)
    corrupted_files_detailed: List[Dict[str, str]] = field(default_factory=list)
    total_tables_extracted: int = 0  # Количество таблиц (операций) во всех листах
    split_stats: Optional[SplitStatistics] = None


@dataclass
class FileSplitStats:
    """Статистика разделения для одного файла."""
    file_path: str
    file_name: str
    card_number: str
    is_xlsx: bool
    is_service_file: bool
    total_sheets: int
    sheets_split: int
    sheets_skipped: int
    split_reason: str = ""  # Почему файл был пропущен или что с ним произошло
    skip_reasons: Dict[str, List[str]] = field(default_factory=dict)  # причина -> [имена листов]
    created_files: int = 0
    has_error: bool = False
    error_message: str = ""


@dataclass
class SplitStatistics:
    """Агрегированная статистика разделения всех файлов."""
    file_stats: List[FileSplitStats] = field(default_factory=list)
    total_xlsx: int = 0
    total_xls: int = 0
    total_service_files: int = 0
    total_sheets_all: int = 0
    total_sheets_split: int = 0
    total_sheets_skipped: int = 0
    total_files_created: int = 0
    total_errors: int = 0
    openpyxl_fallback_count: int = 0
    openpyxl_fallback_files: List[str] = field(default_factory=list)

    def get_top_skip_reasons(self, n: int = 5) -> List[Tuple[str, int]]:
        """Топ-N причин пропуска листов по всем файлам."""
        from collections import Counter
        counter: Counter = Counter()
        for fs in self.file_stats:
            for reason, sheets in fs.skip_reasons.items():
                counter[reason] += len(sheets)
        return counter.most_common(n)

    def get_files_with_most_skips(self, n: int = 5) -> List[Tuple[str, int, int]]:
        """Топ-N файлов по количеству пропущенных листов.
        Returns: [(file_name, total_sheets, sheets_skipped)]
        """
        sorted_stats = sorted(
            [fs for fs in self.file_stats if fs.sheets_skipped > 0],
            key=lambda x: x.sheets_skipped,
            reverse=True,
        )
        return [(s.file_name, s.total_sheets, s.sheets_skipped) for s in sorted_stats[:n]]


# ─── Вспомогательные функции ─────────────────────────────────────────────────


def _extract_card_number(file_path: str, ws: "ExcelSheet") -> str:
    """Извлечь номер карты: сначала из содержимого листа, затем из имени файла."""
    # Используем эвристический анализатор
    card_no = extract_card_number(file_path, ws)
    if card_no:
        return card_no

    # Абсолютный fallback: базовое имя файла
    basename = os.path.basename(file_path)
    return os.path.splitext(basename)[0]


def _merge_multiline_part_numbers(
    rows: List[Tuple[int, str, float, str, int]],
) -> List[Tuple[str, float, str, int]]:
    """Склеить парт-номера, перенесённые на следующую строку."""
    merged: List[Tuple[str, float, str, int]] = []
    buffer = ""
    buffer_qty: Optional[float] = None
    buffer_name = ""
    buffer_row = 0
    last_was_continued = False

    for row_idx, raw_part_no, qty, name, _ in rows:
        if last_was_continued:
            buffer += clean_part_number(raw_part_no)
            last_was_continued = False
        elif raw_part_no.rstrip().endswith("-") or raw_part_no.rstrip().endswith("—") or raw_part_no.rstrip().endswith("–"):
            buffer = clean_part_number(raw_part_no.rstrip("-—–"))
            buffer_qty = qty
            buffer_name = name
            buffer_row = row_idx
            last_was_continued = True
            continue
        else:
            buffer = clean_part_number(raw_part_no)
            buffer_qty = qty
            buffer_name = name
            buffer_row = row_idx

        if buffer and buffer_qty is not None:
            if is_valid_part_number(buffer):
                merged.append((buffer, buffer_qty, buffer_name, buffer_row))
            buffer = ""
            buffer_qty = None
            buffer_name = ""

    if buffer and buffer_qty is not None:
        if is_valid_part_number(buffer):
            merged.append((buffer, buffer_qty, buffer_name, buffer_row))

    return merged


# ─── Парсинг одного файла ────────────────────────────────────────────────────


def parse_card_file(
    file_path: str,
    is_service_file: bool = False,
) -> CardParseResult:
    """Разобрать один файл операционной карты.

    Использует эвристический анализатор для поиска таблицы деталей
    и извлечения номеров карт без привязки к конкретным брендам.

    Args:
        file_path: Путь к .xlsx или .xls файлу.
        is_service_file: True если файл классифицирован как служебный.

    Returns:
        CardParseResult с данными всех непустых листов.
    """
    basename = os.path.basename(file_path)
    logger.debug("Обработка файла: %s", basename)

    reader = ExcelReader(file_path)
    try:
        card_parts: List[CardPart] = []
        sheets_info: List[CardSheetInfo] = []
        aggregated: Dict[str, float] = {}
        card_number = ""

        # Если файл служебный — только собираем информацию о листах, без парсинга деталей
        if is_service_file:
            for sheet_name in reader.sheet_names:
                ws = reader.get_sheet(sheet_name)
                sheet_has_data = _check_sheet_has_data(ws)
                sheets_info.append(CardSheetInfo(
                    card_number=card_number or basename,
                    sheet_name=sheet_name,
                    is_valid=False,
                    has_data=sheet_has_data,
                    max_data_row=ws.max_row,
                ))
            return CardParseResult(
                card_number=basename,
                file_path=file_path,
                sheets=sheets_info,
                parts=card_parts,
                aggregated_parts=aggregated,
                is_service_file=True,
            )

        tables_extracted = 0  # Счётчик таблиц (операций) на всех листах файла

        for sheet_name in reader.sheet_names:
            ws = reader.get_sheet(sheet_name)
            max_row = ws.max_row
            max_col = ws.max_column

            # Пропускаем пустые листы
            if max_row == 0 or max_col == 0:
                sheets_info.append(CardSheetInfo(
                    card_number=card_number or basename,
                    sheet_name=sheet_name,
                    is_valid=False,
                    has_data=False,
                    max_data_row=max_row,
                ))
                continue

            sheet_has_data = _check_sheet_has_data(ws)

            if not sheet_has_data:
                sheets_info.append(CardSheetInfo(
                    card_number=card_number or basename,
                    sheet_name=sheet_name,
                    is_valid=False,
                    has_data=False,
                    max_data_row=max_row,
                ))
                continue

            # Извлекаем номер карты из первого непустого листа (эвристически)
            if not card_number:
                card_number = _extract_card_number(file_path, ws)

            operation_name = ""

            # Ищем таблицы с деталями через эвристический анализатор
            # Поддерживает многооперационные листы (SWM карты)
            first_table_info = HeuristicAnalyzer.find_part_table(ws)
            if first_table_info is None:
                # Fallback: Changan-формат с 图示编号 (Graphic Number)
                header_rows = HeuristicAnalyzer.find_header_rows(ws)
                col_types = HeuristicAnalyzer.detect_column_types(
                    ws, header_rows, allow_qty_fallback=True,
                ) if header_rows else {}

                # Путь 1: Changan с graphic_number колонкой
                graphic_parts = _collect_parts_with_graphic_number(
                    ws, max_row, max_col, header_rows, col_types, basename,
                )

                # Путь 2: Если graphic_number не найден — прямое извлечение
                # по part_no + qty колонкам (если они определены)
                if not graphic_parts:
                    part_no_col = col_types.get("part_no", 0)
                    qty_col = col_types.get("qty", 0)
                    name_col = col_types.get("name_cn", 0) or col_types.get("name_en", 0)

                    # Fallback: ищем part_no по содержимому если не нашли по заголовку
                    data_start = header_rows[-1] + 1 if header_rows else 2
                    if part_no_col == 0:
                        part_no_col = HeuristicAnalyzer._find_part_no_by_content(
                            ws, data_start, max_row + 1, max_col,
                        )
                    if qty_col == 0:
                        known = set()
                        if part_no_col > 0:
                            known.add(part_no_col)
                        if name_col > 0:
                            known.add(name_col)
                        # Собираем тексты заголовков для исключения meta-колонок
                        header_texts_changan: Dict[int, str] = {}
                        for c in range(1, min(max_col + 1, 50)):
                            for hr in header_rows:
                                v = ws.cell_value(hr, c)
                                if v is not None:
                                    header_texts_changan[c] = str(v).strip().lower()
                                    break
                        qty_col = HeuristicAnalyzer._find_qty_by_content(
                            ws, data_start, max_row + 1, max_col,
                            header_texts=header_texts_changan,
                            known_cols=known,
                        )

                    logger.debug(
                        "Sheet '%s' Direct fallback: part_no=%d, qty=%d, name=%d",
                        sheet_name, part_no_col, qty_col, name_col,
                    )

                    if part_no_col > 0:
                        raw_rows = _collect_raw_rows(
                            ws, data_start - 1, max_row, max_col,
                            part_no_col, qty_col, name_col, basename,
                        )
                        merged_parts = _merge_multiline_part_numbers(raw_rows)
                        if merged_parts:
                            graphic_parts = [
                                (pn, qty, name, "") for pn, qty, name, _ in merged_parts
                            ]

                if graphic_parts:
                    tables_extracted += 1
                    for part_no, qty, name, graphic_number in graphic_parts:
                        card_parts.append(CardPart(
                            part_number=part_no,
                            quantity=qty,
                            source_card=graphic_number or card_number or basename,
                            source_sheet=sheet_name,
                        ))
                        aggregated[part_no] = aggregated.get(part_no, 0.0) + qty

                    sheets_info.append(CardSheetInfo(
                        card_number=card_number or basename,
                        sheet_name=sheet_name,
                        operation_name=f"Changan fallback ({len(graphic_parts)} parts)",
                        is_valid=True,
                        has_data=True,
                        max_data_row=max_row,
                    ))
                else:
                    # ── Fallback: Inspection format (检验作业指导书) ──
                    # JC-031 style: sheets with "检验项目" headers,
                    # no standard part table, but multiple inspection operations
                    inspection_ops = _detect_inspection_operations(ws, max_row)
                    if inspection_ops > 1:
                        tables_extracted += inspection_ops
                        operation_name = "Inspection" if not operation_name else operation_name
                        sheets_info.append(CardSheetInfo(
                            card_number=card_number or basename,
                            sheet_name=sheet_name,
                            operation_name=(
                                f"Inspection ({inspection_ops} ops)"
                            ),
                            is_valid=True,
                            has_data=True,
                            max_data_row=max_row,
                        ))
                    else:
                        sheets_info.append(CardSheetInfo(
                            card_number=card_number or basename,
                            sheet_name=sheet_name,
                            operation_name="Лист без таблицы деталей",
                            is_valid=False,
                            has_data=True,
                            max_data_row=max_row,
                        ))
                continue

            header_row, part_no_col, qty_col, name_col = first_table_info

            # ── Fallback: если find_part_table не нашёл qty — ищем по содержимому ──
            if qty_col == 0:
                data_start = header_row + 1
                known = set()
                if part_no_col > 0:
                    known.add(part_no_col)
                if name_col > 0:
                    known.add(name_col)
                # Собираем тексты заголовков для исключения meta-колонок
                header_texts: Dict[int, str] = {}
                for c in range(1, min(max_col + 1, 50)):
                    v = ws.cell_value(header_row, c)
                    if v is not None:
                        header_texts[c] = str(v).strip().lower()
                qty_col = HeuristicAnalyzer._find_qty_by_content(
                    ws, data_start, max_row + 1, max_col,
                    header_texts=header_texts,
                    known_cols=known,
                )
                if qty_col > 0:
                    logger.debug(
                        "Sheet '%s': qty_col=%d found by content fallback (header_row=%d)",
                        sheet_name, qty_col, header_row,
                    )

            # Извлекаем название операции
            operation_name = HeuristicAnalyzer.extract_operation_name(ws, header_row)

            # Собираем детали из ВСЕХ таблиц на листе (многооперационные карты)
            merged_parts, sheet_tables = _collect_all_tables(
                ws, max_row, max_col, basename,
            )
            tables_extracted += sheet_tables

            # Добавляем в результаты
            for part_no, qty, name, _ in merged_parts:
                card_parts.append(CardPart(
                    part_number=part_no,
                    quantity=qty,
                    source_card=card_number or basename,
                    source_sheet=sheet_name,
                ))
                aggregated[part_no] = aggregated.get(part_no, 0.0) + qty

            sheets_info.append(CardSheetInfo(
                card_number=card_number or basename,
                sheet_name=sheet_name,
                operation_name=operation_name,
                is_valid=len(merged_parts) > 0,
                has_data=True,
                max_data_row=max_row,
            ))

    finally:
        reader.close()

    return CardParseResult(
        card_number=card_number or basename,
        file_path=file_path,
        sheets=sheets_info,
        parts=card_parts,
        aggregated_parts=aggregated,
        tables_extracted=tables_extracted,
    )


def _check_sheet_has_data(ws: ExcelSheet) -> bool:
    """Проверить, есть ли данные на листе (проверка начала и сэмплирование)."""
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, min(max_row, 10) + 1):
        for c in range(1, min(max_col, 30) + 1):
            if ws.cell_value(r, c) is not None:
                return True

    if max_row > 10:
        step = max(1, (max_row - 15) // 10)
        for r in range(15, max_row + 1, step):
            for c in range(1, min(max_col, 30) + 1):
                if ws.cell_value(r, c) is not None:
                    return True

    return False


_INSPECTION_HEADER_KW = '检验项目'


def _detect_inspection_operations(ws: ExcelSheet, max_row: int) -> int:
    """Detect inspection-format operations (检验作业指导书) in a sheet.

    JC-031 style format:
      - Row with "检验项目" in column B marks each operation
      - Operation name follows in column D
      - Operations repeat every ~20 rows

    Returns the number of inspection operations found, or 0 if not
    an inspection-format sheet.
    """
    if max_row < 20:
        return 0

    # Find all rows with "检验项目" in column B (col 2)
    header_rows: List[int] = []
    for r in range(1, max_row + 1):
        val = ws.cell_value(r, 2)
        if val is not None and _INSPECTION_HEADER_KW in str(val):
            header_rows.append(r)

    if len(header_rows) < 2:
        return 0

    # Validate consistency: spacings should be roughly uniform
    spacings = [header_rows[i + 1] - header_rows[i]
                for i in range(len(header_rows) - 1)]
    if not spacings:
        return 0

    step = sorted(spacings)[len(spacings) // 2]  # median
    consistent = sum(1 for s in spacings if abs(s - step) <= 3)
    if consistent < len(spacings) * 0.5:
        return 0

    # Validate data density: at least some rows between headers have content
    dataful_blocks = 0
    for i, hr in enumerate(header_rows):
        next_hr = header_rows[i + 1] if i + 1 < len(header_rows) else max_row + 1
        has_content = False
        for r in range(hr + 1, min(hr + 6, next_hr)):
            for c in range(1, 9):
                val = ws.cell_value(r, c)
                if val is not None and str(val).strip():
                    has_content = True
                    break
            if has_content:
                break
        if has_content:
            dataful_blocks += 1

    if dataful_blocks < len(header_rows) * 0.5:
        logger.debug(
            "Inspection detection: only %d/%d blocks have data, rejecting",
            dataful_blocks, len(header_rows),
        )
        return 0

    logger.info(
        "Inspection format detected: %d operations (spacing ~%d rows)",
        len(header_rows), step,
    )
    return len(header_rows)


def _collect_raw_rows(
    ws: ExcelSheet,
    header_row: int,
    max_row: int,
    max_col: int,
    part_no_col: int,
    qty_col: int,
    name_col: int,
    basename: str,
) -> List[Tuple[int, str, float, str, int]]:
    """Собрать сырые строки таблицы деталей.

    Останавливается при обнаружении границы секции:
      - Строка с >= 2 непустыми ячейками, содержащая PART_NO_KEYWORD (новый заголовок)
      - 3+ последовательных пустых строк в колонке part_no

    Returns:
        Список кортежей (row_idx, raw_part_no, qty, name, part_no_col).
    """
    raw_rows: List[Tuple[int, str, float, str, int]] = []
    max_data_row = max_row
    consecutive_empty_pn = 0

    for row_idx in range(header_row + 1, max_data_row + 1):
        try:
            if HeuristicAnalyzer.is_cell_strike(ws, row_idx, part_no_col):
                continue
            raw_part_no = ws.cell_value(row_idx, part_no_col)

            # ── Проверка на границу секции: новый заголовок ──
            # Строка с >= 2 непустыми ячейками, содержащая PART_NO_KEYWORD
            non_empty = 0
            row_values_check: List[str] = []
            for c in range(1, min(max_col + 1, 25)):
                v = ws.cell_value(row_idx, c)
                if v is not None:
                    non_empty += 1
                    row_values_check.append(str(v).strip().lower())

            if non_empty >= 2:
                # Проверяем, есть ли ячейка с PART_NO_KEYWORD И длина < 50 символов
                # (чтобы не спутать с длинными описаниями, содержащими "деталь")
                has_part_no_keyword_short = any(
                    len(rv) < 50 and any(kw in rv for kw in HeuristicAnalyzer._get_part_no_keywords())
                    for rv in row_values_check
                )
                if has_part_no_keyword_short:
                    # Это новый заголовок таблицы — останавливаем сбор
                    logger.debug(
                        "Граница секции на строке %d (новый заголовок с PART_NO_KEYWORD)",
                        row_idx,
                    )
                    break

            if raw_part_no is None:
                # Проверка на пустую строку
                all_empty = True
                for c in range(1, min(max_col + 1, 20)):
                    if ws.cell_value(row_idx, c) is not None:
                        all_empty = False
                        break
                if all_empty:
                    consecutive_empty_pn += 1
                    if consecutive_empty_pn >= 3:
                        logger.debug(
                            "Граница секции на строке %d (3+ пустых строк)", row_idx,
                        )
                        break
                    continue
                continue

            # Сброс счётчика пустых строк
            consecutive_empty_pn = 0

            raw_part_no_str = str(raw_part_no).strip()
            if not raw_part_no_str:
                continue

            skip_keywords = [
                "物料清单", "变更记录", "编制", "校对", "审核", "批准",
                "说明性符号", "工具", "夹具", "文件编号", "文件版次", "无",
            ]
            if any(kw in raw_part_no_str.lower() for kw in skip_keywords):
                continue

            if qty_col > 0:
                if HeuristicAnalyzer.is_cell_strike(ws, row_idx, qty_col):
                    continue
                raw_qty = ws.cell_value(row_idx, qty_col)
                try:
                    qty = float(raw_qty) if raw_qty is not None else 0.0
                except (ValueError, TypeError):
                    # Извлекаем число из строки ("1个" → 1.0, "2pcs" → 2.0)
                    qty = 0.0
                    if raw_qty is not None:
                        import re
                        m = re.match(r'^\s*(\d+(?:[.,]\d+)?)', str(raw_qty).strip())
                        if m:
                            try:
                                qty = float(m.group(1).replace(",", "."))
                            except ValueError:
                                qty = 0.0
            else:
                qty = 0.0

            name = ""
            if name_col > 0:
                name_val = ws.cell_value(row_idx, name_col)
                name = str(name_val).strip() if name_val is not None else ""

            raw_rows.append((row_idx, raw_part_no_str, qty, name, part_no_col))

        except Exception as e:
            logger.debug("Ошибка при обработке строки %d в %s: %s", row_idx, basename, e)
            continue

    return raw_rows


def _collect_all_tables(
    ws: ExcelSheet,
    max_row: int,
    max_col: int,
    basename: str,
) -> Tuple[List[Tuple[str, float, str, int]], int]:
    """Собрать детали из ВСЕХ таблиц на листе (многооперационные карты).

    Последовательно находит таблицы деталей через find_part_table(),
    собирает строки из каждой, и агрегирует результаты.
    Пропускает найденные таблицы, если в них нет валидных part-number.

    Returns:
        Кортеж (parts, table_count):
          - parts: список кортежей (part_no, qty, name, source_row)
          - table_count: количество найденных таблиц (включая пустые)
    """
    all_parts: List[Tuple[str, float, str, int]] = []
    total_part_nos_collected = 0
    start_search = 1
    tables_found = 0
    max_tables = 500  # поддержка больших файлов (G01 SWM: 217 таблиц)

    for table_idx in range(max_tables):
        table_info = HeuristicAnalyzer.find_part_table(ws, start_row=start_search)
        if table_info is None:
            break

        header_row, part_no_col, qty_col, name_col = table_info

        # Если заголовок уже обработан — выходим
        if header_row < start_search:
            break

        # ── Fallback: если find_part_table не нашёл qty — ищем по содержимому ──
        if qty_col == 0:
            data_start = header_row + 1
            known = set()
            if part_no_col > 0:
                known.add(part_no_col)
            if name_col > 0:
                known.add(name_col)
            header_texts_tbl: Dict[int, str] = {}
            for c in range(1, min(max_col + 1, 50)):
                v = ws.cell_value(header_row, c)
                if v is not None:
                    header_texts_tbl[c] = str(v).strip().lower()
            qty_col = HeuristicAnalyzer._find_qty_by_content(
                ws, data_start, max_row + 1, max_col,
                header_texts=header_texts_tbl,
                known_cols=known,
            )

        raw_rows = _collect_raw_rows(
            ws, header_row, max_row, max_col,
            part_no_col, qty_col, name_col, basename,
        )

        merged_parts = _merge_multiline_part_numbers(raw_rows)

        if merged_parts:
            tables_found += 1
            total_part_nos_collected += len(merged_parts)
            all_parts.extend(merged_parts)
            logger.debug(
                "Таблица #%d (R%d): %d деталей",
                table_idx + 1, header_row, len(merged_parts),
            )

        # Продолжаем поиск со следующей строки после последней собранной
        # (или после заголовка, если данных нет)
        last_data_row = header_row
        if raw_rows:
            last_data_row = max(r[0] for r in raw_rows)
        start_search = last_data_row + 1

        # Если таблица оказалась пустой — выходим (защита от цикла)
        if start_search >= max_row:
            break

    if total_part_nos_collected == 0:
        logger.debug("Не найдено таблиц с деталями")

    return all_parts, tables_found


def _collect_parts_with_graphic_number(
    ws: ExcelSheet,
    max_row: int,
    max_col: int,
    header_rows: List[int],
    col_types: Dict[str, int],
    basename: str,
) -> List[Tuple[str, float, str, str]]:
    """Собрать детали из таблицы с колонкой 图示编号 (Graphic Number).

    Используется для Changan-формата, где детали привязаны к операционным картам
    через колонку с номером схемы/операции (например, DP-CH-A01).

    Args:
        ws: Лист Excel.
        max_row: Максимальная строка.
        max_col: Максимальная колонка.
        header_rows: Найденные строки заголовков.
        col_types: Определённые типы колонок.
        basename: Имя файла.

    Returns:
        Список кортежей (part_no, qty, name, graphic_number).
    """
    graphic_col = HeuristicAnalyzer.find_graphic_number_column(ws, header_rows, col_types)
    if graphic_col == 0:
        return []

    part_no_col = col_types.get("part_no", 0)
    name_col = col_types.get("name_cn", 0) or col_types.get("name_en", 0)

    data_start = header_rows[-1] + 1 if header_rows else 2

    # Fallback: ищем part_no по содержимому если не нашли по заголовку
    if part_no_col == 0:
        part_no_col = HeuristicAnalyzer._find_part_no_by_content(
            ws, data_start, max_row + 1, max_col,
        )
        if part_no_col == 0:
            logger.debug(
                "Graphic number extraction: part_no column not found "
                "(col_types=%s, graphic_col=%d)",
                col_types, graphic_col,
            )
            return []

    # Находим количественную колонку (qty или config columns)
    qty_col = col_types.get("qty", 0)

    # Fallback: ищем qty по содержимому если не нашли по заголовку
    if qty_col == 0:
        known = {graphic_col, part_no_col}
        if name_col > 0:
            known.add(name_col)
        qty_col = HeuristicAnalyzer._find_qty_by_content(
            ws, data_start, max_row + 1, max_col,
            known_cols=known,
        )
        if qty_col > 0:
            logger.debug(
                "Graphic number extraction: qty column found by content: %d",
                qty_col,
            )

    parts: List[Tuple[str, float, str, str]] = []

    for r in range(data_start, max_row + 1):
        if HeuristicAnalyzer.is_cell_strike(ws, r, part_no_col):
            continue
        if qty_col > 0 and HeuristicAnalyzer.is_cell_strike(ws, r, qty_col):
            continue
        pn = ws.cell_value(r, part_no_col)
        if pn is None:
            continue
        pn_str = str(pn).strip()
        if not pn_str or not is_valid_part_number(pn_str):
            continue

        pn_clean = clean_part_number(pn_str)

        # Количество
        qty = 0.0
        if qty_col > 0:
            qty_val = ws.cell_value(r, qty_col)
            qty = normalize_quantity(qty_val, default=0.0)
            if qty <= 0:
                qty = 0.0

        # Название
        name = ""
        if name_col > 0:
            name_val = ws.cell_value(r, name_col)
            name = str(name_val).strip() if name_val is not None else ""

        # Номер схемы/операции
        graphic_val = ws.cell_value(r, graphic_col)
        graphic_number = str(graphic_val).strip() if graphic_val is not None else ""

        if graphic_number:
            parts.append((pn_clean, qty, name, graphic_number))

    logger.debug(
        "Graphic number extraction: %d parts with graphic_col=%d",
        len(parts), graphic_col,
    )
    return parts


# ─── Поиск файлов ────────────────────────────────────────────────────────────


def _is_os_temp_file(filename: str) -> bool:
    """Проверить, является ли файл служебным файлом ОС.

    Фильтрует:
      - ~$filename.xlsx — файлы блокировки Excel
      - ._filename.xlsx — метаданные macOS (AppleDouble)
      - ~filename.xlsx — временные файлы
    """
    basename = os.path.basename(filename)
    if basename.startswith("~$"):
        return True
    if basename.startswith("._"):
        return True
    if basename.startswith("~") and basename.endswith((".xlsx", ".xls")):
        return True
    return False


def _find_excel_files(path: str, extract_dir: Optional[str] = None, _seen_names: Optional[set] = None) -> List[str]:
    """Найти все .xlsx и .xls файлы рекурсивно (папка или ZIP).

    Поддерживает вложенные ZIP-архивы (рекурсивно) с извлечением
    в отдельные поддиректории.
    Проверяет дубликаты по ИМЕНИ файла (базовое имя без пути).
    Фильтрует временные файлы (~$, ._*) и не-Excel форматы.
    Удаляет мусор только из временных директорий извлечения.
    """
    files: List[str] = []
    if _seen_names is None:
        _seen_names = set()

    if os.path.isfile(path) and path.lower().endswith(".zip"):
        if extract_dir is None:
            extract_dir = tempfile.mkdtemp(prefix="burlak_cards_")
        else:
            os.makedirs(extract_dir, exist_ok=True)

        logger.info("Распаковка архива %s в %s...", path, extract_dir)
        # Try multiple ZIP metadata encodings for compatibility
        _zip_opened = False
        for enc in ("gbk", "utf-8", "cp1251", "latin-1"):
            try:
                with zipfile.ZipFile(path, "r", metadata_encoding=enc) as z:
                    _safe_extractall(z, extract_dir)
                _zip_opened = True
                break
            except (UnicodeDecodeError, zipfile.BadZipFile) as e:
                logger.debug("ZIP open with %s failed: %s", enc, e)
                continue
        if not _zip_opened:
            logger.warning("Failed to open ZIP with any encoding, trying default")
            with zipfile.ZipFile(path, "r") as z:
                _safe_extractall(z, extract_dir)

        _walk_extracted_dir(extract_dir, extract_dir, files, _seen_names, is_temp=True)

    elif os.path.isdir(path):
        _walk_extracted_dir(path, extract_dir or path, files, _seen_names, is_temp=False)

    elif os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls")):
        if not _is_os_temp_file(path):
            files.append(path)

    return files


def _walk_extracted_dir(walk_root: str, extract_base: str, files: List[str], _seen_names: set, is_temp: bool, is_nested: bool = False) -> None:
    """Обойти директорию, фильтруя только .xlsx/.xls/.zip.

    Основные файлы (is_nested=False) собираются ВСЕ без дедупликации.
    Вложенные файлы (is_nested=True) проверяются на дубликат по имени
    относительно уже собранных основных.
    """
    nested_zips: List[str] = []

    # Детерминированный обход: сортируем корни и имена файлов
    for root, _, filenames in sorted(os.walk(walk_root), key=lambda x: x[0]):
        for fn in sorted(filenames):
            # Фильтруем файлы блокировки Excel (~$), метаданные macOS (._*), и мусор
            if fn.startswith("~$") or fn.startswith("._"):
                if is_temp:
                    _safe_remove(os.path.join(root, fn))
                continue
            full_path = os.path.join(root, fn)
            ext = os.path.splitext(fn)[1].lower()

            if ext in (".xlsx", ".xls"):
                if is_nested and fn in _seen_names:
                    logger.debug("Пропуск дубликата из вложенного архива: %s", fn)
                    if is_temp:
                        _safe_remove(full_path)
                else:
                    _seen_names.add(fn)
                    files.append(full_path)
            elif ext == ".zip":
                nested_zips.append(full_path)
            else:
                if is_temp:
                    _safe_remove(full_path)

    # Обрабатываем вложенные ZIP ПОСЛЕ основных файлов (отсортировано)
    for full_path in sorted(nested_zips):
        fn = os.path.basename(full_path)
        nested_dir = os.path.join(extract_base, f"_nested_{_safe_name(fn)}")
        os.makedirs(nested_dir, exist_ok=True)
        try:
            # Try multiple ZIP metadata encodings
            _nested_opened = False
            for enc in ("gbk", "utf-8", "cp1251", "latin-1"):
                try:
                    with zipfile.ZipFile(full_path, "r", metadata_encoding=enc) as z:
                        _safe_extractall(z, nested_dir)
                    _nested_opened = True
                    break
                except (UnicodeDecodeError, zipfile.BadZipFile):
                    continue
            if not _nested_opened:
                with zipfile.ZipFile(full_path, "r") as z:
                    _safe_extractall(z, nested_dir)
            _walk_extracted_dir(nested_dir, extract_base, files, _seen_names, is_temp=True, is_nested=True)
        except Exception as e:
            logger.warning("Не удалось распаковать вложенный архив %s: %s", fn, e)
        if is_temp:
            _safe_remove(full_path)


def _safe_remove(file_path: str) -> None:
    """Безопасно удалить файл."""
    try:
        os.remove(file_path)
    except Exception as e:
        logger.debug("Failed to remove %s: %s", file_path, e)


def _safe_name(filename: str) -> str:
    """Безопасное имя для поддиректории."""
    return re.sub(r"[^\w\-]", "_", os.path.splitext(filename)[0], flags=re.ASCII)[:50]


# ─── Парсинг всех карт ───────────────────────────────────────────────────────


def parse_cards(
    input_path: str,
    extract_dir: Optional[str] = None,
    show_progress: bool = True,
    max_workers: Optional[int] = None,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
) -> CardsData:
    """Разобрать все операционные карты из указанного источника.

    Выполняет классификацию файлов (операционные vs служебные) и
    параллельный парсинг с использованием ProcessPoolExecutor.

    Args:
        input_path: Путь к папке, ZIP-архиву или одному .xlsx/.xls файлу.
        extract_dir: Директория для извлечения ZIP.
        show_progress: Показывать прогресс-бар.
        max_workers: Количество процессов для параллельного парсинга.
        on_progress: Необязательный callback для отслеживания прогресса.
                     Вызывается после каждого файла: on_progress(processed, total, filename).

    Returns:
        CardsData с агрегированными данными всех карт.
    """
    from burlak_parser.file_classifier import filter_operational_cards

    all_files = _find_excel_files(input_path, extract_dir)
    all_files.sort()  # Детерминированный порядок
    logger.info("Найдено .xlsx/.xls файлов: %d", len(all_files))

    if not all_files:
        raise FileNotFoundError(f"Не найдено .xlsx/.xls файлов в '{input_path}'")

    # ── Конвертация .xls → .xlsx через LibreOffice ──
    xls_files = [f for f in all_files if os.path.splitext(f)[1].lower() == ".xls"]
    if xls_files:
        if is_libreoffice_available():
            convert_dir = os.path.join(
                extract_dir or tempfile.mkdtemp(prefix="burlak_xls_"),
                "_converted_xlsx",
            )
            converted_map = convert_xls_files_batch(xls_files, convert_dir)
            if converted_map:
                logger.info(
                    "Конвертировано %d/%d .xls файлов в .xlsx",
                    len(converted_map), len(xls_files),
                )
                # Заменяем .xls пути на сконвертированные .xlsx
                new_files = []
                for f in all_files:
                    if f in converted_map:
                        new_files.append(converted_map[f])
                    else:
                        new_files.append(f)
                all_files = sorted(new_files)
            else:
                logger.warning(
                    "Не удалось сконвертировать ни одного .xls файла. "
                    "Они будут обработаны через xlrd (без изображений)."
                )
        else:
            logger.warning(
                "LibreOffice не установлен. %d .xls файлов будут "
                "обработаны через xlrd (без изображений). "
                "Для конвертации установите LibreOffice.",
                len(xls_files),
            )

    # Классифицируем все файлы
    classifications = filter_operational_cards(all_files)

    operational_files = sorted(
        [c for c in classifications if c.should_parse_parts],
        key=lambda c: c.file_path,
    )
    service_files = sorted(
        [c for c in classifications if c.is_service_file],
        key=lambda c: c.file_path,
    )
    skipped_files = sorted(
        [c for c in classifications if not c.should_parse_parts and not c.is_service_file],
        key=lambda c: c.file_path,
    )
    total_classified = len(operational_files) + len(service_files) + len(skipped_files)
    logger.info(
        "Найдено файлов: %d → операционных: %d, служебных: %d, пропущено: %d",
        total_classified, len(operational_files), len(service_files), len(skipped_files),
    )
    if total_classified != len(all_files):
        logger.warning(
            "Несовпадение подсчёта: найдено %d, классифицировано %d (возможны дубликаты)",
            len(all_files), total_classified,
        )

    # Парсим операционные карты
    card_results: List[CardParseResult] = []
    all_aggregated: Dict[str, float] = {}
    part_sources: Dict[str, List[Tuple[str, str, float]]] = {}
    total_sheets = 0
    total_skipped = 0
    corrupted: List[str] = []
    corrupted_detailed: List[Dict[str, str]] = []

    parse_files = [(c.file_path, False) for c in operational_files]

    workers = max_workers or min(os.cpu_count() or 4, max(1, len(parse_files)))

    if workers > 1 and len(parse_files) > 1:
        # Параллельный парсинг
        with ProcessPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(
                    parse_card_file, fp, is_svc,
                ): fp for fp, is_svc in parse_files
            }
            iterator = tqdm(
                as_completed(futures),
                total=len(futures),
                desc="Парсинг карт",
                unit="файл",
            ) if show_progress else as_completed(futures)

            _completed_count = 0
            for future in iterator:
                file_path = futures[future]
                try:
                    result = future.result()
                    card_results.append(result)
                    total_sheets += len(result.sheets)
                    total_skipped += sum(1 for s in result.sheets if not s.is_valid)

                    # Агрегируем
                    for part_no, qty in result.aggregated_parts.items():
                        all_aggregated[part_no] = all_aggregated.get(part_no, 0.0) + qty
                        if part_no not in part_sources:
                            part_sources[part_no] = []
                        part_sources[part_no].append(
                            (result.card_number, result.file_path, qty),
                        )

                except Exception as e:
                    logger.warning("Ошибка при обработке %s: %s", file_path, e)
                    corrupted.append(file_path)
                    corrupted_detailed.append({
                        "file": os.path.basename(file_path),
                        "error": str(e),
                        "phase": "parse",
                    })
                    if show_progress:
                        tqdm.write(f"⚠️  Ошибка: {e}")

                _completed_count += 1
                if on_progress is not None:
                    try:
                        on_progress(_completed_count, len(parse_files), os.path.basename(file_path))
                    except Exception:
                        pass
    else:
        # Последовательный парсинг
        iterator = tqdm(parse_files, desc="Парсинг карт", unit="файл") if show_progress else parse_files
        _completed_count = 0
        for file_path, is_svc in iterator:
            try:
                result = parse_card_file(file_path, is_service_file=is_svc)
                card_results.append(result)
                total_sheets += len(result.sheets)
                total_skipped += sum(1 for s in result.sheets if not s.is_valid)

                for part_no, qty in result.aggregated_parts.items():
                    all_aggregated[part_no] = all_aggregated.get(part_no, 0.0) + qty
                    if part_no not in part_sources:
                        part_sources[part_no] = []
                    part_sources[part_no].append(
                        (result.card_number, result.file_path, qty),
                    )
            except Exception as e:
                logger.warning("Ошибка при обработке %s: %s", file_path, e)
                corrupted.append(file_path)
                corrupted_detailed.append({
                    "file": os.path.basename(file_path),
                    "error": str(e),
                    "phase": "parse",
                })
                if show_progress:
                    tqdm.write(f"⚠️  Ошибка: {e}")

            _completed_count += 1
            if on_progress is not None:
                try:
                    on_progress(_completed_count, len(parse_files), os.path.basename(file_path))
                except Exception:
                    pass

    # Добавляем служебные файлы в card_results (для split_cards)
    for svc in service_files:
        try:
            result = parse_card_file(
                svc.file_path,
                is_service_file=True,
            )
            card_results.append(result)
            total_sheets += len(result.sheets)
            total_skipped += sum(1 for s in result.sheets if not s.is_valid)
        except Exception as e:
            logger.warning("Ошибка при обработке служебного файла %s: %s", svc.file_path, e)

    processed = len(card_results)
    logger.info("Обработано карт: %d", processed)
    logger.info("Всего листов: %d, пропущено (пустых): %d", total_sheets, total_skipped)
    logger.info("Уникальных деталей найдено: %d", len(all_aggregated))

    # Суммируем количество таблиц (операций) во всех результатах
    total_tables = sum(r.tables_extracted for r in card_results)

    # Сортируем card_results для детерминированного порядка
    card_results.sort(key=lambda r: r.file_path)

    # Строим словарь оригинальных номеров деталей из карт
    original_part_numbers: Dict[str, str] = {}
    for result in card_results:
        for cp in result.parts:
            clean_pn = clean_part_number(cp.part_number)
            if clean_pn not in original_part_numbers:
                original_part_numbers[clean_pn] = cp.part_number

    return CardsData(
        all_parts=all_aggregated,
        original_part_numbers=original_part_numbers,
        part_sources=part_sources,
        card_results=card_results,
        total_cards_processed=processed,
        total_sheets_processed=total_sheets - total_skipped,
        total_sheets_skipped=total_skipped,
        service_files_skipped=len(service_files),
        corrupted_files=corrupted,
        corrupted_files_detailed=corrupted_detailed,
        total_tables_extracted=total_tables,
    )


# ─── Сервис ──────────────────────────────────────────────────────────────────


class CardService:
    """Сервис парсинга операционных карт.

    Готов к использованию в серверной архитектуре (FastAPI).
    Поддерживает:
      - Загрузку из файла/папки/ZIP (load)
      - Загрузку из памяти (load_from_bytes) — для HTTP upload
      - Асинхронную загрузку (load_async) — не блокирует event loop
      - Автоочистку временных файлов (cleanup / context manager)
    """

    def __init__(self, max_workers: Optional[int] = None):
        self._cards: Optional[CardsData] = None
        self._temp_paths: List[str] = []
        self._temp_dirs: List[str] = []
        self.max_workers = max_workers or os.cpu_count() or 4

    @property
    def cards(self) -> Optional[CardsData]:
        return self._cards

    @property
    def is_loaded(self) -> bool:
        return self._cards is not None

    def load(
        self,
        input_path: str,
        extract_dir: Optional[str] = None,
        on_progress: Optional[Callable[[int, int, str], None]] = None,
    ) -> CardsData:
        """Загрузить и распарсить операционные карты.

        Args:
            input_path: Путь к папке/ZIP-архиву с картами.
            extract_dir: Директория для извлечения ZIP.
            on_progress: Необязательный callback для отслеживания прогресса.
                         Вызывается после каждого файла: on_progress(processed, total, filename).

        Returns:
            Данные CardsData.
        """
        self._cards = parse_cards(
            input_path,
            extract_dir=extract_dir,
            max_workers=self.max_workers,
            show_progress=False,
            on_progress=on_progress,
        )
        return self._cards

    def load_from_bytes(self, data: bytes, filename: str) -> CardsData:
        """Загрузить карты из байтового содержимого (in-memory upload).

        Принимает ZIP-архив, .xlsx или .xls файл как байты.
        Сохраняет во временный файл, извлекает/парсит, возвращает результат.
        Временные файлы будут удалены при вызове cleanup() или выходе из
        контекстного менеджера.

        Args:
            data: Байтовое содержимое (ZIP, .xlsx или .xls).
            filename: Имя файла для определения расширения и формата.

        Returns:
            Данные CardsData.
        """
        suffix = os.path.splitext(filename)[1] or ".zip"
        fd, path = tempfile.mkstemp(suffix=suffix, prefix="cards_upload_")
        os.close(fd)
        with open(path, "wb") as f:
            f.write(data)
        self._temp_paths.append(path)

        # Для ZIP — создаём отдельную директорию извлечения
        extract_dir: Optional[str] = None
        if suffix.lower() == ".zip":
            extract_dir = tempfile.mkdtemp(prefix="cards_extract_")
            self._temp_dirs.append(extract_dir)

        return self.load(path, extract_dir=extract_dir)

    async def load_async(self, data: bytes, filename: str) -> CardsData:
        """Асинхронная загрузка карт из байтов.

        Парсинг CPU-bound — выполняется в отдельном потоке,
        не блокируя event loop.

        Args:
            data: Байтовое содержимое (ZIP, .xlsx или .xls).
            filename: Имя файла для определения расширения.

        Returns:
            Данные CardsData.
        """
        import asyncio
        return await asyncio.to_thread(self.load_from_bytes, data, filename)

    def cleanup(self) -> None:
        """Удалить все временные файлы и директории."""
        for path in self._temp_paths:
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except Exception as e:
                logger.debug("Failed to remove temp file %s: %s", path, e)
        for d in self._temp_dirs:
            try:
                if os.path.isdir(d):
                    import shutil
                    shutil.rmtree(d, ignore_errors=True)
            except Exception as e:
                logger.debug("Failed to remove temp dir %s: %s", d, e)
        self._temp_paths.clear()
        self._temp_dirs.clear()

    def __enter__(self) -> CardService:
        return self

    def __exit__(self, *args: object) -> None:
        self.cleanup()

    def get_all_parts(self) -> Dict[str, float]:
        """Получить все агрегированные детали из карт."""
        if not self._cards:
            raise RuntimeError("Карты не загружены. Вызовите load() сначала.")
        return dict(self._cards.all_parts)

    def get_part_sources(self) -> Dict[str, List[Tuple[str, str, float]]]:
        """Получить источники для каждой детали."""
        if not self._cards:
            raise RuntimeError("Карты не загружены.")
        return dict(self._cards.part_sources)

    def get_card_results(self) -> List[CardParseResult]:
        """Получить результаты парсинга каждой карты."""
        if not self._cards:
            raise RuntimeError("Карты не загружены.")
        return list(self._cards.card_results)


# ─── Разделение листов (делегировано в splitter.py) ─────────────────────────


TEMPLATE_SHEET_KEYWORDS = ["空表", "填写范本", "范本"]

# Ключевые слова для определения "инспекционного" формата файлов
# (检验作业指导书 — инструкция по проверке качества)
_INSPECTION_FILE_KEYWORDS = ["检验作业指导书", "检验指导书", "检验项目"]
# Ключевые слова для определения "主要内容" листов в инспекционных файлах
_INSPECTION_DATA_SHEET_KEYWORDS = ["内容", "内容页", "数据"]
# Ключевые слова для "мусорных" листов в инспекционных файлах
_INSPECTION_SERVICE_SHEET_KEYWORDS = ["封面", "目录", "Macro", "Sheet2", "Sheet3", "更改"]


def _is_inspection_format_file(file_path: str, sheets_info: List[CardSheetInfo]) -> bool:
    """Определить, является ли файл инспекционным (检验作业指导书).

    Инспекционные файлы содержат:
      - Заголовок "检验项目" в ячейках листов
      - Много операций проверки на одном листе
      - Мало данных в каждой операции (1-3 строки на операцию)

    Returns:
        True если файл инспекционного формата.
    """
    basename = os.path.basename(file_path).lower()
    # Проверяем имя файла на ключевые слова
    for kw in _INSPECTION_FILE_KEYWORDS:
        if kw in basename:
            return True
    # Фоллбэк: проверяем содержимое листов — если хотя бы один лист
    # содержит >10 inspection операций, файл инспекционный
    for s in sheets_info:
        if s.operation_name and s.operation_name.startswith("Inspection ("):
            try:
                ops_str = s.operation_name.split("(")[1].split(")")[0].split()[0]
                if int(ops_str) >= 10:
                    return True
            except (IndexError, ValueError):
                pass
    return False


def _select_best_data_sheet(
    sheets_info: List[CardSheetInfo],
) -> Optional[CardSheetInfo]:
    """Выбрать лучший лист с данными из списка листов.

    Приоритет:
      1. Лист с именем из _INSPECTION_DATA_SHEET_KEYWORDS (для инспекционных файлов)
      2. Лист с наибольшим количеством строк данных
      3. Первый валидный лист
    """
    if not sheets_info:
        return None

    # Приоритет 1: Имя листа совпадает с ключевыми словами данных
    for s in sheets_info:
        if not s.has_data:
            continue
        for kw in _INSPECTION_DATA_SHEET_KEYWORDS:
            if kw in s.sheet_name:
                return s

    # Приоритет 2: Наибольшее количество строк данных
    best = None
    for s in sheets_info:
        if not s.has_data:
            continue
        # Пропускаем листы-мусор
        is_service = False
        for kw in _INSPECTION_SERVICE_SHEET_KEYWORDS:
            if kw in s.sheet_name:
                is_service = True
                break
        if is_service and len(sheets_info) > 1:
            continue
        if best is None or s.max_data_row > best.max_data_row:
            best = s

    return best


def _find_main_data_sheet(
    sheets: List[CardSheetInfo],
    service_keywords: List[str],
) -> Optional[CardSheetInfo]:
    """Return the data sheet with the most rows, excluding service sheets
    when multiple sheets exist.

    Used to find the primary operational sheet in files that may have
    service sheets (封面, 目录) alongside real data sheets.

    УЛУЧШЕНИЕ: фильтрует листы с очень малым количеством данных (< 5 строк)
    и листы с типичными именами мусора (Macro, Sheet2, Sheet3).
    """
    best = None
    for s in sheets:
        if not s.has_data:
            continue
        # Пропускаем листы с очень малым количеством данных
        if s.max_data_row < 5 and len(sheets) > 1:
            continue
        is_svc = any(kw in s.sheet_name for kw in service_keywords)
        if is_svc and len(sheets) > 1:
            continue
        # Пропускаем типичные мусорные листы
        sheet_lower = s.sheet_name.lower()
        if any(kw in sheet_lower for kw in ["macro", "sheet2", "sheet3", "module"]):
            if len(sheets) > 1:
                continue
        if best is None or s.max_data_row > best.max_data_row:
            best = s
    return best


# Причины пропуска листов
SKIP_REASON_TEMPLATE = "Имя листа содержит ключевое слово шаблона"
SKIP_REASON_NO_DATA = "Пустой лист (нет данных)"
SKIP_REASON_NOT_VALID = "Нет валидных деталей"

# Причины пропуска файлов
FILE_SKIP_NOT_XLSX = "Не .xlsx формат (пропущен)"
FILE_SKIP_SERVICE = "Служебный файл (пропущен)"

# Ключевые слова служебных листов — пропускаем при разделении
# (空表 обрабатывается отдельно через TEMPLATE_SHEET_KEYWORDS с проверкой has_data)
_SPLITTER_SERVICE_SHEET_KEYWORDS = ["封面", "目录"]


def split_cards_to_files(
    cards_data: CardsData,
    output_dir: str,
    split_all_non_empty: bool = True,
    max_workers: Optional[int] = None,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
) -> List[str]:
    """Разделить многолистовые файлы на отдельные .xlsx файлы.

    Использует CardSplitter из burlak_parser.splitter для ZIP-разделения
    с очисткой named ranges.

    Args:
        cards_data: Данные распарсенных карт.
        output_dir: Директория для сохранения разделённых файлов.
        split_all_non_empty: Если True, разделяет все непустые листы;
                             если False — только листы с таблицей деталей.
        max_workers: Количество процессов для параллельного разделения.
        on_progress: Необязательный callback для отслеживания прогресса.
                     Вызывается после каждого файла: on_progress(processed, total, filename).

    Returns:
        Список путей к созданным файлам.
    """
    from burlak_parser.splitter import (
        CardSplitter,
        _safe_filename,
        preallocate_split_paths,
        _extract_to_path_worker,
        find_table_boundaries,
        _vertical_split_worker,
        TableBoundary,
    )

    os.makedirs(output_dir, exist_ok=True)
    splitter = CardSplitter(max_workers=max_workers)

    # ── Собираем per-file статистику ──
    tasks: List[Tuple[str, str, List[str], str]] = []
    file_stats: List[FileSplitStats] = []
    total_xlsx = 0
    total_xls = 0
    total_service = 0
    total_sheets_all = 0
    total_sheets_split = 0
    total_sheets_skipped = 0

    for result in cards_data.card_results:
        file_name = os.path.basename(result.file_path)
        ext = os.path.splitext(file_name)[1].lower()
        is_xlsx = ext == ".xlsx"
        is_xls = ext == ".xls"

        total_sheets_all += len(result.sheets)

        # ── Защита: пропуск служебных файлов ДО любых операций ──
        if result.is_service_file:
            total_service += 1
            if is_xlsx:
                total_xlsx += 1
            else:
                total_xls += 1
            file_stats.append(FileSplitStats(
                file_path=result.file_path,
                file_name=file_name,
                card_number=result.card_number,
                is_xlsx=is_xlsx,
                is_service_file=True,
                total_sheets=len(result.sheets),
                sheets_split=0,
                sheets_skipped=len(result.sheets),
                split_reason=FILE_SKIP_SERVICE,
            ))
            continue

        # ── Пропуск не-.xlsx файлов ──
        if not is_xlsx:
            total_xls += 1
            file_stats.append(FileSplitStats(
                file_path=result.file_path,
                file_name=file_name,
                card_number=result.card_number,
                is_xlsx=False,
                is_service_file=False,
                total_sheets=len(result.sheets),
                sheets_split=0,
                sheets_skipped=len(result.sheets),
                split_reason=FILE_SKIP_NOT_XLSX,
            ))
            continue

        total_xlsx += 1

        # Анализируем листы — какие будут разделены, какие пропущены
        sheets_to_split = []
        skip_reasons: Dict[str, List[str]] = {}
        sheets_skipped = 0

        for sheet_info in result.sheets:
            # ── Пропуск служебных листов (封面, 目录, 空表) ──
            is_service_sheet = any(
                kw in sheet_info.sheet_name for kw in _SPLITTER_SERVICE_SHEET_KEYWORDS
            )
            if is_service_sheet:
                skip_reasons.setdefault("Служебный лист (пропущен)", []).append(
                    sheet_info.sheet_name,
                )
                sheets_skipped += 1
                continue

            is_template_name = any(kw in sheet_info.sheet_name for kw in TEMPLATE_SHEET_KEYWORDS)
            if is_template_name and not sheet_info.has_data:
                skip_reasons.setdefault(SKIP_REASON_TEMPLATE, []).append(sheet_info.sheet_name)
                sheets_skipped += 1
                continue
            if split_all_non_empty:
                if sheet_info.has_data:
                    sheets_to_split.append(sheet_info.sheet_name)
                else:
                    skip_reasons.setdefault(SKIP_REASON_NO_DATA, []).append(sheet_info.sheet_name)
                    sheets_skipped += 1
            else:
                if sheet_info.is_valid:
                    sheets_to_split.append(sheet_info.sheet_name)
                else:
                    skip_reasons.setdefault(SKIP_REASON_NOT_VALID, []).append(sheet_info.sheet_name)
                    sheets_skipped += 1

        total_sheets_skipped += sheets_skipped

        if sheets_to_split:
            tasks.append((
                result.file_path,
                output_dir,
                sheets_to_split,
                result.card_number,
            ))
            total_sheets_split += len(sheets_to_split)

        file_stats.append(FileSplitStats(
            file_path=result.file_path,
            file_name=file_name,
            card_number=result.card_number,
            is_xlsx=True,
            is_service_file=result.is_service_file,
            total_sheets=len(result.sheets),
            sheets_split=len(sheets_to_split),
            sheets_skipped=sheets_skipped,
            skip_reasons=skip_reasons,
            split_reason="",
        ))

    if not tasks:
        cards_data.split_stats = SplitStatistics(
            file_stats=file_stats,
            total_xlsx=total_xlsx,
            total_xls=total_xls,
            total_service_files=total_service,
            total_sheets_all=total_sheets_all,
            total_sheets_split=total_sheets_split,
            total_sheets_skipped=total_sheets_skipped,
            total_files_created=0,
        )
        return []

    # ── ШАГ 0: Обнаружение файлов с вертикальными таблицами (SWM-стиль) ──
    # Файлы с 1 листом и >1 таблицами (операциями) на этом листе
    # требуют вертикального разделения: каждая операция → отдельный .xlsx.
    #
    # УЛУЧШЕННАЯ логика: строгие пороги для предотвращения false positives.
    # Каждое условие требует ДОПОЛНИТЕЛЬНЫХ подтверждений:
    #   - Наличие маркерных паттернов (鑫源汽车, 检验项目)
    #   - Стабильные интервалы между операциями
    #   - Достаточное количество данных в каждой операции
    vertical_split_files: Dict[str, int] = {}  # file_path -> tables_extracted
    for result in cards_data.card_results:
        is_xlsx = os.path.splitext(result.file_path)[1].lower() == ".xlsx"

        if not is_xlsx:
            continue
        if result.is_service_file:
            continue

        # Find the MAIN data sheet (the one with the most rows)
        main_sheet = _select_best_data_sheet(result.sheets)
        if main_sheet is None:
            main_sheet = _find_main_data_sheet(
                result.sheets, _SPLITTER_SERVICE_SHEET_KEYWORDS,
            )
        if main_sheet is None:
            continue

        max_data_rows = main_sheet.max_data_row
        is_inspection = _is_inspection_format_file(result.file_path, result.sheets)

        # Условие 1: Много таблиц на одном листе (базовый детектор)
        # Требуем минимум 2 таблицы с данными, и чтобы файл был ОДНОЛИСТОВЫМ
        multi_table = (
            result.tables_extracted >= 2
            and len(result.sheets) == 1
        )

        # Условие 2: Мегалист — ОЧЕНЬ строгие пороги
        # Только для SWM-формата с 鑫源汽车 маркерами
        mega_sheet = (
            len(result.sheets) == 1
            and max_data_rows > 800
            and len(result.parts) > 200
            and result.tables_extracted > 1
        )

        # Условие 3: Инспекционный формат — пороги повышены
        # Требуем минимум 8 таблиц (было 5) и 200+ строк (было 100)
        inspection_mega = (
            is_inspection
            and result.tables_extracted >= 8
            and max_data_rows > 200
            and len(result.sheets) == 1
        )

        # Условие 4: Инспекционный формат с 15+ операциями на листе
        # (检验作业指导书 с проверкой качества: 15 станций × ~20 строк = 300 строк)
        # Допускаем файлы с любым количеством листов — вертикальный split
        # работает только с основным листом данных (内容).
        inspection_large = (
            is_inspection
            and result.tables_extracted >= 12
            and max_data_rows > 150
        )

        if multi_table or mega_sheet or inspection_mega or inspection_large:
            vertical_split_files[result.file_path] = max(result.tables_extracted, 2)
            if multi_table:
                reason = "много таблиц"
            elif mega_sheet:
                reason = "мегалист (SWM)"
            elif inspection_large:
                reason = "инспекционный формат (большой)"
            else:
                reason = "инспекционный формат"
            logger.info(
                "Обнаружен многооперационный файл: %s (%s, %d таблиц, %d строк, %d деталей) "
                "— будет разделён вертикально",
                os.path.basename(result.file_path),
                reason, result.tables_extracted, max_data_rows, len(result.parts),
            )

    # ── ШАГ 1: Детерминированная предварительная разметка путей (ГЛАВНЫЙ ПОТОК) ──
    # Сортируем задачи по имени исходного файла для детерминированного порядка
    tasks.sort(key=lambda t: t[0])

    # Исключаем из path_map файлы, которые будут разделены вертикально
    # (у них другой механизм именования — по операциям)
    normal_tasks = [
        t for t in tasks
        if t[0] not in vertical_split_files
    ]

    # Вычисляем все целевые пути ДО запуска рабочих процессов
    # Это гарантирует 100% детерминизм: список отсортирован, коллизии
    # разрешаются последовательно (_1, _2, ...) в одном потоке.
    path_map = preallocate_split_paths(normal_tasks, output_dir)

    # ── ШАГ 2: Собираем индивидуальные задачи (один лист → один путь) ──
    sheet_tasks: List[Tuple[str, str, str]] = []  # (source_path, output_path, sheet_name)
    for source_path, _out_dir, sheet_names, file_label in normal_tasks:
        for sheet_name in sheet_names:
            output_path = path_map.get((source_path, sheet_name))
            if output_path:
                sheet_tasks.append((source_path, output_path, sheet_name))

    # Сортируем индивидуальные задачи для детерминированного порядка
    sheet_tasks.sort(key=lambda t: (t[0], t[2]))

    workers = max_workers or os.cpu_count() or 4
    all_created: List[str] = []
    corrupted: List[str] = []
    openpyxl_count = 0
    openpyxl_files: List[str] = []
    manifest: Dict[str, List[str]] = {}
    _total_sheet_tasks = len(sheet_tasks)

    # ── ШАГ 3: Параллельное или последовательное выполнение ──
    # Рабочие процессы НЕ проверяют существование файла —
    # уникальность пути уже гарантирована preallocate_split_paths.

    if workers > 1 and len(sheet_tasks) > 1:
        with ProcessPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(
                    _extract_to_path_worker, src, out, sheet,
                ): (src, out, sheet)
                for src, out, sheet in sheet_tasks
            }

            _split_completed = 0
            for future in as_completed(futures):
                src, out, sheet = futures[future]
                try:
                    worker_result = future.result()
                    result_path = worker_result.get("path")
                    err_msg = worker_result.get("error")

                    if result_path:
                        all_created.append(result_path)
                        if worker_result.get("used_fallback"):
                            openpyxl_count += 1
                            source_basename = worker_result.get("source_basename", "")
                            if source_basename:
                                openpyxl_files.append(source_basename)
                        # Обновляем манифест
                        original_name = worker_result.get("source_basename", "")
                        if original_name:
                            manifest.setdefault(original_name, []).append(
                                os.path.basename(result_path),
                            )
                    else:
                        corrupted.append(src)
                        for fs in file_stats:
                            if fs.file_path == src:
                                fs.has_error = True
                                fs.error_message = err_msg or "Unknown error"
                                break
                except Exception as e:
                    err_msg = str(e)
                    logger.error(
                        "Критическая ошибка рабочего процесса для %s: %s",
                        os.path.basename(src), err_msg,
                    )
                    corrupted.append(src)
                    for fs in file_stats:
                        if fs.file_path == src:
                            fs.has_error = True
                            fs.error_message = err_msg
                            break

                _split_completed += 1
                if on_progress is not None:
                    try:
                        on_progress(_split_completed, _total_sheet_tasks, os.path.basename(src))
                    except Exception:
                        pass
    else:
        # Последовательное выполнение (тоже использует предвычисленные пути)
        _split_completed = 0
        for src, out, sheet in sheet_tasks:
            try:
                worker_result = _extract_to_path_worker(src, out, sheet)
                result_path = worker_result.get("path")
                err_msg = worker_result.get("error")

                if result_path:
                    all_created.append(result_path)
                    if worker_result.get("used_fallback"):
                        openpyxl_count += 1
                        source_basename = worker_result.get("source_basename", "")
                        if source_basename:
                            openpyxl_files.append(source_basename)
                    # Обновляем манифест
                    original_name = worker_result.get("source_basename", "")
                    if original_name:
                        manifest.setdefault(original_name, []).append(
                            os.path.basename(result_path),
                        )
                else:
                    logger.warning(
                        "Повреждённый файл при разделении %s: %s",
                        os.path.basename(src), err_msg,
                    )
                    corrupted.append(src)
                    for fs in file_stats:
                        if fs.file_path == src:
                            fs.has_error = True
                            fs.error_message = err_msg or "Unknown error"
                            break
            except Exception as e:
                err_msg = str(e)
                logger.warning(
                    "Повреждённый файл при разделении %s: %s",
                    os.path.basename(src), err_msg,
                )
                corrupted.append(src)
                for fs in file_stats:
                    if fs.file_path == src:
                        fs.has_error = True
                        fs.error_message = err_msg
                        break

            _split_completed += 1
            if on_progress is not None:
                try:
                    on_progress(_split_completed, _total_sheet_tasks, os.path.basename(src))
                except Exception:
                    pass

    # ── ШАГ 4.5: Вертикальное разделение многооперационных файлов ──
    # Для файлов с 1 листом и N таблицами (SWM-стиль),
    # находим границы таблиц и создаём отдельный .xlsx для каждой.
    #
    # Если границы не найдены (false positive detection) — файл
    # возвращается в normal_tasks для обычного горизонтального split.
    vertical_fallback_to_normal: List[str] = []
    already_vertically_split: List[str] = []  # Файлы, созданные вертикальным split — НЕ обрабатывать в ШАГ 4.6
    for result in cards_data.card_results:
        if result.file_path not in vertical_split_files:
            continue

        source_path = result.file_path
        file_name = os.path.basename(source_path)

        # Safety: .xls files cannot be processed by openpyxl splitter
        if os.path.splitext(source_path)[1].lower() != ".xlsx":
            logger.info(
                "Вертикальный split: пропуск .xls файла (не поддерживается): %s",
                file_name,
            )
            continue
        # Find the MAIN data sheet for vertical split
        # Используем _select_best_data_sheet для инспекционных файлов
        # (берёт лист "内容" вместо первого попавшегося)
        main_sheet_v = _select_best_data_sheet(result.sheets)
        if main_sheet_v is None:
            main_sheet_v = _find_main_data_sheet(
                result.sheets, _SPLITTER_SERVICE_SHEET_KEYWORDS,
            )
        if main_sheet_v is None:
            continue
        sheet_name = main_sheet_v.sheet_name

        try:
            # Находим границы всех таблиц на листе
            boundaries = find_table_boundaries(source_path, sheet_name)
            if not boundaries:
                logger.warning(
                    "Вертикальный split: не найдены границы таблиц в %s "
                    "— возврат к горизонтальному split",
                    file_name,
                )
                vertical_fallback_to_normal.append(source_path)
                continue

            logger.info(
                "Вертикальный split %s: найдено %d границ таблиц",
                file_name, len(boundaries),
            )

            # Создаём файлы для каждой операции
            created_vertical = _vertical_split_worker(
                source_path=source_path,
                output_dir=output_dir,
                sheet_name=sheet_name,
                boundaries=boundaries,
                card_label=result.card_number or file_name,
            )

            all_created.extend(created_vertical)
            already_vertically_split.extend(created_vertical)
            for vpath in created_vertical:
                manifest.setdefault(file_name, []).append(
                    os.path.basename(vpath),
                )

            # Обновляем статистику файла
            for fs in file_stats:
                if fs.file_path == source_path:
                    fs.created_files = len(created_vertical)
                    fs.sheets_split = len(boundaries)
                    break

            logger.info(
                "Вертикальный split %s: создано %d файлов",
                file_name, len(created_vertical),
            )
        except Exception as e:
            err_msg = f"Вертикальный split {file_name}: {e}"
            logger.error(err_msg)
            corrupted.append(source_path)
            for fs in file_stats:
                if fs.file_path == source_path:
                    fs.has_error = True
                    fs.error_message = err_msg
                    break

    # ── Возвращаем false-positive вертикальные файлы в normal_tasks ──
    # ВАЖНО: удаляем файлы из vertical_split_files чтобы они не
    # фильтровались повторно при пересчёте normal_tasks.
    # НЕ добавляем новые tasks — оригинальные записи уже в tasks.
    for fb_path in vertical_fallback_to_normal:
        del vertical_split_files[fb_path]
        logger.info(
            "Файл %s возвращён в горизонтальный split",
            os.path.basename(fb_path),
        )

    if vertical_fallback_to_normal:
        # Recompute path_map with updated normal_tasks (fallback files now included)
        tasks.sort(key=lambda t: t[0])
        normal_tasks = [t for t in tasks if t[0] not in vertical_split_files]
        additional_path_map = preallocate_split_paths(normal_tasks, output_dir)
        path_map.update(additional_path_map)

        # Build sheet_tasks for the fallback files only (use original task entries)
        additional_sheet_tasks: List[Tuple[str, str, str]] = []
        for source_path, _out_dir, sheet_names, file_label in normal_tasks:
            if source_path not in vertical_fallback_to_normal:
                continue
            for sheet_name in sheet_names:
                output_path = path_map.get((source_path, sheet_name))
                if output_path:
                    additional_sheet_tasks.append((source_path, output_path, sheet_name))

        # Process additional tasks (sequential, since fallback is rare)
        for src, out, sheet in additional_sheet_tasks:
            try:
                worker_result = _extract_to_path_worker(src, out, sheet)
                result_path = worker_result.get("path")
                err_msg = worker_result.get("error")

                if result_path:
                    all_created.append(result_path)
                    if worker_result.get("used_fallback"):
                        openpyxl_count += 1
                        source_basename = worker_result.get("source_basename", "")
                        if source_basename:
                            openpyxl_files.append(source_basename)
                    original_name = worker_result.get("source_basename", "")
                    if original_name:
                        manifest.setdefault(original_name, []).append(
                            os.path.basename(result_path),
                        )
                else:
                    logger.warning(
                        "Повреждённый файл при разделении %s: %s",
                        os.path.basename(src), err_msg,
                    )
                    corrupted.append(src)
                    for fs in file_stats:
                        if fs.file_path == src:
                            fs.has_error = True
                            fs.error_message = err_msg or "Unknown error"
                            break
            except Exception as e:
                err_msg = str(e)
                logger.warning(
                    "Повреждённый файл при разделении %s: %s",
                    os.path.basename(src), err_msg,
                )
                corrupted.append(src)
                for fs in file_stats:
                    if fs.file_path == src:
                        fs.has_error = True
                        fs.error_message = err_msg
                        break

    # ── ШАГ 4.6: Вертикальный split для файлов, созданных горизонтальным split ──
    # После горизонтального split каждый файл содержит 1 лист.
    # Если на этом листе >1 операции — разделяем вертикально.
    #
    # ВАЖНО: Исключаем файлы, уже созданные вертикальным split в ШАГ 4.5,
    # чтобы не разделять повторно корректно разрезанные карты.
    already_vertically_split_set = set(already_vertically_split)
    post_split_files = [
        f for f in all_created
        if f not in already_vertically_split_set
    ]
    for split_path in post_split_files:
        if os.path.splitext(split_path)[1].lower() != ".xlsx":
            continue
        if not os.path.isfile(split_path):
            continue

        try:
            import openpyxl
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                _wb = openpyxl.load_workbook(split_path, read_only=True, data_only=True)
                _sheet_count = len(_wb.sheetnames)
                # Use sheets[0] as fallback: ZIP-splitter doesn't set activeTab
                _sheet_name = (
                    _wb.active.title if _wb.active else ""
                ) or (_wb.sheetnames[0] if _wb.sheetnames else "")
                _max_row = 0
                if _sheet_name and _sheet_name in _wb.sheetnames:
                    _ws = _wb[_sheet_name]
                    _max_row = _ws.max_row or 0
                _wb.close()

            if _sheet_count != 1 or not _sheet_name:
                continue

            # Skip small files: < 100 rows can't be meaningfully split vertically
            # (Jetour 作业指导书 files have ~79 rows with steps + materials as
            # separate "tables" — these are NOT separate operations)
            if _max_row < 100:
                continue

            boundaries = find_table_boundaries(split_path, _sheet_name)
            if not boundaries or len(boundaries) <= 1:
                continue

            # Фильтруем пустые границы
            valid_boundaries = [b for b in boundaries if b.data_end > b.header_row]
            if len(valid_boundaries) <= 1:
                continue

            # Читаем ZIP один раз для всех операций (оптимизация I/O)
            try:
                with open(split_path, 'rb') as _f:
                    _preloaded_zip = _f.read()
            except OSError:
                _preloaded_zip = None

            file_basename = os.path.basename(split_path)
            logger.info(
                "Вертикальный split (post-horizontal) %s: %d операций на листе '%s'",
                file_basename, len(valid_boundaries), _sheet_name,
            )

            created_vertical = _vertical_split_worker(
                source_path=split_path,
                output_dir=output_dir,
                sheet_name=_sheet_name,
                boundaries=valid_boundaries,
                card_label=file_basename,
                preloaded_zip=_preloaded_zip,
            )

            if created_vertical:
                all_created.extend(created_vertical)
                manifest.setdefault(file_basename, []).extend(
                    os.path.basename(v) for v in created_vertical
                )
                # Удаляем оригинальный файл (заменён вертикальными)
                try:
                    os.remove(split_path)
                    all_created.remove(split_path)
                except OSError:
                    pass

                logger.info(
                    "Вертикальный split (post-horizontal) %s: создано %d файлов",
                    file_basename, len(created_vertical),
                )
        except Exception as e:
            logger.debug("Post-horizontal vertical split skipped for %s: %s",
                         os.path.basename(split_path), e)

    # ── ШАГ 4: Пост-обработка ──

    # Сортируем результаты для детерминированного порядка
    all_created.sort()
    corrupted.sort()

    if corrupted:
        logger.warning("Повреждённых файлов при разделении: %d", len(corrupted))
        # Выводим имена повреждённых файлов в лог
        for cf in corrupted:
            logger.warning("  ⚠️  %s", os.path.basename(cf))
    cards_data.corrupted_files.extend(corrupted)

    # ── ШАГ 4.7: Изоляция повреждённых файлов ──
    # Копируем каждый действительно повреждённый файл в corrupted_cards/ с описанием ошибки
    corrupted_detailed: List[Dict[str, str]] = []
    for cf in corrupted:
        try:
            fname = os.path.basename(cf)
            error_msg = "Файл не удалось разделить: критическая ошибка при извлечении листа"
            # Ищем описание ошибки в file_stats
            for fs in file_stats:
                if fs.file_path == cf and fs.error_message:
                    error_msg = fs.error_message
                    break

            # Копируем в output_dir/corrupted_cards/
            corrupted_dir = os.path.join(output_dir, "corrupted_cards")
            os.makedirs(corrupted_dir, exist_ok=True)

            dest_path = os.path.join(corrupted_dir, fname)
            counter = 1
            while os.path.exists(dest_path):
                name_part, ext = os.path.splitext(fname)
                dest_path = os.path.join(corrupted_dir, f"{name_part}_{counter}{ext}")
                counter += 1

            shutil.copy2(cf, dest_path)

            # Записываем .error файл
            error_path = dest_path + ".error"
            with open(error_path, "w", encoding="utf-8") as f:
                f.write(f"Source: {cf}\n")
                f.write(f"Error: {error_msg}\n")

            corrupted_detailed.append({
                "file_name": fname,
                "folder": corrupted_dir,
                "error": error_msg,
            })
            logger.info("Повреждённый файл изолирован: %s → corrupted_cards/", fname)
        except Exception as e:
            logger.warning("Не удалось изолировать повреждённый файл %s: %s", cf, e)

    cards_data.corrupted_files_detailed = corrupted_detailed

    # Подсчёт created_files на задачу: для каждого task сопоставляем
    # созданные файлы по префиксу из safe_label (совпадает с именованием splitter)
    all_created_basenames: List[str] = [os.path.basename(f) for f in all_created]
    task_file_to_stats: Dict[str, FileSplitStats] = {}
    for fs in file_stats:
        if not fs.split_reason:
            task_file_to_stats[fs.file_path] = fs
    for source_path, _out_dir, _sheets, file_label in tasks:
        if source_path in task_file_to_stats:
            safe_label = _safe_filename(file_label)[:50]
            created_count = sum(1 for bn in all_created_basenames if bn.startswith(safe_label))
            task_file_to_stats[source_path].created_files = created_count

    total_errors = len(corrupted)

    cards_data.split_stats = SplitStatistics(
        file_stats=sorted(file_stats, key=lambda fs: fs.file_path),
        total_xlsx=total_xlsx,
        total_xls=total_xls,
        total_service_files=total_service,
        total_sheets_all=total_sheets_all,
        total_sheets_split=total_sheets_split,
        total_sheets_skipped=total_sheets_skipped,
        total_files_created=len(all_created),
        total_errors=total_errors,
        openpyxl_fallback_count=openpyxl_count,
        openpyxl_fallback_files=sorted(set(openpyxl_files)),
    )

    logger.info("Создано отдельных файлов: %d", len(all_created))
    if openpyxl_count > 0:
        logger.info(
            "Успешно спасены через openpyxl (метод fallback): %d файлов: %s",
            openpyxl_count, sorted(set(openpyxl_files)),
        )

    # Сохраняем манифест генерации файлов
    if manifest:
        import json
        manifest_path = os.path.join(output_dir, "split_manifest.json")
        # Очищаем манифест: удаляем записи для файлов, которые были
        # удалены при вертикальном split (заменены операционными файлами)
        cleaned_manifest: Dict[str, List[str]] = {}
        for key, values in manifest.items():
            surviving = [
                v for v in values
                if os.path.isfile(os.path.join(output_dir, v))
            ]
            if surviving:
                cleaned_manifest[key] = surviving
        # Сортируем для детерминированного вывода
        sorted_manifest = {k: sorted(v) for k, v in sorted(cleaned_manifest.items())}
        try:
            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(sorted_manifest, f, ensure_ascii=False, indent=2)
            logger.info("Журнал генерации файлов сохранен: %s", manifest_path)
        except Exception as e:
            logger.warning("Не удалось сохранить манифест: %s", e)

    return all_created
